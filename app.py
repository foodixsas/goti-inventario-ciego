"""
Backend Flask para Inventario Ciego - Render Deploy
Conecta a Azure PostgreSQL
"""
from flask import Flask, request, jsonify, send_from_directory, send_file, render_template_string
from flask_cors import CORS
import psycopg2
from psycopg2.pool import SimpleConnectionPool
from psycopg2.extras import RealDictCursor, execute_values
import os, re, secrets, smtplib, json
import requests   # lo usan los endpoints de Telegram (probar / traer nombres)
from decimal import Decimal
from datetime import datetime, timedelta, timezone
from io import BytesIO

# Zona horaria Ecuador (UTC-5)
TZ_ECUADOR = timezone(timedelta(hours=-5))
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from flask.json.provider import DefaultJSONProvider

class CustomJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)

app = Flask(__name__, static_folder='static')
app.json_provider_class = CustomJSONProvider
app.json = CustomJSONProvider(app)
CORS(app, origins=['https://inventario-ciego-5bdr.onrender.com'])

@app.after_request
def add_no_cache_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# Configuracion de la base de datos Azure PostgreSQL
DB_CONFIG = {
    'host': os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
    'database': os.environ.get('DB_NAME', 'InventariosLocales'),
    'user': os.environ.get('DB_USER', 'adminChios'),
    'password': os.environ.get('DB_PASSWORD', 'Burger2023'),
    'port': os.environ.get('DB_PORT', '5432'),
    'sslmode': 'require',
    'keepalives': 1,
    'keepalives_idle': 30,
    'keepalives_interval': 10,
    'keepalives_count': 5,
    'connect_timeout': 10
}

_connection_pool = None
_movimientos_pool = None

def _get_pool():
    global _connection_pool
    if _connection_pool is None:
        _connection_pool = SimpleConnectionPool(
            minconn=1, maxconn=5,
            **DB_CONFIG, cursor_factory=RealDictCursor
        )
    return _connection_pool

def _get_movimientos_pool():
    """Pool separado para BD movimientos"""
    global _movimientos_pool
    if _movimientos_pool is None:
        _movimientos_pool = SimpleConnectionPool(
            minconn=1, maxconn=3,
            host=os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
            database='movimientos',
            user=os.environ.get('DB_USER', 'adminChios'),
            password=os.environ.get('DB_PASSWORD', 'Burger2023'),
            port=os.environ.get('DB_PORT', '5432'),
            sslmode='require',
            connect_timeout=10
        )
    return _movimientos_pool

def get_db():
    """Obtiene conexion del pool, validando que este viva"""
    conn = _get_pool().getconn()
    try:
        conn.cursor().execute("SELECT 1")
        conn.rollback()
    except Exception:
        # Conexion stale - cerrar y crear nueva
        try:
            _get_pool().putconn(conn, close=True)
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
        conn = psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)
    return conn

def release_db(conn):
    try:
        if conn.closed:
            return
        _get_pool().putconn(conn)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass


def init_db():
    """Crea tabla merma_operativa y migra asignacion_diferencias al startup"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.merma_operativa (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150) NOT NULL,
                unidad VARCHAR(20) NOT NULL,
                cantidad NUMERIC(12,4) NOT NULL,
                motivo TEXT,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                costo_total NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            ALTER TABLE goti.asignacion_diferencias
                ADD COLUMN IF NOT EXISTS codigo VARCHAR(50),
                ADD COLUMN IF NOT EXISTS nombre VARCHAR(150),
                ADD COLUMN IF NOT EXISTS unidad VARCHAR(20),
                ADD COLUMN IF NOT EXISTS local VARCHAR(50),
                ADD COLUMN IF NOT EXISTS fecha DATE
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.bajas_directas (
                id SERIAL PRIMARY KEY,
                baja_grupo BIGINT,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150) NOT NULL,
                unidad VARCHAR(20) NOT NULL,
                cantidad NUMERIC(12,4) NOT NULL,
                persona VARCHAR(100),
                motivo TEXT,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                costo_total NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS baja_grupo BIGINT
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS documento VARCHAR(100)
        """)
        cur.execute("""
            ALTER TABLE goti.bajas_directas
                ADD COLUMN IF NOT EXISTS codigo_baja VARCHAR(50)
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.bajas_asignaciones (
                id SERIAL PRIMARY KEY,
                baja_grupo BIGINT NOT NULL,
                persona VARCHAR(100) NOT NULL,
                monto NUMERIC(12,2) NOT NULL,
                fecha DATE,
                local VARCHAR(50),
                motivo TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # ---- Tablas para Asignación por Sección (prototipo) ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_seccion (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                nombre VARCHAR(100),
                total_valor NUMERIC(12,2) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asig_seccion_productos (
                id SERIAL PRIMARY KEY,
                seccion_id INT NOT NULL,
                conteo_id INT NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(150),
                diferencia NUMERIC(12,4),
                costo_unitario NUMERIC(12,4),
                cantidad_asignada NUMERIC(12,4),
                valor NUMERIC(12,2)
            )
        """)
        cur.execute("""
            ALTER TABLE goti.asig_seccion_productos
                ADD COLUMN IF NOT EXISTS cantidad_asignada NUMERIC(12,4)
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asig_seccion_personas (
                id SERIAL PRIMARY KEY,
                seccion_id INT NOT NULL,
                persona VARCHAR(100),
                monto NUMERIC(12,2)
            )
        """)
        # ---- Tablas para Asignacion Semanal ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.semanas_inventario (
                id SERIAL PRIMARY KEY,
                fecha_inicio DATE NOT NULL,
                fecha_fin DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                estado VARCHAR(20) DEFAULT 'abierta' CHECK (estado IN ('abierta', 'cerrada')),
                cerrada_por VARCHAR(100),
                cerrada_at TIMESTAMP,
                notas TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(fecha_inicio, local)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_semanal (
                id SERIAL PRIMARY KEY,
                semana_id INT NOT NULL,
                codigo VARCHAR(50) NOT NULL,
                nombre VARCHAR(150),
                unidad VARCHAR(20),
                local VARCHAR(50),
                diferencia_semanal NUMERIC(12,4) DEFAULT 0,
                costo_unitario NUMERIC(12,4) DEFAULT 0,
                grupo_idx INT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.asignacion_semanal_personas (
                id SERIAL PRIMARY KEY,
                asignacion_semanal_id INT NOT NULL,
                persona VARCHAR(100) NOT NULL,
                cantidad NUMERIC(12,4) DEFAULT 0,
                monto NUMERIC(12,2) DEFAULT 0
            )
        """)
        # ---- Columna grupo_idx para preservar estructura de grupos ----
        cur.execute("""
            ALTER TABLE goti.asignacion_semanal
                ADD COLUMN IF NOT EXISTS grupo_idx INT DEFAULT 0
        """)
        # ---- Columnas de auditoria: quien contó y quien modificó ----
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
                ADD COLUMN IF NOT EXISTS contado_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS contado_at TIMESTAMP,
                ADD COLUMN IF NOT EXISTS contado2_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS contado2_at TIMESTAMP,
                ADD COLUMN IF NOT EXISTS modificado_por VARCHAR(50),
                ADD COLUMN IF NOT EXISTS modificado_at TIMESTAMP
        """)
        # ---- Tabla de permisos por ROL (ver + editar) ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.rol_modulos (
                id SERIAL PRIMARY KEY,
                rol VARCHAR(20) NOT NULL,
                modulo VARCHAR(30) NOT NULL,
                puede_ver BOOLEAN DEFAULT TRUE,
                puede_editar BOOLEAN DEFAULT FALSE,
                UNIQUE(rol, modulo)
            )
        """)
        # Migrar: agregar columnas si tabla ya existia sin ellas
        cur.execute("""
            ALTER TABLE goti.rol_modulos
                ADD COLUMN IF NOT EXISTS puede_ver BOOLEAN DEFAULT TRUE,
                ADD COLUMN IF NOT EXISTS puede_editar BOOLEAN DEFAULT FALSE,
                ADD COLUMN IF NOT EXISTS puede_eliminar BOOLEAN DEFAULT FALSE
        """)
        # Seed defaults si la tabla esta vacia
        cur.execute("SELECT COUNT(*) as cnt FROM goti.rol_modulos")
        if cur.fetchone()['cnt'] == 0:
            # Subgerente: conteo, observaciones, historico, dashboard
            for mod in ['conteo','observaciones','historico','dashboard']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('subgerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Supervisor: ve todos los locales, ve todo pero no edita usuarios
            for mod in ['conteo','observaciones','historico','dashboard','cruce','bajas','semanal','correccion']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('supervisor', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Gerente: todo lo del subgerente + semanal, cruce, bajas
            for mod in ['conteo','observaciones','historico','dashboard']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('gerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            for mod in ['cruce','bajas','semanal','correccion']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('gerente', %s, TRUE, TRUE, FALSE) ON CONFLICT DO NOTHING", (mod,))
            # Admin: ve, edita y elimina todo
            for mod in ['conteo','observaciones','historico','dashboard','cruce','bajas','semanal','correccion','usuarios']:
                cur.execute("INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar) VALUES ('admin', %s, TRUE, TRUE, TRUE) ON CONFLICT DO NOTHING", (mod,))
        # Actualizar registros existentes que no tengan puede_ver seteado (migracion)
        cur.execute("UPDATE goti.rol_modulos SET puede_ver = TRUE WHERE puede_ver IS NULL")
        cur.execute("UPDATE goti.rol_modulos SET puede_editar = TRUE WHERE puede_editar IS NULL")

        # Migrar roles: empleado → subgerente, supervisor → gerente
        try:
            cur.execute("SAVEPOINT migrate_roles")
            cur.execute("UPDATE goti.usuarios SET rol = 'subgerente' WHERE rol = 'empleado'")
            cur.execute("UPDATE goti.usuarios SET rol = 'gerente' WHERE rol = 'supervisor'")
            cur.execute("UPDATE goti.rol_modulos SET rol = 'subgerente' WHERE rol = 'empleado'")
            cur.execute("UPDATE goti.rol_modulos SET rol = 'gerente' WHERE rol = 'supervisor'")
            cur.execute("RELEASE SAVEPOINT migrate_roles")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT migrate_roles")

        # Garantizar que gerente siempre tenga acceso a semanal con edicion
        for mod in ['semanal', 'cruce', 'bajas', 'correccion']:
            cur.execute("""
                INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar)
                VALUES ('gerente', %s, TRUE, TRUE, FALSE)
                ON CONFLICT (rol, modulo) DO UPDATE SET puede_ver = TRUE, puede_editar = TRUE
            """, (mod,))

        # ---- Tabla Cuadres de Caja ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.cuadres_caja (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                venta_sistema NUMERIC(12,2) DEFAULT 0,
                efectivo_contado NUMERIC(12,2) DEFAULT 0,
                venta_tarjeta NUMERIC(12,2) DEFAULT 0,
                venta_transferencia NUMERIC(12,2) DEFAULT 0,
                venta_plataformas NUMERIC(12,2) DEFAULT 0,
                otros_ingresos NUMERIC(12,2) DEFAULT 0,
                gastos_retiros NUMERIC(12,2) DEFAULT 0,
                efectivo_esperado NUMERIC(12,2) DEFAULT 0,
                diferencia NUMERIC(12,2) DEFAULT 0,
                observacion TEXT,
                registrado_por VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(fecha, local)
            )
        """)

        # ---- Tabla Delivery Liquidaciones ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.delivery_liquidaciones (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                plataforma VARCHAR(30) NOT NULL,
                total_pedidos INT DEFAULT 0,
                venta_bruta NUMERIC(12,2) DEFAULT 0,
                comision_pct NUMERIC(5,2) DEFAULT 0,
                comision_monto NUMERIC(12,2) DEFAULT 0,
                iva_comision NUMERIC(12,2) DEFAULT 0,
                propinas NUMERIC(12,2) DEFAULT 0,
                ajustes NUMERIC(12,2) DEFAULT 0,
                neto_recibir NUMERIC(12,2) DEFAULT 0,
                depositado_real NUMERIC(12,2) DEFAULT 0,
                diferencia NUMERIC(12,2) DEFAULT 0,
                referencia VARCHAR(100),
                observacion TEXT,
                registrado_por VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # ---- Tabla Registro de Facturas ----
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.facturas_registro (
                id SERIAL PRIMARY KEY,
                fecha_emision DATE NOT NULL,
                local VARCHAR(50) NOT NULL,
                proveedor VARCHAR(200) NOT NULL,
                ruc VARCHAR(20),
                numero_factura VARCHAR(50),
                autorizacion VARCHAR(60),
                subtotal_0 NUMERIC(12,2) DEFAULT 0,
                subtotal_iva NUMERIC(12,2) DEFAULT 0,
                iva NUMERIC(12,2) DEFAULT 0,
                total NUMERIC(12,2) DEFAULT 0,
                categoria VARCHAR(50) DEFAULT 'Otros',
                forma_pago VARCHAR(30) DEFAULT 'Transferencia',
                estado_pago VARCHAR(20) DEFAULT 'Pendiente',
                observacion TEXT,
                registrado_por VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        conn.commit()
        print('init_db: tablas OK')
    except Exception as e:
        print(f'init_db error: {e}')
    finally:
        if conn:
            release_db(conn)


# Helper: mapeo de IDs de bodega a nombres legibles
BODEGAS_NOMBRES = {
    'real_audiencia': 'Real Audiencia',
    'floreana': 'Floreana',
    'portugal': 'Portugal',
    'santo_cachon_real': 'Santo Cachon Real',
    'santo_cachon_portugal': 'Santo Cachon Portugal',
    'simon_bolon': 'Simon Bolon',
    'bodega_principal': 'Bodega Principal',
    'materia_prima': 'Materia Prima',
    'planta': 'Planta de Produccion'
}

# (USUARIO_BODEGA eliminado — permisos de bodega ahora se manejan desde BD tabla usuario_bodegas)

# ==================== RUTAS ESTATICAS ====================

@app.route('/')
def index():
    import json as json_lib, base64
    # Inyectar personas directamente en el HTML como JSON en data attribute (evita problemas de encoding en script)
    try:
        personas = _obtener_personas()
    except Exception:
        personas = _personas_cache['datos'] if _personas_cache['datos'] else []
    html_path = os.path.join(app.static_folder, 'index.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    # Usar base64 para evitar cualquier problema de encoding/caracteres especiales
    personas_json = json_lib.dumps(personas, ensure_ascii=True)
    personas_b64 = base64.b64encode(personas_json.encode('utf-8')).decode('ascii')
    inject = f'<script id="personas-data" type="application/json">{personas_json}</script>\n'
    inject += f'<meta name="personas-b64" content="{personas_b64}">\n'
    html = html.replace('</head>', inject + '</head>')
    return html

@app.route('/costos')
def pagina_costos():
    """El tablero de costos, en su propia pagina."""
    with open(os.path.join(app.static_folder, 'costos.html'), 'r', encoding='utf-8') as f:
        return f.read()


@app.route('/precios')
def pagina_precios():
    """Precios de compra, en su propia pagina.

    Aparte del index a proposito: no depende del menu, ni de app.js, ni de que
    el navegador suelte el cache de un archivo de 500 KB.
    """
    ruta = os.path.join(app.static_folder, 'precios.html')
    with open(ruta, 'r', encoding='utf-8') as f:
        return f.read()


@app.route('/establecer-clave')
def pagina_establecer_clave():
    """Pagina publica donde el usuario establece su contrasena."""
    token = request.args.get('token', '')
    if not token:
        return PAGINA_TOKEN_INVALIDO, 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT username, nombre, invite_token_expires FROM goti.usuarios
                       WHERE invite_token = %s AND activo = TRUE""", (token,))
        user = cur.fetchone()
        if not user:
            return PAGINA_TOKEN_INVALIDO, 404
        if user['invite_token_expires'] and user['invite_token_expires'] < datetime.utcnow():
            return PAGINA_TOKEN_INVALIDO, 410
        html = PAGINA_ESTABLECER_CLAVE.replace('{{ nombre }}', user['nombre']).replace('{{ username }}', user['username']).replace('{{ token }}', token)
        return html
    except Exception as e:
        print(f"Error en /establecer-clave: {e}")
        return PAGINA_TOKEN_INVALIDO, 500
    finally:
        if conn:
            release_db(conn)


@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('static', path)

# ==================== API ====================

_login_attempts = {}

def _check_rate_limit(ip, max_attempts=5, window=60):
    now = _time.time()
    attempts = _login_attempts.get(ip, [])
    attempts = [t for t in attempts if now - t < window]
    _login_attempts[ip] = attempts
    return len(attempts) < max_attempts

def _record_login_attempt(ip):
    now = _time.time()
    if ip not in _login_attempts:
        _login_attempts[ip] = []
    _login_attempts[ip].append(now)

@app.route('/api/login', methods=['POST'])
def login():
    ip = request.remote_addr
    if not _check_rate_limit(ip):
        return jsonify({'success': False, 'error': 'Demasiados intentos. Espera 60 segundos.'}), 429

    data = request.json
    username = data.get('username')
    password = data.get('password')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT username, nombre, rol FROM goti.usuarios
            WHERE username = %s AND password = %s AND activo = TRUE
        """, (username, password))
        user = cur.fetchone()

        if user:
            # Cargar bodegas desde BD
            cur.execute("""
                SELECT ub.bodega FROM goti.usuario_bodegas ub
                JOIN goti.usuarios u ON u.id = ub.usuario_id
                WHERE u.username = %s
                ORDER BY ub.bodega
            """, (user['username'],))
            bodegas_user = [r['bodega'] for r in cur.fetchall()]
            # Compatibilidad: si tiene 1 sola bodega de ventas, enviar como string
            bodegas_ventas = [b for b in bodegas_user if b not in ('bodega_principal', 'materia_prima', 'planta')]
            bodega_asignada = bodegas_ventas[0] if len(bodegas_ventas) == 1 else None
            # Cargar modulos permitidos segun el ROL (con nivel ver/editar)
            cur.execute("""
                SELECT modulo, puede_ver, puede_editar FROM goti.rol_modulos
                WHERE rol = %s ORDER BY modulo
            """, (user['rol'],))
            modulos_user = [r['modulo'] for r in cur.fetchall() if r['puede_ver']]
            permisos_user = {}
            cur.execute("""
                SELECT modulo, puede_ver, puede_editar, COALESCE(puede_eliminar, FALSE) as puede_eliminar
                FROM goti.rol_modulos WHERE rol = %s
            """, (user['rol'],))
            for r in cur.fetchall():
                permisos_user[r['modulo']] = {'ver': r['puede_ver'], 'editar': r['puede_editar'], 'eliminar': r['puede_eliminar']}
            return jsonify({
                'success': True,
                'user': {
                    'username': user['username'],
                    'nombre': user['nombre'],
                    'rol': user['rol'],
                    'bodega': bodega_asignada,
                    'bodegas': bodegas_user,
                    'modulos': modulos_user,
                    'permisos': permisos_user
                }
            })

        _record_login_attempt(ip)
        return jsonify({'success': False, 'error': 'Credenciales invalidas'}), 401
    except Exception as e:
        print(f"Error en /api/login: {e}")
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/categorias', methods=['GET'])
def get_categorias():
    # Categorias estaticas
    categorias = [
        {'id': 1, 'nombre': 'Bebidas'},
        {'id': 2, 'nombre': 'Carnes'},
        {'id': 3, 'nombre': 'Lacteos'},
        {'id': 4, 'nombre': 'Congelados'},
        {'id': 5, 'nombre': 'Otros'}
    ]
    return jsonify(categorias)

@app.route('/api/bodegas', methods=['GET'])
def get_bodegas():
    bodegas = [
        {'id': 'real_audiencia', 'nombre': 'Real Audiencia'},
        {'id': 'floreana', 'nombre': 'Floreana'},
        {'id': 'portugal', 'nombre': 'Portugal'},
        {'id': 'santo_cachon_real', 'nombre': 'Santo Cachon Real'},
        {'id': 'santo_cachon_portugal', 'nombre': 'Santo Cachon Portugal'},
        {'id': 'simon_bolon', 'nombre': 'Simon Bolon'},
        {'id': 'bodega_principal', 'nombre': 'Bodega Principal'},
        {'id': 'materia_prima', 'nombre': 'Materia Prima'},
        {'id': 'planta', 'nombre': 'Planta de Produccion'}
    ]
    return jsonify(bodegas)

@app.route('/api/personas', methods=['GET'])
def api_personas():
    """Personas que han realizado conteos (para filtro de contador en dashboard)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT c.contado_por as username,
                   COALESCE(u.nombre, c.contado_por) as nombre
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u ON u.username = c.contado_por
            WHERE c.contado_por IS NOT NULL AND c.contado_por != ''
            ORDER BY 2
        """)
        rows = cur.fetchall()
        return jsonify([{'username': r['username'], 'nombre': r['nombre']} for r in rows])
    except Exception as e:
        print(f"Error en /api/personas: {e}")
        return jsonify([]), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/inventario/consultar', methods=['GET'])
def consultar_inventario():
    fecha = request.args.get('fecha')
    local = request.args.get('local')

    if not fecha or not local:
        return jsonify({'error': 'Fecha y local son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Asegurar columnas: observaciones, motivo, corregido (auditoría) y justificado (no descontar)
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS observaciones TEXT;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS motivo TEXT;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS corregido BOOLEAN DEFAULT FALSE;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS justificado BOOLEAN DEFAULT FALSE;
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS cantidad_justificada NUMERIC(12,4) DEFAULT 0;
        """)
        conn.commit()

        cur.execute("""
            SELECT c.id, c.codigo, c.nombre, c.unidad, c.cantidad, c.cantidad_contada, c.cantidad_contada_2,
                   c.observaciones,
                   COALESCE(c.motivo, '') as motivo,
                   COALESCE(c.corregido, FALSE) as corregido,
                   COALESCE(c.justificado, FALSE) as justificado,
                   COALESCE(c.cantidad_justificada, 0) as cantidad_justificada,
                   COALESCE(c.costo_unitario, 0) as costo_unitario,
                   c.contado_por,
                   c.contado2_por,
                   u1.nombre as contado_por_nombre,
                   u2.nombre as contado2_por_nombre,
                   c.contado_at,
                   c.contado2_at
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u1 ON u1.username = c.contado_por
            LEFT JOIN goti.usuarios u2 ON u2.username = c.contado2_por
            WHERE c.fecha = %s AND c.local = %s
            ORDER BY c.codigo
        """, (fecha, local))

        productos = cur.fetchall()

        # Incluir personas del cache (nunca bloquea, solo datos en memoria)
        personas = _personas_cache['datos']

        return jsonify({'productos': productos, 'personas': personas})
    except Exception as e:
        print(f"Error en /api/inventario/consultar: {e}")
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/inventario/autofill-conteo2', methods=['POST'])
def autofill_conteo2():
    """Auto-llena conteo 2 con conteo 1 para productos donde conteo1 == sistema"""
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')

    if not fecha or not local:
        return jsonify({'error': 'fecha y local son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.inventario_ciego_conteos
            SET cantidad_contada_2 = cantidad_contada
            WHERE fecha = %s AND local = %s
              AND cantidad_contada IS NOT NULL
              AND cantidad_contada_2 IS NULL
              AND cantidad_contada = cantidad
        """, (fecha, local))
        actualizados = cur.rowcount
        conn.commit()

        return jsonify({'success': True, 'actualizados': actualizados})
    except Exception as e:
        print(f"Error en /api/inventario/autofill-conteo2: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/inventario/guardar-conteo', methods=['POST'])
def guardar_conteo():
    data = request.json
    id_producto = data.get('id')
    cantidad = data.get('cantidad_contada')
    conteo = data.get('conteo', 1)
    usuario = data.get('usuario', '')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if conteo == 2:
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET cantidad_contada_2 = %s, contado2_por = %s, contado2_at = NOW()
                WHERE id = %s
            """, (cantidad, usuario or None, id_producto))
        else:
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET cantidad_contada = %s, contado_por = %s, contado_at = NOW()
                WHERE id = %s
            """, (cantidad, usuario or None, id_producto))

        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/inventario/guardar-conteo: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/inventario/guardar-observacion', methods=['POST'])
def guardar_observacion():
    data = request.json
    id_producto = data.get('id')
    observaciones = data.get('observaciones', None)
    motivo = data.get('motivo', None)
    corregido = data.get('corregido', None)
    justificado = data.get('justificado', None)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Construir SET dinámico según los campos enviados
        sets = []
        params = []
        if observaciones is not None:
            sets.append("observaciones = %s")
            params.append(observaciones)
        if motivo is not None:
            sets.append("motivo = %s")
            params.append(motivo)
        if corregido is not None:
            sets.append("corregido = %s")
            params.append(bool(corregido))
        if justificado is not None:
            sets.append("justificado = %s")
            params.append(bool(justificado))
        cantidad_justificada = data.get('cantidad_justificada', None)
        if cantidad_justificada is not None:
            sets.append("cantidad_justificada = %s")
            params.append(float(cantidad_justificada))
            # Si pone cantidad > 0, marcar justificado=TRUE automaticamente
            if float(cantidad_justificada) > 0:
                sets.append("justificado = TRUE")
            else:
                sets.append("justificado = FALSE")

        if sets:
            params.append(id_producto)
            cur.execute(f"""
                UPDATE goti.inventario_ciego_conteos
                SET {', '.join(sets)}
                WHERE id = %s
            """, params)
        conn.commit()

        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/inventario/guardar-observacion: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

# ==================== REPORTE MOTIVOS ====================

@app.route('/api/reportes/motivos-lista', methods=['GET'])
def reporte_motivos_lista():
    """Devuelve lista de motivos unicos disponibles."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT motivo FROM goti.inventario_ciego_conteos
            WHERE motivo IS NOT NULL AND motivo != ''
            ORDER BY motivo
        """)
        motivos = [r['motivo'] for r in cur.fetchall()]
        return jsonify(motivos)
    except Exception as e:
        return jsonify([])
    finally:
        if conn: release_db(conn)


@app.route('/api/reportes/motivos', methods=['GET'])
def reporte_motivos():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]
    producto = request.args.get('producto', '')
    contador = request.args.get('contador', '').strip()

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Asegurar columna motivo existe en conteos
        cur.execute("""
            ALTER TABLE goti.inventario_ciego_conteos
            ADD COLUMN IF NOT EXISTS motivo TEXT
        """)
        conn.commit()

        # Asegurar tabla manuales existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.observaciones_manuales (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(100) NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(255) NOT NULL,
                diferencia NUMERIC(12,3) DEFAULT 0,
                motivo TEXT,
                observaciones TEXT,
                corregido BOOLEAN DEFAULT FALSE,
                creado_por VARCHAR(100),
                creado_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()

        # Historico de versiones del cruce operativo. Cada vez que se vuelve
        # a cruzar, la version anterior se guarda aqui antes de pisarla.
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.cruce_operativo_versiones (
                id SERIAL PRIMARY KEY,
                ejecucion_id INTEGER NOT NULL,
                version INTEGER NOT NULL,
                bodega VARCHAR(50) NOT NULL,
                fecha_toma DATE NOT NULL,
                fecha_corte_contifico DATE,
                solicitado_por VARCHAR(150),
                solicitado_at TIMESTAMP,
                timestamp_cruce TIMESTAMP,
                total_productos_toma INTEGER,
                total_productos_contifico INTEGER,
                total_cruzados INTEGER,
                -- filas que tenia el detalle completo, del que solo se guardan
                -- las que descuadran
                total_filas_detalle INTEGER,
                total_con_diferencia INTEGER,
                valor_total_dif NUMERIC(18,4),
                archivado_at TIMESTAMP DEFAULT NOW(),
                archivado_por VARCHAR(150)
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_cruce_ver_bod_fecha
            ON goti.cruce_operativo_versiones (bodega, fecha_toma, version)
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.cruce_operativo_versiones_detalle (
                id SERIAL PRIMARY KEY,
                version_id INTEGER NOT NULL
                    REFERENCES goti.cruce_operativo_versiones(id) ON DELETE CASCADE,
                codigo VARCHAR(50),
                cantidad_toma NUMERIC(18,4),
                cantidad_sistema NUMERIC(18,4),
                diferencia NUMERIC(18,4),
                costo_unitario NUMERIC(18,4),
                valor_diferencia NUMERIC(18,4),
                factor NUMERIC(14,4)
            )
        """)
        cur.execute("""
            CREATE INDEX IF NOT EXISTS idx_cruce_verdet_version
            ON goti.cruce_operativo_versiones_detalle (version_id)
        """)
        # CREATE TABLE IF NOT EXISTS no toca una tabla que ya existe, asi que
        # una columna anadida despues no llegaria nunca a las bases donde la
        # tabla ya estaba creada.
        cur.execute("""
            ALTER TABLE goti.cruce_operativo_versiones
            ADD COLUMN IF NOT EXISTS total_filas_detalle INTEGER
        """)
        # Los importes tienen que tener EXACTAMENTE los decimales del original,
        # numeric(18,4). Con menos, la copia redondea: una diferencia de 0.0004
        # se guarda como 0.00 y el recuento de productos descuadrados deja de
        # cuadrar con el cruce del que salio.
        cur.execute("""
            ALTER TABLE goti.cruce_operativo_versiones
            ALTER COLUMN valor_total_dif TYPE NUMERIC(18,4)
        """)
        for _col in ('cantidad_toma', 'cantidad_sistema', 'diferencia',
                     'costo_unitario', 'valor_diferencia'):
            cur.execute(
                'ALTER TABLE goti.cruce_operativo_versiones_detalle '
                'ALTER COLUMN ' + _col + ' TYPE NUMERIC(18,4)')
        cur.execute("""
            ALTER TABLE goti.cruce_operativo_versiones_detalle
            ALTER COLUMN factor TYPE NUMERIC(14,4)
        """)
        conn.commit()

        # Motivos de conteos
        query1 = """
            SELECT motivo, COUNT(*) as cantidad
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
              AND motivo IS NOT NULL AND motivo != ''
        """
        params1 = [fecha_desde, fecha_hasta]
        if producto:
            query1 += " AND codigo = %s"
            params1.append(producto)
        if len(bodegas) == 1:
            query1 += " AND local = %s"
            params1.append(bodegas[0])
        elif len(bodegas) > 1:
            query1 += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params1.extend(bodegas)
        if contador:
            query1 += " AND contado_por = %s"
            params1.append(contador)
        query1 += " GROUP BY motivo"

        cur.execute(query1, params1)
        motivos_conteo = cur.fetchall()

        # Motivos de observaciones manuales
        query2 = """
            SELECT motivo, COUNT(*) as cantidad
            FROM goti.observaciones_manuales
            WHERE fecha >= %s AND fecha <= %s
              AND motivo IS NOT NULL AND motivo != ''
        """
        params2 = [fecha_desde, fecha_hasta]
        if producto:
            query2 += " AND codigo = %s"
            params2.append(producto)
        if len(bodegas) == 1:
            query2 += " AND local = %s"
            params2.append(bodegas[0])
        elif len(bodegas) > 1:
            query2 += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params2.extend(bodegas)
        query2 += " GROUP BY motivo"

        cur.execute(query2, params2)
        motivos_manual = cur.fetchall()

        # Combinar ambos
        totales = {}
        for m in motivos_conteo:
            totales[m['motivo']] = totales.get(m['motivo'], 0) + m['cantidad']
        for m in motivos_manual:
            totales[m['motivo']] = totales.get(m['motivo'], 0) + m['cantidad']

        resultado = [{'motivo': k, 'cantidad': v} for k, v in totales.items()]
        resultado.sort(key=lambda x: x['cantidad'], reverse=True)

        return jsonify(resultado)
    except Exception as e:
        print(f"Error en /api/reportes/motivos: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/reportes/personas-errores', methods=['GET'])
def reporte_personas_errores():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT
                contado_por,
                COUNT(*) as total_conteos,
                SUM(CASE
                    WHEN cantidad_contada IS NOT NULL
                     AND cantidad_contada != cantidad
                     AND COALESCE(justificado, FALSE) = FALSE
                    THEN 1 ELSE 0
                END) as total_errores
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
              AND contado_por IS NOT NULL AND contado_por != ''
              AND cantidad_contada IS NOT NULL
        """
        params = [fecha_desde, fecha_hasta]

        if len(bodegas) == 1:
            query += " AND local = %s"
            params.append(bodegas[0])
        elif len(bodegas) > 1:
            query += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params.extend(bodegas)

        query += """
            GROUP BY contado_por
            HAVING COUNT(*) >= 5
            ORDER BY (SUM(CASE
                WHEN cantidad_contada IS NOT NULL
                 AND cantidad_contada != cantidad
                 AND COALESCE(justificado, FALSE) = FALSE
                THEN 1 ELSE 0
            END) * 100.0 / COUNT(*)) DESC
            LIMIT 10
        """

        cur.execute(query, params)
        rows = cur.fetchall()

        # Obtener nombres reales si existen
        nombres_query = "SELECT username, nombre FROM goti.usuarios WHERE username = ANY(%s)"
        usernames = [r['contado_por'] for r in rows]
        nombres_map = {}
        if usernames:
            cur.execute(nombres_query, (usernames,))
            for u in cur.fetchall():
                nombres_map[u['username']] = u['nombre']

        resultado = []
        for r in rows:
            total = r['total_conteos']
            errores = r['total_errores']
            pct = round(errores * 100.0 / total, 1) if total > 0 else 0
            resultado.append({
                'persona': nombres_map.get(r['contado_por'], r['contado_por']),
                'username': r['contado_por'],
                'total_conteos': total,
                'total_errores': errores,
                'porcentaje_error': pct
            })

        return jsonify(resultado)
    except Exception as e:
        print(f"Error en /api/reportes/personas-errores: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/reportes/diferencias-fecha', methods=['GET'])
def reporte_diferencias_fecha():
    """Productos con diferencia para una fecha y bodega específica"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')
    excluir_justificados = request.args.get('excluir_justificados', '0') == '1'

    if not fecha:
        return jsonify({'error': 'fecha requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT c.nombre, c.unidad,
                   c.cantidad as sistema,
                   COALESCE(c.cantidad_contada_2, c.cantidad_contada) as conteo,
                   COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad as diferencia,
                   COALESCE(c.motivo, '') as motivo,
                   COALESCE(u1.nombre, c.contado_por, '') as responsable
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u1 ON u1.username = c.contado_por
            WHERE c.fecha = %s
              AND COALESCE(c.cantidad_contada_2, c.cantidad_contada) IS NOT NULL
              AND COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad != 0
        """
        params = [fecha]
        if bodega:
            query += " AND c.local = %s"
            params.append(bodega)
        if excluir_justificados:
            query += " AND (c.justificado IS NULL OR c.justificado = FALSE)"
        query += " ORDER BY ABS(COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad) DESC LIMIT 50"

        cur.execute(query, params)
        productos = [{
            'nombre': r['nombre'],
            'unidad': r['unidad'],
            'sistema': float(r['sistema']),
            'conteo': float(r['conteo']),
            'diferencia': float(r['diferencia']),
            'motivo': r['motivo'],
            'responsable': r['responsable']
        } for r in cur.fetchall()]

        return jsonify(productos)
    except Exception as e:
        print(f"Error en /api/reportes/diferencias-fecha: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/reportes/motivos/detalle', methods=['GET'])
def reporte_motivo_detalle():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    motivo = request.args.get('motivo', '')
    bodegas_f = request.args.getlist('bodega')
    bodegas_f = [b for b in bodegas_f if b]
    contador = request.args.get('contador', '').strip()
    excluir_justificados = request.args.get('excluir_justificados', '0') == '1'

    if not fecha_desde or not fecha_hasta or not motivo:
        return jsonify({'error': 'fecha_desde, fecha_hasta y motivo requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Ocurrencias individuales de conteos con ese motivo
        query1 = """
            SELECT c.fecha, c.nombre, c.local,
                   COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad as diferencia,
                   COALESCE(u.nombre, c.contado_por, '') as responsable,
                   COALESCE(c.observaciones, '') as observacion
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u ON u.username = c.contado_por
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.motivo = %s
        """
        params1 = [fecha_desde, fecha_hasta, motivo]
        if len(bodegas_f) == 1:
            query1 += " AND c.local = %s"
            params1.append(bodegas_f[0])
        elif len(bodegas_f) > 1:
            query1 += " AND c.local IN (" + ",".join(["%s"] * len(bodegas_f)) + ")"
            params1.extend(bodegas_f)
        if contador:
            query1 += " AND c.contado_por = %s"
            params1.append(contador)
        if excluir_justificados:
            query1 += " AND (c.justificado IS NULL OR c.justificado = FALSE)"
        query1 += " ORDER BY c.fecha DESC, c.local"
        cur.execute(query1, params1)
        rows_conteo = cur.fetchall()

        # Ocurrencias individuales de observaciones manuales con ese motivo
        query2 = """
            SELECT fecha, nombre, local, diferencia, '' as responsable
            FROM goti.observaciones_manuales
            WHERE fecha >= %s AND fecha <= %s AND motivo = %s
        """
        params2 = [fecha_desde, fecha_hasta, motivo]
        if len(bodegas_f) == 1:
            query2 += " AND local = %s"
            params2.append(bodegas_f[0])
        elif len(bodegas_f) > 1:
            query2 += " AND local IN (" + ",".join(["%s"] * len(bodegas_f)) + ")"
            params2.extend(bodegas_f)
        query2 += " ORDER BY fecha DESC, local"
        cur.execute(query2, params2)
        rows_manual = cur.fetchall()

        resultado = []
        for r in rows_conteo:
            resultado.append({
                'fecha': r['fecha'].strftime('%d/%m/%Y'),
                'nombre': r['nombre'],
                'local': BODEGAS_NOMBRES.get(r['local'], r['local']),
                'diferencia': round(float(r['diferencia'] or 0), 3),
                'responsable': r['responsable'],
                'observacion': r['observacion']
            })
        for r in rows_manual:
            resultado.append({
                'fecha': r['fecha'].strftime('%d/%m/%Y'),
                'nombre': r['nombre'],
                'local': BODEGAS_NOMBRES.get(r['local'], r['local']),
                'diferencia': round(float(r['diferencia'] or 0), 3),
                'responsable': '',
                'observacion': ''
            })
        resultado.sort(key=lambda x: x['fecha'], reverse=True)

        return jsonify(resultado)
    except Exception as e:
        print(f"Error en /api/reportes/motivos/detalle: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

# ==================== OBSERVACIONES MANUALES ====================

@app.route('/api/observaciones-manuales', methods=['GET'])
def listar_obs_manuales():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify({'error': 'fecha y local requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.observaciones_manuales (
                id SERIAL PRIMARY KEY,
                fecha DATE NOT NULL,
                local VARCHAR(100) NOT NULL,
                codigo VARCHAR(50),
                nombre VARCHAR(255) NOT NULL,
                diferencia NUMERIC(12,3) DEFAULT 0,
                motivo TEXT,
                observaciones TEXT,
                corregido BOOLEAN DEFAULT FALSE,
                justificado BOOLEAN DEFAULT FALSE,
                creado_por VARCHAR(100),
                creado_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Asegurar columna justificado (migracion)
        cur.execute("ALTER TABLE goti.observaciones_manuales ADD COLUMN IF NOT EXISTS justificado BOOLEAN DEFAULT FALSE")
        conn.commit()

        cur.execute("""
            SELECT id, codigo, nombre, diferencia, motivo, observaciones, corregido, COALESCE(justificado, FALSE) as justificado, creado_por
            FROM goti.observaciones_manuales
            WHERE fecha = %s AND local = %s
            ORDER BY creado_at
        """, (fecha, local))
        return jsonify(cur.fetchall())
    except Exception as e:
        print(f"Error en /api/observaciones-manuales GET: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales', methods=['POST'])
def crear_obs_manual():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    codigo = data.get('codigo', '')
    nombre = data.get('nombre', '')
    diferencia = data.get('diferencia', 0)
    motivo = data.get('motivo', '')
    observaciones = data.get('observaciones', '')
    creado_por = data.get('creado_por', '')

    if not fecha or not local or not nombre:
        return jsonify({'error': 'fecha, local y nombre son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.observaciones_manuales
            (fecha, local, codigo, nombre, diferencia, motivo, observaciones, creado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (fecha, local, codigo, nombre, float(diferencia), motivo, observaciones, creado_por))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': new_id})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales POST: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales/<int:obs_id>', methods=['PUT'])
def actualizar_obs_manual(obs_id):
    data = request.json
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sets = []
        params = []
        for campo in ['motivo', 'observaciones', 'diferencia']:
            if campo in data:
                sets.append(f"{campo} = %s")
                params.append(data[campo])
        if 'corregido' in data:
            sets.append("corregido = %s")
            params.append(bool(data['corregido']))
        if 'justificado' in data:
            sets.append("justificado = %s")
            params.append(bool(data['justificado']))
        if sets:
            params.append(obs_id)
            cur.execute(f"""
                UPDATE goti.observaciones_manuales
                SET {', '.join(sets)}
                WHERE id = %s
            """, params)
            conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales PUT: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/observaciones-manuales/<int:obs_id>', methods=['DELETE'])
def eliminar_obs_manual(obs_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.observaciones_manuales WHERE id = %s", (obs_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/observaciones-manuales DELETE: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)

@app.route('/api/admin/corregir-conteo', methods=['PUT'])
def corregir_conteo():
    """Permite al admin corregir conteo1 y/o conteo2 de un producto"""
    data = request.json
    id_producto = data.get('id')
    cantidad_contada = data.get('cantidad_contada')
    cantidad_contada_2 = data.get('cantidad_contada_2')
    cantidad_sistema = data.get('cantidad')
    usuario = data.get('usuario', '')

    if id_producto is None:
        return jsonify({'error': 'id es requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.inventario_ciego_conteos
            SET cantidad = COALESCE(%s, cantidad),
                cantidad_contada = %s,
                cantidad_contada_2 = %s,
                modificado_por = %s,
                modificado_at = CURRENT_TIMESTAMP,
                corregido = TRUE
            WHERE id = %s
        """, (cantidad_sistema, cantidad_contada, cantidad_contada_2, usuario or None, id_producto))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error en /api/admin/corregir-conteo: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/inventario/cargar', methods=['POST'])
def cargar_inventario():
    """Endpoint para cargar datos desde el script de Selenium"""
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    productos = data.get('productos', [])

    if not fecha or not local or not productos:
        return jsonify({'error': 'Datos incompletos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        registros = 0
        for prod in productos:
            cur.execute("""
                INSERT INTO goti.inventario_ciego_conteos
                (fecha, local, codigo, nombre, unidad, cantidad)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (fecha, local, codigo)
                DO UPDATE SET cantidad = EXCLUDED.cantidad, nombre = EXCLUDED.nombre
            """, (fecha, local, prod['codigo'], prod['nombre'], prod['unidad'], prod['cantidad']))
            registros += 1

        conn.commit()

        return jsonify({'success': True, 'registros': registros})
    except Exception as e:
        print(f"Error en /api/inventario/cargar: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)

MARCA_BODEGA = {
    'bodega_principal': 'BODEGA_PRINCIPAL',
    'materia_prima': 'MATERIA_PRIMA',
    'planta': 'PLANTA',
}

def _get_lunes(fecha):
    """Retorna el lunes de la semana de la fecha dada"""
    from datetime import date
    if isinstance(fecha, str):
        fecha = date.fromisoformat(fecha)
    return fecha - timedelta(days=fecha.weekday())

def _seleccionar_productos_semana(cur, bodega, fecha_lunes):
    """Selecciona productos para conteo semanal.
    BP: 14 fijos (tipo_conteo='fijo')
    MP/Planta: 10 aleatorios que no se hayan usado recientemente"""
    marca = MARCA_BODEGA.get(bodega)
    if not marca:
        return []

    if bodega == 'bodega_principal':
        # Fijos: siempre los mismos
        cur.execute("""
            SELECT codigo, nombre, unidad FROM goti.productos_por_marca
            WHERE marca = %s AND tipo_conteo = 'fijo' AND activo = TRUE
            ORDER BY nombre
        """, (marca,))
        return cur.fetchall()

    # Si la semana YA tiene rotacion, reutilizarla. Sin esto cada llamada
    # sorteaba 10 productos nuevos y pisaba la seleccion de la semana, incluso
    # cuando el endpoint terminaba devolviendo 409 'ya existe'.
    cur.execute("""
        SELECT codigos FROM goti.rotacion_semanal_bodegas
        WHERE bodega = %s AND semana_inicio = %s
    """, (bodega, fecha_lunes))
    rotacion = cur.fetchone()
    if rotacion and rotacion['codigos']:
        cur.execute("""
            SELECT codigo, nombre, unidad FROM goti.productos_por_marca
            WHERE marca = %s AND codigo = ANY(%s) AND activo = TRUE
            ORDER BY nombre
        """, (marca, list(rotacion['codigos'])))
        return cur.fetchall()

    # MP / Planta: 10 aleatorios sin repetir hasta agotar todos
    cur.execute("""
        SELECT codigo FROM goti.productos_por_marca
        WHERE marca = %s AND activo = TRUE
    """, (marca,))
    todos = [r['codigo'] for r in cur.fetchall()]

    # Obtener codigos usados en rotaciones recientes
    cur.execute("""
        SELECT codigos FROM goti.rotacion_semanal_bodegas
        WHERE bodega = %s ORDER BY semana_inicio DESC
    """, (bodega,))
    usados = set()
    for r in cur.fetchall():
        if r['codigos']:
            usados.update(r['codigos'])
        if len(usados) >= len(todos) - 10:
            break  # Ya agotamos casi todos, reset

    # Si ya se usaron casi todos, resetear
    disponibles = [c for c in todos if c not in usados]
    if len(disponibles) < 10:
        disponibles = todos  # Reset: todos disponibles de nuevo

    # Seleccionar 10 aleatorios
    import random
    seleccion = random.sample(disponibles, min(10, len(disponibles)))

    # Guardar rotacion
    domingo = fecha_lunes + timedelta(days=6)
    cur.execute("""
        INSERT INTO goti.rotacion_semanal_bodegas (bodega, semana_inicio, semana_fin, codigos)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (bodega, semana_inicio) DO UPDATE SET codigos = EXCLUDED.codigos
    """, (bodega, fecha_lunes, domingo, seleccion))

    # Retornar datos completos
    cur.execute("""
        SELECT codigo, nombre, unidad FROM goti.productos_por_marca
        WHERE marca = %s AND codigo = ANY(%s) AND activo = TRUE
        ORDER BY nombre
    """, (marca, seleccion))
    return cur.fetchall()


@app.route('/api/inventario/generar-conteo-operativo', methods=['POST'])
def generar_conteo_operativo():
    """Genera conteo DIARIO para bodegas operativas.
    Los productos se seleccionan UNA vez por semana (lunes) y se usan todos los dias.
    BP: 14 fijos | MP/Planta: 10 aleatorios semanales.
    Crea una tarea POR DIA para que el worker descargue stock de Contifico.
    Body: {bodega, fecha}"""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha = data.get('fecha')

    BODEGAS_VALIDAS = ('bodega_principal', 'materia_prima', 'planta')
    if bodega not in BODEGAS_VALIDAS:
        return jsonify({'error': f'bodega invalida. Validas: {BODEGAS_VALIDAS}'}), 400
    if not fecha:
        return jsonify({'error': 'fecha requerida'}), 400

    from datetime import date
    if isinstance(fecha, str):
        fecha_date = date.fromisoformat(fecha)
    else:
        fecha_date = fecha
    fecha_lunes = _get_lunes(fecha_date)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Seleccionar productos para la semana (solo genera rotacion nueva si es la primera vez)
        productos = _seleccionar_productos_semana(cur, bodega, fecha_lunes)
        if not productos:
            return jsonify({'error': 'No hay productos configurados para esta bodega'}), 400

        n_fijos = len(productos) if bodega == 'bodega_principal' else 0
        n_aleatorios = 0 if bodega == 'bodega_principal' else len(productos)

        # Crear tabla si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.conteo_operativo_tareas (
                id SERIAL PRIMARY KEY,
                bodega VARCHAR(50) NOT NULL,
                fecha DATE NOT NULL,
                estado VARCHAR(20) DEFAULT 'pendiente',
                solicitado_at TIMESTAMP DEFAULT NOW(),
                worker_lock VARCHAR(50),
                timestamp_inicio TIMESTAMP,
                timestamp_fin TIMESTAMP,
                total_productos INT,
                fijos INT,
                aleatorios INT,
                error_msg TEXT,
                codigos_seleccionados TEXT[],
                semana_inicio DATE,
                semana_fin DATE,
                UNIQUE(bodega, fecha)
            )
        """)
        for col in ['codigos_seleccionados TEXT[]', 'semana_inicio DATE', 'semana_fin DATE']:
            try:
                cur.execute(f"ALTER TABLE goti.conteo_operativo_tareas ADD COLUMN IF NOT EXISTS {col}")
            except Exception:
                conn.rollback()
        conn.commit()

        codigos = [p['codigo'] for p in productos]
        domingo = fecha_lunes + timedelta(days=6)

        # Verificar si ya hay tarea para este DIA+bodega
        cur.execute("""
            SELECT id, estado FROM goti.conteo_operativo_tareas
            WHERE bodega = %s AND fecha = %s
        """, (bodega, str(fecha_date)))
        existente = cur.fetchone()
        if existente:
            if existente['estado'] == 'completado':
                return jsonify({'error': f'Ya se genero el conteo para {fecha} en {bodega}', 'ya_existe': True}), 409
            # Resetear si pendiente/error
            cur.execute("""
                UPDATE goti.conteo_operativo_tareas
                SET estado='pendiente', solicitado_at=NOW(), worker_lock=NULL, error_msg=NULL,
                    timestamp_inicio=NULL, timestamp_fin=NULL,
                    total_productos=%s, fijos=%s, aleatorios=%s,
                    codigos_seleccionados=%s, semana_inicio=%s, semana_fin=%s
                WHERE id = %s
            """, (len(codigos), n_fijos, n_aleatorios, codigos, str(fecha_lunes), str(domingo), existente['id']))
            conn.commit()
            return jsonify({
                'id': existente['id'], 'estado': 'pendiente', 'reset': True,
                'fecha': str(fecha_date),
                'productos': len(codigos), 'fijos': n_fijos, 'aleatorios': n_aleatorios,
                'semana': f'{fecha_lunes} - {domingo}',
                'codigos': codigos
            })

        cur.execute("""
            INSERT INTO goti.conteo_operativo_tareas
                (bodega, fecha, total_productos, fijos, aleatorios, codigos_seleccionados, semana_inicio, semana_fin)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
        """, (bodega, str(fecha_date), len(codigos), n_fijos, n_aleatorios, codigos, str(fecha_lunes), str(domingo)))
        new_id = cur.fetchone()['id']
        conn.commit()

        return jsonify({
            'id': new_id, 'estado': 'pendiente',
            'fecha': str(fecha_date),
            'productos': len(codigos), 'fijos': n_fijos, 'aleatorios': n_aleatorios,
            'semana': f'{fecha_lunes} - {domingo}',
            'codigos': codigos,
            'detalle': [{'codigo': p['codigo'], 'nombre': p['nombre'], 'unidad': p['unidad']} for p in productos]
        })
    except Exception as e:
        print(f"Error en generar-conteo-operativo: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/productos-semana', methods=['GET'])
def conteo_op_productos_semana():
    """Retorna los productos seleccionados para conteo en una semana+bodega.
    Usado por el worker para saber que descargar de Contifico."""
    bodega = request.args.get('bodega')
    fecha = request.args.get('fecha')
    if not bodega or not fecha:
        return jsonify({'error': 'bodega y fecha requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Buscar productos: primero por fecha exacta, luego por cualquier dia de la semana
        lunes = _get_lunes(fecha)
        cur.execute("""
            SELECT codigos_seleccionados, semana_inicio, semana_fin
            FROM goti.conteo_operativo_tareas
            WHERE bodega = %s AND (fecha = %s OR semana_inicio = %s)
            ORDER BY fecha DESC LIMIT 1
        """, (bodega, str(fecha), str(lunes)))
        r = cur.fetchone()
        if not r or not r['codigos_seleccionados']:
            return jsonify({'error': 'No hay productos seleccionados para esta semana'}), 404
        marca = MARCA_BODEGA.get(bodega, '')
        cur.execute("""
            SELECT codigo, nombre, unidad, equivalencia FROM goti.productos_por_marca
            WHERE marca = %s AND codigo = ANY(%s) AND activo = TRUE
            ORDER BY nombre
        """, (marca, r['codigos_seleccionados']))
        productos = cur.fetchall()
        return jsonify({
            'bodega': bodega, 'semana_inicio': str(r['semana_inicio']), 'semana_fin': str(r['semana_fin']),
            'productos': [dict(p) for p in productos]
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# Quien puede coger tareas de las colas. Por defecto solo Render: la PC de
# Finanzas arranca su worker sola al iniciar sesion y, mientras pudo, se llevo
# el 100% de las cargas de ajuste y las hizo con el codigo viejo -de uno en uno,
# mas de una hora por bodega-. Se deja configurable por si hiciera falta
# reactivarla como respaldo.
WORKERS_PERMITIDOS = {w.strip().lower()
                      for w in os.environ.get('WORKERS_PERMITIDOS', 'render').split(',')
                      if w.strip()}


def worker_autorizado(worker_id):
    """True si ese worker puede llevarse tareas."""
    return (worker_id or '').strip().lower() in WORKERS_PERMITIDOS


@app.route('/api/conteo-op/pendientes', methods=['GET'])
def conteo_op_pendientes():
    """Worker toma tareas de conteo operativo."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    worker_id = request.args.get('worker_id', 'pc-finanzas')
    if not worker_autorizado(worker_id):
        # Silencio en vez de error: el worker viejo reintentaria en bucle y
        # llenaria el log. Con la lista vacia simplemente no hace nada.
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.conteo_operativo_tareas
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.conteo_operativo_tareas
                WHERE estado = 'pendiente' ORDER BY solicitado_at ASC LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'], 'bodega': r['bodega'],
            'fecha': r['fecha'].isoformat() if r['fecha'] else None,
            'tipo': 'conteo_operativo',
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/resultado', methods=['POST'])
def conteo_op_resultado():
    """Worker reporta resultado del conteo operativo."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.conteo_operativo_tareas
            SET estado = %s, timestamp_fin = NOW(),
                total_productos = %s, fijos = %s, aleatorios = %s, error_msg = %s
            WHERE id = %s
        """, (estado, data.get('total_productos'), data.get('fijos'), data.get('aleatorios'),
              data.get('error_msg'), ejec_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo-op/estado/<int:ejec_id>', methods=['GET'])
def conteo_op_estado(ejec_id):
    """Polling del panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM goti.conteo_operativo_tareas WHERE id = %s", (ejec_id,))
        r = cur.fetchone()
        if not r: return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'], 'bodega': r['bodega'], 'estado': r['estado'],
            'total_productos': r['total_productos'], 'fijos': r['fijos'],
            'aleatorios': r['aleatorios'], 'error_msg': r['error_msg'],
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/historico', methods=['GET'])
def historico():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT
                fecha,
                local,
                COUNT(*) as total_productos,
                COUNT(cantidad_contada) as total_contados,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia,
                COUNT(CASE WHEN cantidad_contada IS NOT NULL THEN 1 END) as total_con_conteo1,
                COUNT(CASE WHEN cantidad_contada_2 IS NOT NULL THEN 1 END) as total_con_conteo2
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """
        params = [fecha_desde, fecha_hasta]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " GROUP BY fecha, local ORDER BY fecha DESC, local"

        cur.execute(query, params)
        resultados = cur.fetchall()

        # Calcular estado para cada registro
        datos = []
        for r in resultados:
            total = r['total_productos']
            contados = r['total_contados']
            con_conteo2 = r['total_con_conteo2']

            if con_conteo2 > 0 or (contados == total and r['total_con_diferencia'] == 0):
                estado = 'completo'
            elif contados > 0:
                estado = 'en_proceso'
            else:
                estado = 'pendiente'

            porcentaje = round((contados / total * 100) if total > 0 else 0)

            datos.append({
                'fecha': str(r['fecha']),
                'local': r['local'],
                'total_productos': total,
                'total_contados': contados,
                'total_con_diferencia': r['total_con_diferencia'],
                'estado': estado,
                'porcentaje': porcentaje
            })

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/historico: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/historico/pivot', methods=['GET'])
def historico_pivot():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('bodega')
    if not fecha_desde or not fecha_hasta or not local:
        return jsonify({'error': 'fecha_desde, fecha_hasta y bodega son requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT
                c.id, c.codigo, c.nombre, c.unidad,
                c.fecha,
                c.cantidad AS stock,
                COALESCE(c.cantidad_contada_2, c.cantidad_contada) AS contado,
                COALESCE(c.cantidad_contada_2, c.cantidad_contada) - c.cantidad AS diferencia,
                c.costo_unitario
            FROM goti.inventario_ciego_conteos c
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
            ORDER BY c.codigo, c.fecha
        """, (fecha_desde, fecha_hasta, local))
        rows = cur.fetchall()

        # Obtener personas asignadas con cantidades y costos para el periodo/bodega
        cur.execute("""
            SELECT c.codigo, a.persona,
                   SUM(ABS(a.cantidad)) AS cantidad_neta,
                   SUM(a.cantidad)      AS cantidad_ajustada,
                   MAX(c.costo_unitario) AS costo_unitario
            FROM goti.asignacion_diferencias a
            JOIN goti.inventario_ciego_conteos c ON a.conteo_id = c.id
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
              AND a.persona IS NOT NULL AND a.persona <> ''
            GROUP BY c.codigo, a.persona
        """, (fecha_desde, fecha_hasta, local))
        asig_rows = cur.fetchall()

        # Obtener contadores (quién contó) por fecha
        cur.execute("""
            SELECT c.fecha,
                   u.nombre as contador_nombre,
                   MIN(c.contado_at) as hora_inicio,
                   MAX(c.contado_at) as hora_fin,
                   'conteo1' as tipo
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u ON u.username = c.contado_por
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
              AND c.contado_por IS NOT NULL
            GROUP BY c.fecha, u.nombre

            UNION ALL

            SELECT c.fecha,
                   u.nombre as contador_nombre,
                   MIN(c.contado2_at) as hora_inicio,
                   MAX(c.contado2_at) as hora_fin,
                   'conteo2' as tipo
            FROM goti.inventario_ciego_conteos c
            LEFT JOIN goti.usuarios u ON u.username = c.contado2_por
            WHERE c.fecha >= %s AND c.fecha <= %s AND c.local = %s
              AND c.contado2_por IS NOT NULL
            GROUP BY c.fecha, u.nombre

            ORDER BY fecha, tipo
        """, (fecha_desde, fecha_hasta, local, fecha_desde, fecha_hasta, local))
        cont_rows = cur.fetchall()

        release_db(conn)
        conn = None

        # Mapa codigo -> {persona: {cant_neta, desc_neto, cant_ajustada, desc_ajustado}}
        personas_por_codigo = {}
        for ar in asig_rows:
            cod = ar['codigo']
            if cod not in personas_por_codigo:
                personas_por_codigo[cod] = {}
            costo = float(ar['costo_unitario'] or 0)
            cant_neta = float(ar['cantidad_neta'] or 0)          # SUM(ABS) siempre positivo
            cant_ajust = float(ar['cantidad_ajustada'] or 0)     # SUM real, puede ser +/-
            personas_por_codigo[cod][ar['persona']] = {
                'cant_neta':       cant_neta,
                'desc_neto':       round(cant_neta * costo, 4),          # Valor Neto
                'cant_ajustada':   abs(cant_ajust),                       # ABS del neto
                'desc_ajustado':   round(abs(cant_ajust) * costo, 4)     # Valor Ajustado
            }

        productos = {}
        fechas = set()
        for r in rows:
            codigo = r['codigo']
            fecha = str(r['fecha'])
            fechas.add(fecha)
            if codigo not in productos:
                personas_cod = personas_por_codigo.get(codigo, {})
                productos[codigo] = {
                    'codigo': codigo,
                    'nombre': r['nombre'],
                    'unidad': r['unidad'],
                    'porFecha': {},
                    'personas': sorted(personas_cod.keys()),
                    'descuentosPorPersona': personas_cod
                }
            productos[codigo]['porFecha'][fecha] = {
                'stock': float(r['stock'] or 0),
                'contado': float(r['contado']) if r['contado'] is not None else None,
                'diferencia': float(r['diferencia']) if r['diferencia'] is not None else None,
                'costo_unitario': float(r['costo_unitario'] or 0)
            }

        # Lista de todas las personas únicas del periodo
        todas_personas = sorted({p for ps in personas_por_codigo.values() for p in ps.keys()})

        contadores_por_fecha = {}
        for cr in cont_rows:
            f = str(cr['fecha'])
            if f not in contadores_por_fecha:
                contadores_por_fecha[f] = []
            hi = cr['hora_inicio'].strftime('%H:%M') if cr['hora_inicio'] else ''
            hf = cr['hora_fin'].strftime('%H:%M') if cr['hora_fin'] else ''
            contadores_por_fecha[f].append({
                'nombre': cr['contador_nombre'] or '',
                'hora_inicio': hi,
                'hora_fin': hf,
                'tipo': 'Conteo 1' if cr['tipo'] == 'conteo1' else 'Conteo 2'
            })

        return jsonify({
            'fechas': sorted(fechas),
            'productos': list(productos.values()),
            'personas': todas_personas,
            'contadores': contadores_por_fecha
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/diferencias', methods=['GET'])
def reporte_diferencias():
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega')

    if not fecha:
        return jsonify({'error': 'fecha es requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT codigo, nombre, unidad, cantidad as sistema,
                   cantidad_contada as conteo1,
                   cantidad_contada_2 as conteo2,
                   COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as diferencia,
                   COALESCE(motivo, '') as motivo,
                   observaciones,
                   COALESCE(corregido, FALSE) as corregido,
                   local
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """
        params = [fecha]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " ORDER BY ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) DESC"

        cur.execute(query, params)
        productos = cur.fetchall()

        # Convertir Decimal a float
        datos = []
        for p in productos:
            item = {
                'codigo': p['codigo'],
                'nombre': p['nombre'],
                'unidad': p['unidad'],
                'sistema': float(p['sistema']) if p['sistema'] is not None else 0,
                'conteo1': float(p['conteo1']) if p['conteo1'] is not None else None,
                'conteo2': float(p['conteo2']) if p['conteo2'] is not None else None,
                'diferencia': float(p['diferencia']) if p['diferencia'] is not None else 0,
                'motivo': p['motivo'] or '',
                'observaciones': p['observaciones'] or '',
                'corregido': bool(p['corregido'])
            }
            if not bodega:
                item['local'] = p['local']
                item['local_nombre'] = BODEGAS_NOMBRES.get(p['local'], p['local'])
            datos.append(item)

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/reportes/diferencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/exportar-excel', methods=['GET'])
def exportar_excel():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT fecha, local, codigo, nombre, unidad,
                   cantidad as sistema,
                   cantidad_contada as conteo1,
                   cantidad_contada_2 as conteo2,
                   COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as diferencia,
                   COALESCE(motivo, '') as motivo,
                   observaciones,
                   COALESCE(corregido, FALSE) as corregido
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """
        params = [fecha_desde, fecha_hasta]

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += " ORDER BY fecha, local, codigo"

        cur.execute(query, params)
        registros = cur.fetchall()

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Estilos
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        dif_neg_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        dif_neg_font = Font(name='Calibri', bold=True, color='B91C1C')
        dif_pos_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        dif_pos_font = Font(name='Calibri', bold=True, color='059669')

        # Agrupar por fecha+local
        grupos = {}
        for r in registros:
            key = (str(r['fecha']), r['local'])
            if key not in grupos:
                grupos[key] = []
            grupos[key].append(r)

        headers = ['Codigo', 'Producto', 'Unidad', 'Sistema', 'Conteo 1', 'Conteo 2', 'Diferencia', 'Motivo', 'Observaciones', 'Corregido']

        for (fecha, local), items in grupos.items():
            sheet_name = f"{fecha}_{local}"[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Datos
            for row_idx, item in enumerate(items, 2):
                vals = [
                    item['codigo'],
                    item['nombre'],
                    item['unidad'],
                    float(item['sistema']) if item['sistema'] is not None else 0,
                    float(item['conteo1']) if item['conteo1'] is not None else '',
                    float(item['conteo2']) if item['conteo2'] is not None else '',
                    float(item['diferencia']) if item['diferencia'] is not None else '',
                    item.get('motivo') or '',
                    item['observaciones'] or '',
                    'Sí' if item.get('corregido') else 'No'
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                    # Colorear diferencias
                    if col_idx == 7 and val != '' and val != 0:
                        if val < 0:
                            cell.fill = dif_neg_fill
                            cell.font = dif_neg_font
                        else:
                            cell.fill = dif_pos_fill
                            cell.font = dif_pos_font

            # Auto-width
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

        if not wb.sheetnames:
            ws = wb.create_sheet(title='Sin datos')
            ws.cell(row=1, column=1, value='No se encontraron registros para el rango seleccionado')

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_{fecha_desde}_a_{fecha_hasta}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error en /api/reportes/exportar-excel: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/tendencias', methods=['GET'])
def reporte_tendencias():
    bodega = request.args.get('bodega')
    limite = request.args.get('limite', 20, type=int)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT
                codigo,
                nombre,
                COUNT(*) as frecuencia,
                ROUND(AVG(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad))::numeric, 3) as promedio_desviacion,
                ROUND(SUM(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)::numeric, 3) as diferencia_acumulada
            FROM goti.inventario_ciego_conteos
            WHERE COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """
        params = []

        if bodega:
            query += " AND local = %s"
            params.append(bodega)

        query += """
            GROUP BY codigo, nombre
            ORDER BY frecuencia DESC, promedio_desviacion DESC
            LIMIT %s
        """
        params.append(limite)

        cur.execute(query, params)
        productos = cur.fetchall()

        datos = []
        for i, p in enumerate(productos, 1):
            datos.append({
                'ranking': i,
                'codigo': p['codigo'],
                'nombre': p['nombre'],
                'frecuencia': p['frecuencia'],
                'promedio_desviacion': float(p['promedio_desviacion']) if p['promedio_desviacion'] else 0,
                'diferencia_acumulada': float(p['diferencia_acumulada']) if p['diferencia_acumulada'] else 0
            })

        return jsonify(datos)
    except Exception as e:
        print(f"Error en /api/reportes/tendencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/productos-disponibles', methods=['GET'])
def productos_disponibles():
    """Devuelve productos distintos para un rango de fechas y bodega."""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')
    if not fecha_desde or not fecha_hasta:
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        query = "SELECT DISTINCT codigo, nombre FROM goti.inventario_ciego_conteos WHERE fecha >= %s AND fecha <= %s"
        params = [fecha_desde, fecha_hasta]
        if bodega:
            query += " AND local = %s"
            params.append(bodega)
        query += " ORDER BY nombre"
        cur.execute(query, params)
        return jsonify([{'codigo': r['codigo'], 'nombre': r['nombre']} for r in cur.fetchall()])
    except Exception as e:
        return jsonify([])
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/dashboard', methods=['GET'])
def reporte_dashboard():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]  # filtrar vacíos
    producto = request.args.get('producto', '').strip()
    contador = request.args.get('contador', '').strip()
    excluir_justificados = request.args.get('excluir_justificados', '0') == '1'

    if not fecha_desde or not fecha_hasta:
        return jsonify({'error': 'fecha_desde y fecha_hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Filtros comunes
        filtro_extra = ""
        params = [fecha_desde, fecha_hasta]
        if len(bodegas) == 1:
            filtro_extra += " AND local = %s"
            params.append(bodegas[0])
        elif len(bodegas) > 1:
            filtro_extra += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params.extend(bodegas)
        if producto:
            filtro_extra += " AND codigo = %s"
            params.append(producto)
        if contador:
            filtro_extra += " AND (contado_por = %s OR contado2_por = %s)"
            params.extend([contador, contador])
        if excluir_justificados:
            filtro_extra += " AND (justificado IS NULL OR justificado = FALSE)"

        # Resumen por bodega
        query = """
            SELECT
                local,
                COUNT(*) as total_productos,
                COUNT(cantidad_contada) as total_contados,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia,
                COALESCE(ROUND(AVG(ABS(
                    CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                         AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN COALESCE(cantidad_contada_2, cantidad_contada) - cantidad END
                ))::numeric, 3), 0) as promedio_diferencia_abs,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad < 0
                    THEN 1 END) as total_faltantes,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad > 0
                    THEN 1 END) as total_sobrantes,
                COALESCE(SUM(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad < 0
                    THEN ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0) END), 0) as valor_faltantes,
                COALESCE(SUM(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad > 0
                    THEN ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0) END), 0) as valor_sobrantes
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
        """ + filtro_extra + " GROUP BY local ORDER BY local"
        cur.execute(query, params)

        resultados = cur.fetchall()

        bodegas_data = []
        for r in resultados:
            bodegas_data.append({
                'local': r['local'],
                'local_nombre': BODEGAS_NOMBRES.get(r['local'], r['local']),
                'total_productos': r['total_productos'],
                'total_contados': r['total_contados'],
                'total_con_diferencia': r['total_con_diferencia'],
                'promedio_diferencia_abs': float(r['promedio_diferencia_abs']),
                'total_faltantes': r['total_faltantes'],
                'total_sobrantes': r['total_sobrantes'],
                'valor_faltantes': float(r['valor_faltantes']),
                'valor_sobrantes': float(r['valor_sobrantes'])
            })

        # Top 10 productos con mayor descuadre en valor (agrupados por producto)
        query_top = """
            SELECT codigo, nombre, unidad,
                   SUM(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)) as diferencia_total,
                   AVG(COALESCE(costo_unitario, 0)) as costo_unitario,
                   SUM(ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad) * COALESCE(costo_unitario, 0)) as valor_descuadre
            FROM goti.inventario_ciego_conteos
            WHERE fecha >= %s AND fecha <= %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
        """ + filtro_extra
        params_top = list(params)
        query_top += " GROUP BY codigo, nombre, unidad ORDER BY valor_descuadre DESC LIMIT 10"
        cur.execute(query_top, params_top)
        top_descuadre = []
        for r in cur.fetchall():
            top_descuadre.append({
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'unidad': r['unidad'],
                'diferencia': float(r['diferencia_total']),
                'costo_unitario': float(r['costo_unitario']),
                'valor_descuadre': float(r['valor_descuadre'])
            })

        # % cumplimiento por bodega (contados / total)
        cumplimiento = []
        for b in bodegas_data:
            pct = round(b['total_contados'] / b['total_productos'] * 100, 1) if b['total_productos'] > 0 else 0
            cumplimiento.append({
                'local': b['local'],
                'local_nombre': b['local_nombre'],
                'porcentaje': pct,
                'exactos': b['total_contados'] - b['total_con_diferencia'],
                'con_diferencia': b['total_con_diferencia']
            })

        # Promedio diario de exactitud (items contados sin error / items contados)
        query_prom = """
            SELECT AVG(exactitud_dia) as promedio_exactitud,
                   AVG(cumplimiento_dia) as promedio_cumplimiento,
                   COUNT(*) as total_dias
            FROM (
                SELECT fecha,
                       CASE WHEN COUNT(cantidad_contada) > 0
                            THEN (COUNT(cantidad_contada) - COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                                AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0 THEN 1 END))::float
                                / COUNT(cantidad_contada) * 100
                            ELSE 0 END as exactitud_dia,
                       CASE WHEN COUNT(*) > 0
                            THEN COUNT(cantidad_contada)::float / COUNT(*) * 100
                            ELSE 0 END as cumplimiento_dia
                FROM goti.inventario_ciego_conteos
                WHERE fecha >= %s AND fecha <= %s
        """ + filtro_extra + """
                GROUP BY fecha
            ) dias
        """
        cur.execute(query_prom, params)
        prom = cur.fetchone()
        promedios = {
            'exactitud_promedio': round(float(prom['promedio_exactitud'] or 0), 1),
            'cumplimiento_promedio': round(float(prom['promedio_cumplimiento'] or 0), 1),
            'total_dias': prom['total_dias'] or 0
        }

        # Actividad de contadores en el periodo
        query_cont = """
            SELECT
                u.nombre as contador,
                c.contado_por as username,
                COUNT(DISTINCT c.fecha) as dias_contados,
                COUNT(*) as total_items,
                COUNT(DISTINCT c.local) as bodegas_cubiertas,
                MAX(c.contado_at) as ultima_actividad
            FROM goti.inventario_ciego_conteos c
            JOIN goti.usuarios u ON u.username = c.contado_por
            WHERE c.fecha >= %s AND c.fecha <= %s
              AND c.contado_por IS NOT NULL
        """ + filtro_extra + """
            GROUP BY u.nombre, c.contado_por
            ORDER BY total_items DESC
        """
        cur.execute(query_cont, params)
        contadores_data = []
        for r in cur.fetchall():
            ua = r['ultima_actividad']
            contadores_data.append({
                'nombre': r['contador'],
                'username': r['username'],
                'dias_contados': r['dias_contados'],
                'total_items': r['total_items'],
                'bodegas_cubiertas': r['bodegas_cubiertas'],
                'ultima_actividad': ua.strftime('%d/%m %H:%M') if ua else ''
            })

        return jsonify({
            'bodegas': bodegas_data,
            'top_descuadre': top_descuadre,
            'cumplimiento': cumplimiento,
            'promedios': promedios,
            'contadores': contadores_data
        })
    except Exception as e:
        print(f"Error en /api/reportes/dashboard: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/reportes/tendencias-temporal', methods=['GET'])
def reporte_tendencias_temporal():
    bodegas = request.args.getlist('bodega')
    bodegas = [b for b in bodegas if b]
    dias = request.args.get('dias', 30, type=int)
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    motivo = request.args.get('motivo', '')
    producto = request.args.get('producto', '')
    contador = request.args.get('contador', '').strip()
    excluir_justificados = request.args.get('excluir_justificados', '0') == '1'

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if fecha_desde and fecha_hasta:
            where_fecha = "fecha >= %s AND fecha <= %s"
            params = [fecha_desde, fecha_hasta]
        else:
            where_fecha = "fecha >= CURRENT_DATE - %s"
            params = [dias]

        motivo_filter = ""
        if motivo:
            motivo_filter = " AND motivo = %s"
            params.append(motivo)

        if producto:
            motivo_filter += " AND codigo = %s"
            params.append(producto)

        if contador:
            motivo_filter += " AND (contado_por = %s OR contado2_por = %s)"
            params.extend([contador, contador])

        if excluir_justificados:
            motivo_filter += " AND (justificado IS NULL OR justificado = FALSE)"

        query = f"""
            SELECT
                fecha,
                local,
                COUNT(CASE WHEN COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
                    AND COALESCE(cantidad_contada_2, cantidad_contada) - cantidad != 0
                    THEN 1 END) as total_con_diferencia
            FROM goti.inventario_ciego_conteos
            WHERE {where_fecha}{motivo_filter}
        """

        if len(bodegas) == 1:
            query += " AND local = %s"
            params.append(bodegas[0])
        elif len(bodegas) > 1:
            query += " AND local IN (" + ",".join(["%s"] * len(bodegas)) + ")"
            params.extend(bodegas)

        query += " GROUP BY fecha, local ORDER BY fecha, local"

        cur.execute(query, params)
        resultados = cur.fetchall()

        # Agrupar por fecha y series por bodega
        fechas_set = set()
        series_dict = {}
        for r in resultados:
            fecha_str = str(r['fecha'])
            local = r['local']
            fechas_set.add(fecha_str)
            if local not in series_dict:
                series_dict[local] = {}
            series_dict[local][fecha_str] = r['total_con_diferencia']

        fechas = sorted(fechas_set)
        series = {}
        for local, valores in series_dict.items():
            series[local] = {
                'nombre': BODEGAS_NOMBRES.get(local, local),
                'datos': [valores.get(f, 0) for f in fechas]
            }

        return jsonify({
            'fechas': fechas,
            'series': series
        })
    except Exception as e:
        print(f"Error en /api/reportes/tendencias-temporal: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Asignacion de Diferencias
# ============================================================

import base64 as _b64
_AIRTABLE_FB = _b64.b64decode('cGF0TVYzOFJhOTBhQXprRlAuZWRhNTE1Y2E4MjkzYjI1ODJjYTdmODVmYzNlMGE4NTllNzRjMjhhNWZkOTY0YjA4Zjg2NTJiMjk3MzRjNTg0Nw==').decode()
def _get_airtable_token():
    return os.environ.get('AIRTABLE_TOKEN', '') or _AIRTABLE_FB
AIRTABLE_BASE = os.environ.get('AIRTABLE_BASE', 'appzTllAjxu4TOs1a')
AIRTABLE_TABLE = os.environ.get('AIRTABLE_TABLE', 'tbldYTLfQ3DoEK0WA')

# Catálogo de productos desde Airtable (base app5zYXr1GmF2bmVF)
CATALOGO_BASE = 'app5zYXr1GmF2bmVF'
CATALOGO_TABLE = 'tbl8hyvwwfSnrspAt'
CATALOGO_VIEW = 'viwxcPxcde6c3JhbE'  # "Matriz Sis Inventarios (No tocar)"
_catalogo_cache = {'datos': [], 'ts': 0}

def _cargar_catalogo_airtable():
    import time, urllib.request, json as json_lib
    token = _get_airtable_token()
    all_records = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{CATALOGO_BASE}/{CATALOGO_TABLE}?view={CATALOGO_VIEW}&pageSize=100'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {token}'})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json_lib.loads(r.read())
        for rec in data['records']:
            f = rec['fields']
            codigo = f.get('Código', '').strip()
            nombre = f.get('Nombre Producto', f.get('Nombre Copia', '')).strip()
            unidad = f.get('Unidad Contifico', '').strip()
            if codigo and nombre:
                all_records.append({'codigo': codigo, 'nombre': nombre, 'unidad': unidad})
        offset = data.get('offset')
        if not offset:
            break
    _catalogo_cache['datos'] = all_records
    _catalogo_cache['ts'] = time.time()
    return all_records

@app.route('/api/catalogo-productos', methods=['GET'])
def get_catalogo_productos():
    import time, urllib.request, json as json_lib
    # Cache de 1 hora
    if time.time() - _catalogo_cache['ts'] < 3600 and _catalogo_cache['datos']:
        return jsonify(_catalogo_cache['datos'])
    try:
        datos = _cargar_catalogo_airtable()
        return jsonify(datos)
    except Exception as e:
        # Si falla pero hay cache viejo, devolver igual
        if _catalogo_cache['datos']:
            return jsonify(_catalogo_cache['datos'])
        return jsonify({'error': str(e)}), 500

# Cache de personas en memoria del servidor
import time as _time
_personas_cache = {'datos': [], 'timestamp': 0}
PERSONAS_CACHE_TTL = 300  # 5 minutos

# Mapeo de bodega a centros de costo de Airtable
BODEGA_CENTROS = {
    'real_audiencia': ['Chios Real Audiencia'],
    'floreana': ['Chios Floreana'],
    'portugal': ['Chios Portugal'],
    'santo_cachon_real': ['Santo Cachon Real Audiencia', 'Santo Cach\u00f3n Real Audiencia'],
    'santo_cachon_portugal': ['Santo Cachon Portugal', 'Santo Cach\u00f3n Portugal'],
    'simon_bolon': ['Simon Bolon Real Audiencia', 'Sim\u00f3n Bol\u00f3n Real Audiencia'],
}

# ============================================================
# MODULO: Cruce Operativo (bodegas operativas)
# ============================================================

BODEGAS_OPERATIVAS = {
    'bodega_principal': 'Bodega Principal',
    'materia_prima': 'Materia Prima',
    'planta': 'Planta de Produccion',
    'real_audiencia': 'Real Audiencia (Chios)',
    'floreana': 'Floreana (Chios)',
    'portugal': 'Portugal (Chios)',
    'santo_cachon_real': 'Santo Cachon Real',
    'santo_cachon_portugal': 'Santo Cachon Portugal',
    'simon_bolon': 'Simon Bolon',
}

@app.route('/api/cruce/ejecuciones', methods=['GET'])
def cruce_ejecuciones():
    """Lista ejecuciones del cruce operativo con filtros"""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    bodega = request.args.get('bodega')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = """SELECT * FROM goti.cruce_operativo_ejecuciones WHERE 1=1"""
        params = []
        if fecha_desde:
            sql += " AND fecha_toma >= %s"
            params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND fecha_toma <= %s"
            params.append(fecha_hasta)
        if bodega:
            sql += " AND bodega = %s"
            params.append(bodega)
        sql += " ORDER BY fecha_toma DESC, bodega"
        cur.execute(sql, params)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
                'bodega': r['bodega'],
                'bodega_nombre': BODEGAS_OPERATIVAS.get(r['bodega'], r['bodega']),
                'estado': r['estado'],
                'total_productos_toma': r['total_productos_toma'],
                'total_productos_contifico': r['total_productos_contifico'],
                'total_cruzados': r['total_cruzados'],
                'total_con_diferencia': r['total_con_diferencia'],
                'timestamp_deteccion': r['timestamp_deteccion'].isoformat() if r['timestamp_deteccion'] else None,
                'timestamp_cruce': r['timestamp_cruce'].isoformat() if r['timestamp_cruce'] else None,
                'error_msg': r['error_msg'],
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/ejecuciones: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


MOTIVOS_CRUCE = [
    'Error de sistema',
    'Mal conteo',
    'Mal tipeo',
    'Facturacion cargada fuera de tiempo',
    'Factura mal cargada',
    'Baja cargada fuera de tiempo',
    'Bajas mal ejecutadas',
    'Traslados mal ejecutado de bodega Principal',
    'Traslado mal ejecutado por Planta de Produccion',
    'Traslado entre tiendas Erroneo',
    'Produccion mal ejecutada',
    'Cruce de productos',
    'Compra Extraordinaria',
    'Descuento a trabajador',
    'Producto sin justificacion',
]


def _asegurar_tabla_obs_cruce(cur):
    """La tabla de observaciones del cruce, creada si aun no existe.

    Vive aparte del detalle a proposito: cada CUADRAR borra y reescribe
    goti.cruce_operativo_detalle, asi que una nota guardada ahi se perderia
    en el siguiente cruce. Amarrada a bodega+fecha_toma+codigo sobrevive a
    todos los recruces del mismo dia, que es justo lo que hace falta para ir
    cuadrando: se anota, se vuelve a cruzar, y la nota sigue al lado de la
    diferencia ya actualizada.
    """
    cur.execute("""
        CREATE TABLE IF NOT EXISTS goti.cruce_operativo_observaciones (
            id SERIAL PRIMARY KEY,
            bodega VARCHAR(50) NOT NULL,
            fecha_toma DATE NOT NULL,
            codigo VARCHAR(50) NOT NULL,
            nombre TEXT,
            motivo TEXT,
            observaciones TEXT,
            diferencia_al_anotar NUMERIC(18,4),
            valor_al_anotar NUMERIC(18,4),
            creado_por VARCHAR(150),
            creado_at TIMESTAMP DEFAULT NOW(),
            modificado_por VARCHAR(150),
            modificado_at TIMESTAMP,
            UNIQUE (bodega, fecha_toma, codigo)
        )
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_cruce_obs_bod_fecha
        ON goti.cruce_operativo_observaciones (bodega, fecha_toma)
    """)


@app.route('/api/cruce-op/motivos', methods=['GET'])
def cruce_op_motivos():
    """Los mismos motivos que usan los locales, para que los dos informes se
    lean con el mismo criterio."""
    return jsonify(MOTIVOS_CRUCE)


@app.route('/api/cruce-op/observacion', methods=['POST'])
def cruce_op_observacion():
    """Guarda el motivo y el comentario de un producto del cruce.

    Se guarda contra (bodega, fecha_toma, codigo) y no contra la ejecucion,
    para que sobreviva a los recruces. Junto con el texto queda la diferencia
    que habia en ese momento: si un cruce posterior la cambia, se puede ver
    que la nota hablaba de otro numero.
    """
    data = request.json or {}
    ejec_id = data.get('ejecucion_id')
    codigo = (data.get('codigo') or '').strip()
    motivo = (data.get('motivo') or '').strip() or None
    obs = (data.get('observaciones') or '').strip() or None
    usuario = (data.get('usuario') or '').strip() or 'panel'

    if not ejec_id or not codigo:
        return jsonify({'error': 'ejecucion_id y codigo requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        _asegurar_tabla_obs_cruce(cur)

        cur.execute("""
            SELECT e.bodega, e.fecha_toma, d.nombre, d.diferencia, d.valor_diferencia
            FROM goti.cruce_operativo_ejecuciones e
            LEFT JOIN goti.cruce_operativo_detalle d
                   ON d.ejecucion_id = e.id AND d.codigo = %s
            WHERE e.id = %s
        """, (codigo, ejec_id))
        ctx = cur.fetchone()
        if not ctx:
            return jsonify({'error': 'Ejecucion no encontrada'}), 404

        # Una nota vacia es un borrado: si le quitan el motivo y el texto, la
        # fila no tiene por que quedarse ocupando lugar.
        if not motivo and not obs:
            cur.execute("""
                DELETE FROM goti.cruce_operativo_observaciones
                WHERE bodega = %s AND fecha_toma = %s AND codigo = %s
            """, (ctx['bodega'], ctx['fecha_toma'], codigo))
            conn.commit()
            return jsonify({'ok': True, 'borrado': True})

        cur.execute("""
            INSERT INTO goti.cruce_operativo_observaciones
                (bodega, fecha_toma, codigo, nombre, motivo, observaciones,
                 diferencia_al_anotar, valor_al_anotar, creado_por)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (bodega, fecha_toma, codigo) DO UPDATE SET
                motivo = EXCLUDED.motivo,
                observaciones = EXCLUDED.observaciones,
                diferencia_al_anotar = EXCLUDED.diferencia_al_anotar,
                valor_al_anotar = EXCLUDED.valor_al_anotar,
                modificado_por = EXCLUDED.creado_por,
                modificado_at = NOW()
            RETURNING COALESCE(modificado_por, creado_por) AS por,
                      COALESCE(modificado_at, creado_at) AS cuando
        """, (ctx['bodega'], ctx['fecha_toma'], codigo, ctx['nombre'], motivo,
              obs, ctx['diferencia'], ctx['valor_diferencia'], usuario))
        r = cur.fetchone()
        conn.commit()
        return jsonify({'ok': True, 'por': r['por'],
                        'cuando': r['cuando'].isoformat() if r['cuando'] else None})
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error en /api/cruce-op/observacion: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/detalle', methods=['GET'])
def cruce_detalle():
    """Detalle producto por producto de un cruce"""
    ejec_id = request.args.get('ejecucion_id')
    solo_dif = request.args.get('solo_diferencias', 'false').lower() == 'true'
    if not ejec_id:
        return jsonify({'error': 'ejecucion_id requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        _asegurar_tabla_obs_cruce(cur)
        sql = """SELECT d.*, o.motivo, o.observaciones,
                        o.diferencia_al_anotar,
                        COALESCE(o.modificado_por, o.creado_por) AS obs_por,
                        COALESCE(o.modificado_at, o.creado_at) AS obs_at
                 FROM goti.cruce_operativo_detalle d
                 JOIN goti.cruce_operativo_ejecuciones e ON e.id = d.ejecucion_id
                 LEFT JOIN goti.cruce_operativo_observaciones o
                        ON o.bodega = e.bodega
                       AND o.fecha_toma = e.fecha_toma
                       AND o.codigo = d.codigo
                 WHERE d.ejecucion_id = %s"""
        if solo_dif:
            sql += " AND d.diferencia != 0"
        sql += " ORDER BY ABS(d.valor_diferencia) DESC"
        cur.execute(sql, (ejec_id,))
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'categoria': r['categoria'],
                'unidad': r['unidad'],
                'cantidad_toma': float(r['cantidad_toma']) if r['cantidad_toma'] is not None else None,
                'cantidad_sistema': float(r['cantidad_sistema']) if r['cantidad_sistema'] is not None else None,
                'diferencia': float(r['diferencia']) if r['diferencia'] is not None else None,
                'costo_unitario': float(r['costo_unitario']) if r['costo_unitario'] is not None else 0,
                'valor_diferencia': float(r['valor_diferencia']) if r['valor_diferencia'] is not None else 0,
                'tipo_abc': r['tipo_abc'],
                'origen': r['origen'],
                'motivo': r['motivo'],
                'observaciones': r['observaciones'],
                'obs_por': r['obs_por'],
                'obs_at': r['obs_at'].isoformat() if r['obs_at'] else None,
                # si el cruce se volvio a correr despues de la nota, la
                # diferencia de la que hablaba ya no es la de ahora
                'obs_desactualizada': (
                    r['motivo'] is not None
                    and r['diferencia_al_anotar'] is not None
                    and r['diferencia'] is not None
                    and abs(float(r['diferencia_al_anotar']) - float(r['diferencia'])) > 0.01
                ),
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/detalle: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/resumen', methods=['GET'])
def cruce_resumen():
    """KPIs: ultima ejecucion por bodega, totales, valor diferencias"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            WITH ultimas AS (
                SELECT DISTINCT ON (bodega) id, bodega, fecha_toma,
                       total_productos_toma, total_con_diferencia
                FROM goti.cruce_operativo_ejecuciones
                WHERE estado = 'completado'
                ORDER BY bodega, fecha_toma DESC
            )
            SELECT u.id, u.bodega, u.fecha_toma, u.total_productos_toma, u.total_con_diferencia,
                   COALESCE(SUM(d.valor_diferencia) FILTER (WHERE d.diferencia != 0), 0) as valor_total,
                   COUNT(*) FILTER (WHERE d.diferencia < 0) as faltantes,
                   COUNT(*) FILTER (WHERE d.diferencia > 0) as sobrantes
            FROM ultimas u
            LEFT JOIN goti.cruce_operativo_detalle d ON d.ejecucion_id = u.id
            GROUP BY u.id, u.bodega, u.fecha_toma, u.total_productos_toma, u.total_con_diferencia
            ORDER BY u.bodega
        """)
        rows = cur.fetchall()

        resumen = []
        for r in rows:
            resumen.append({
                'bodega': r['bodega'],
                'bodega_nombre': BODEGAS_OPERATIVAS.get(r['bodega'], r['bodega']),
                'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
                'total_productos_toma': r['total_productos_toma'],
                'total_con_diferencia': r['total_con_diferencia'],
                'valor_total_diferencias': float(r['valor_total']),
                'faltantes': r['faltantes'],
                'sobrantes': r['sobrantes'],
            })
        return jsonify(resumen)
    except Exception as e:
        print(f"Error en /api/cruce/resumen: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/exportar-excel', methods=['GET'])
def cruce_exportar_excel():
    """Exporta detalle de un cruce a Excel"""
    ejec_id = request.args.get('ejecucion_id')
    if not ejec_id:
        return jsonify({'error': 'ejecucion_id requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Info ejecucion
        cur.execute("SELECT * FROM goti.cruce_operativo_ejecuciones WHERE id = %s", (ejec_id,))
        ejec = cur.fetchone()
        if not ejec:
            return jsonify({'error': 'Ejecucion no encontrada'}), 404

        # Detalle
        _asegurar_tabla_obs_cruce(cur)
        cur.execute("""SELECT d.*, o.motivo, o.observaciones
                       FROM goti.cruce_operativo_detalle d
                       JOIN goti.cruce_operativo_ejecuciones e ON e.id = d.ejecucion_id
                       LEFT JOIN goti.cruce_operativo_observaciones o
                              ON o.bodega = e.bodega
                             AND o.fecha_toma = e.fecha_toma
                             AND o.codigo = d.codigo
                       WHERE d.ejecucion_id = %s
                       ORDER BY ABS(d.valor_diferencia) DESC""", (ejec_id,))
        rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        bodega_nombre = BODEGAS_OPERATIVAS.get(ejec['bodega'], ejec['bodega'])
        ws.title = f"{bodega_nombre}"[:31]

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        red_font = Font(color='B91C1C', bold=True)
        green_font = Font(color='059669', bold=True)
        red_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        green_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        gray_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Codigo', 'Producto', 'Categoria', 'Tipo', 'Unidad',
                   'Fisico', 'Sistema', 'Diferencia', 'Costo Unit.', 'Valor Dif.', 'Origen',
                   'Motivo', 'Observacion']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for i, r in enumerate(rows, 2):
            vals = [r['codigo'], r['nombre'], r['categoria'], r['tipo_abc'], r['unidad'],
                    float(r['cantidad_toma']) if r['cantidad_toma'] is not None else 0,
                    float(r['cantidad_sistema']) if r['cantidad_sistema'] is not None else 0,
                    float(r['diferencia']) if r['diferencia'] is not None else 0,
                    float(r['costo_unitario']) if r['costo_unitario'] is not None else 0,
                    float(r['valor_diferencia']) if r['valor_diferencia'] is not None else 0,
                    r['origen'], r['motivo'] or '', r['observaciones'] or '']
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=v)
                cell.border = thin_border
            dif = vals[7]
            origen = vals[10]
            if dif < 0:
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = red_fill
                ws.cell(row=i, column=8).font = red_font
            elif dif > 0:
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = green_fill
                ws.cell(row=i, column=8).font = green_font
            if origen == 'solo_toma':
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = yellow_fill
            elif origen == 'solo_contifico':
                for col in range(1, len(vals) + 1):
                    ws.cell(row=i, column=col).fill = gray_fill

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_str = ejec['fecha_toma'].strftime('%Y-%m-%d') if ejec['fecha_toma'] else 'sin-fecha'
        filename = f"cruce_{ejec['bodega']}_{fecha_str}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Error en /api/cruce/exportar-excel: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce/tendencias', methods=['GET'])
def cruce_tendencias():
    """Top productos con diferencias recurrentes"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT d.codigo, d.nombre, d.categoria,
                   COUNT(*) as veces_con_diferencia,
                   ROUND(AVG(ABS(d.diferencia))::numeric, 2) as promedio_dif_abs,
                   ROUND(SUM(d.valor_diferencia)::numeric, 2) as valor_total
            FROM goti.cruce_operativo_detalle d
            JOIN goti.cruce_operativo_ejecuciones e ON d.ejecucion_id = e.id
            WHERE d.diferencia != 0 AND e.estado = 'completado'
            GROUP BY d.codigo, d.nombre, d.categoria
            HAVING COUNT(*) >= 2
            ORDER BY valor_total DESC
            LIMIT 30
        """)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'categoria': r['categoria'],
                'veces_con_diferencia': r['veces_con_diferencia'],
                'promedio_dif_abs': float(r['promedio_dif_abs']),
                'valor_total': float(r['valor_total']),
            })
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce/tendencias: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/borrar-datos', methods=['POST'])
def borrar_datos():
    """Borra datos de inventario para una bodega y fecha especifica"""
    clave = request.args.get('key', '')
    if clave != 'ChiosCostos2026':
        return jsonify({'error': 'no autorizado'}), 403
    conn = None
    try:
        data = request.get_json() or {}
        fecha = data.get('fecha')
        local = data.get('local')
        if not fecha or not local:
            return jsonify({'error': 'fecha y local son requeridos'}), 400

        conn = get_db()
        cur = conn.cursor()
        # Primero borrar asignaciones relacionadas
        cur.execute("""
            DELETE FROM goti.asignacion_diferencias
            WHERE conteo_id IN (
                SELECT id FROM goti.inventario_ciego_conteos
                WHERE fecha = %s AND local = %s
            )
        """, (fecha, local))
        asig_borradas = cur.rowcount

        cur.execute("""
            DELETE FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s
        """, (fecha, local))
        conteos_borrados = cur.rowcount
        conn.commit()

        return jsonify({
            'success': True,
            'conteos_borrados': conteos_borrados,
            'asignaciones_borradas': asig_borradas
        })
    except Exception as e:
        print(f"Error en /api/admin/borrar-datos: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/actualizar-costos', methods=['POST'])
def actualizar_costos():
    """Actualiza costo_unitario - acepta costos pre-calculados o lista de pendientes"""
    clave = request.args.get('key', '')
    if clave != 'ChiosCostos2026':
        return jsonify({'error': 'no autorizado'}), 403
    try:
        data = request.get_json() or {}

        # Modo 1: costos pre-calculados {nombre: costo}
        costos_directos = data.get('costos', {})
        if costos_directos:
            conn_inv = get_db()
            cur_inv = conn_inv.cursor()
            total = 0
            for nombre, costo in costos_directos.items():
                cur_inv.execute("""
                    UPDATE goti.inventario_ciego_conteos
                    SET costo_unitario = %s
                    WHERE nombre = %s AND (costo_unitario IS NULL OR costo_unitario = 0)
                """, (float(costo), nombre))
                total += cur_inv.rowcount
            conn_inv.commit()
            release_db(conn_inv)
            return jsonify({
                'productos_recibidos': len(costos_directos),
                'registros_actualizados': total
            })

        # Modo 2: devolver lista de productos sin costo
        conn_inv = get_db()
        cur_inv = conn_inv.cursor()
        cur_inv.execute("""
            SELECT DISTINCT nombre FROM goti.inventario_ciego_conteos
            WHERE costo_unitario IS NULL OR costo_unitario = 0
        """)
        nombres = [r['nombre'] for r in cur_inv.fetchall()]
        release_db(conn_inv)
        return jsonify({'pendientes': nombres, 'total': len(nombres)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


_cedulas_cache = {'datos': {}, 'timestamp': 0}

@app.route('/api/personas-cedulas-debug', methods=['GET'])
def debug_personas_airtable():
    """Debug: trae TODOS los campos de los primeros 3 registros"""
    import urllib.request, json as json_lib
    try:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=3'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        return jsonify({'records': [r.get('fields', {}) for r in data.get('records', [])]})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/personas-cedulas', methods=['GET'])
def obtener_personas_cedulas():
    """Retorna mapa {nombre: cedula} desde AirTable"""
    global _cedulas_cache
    ahora = _time.time()
    if _cedulas_cache['datos'] and (ahora - _cedulas_cache['timestamp']) < 600:
        return jsonify(_cedulas_cache['datos'])

    import urllib.request, json as json_lib
    cedulas = {}
    offset = None
    try:
        # Traer TODOS los campos (sin filtrar) para buscar cualquier variante de cédula
        while True:
            url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
            if offset:
                url += f'&offset={offset}'
            req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
            data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
            for r in data.get('records', []):
                f = r.get('fields', {})
                nombre = f.get('nombre') or f.get('Nombre') or ''
                # Buscar cédula en cualquier campo cuyo nombre contenga "ced" o "identif"
                ced = ''
                for k, v in f.items():
                    kl = k.lower().replace('é', 'e').replace('á', 'a').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
                    if 'cedula' in kl or 'identif' in kl or kl == 'ci' or kl == 'dni':
                        ced = str(v).strip()
                        break
                if nombre and ced:
                    cedulas[nombre] = ced
            offset = data.get('offset')
            if not offset:
                break
        _cedulas_cache = {'datos': cedulas, 'timestamp': ahora}
        return jsonify(cedulas)
    except Exception as e:
        print(f"Error cargando cedulas: {e}")
        return jsonify({})

def _cargar_personas_airtable():
    """Carga personas desde Airtable y actualiza cache del servidor"""
    import urllib.request, json as json_lib
    todos = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
        url += '&fields%5B%5D=nombre&fields%5B%5D=estado'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        for r in data.get('records', []):
            f = r.get('fields', {})
            nombre = f.get('nombre', '')
            if nombre:
                todos.append(nombre)
        offset = data.get('offset')
        if not offset:
            break
    resultado = sorted(set(todos))
    _personas_cache['datos'] = resultado
    _personas_cache['timestamp'] = _time.time()
    return resultado


def _obtener_personas():
    """Obtiene personas desde cache o Airtable si cache expirado"""
    ahora = _time.time()
    if _personas_cache['datos'] and (ahora - _personas_cache['timestamp']) < PERSONAS_CACHE_TTL:
        return _personas_cache['datos']
    try:
        return _cargar_personas_airtable()
    except Exception as e:
        print(f'Error cargando personas de Airtable: {e}')
        return _personas_cache['datos'] if _personas_cache['datos'] else []


_personas_correo_cache = {'datos': [], 'timestamp': 0}

def _obtener_personas_con_correo():
    """Obtiene personas activas con nombre y correo desde AirTable."""
    import urllib.request, json as json_lib
    ahora = _time.time()
    if _personas_correo_cache['datos'] and (ahora - _personas_correo_cache['timestamp']) < PERSONAS_CACHE_TTL:
        return _personas_correo_cache['datos']
    todos = []
    offset = None
    while True:
        url = f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}?pageSize=100'
        url += '&fields%5B%5D=nombre&fields%5B%5D=estado&fields%5B%5D=correo'
        if offset:
            url += f'&offset={offset}'
        req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_get_airtable_token()}'})
        data = json_lib.loads(urllib.request.urlopen(req, timeout=10).read())
        for r in data.get('records', []):
            f = r.get('fields', {})
            if f.get('estado') == 'Activo':
                nombre = f.get('nombre', '')
                correo = f.get('correo', '')
                if nombre:
                    todos.append({'nombre': nombre, 'correo': correo or ''})
        offset = data.get('offset')
        if not offset:
            break
    todos.sort(key=lambda x: x['nombre'])
    _personas_correo_cache['datos'] = todos
    _personas_correo_cache['timestamp'] = _time.time()
    return todos


@app.route('/api/personas', methods=['GET'])
def get_personas():
    try:
        if request.args.get('refresh') == '1':
            _personas_cache['timestamp'] = 0
        personas = _obtener_personas()
        return jsonify(personas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inventario/asignaciones', methods=['GET'])
def get_asignaciones():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify({'error': 'fecha y local son requeridos'}), 400
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT a.id, a.conteo_id, a.persona, a.cantidad
            FROM goti.asignacion_diferencias a
            JOIN goti.inventario_ciego_conteos c ON a.conteo_id = c.id
            WHERE c.fecha = %s AND c.local = %s
            ORDER BY a.conteo_id, a.id
        """, (fecha, local))
        rows = cur.fetchall()
        release_db(conn)
        result = {}
        for r in rows:
            cid = str(r['conteo_id'])
            if cid not in result:
                result[cid] = []
            result[cid].append({
                'id': r['id'],
                'persona': r['persona'],
                'cantidad': float(r['cantidad'])
            })
        return jsonify({'asignaciones': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/inventario/guardar-asignaciones', methods=['POST'])
def guardar_asignaciones():
    data = request.json
    conteo_id = data.get('conteo_id')
    asignaciones = data.get('asignaciones', [])
    if not conteo_id:
        return jsonify({'error': 'conteo_id es requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM goti.asignacion_diferencias
            WHERE conteo_id = %s
        """, (conteo_id,))
        # Obtener info del producto para guardar datos auto-contenidos
        cur.execute("""
            SELECT codigo, nombre, unidad, local, fecha
            FROM goti.inventario_ciego_conteos
            WHERE id = %s
        """, (conteo_id,))
        conteo_info = cur.fetchone()
        for a in asignaciones:
            if a.get('persona') and a.get('cantidad') and float(a['cantidad']) > 0:
                if conteo_info:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias
                            (conteo_id, persona, cantidad, codigo, nombre, unidad, local, fecha)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (conteo_id, a['persona'].strip(), float(a['cantidad']),
                          conteo_info['codigo'], conteo_info['nombre'], conteo_info['unidad'],
                          conteo_info['local'], conteo_info['fecha']))
                else:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias (conteo_id, persona, cantidad)
                        VALUES (%s, %s, %s)
                    """, (conteo_id, a['persona'].strip(), float(a['cantidad'])))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MÓDULO: Asignación por Sección (prototipo)
# ============================================================

@app.route('/api/conteo/secciones', methods=['GET'])
def listar_secciones_conteo():
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha or not local:
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre, total_valor
            FROM goti.asignacion_seccion
            WHERE fecha = %s AND local = %s
            ORDER BY created_at
        """, (fecha, local))
        secciones = cur.fetchall()
        result = []
        for s in secciones:
            cur.execute("""
                SELECT conteo_id, codigo, nombre, diferencia, costo_unitario, cantidad_asignada, valor
                FROM goti.asig_seccion_productos
                WHERE seccion_id = %s ORDER BY id
            """, (s['id'],))
            productos = [{'conteo_id': r['conteo_id'], 'codigo': r['codigo'],
                          'nombre': r['nombre'], 'diferencia': float(r['diferencia'] or 0),
                          'costo_unitario': float(r['costo_unitario'] or 0),
                          'cantidad_asignada': float(r['cantidad_asignada'] or 0),
                          'valor': float(r['valor'] or 0)} for r in cur.fetchall()]
            cur.execute("""
                SELECT persona, monto
                FROM goti.asig_seccion_personas
                WHERE seccion_id = %s ORDER BY id
            """, (s['id'],))
            personas = [{'persona': r['persona'], 'monto': float(r['monto'] or 0)} for r in cur.fetchall()]
            result.append({'id': s['id'], 'nombre': s['nombre'] or '',
                           'total_valor': float(s['total_valor'] or 0),
                           'productos': productos, 'personas': personas})
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo/secciones/guardar', methods=['POST'])
def guardar_seccion_conteo():
    """Divide productos equitativamente entre personas y guarda en asignacion_diferencias"""
    data = request.json
    productos = data.get('productos', [])
    personas = data.get('personas', [])  # lista de strings (nombres)
    if not productos:
        return jsonify({'error': 'Sin productos'}), 400
    if not personas:
        return jsonify({'error': 'Sin personas'}), 400
    n_personas = len(personas)
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        for p in productos:
            conteo_id = p['conteo_id']
            cantidad_por_persona = float(p.get('cantidad_asignada', 0)) / n_personas
            # Borrar asignaciones previas para este conteo
            cur.execute("""
                DELETE FROM goti.asignacion_diferencias
                WHERE conteo_id = %s
            """, (conteo_id,))
            # Obtener info del producto para guardar datos auto-contenidos
            cur.execute("""
                SELECT codigo, nombre, unidad, local, fecha
                FROM goti.inventario_ciego_conteos WHERE id = %s
            """, (conteo_id,))
            info = cur.fetchone()
            for nombre_persona in personas:
                if info:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias
                            (conteo_id, persona, cantidad, codigo, nombre, unidad, local, fecha)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (conteo_id, nombre_persona.strip(), cantidad_por_persona,
                          info['codigo'], info['nombre'], info['unidad'],
                          info['local'], info['fecha']))
                else:
                    cur.execute("""
                        INSERT INTO goti.asignacion_diferencias (conteo_id, persona, cantidad)
                        VALUES (%s, %s, %s)
                    """, (conteo_id, nombre_persona.strip(), cantidad_por_persona))
        conn.commit()
        return jsonify({'success': True, 'productos': len(productos), 'personas': n_personas})
    except Exception as e:
        if conn:
            try: conn.rollback()
            except: pass
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/conteo/secciones/<int:seccion_id>', methods=['DELETE'])
def eliminar_seccion_conteo(seccion_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.asig_seccion_productos WHERE seccion_id=%s", (seccion_id,))
        cur.execute("DELETE FROM goti.asig_seccion_personas WHERE seccion_id=%s", (seccion_id,))
        cur.execute("DELETE FROM goti.asignacion_seccion WHERE id=%s", (seccion_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})


@app.route('/api/debug-db', methods=['GET'])
def debug_db():
    """Diagnostico de conexion a BD"""
    import traceback
    result = {'pool_status': 'unknown', 'direct_conn': 'unknown'}
    # Test 1: pool connection
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT 1 as test, current_timestamp as ts, version() as ver")
        row = cur.fetchone()
        result['pool_status'] = 'ok'
        result['pool_data'] = {'test': row['test'], 'ts': str(row['ts']), 'ver': row['ver'][:60]}
        release_db(conn)
    except Exception as e:
        result['pool_status'] = 'error'
        result['pool_error'] = str(e)
        result['pool_traceback'] = traceback.format_exc()
    # Test 2: direct connection (bypass pool)
    try:
        conn2 = psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)
        cur2 = conn2.cursor()
        cur2.execute("SELECT COUNT(*) as cnt FROM goti.usuarios")
        row2 = cur2.fetchone()
        result['direct_conn'] = 'ok'
        result['direct_data'] = {'usuarios_count': row2['cnt']}
        conn2.close()
    except Exception as e:
        result['direct_conn'] = 'error'
        result['direct_error'] = str(e)
        result['direct_traceback'] = traceback.format_exc()
    result['db_config_host'] = DB_CONFIG['host']
    result['db_config_db'] = DB_CONFIG['database']
    return jsonify(result)


@app.route('/api/debug-personas', methods=['GET'])
def debug_personas():
    """Endpoint de diagnostico para el cache de personas"""
    ahora = _time.time()
    cache_age = ahora - _personas_cache['timestamp'] if _personas_cache['timestamp'] > 0 else -1
    token = _get_airtable_token()
    return jsonify({
        'cache_count': len(_personas_cache['datos']),
        'cache_age_seconds': round(cache_age, 1),
        'cache_ttl': PERSONAS_CACHE_TTL,
        'cache_expired': cache_age > PERSONAS_CACHE_TTL if cache_age >= 0 else True,
        'airtable_token_configured': bool(token),
        'token_length': len(token) if token else 0,
        'env_keys_with_air': [k for k in os.environ.keys() if 'AIR' in k.upper()],
        'primeras_3': _personas_cache['datos'][:3] if _personas_cache['datos'] else []
    })

# ==================== MERMA OPERATIVA ====================

@app.route('/api/merma', methods=['GET'])
def listar_mermas():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        filtros = []
        params = []
        if fecha_desde:
            filtros.append("fecha >= %s")
            params.append(fecha_desde)
        if fecha_hasta:
            filtros.append("fecha <= %s")
            params.append(fecha_hasta)
        if local:
            filtros.append("local = %s")
            params.append(local)
        where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
        cur.execute(f"""
            SELECT id, fecha, local, codigo, nombre, unidad, cantidad, motivo,
                   costo_unitario, costo_total, created_at
            FROM goti.merma_operativa
            {where}
            ORDER BY fecha DESC, created_at DESC
        """, params)
        rows = cur.fetchall()
        result = []
        for r in rows:
            result.append({
                'id': r['id'],
                'fecha': str(r['fecha']),
                'local': r['local'],
                'codigo': r['codigo'],
                'nombre': r['nombre'],
                'unidad': r['unidad'],
                'cantidad': float(r['cantidad']),
                'motivo': r['motivo'] or '',
                'costo_unitario': float(r['costo_unitario'] or 0),
                'costo_total': float(r['costo_total'] or 0),
                'created_at': str(r['created_at'])
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/merma/registrar', methods=['POST'])
def registrar_merma():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    codigo = data.get('codigo', '').strip()
    nombre = data.get('nombre', '').strip()
    unidad = data.get('unidad', '').strip()
    cantidad = data.get('cantidad')
    motivo = data.get('motivo', '').strip()
    costo_unitario = float(data.get('costo_unitario') or 0)
    if not all([fecha, local, codigo, nombre, cantidad]):
        return jsonify({'error': 'Faltan campos requeridos: fecha, local, codigo, nombre, cantidad'}), 400
    costo_total = float(cantidad) * costo_unitario
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.merma_operativa
                (fecha, local, codigo, nombre, unidad, cantidad, motivo, costo_unitario, costo_total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (fecha, local, codigo, nombre, unidad, float(cantidad), motivo, costo_unitario, costo_total))
        nuevo_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'success': True, 'id': nuevo_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/merma/<int:merma_id>', methods=['DELETE'])
def eliminar_merma(merma_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.merma_operativa WHERE id = %s", (merma_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas', methods=['GET'])
def listar_bajas():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        filtros = []
        params = []
        if fecha_desde:
            filtros.append("b.fecha >= %s"); params.append(fecha_desde)
        if fecha_hasta:
            filtros.append("b.fecha <= %s"); params.append(fecha_hasta)
        if local:
            filtros.append("b.local = %s"); params.append(local)
        where = ("WHERE " + " AND ".join(filtros)) if filtros else ""
        # Traer grupos con sus productos y asignaciones
        cur.execute(f"""
            SELECT b.baja_grupo,
                   MIN(b.fecha) AS fecha,
                   MIN(b.local) AS local,
                   MIN(b.motivo) AS motivo,
                   MIN(b.documento) AS documento,
                   MIN(b.codigo_baja) AS codigo_baja,
                   SUM(b.costo_total) AS total_costo,
                   MIN(b.created_at) AS created_at
            FROM goti.bajas_directas b
            {where}
            GROUP BY b.baja_grupo
            ORDER BY MIN(b.created_at) DESC
        """, params)
        grupos = cur.fetchall()
        result = []
        for g in grupos:
            grp = g['baja_grupo']
            # Productos del grupo
            cur.execute("""
                SELECT id, codigo, nombre, unidad, cantidad, costo_unitario, costo_total
                FROM goti.bajas_directas
                WHERE baja_grupo = %s ORDER BY id
            """, (grp,))
            items = [{'id': r['id'], 'codigo': r['codigo'], 'nombre': r['nombre'],
                      'unidad': r['unidad'], 'cantidad': float(r['cantidad']),
                      'costo_unitario': float(r['costo_unitario'] or 0),
                      'costo_total': float(r['costo_total'] or 0)} for r in cur.fetchall()]
            # Asignaciones del grupo
            cur.execute("""
                SELECT id, persona, monto FROM goti.bajas_asignaciones
                WHERE baja_grupo = %s ORDER BY id
            """, (grp,))
            asigs = [{'id': r['id'], 'persona': r['persona'], 'monto': float(r['monto'])} for r in cur.fetchall()]
            result.append({
                'baja_grupo': grp,
                'fecha': str(g['fecha']),
                'local': g['local'],
                'motivo': g['motivo'] or '',
                'documento': g['documento'] or '',
                'codigo_baja': g['codigo_baja'] or '',
                'total_costo': float(g['total_costo'] or 0),
                'created_at': str(g['created_at']),
                'items': items,
                'asignaciones': asigs
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas/registrar', methods=['POST'])
def registrar_baja():
    import time as _time_mod
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    motivo = data.get('motivo', '').strip()
    documento = data.get('documento', '').strip()
    codigo_baja = data.get('codigo_baja', '').strip()
    items = data.get('items', [])
    asignaciones = data.get('asignaciones', [])
    if not all([fecha, local]):
        return jsonify({'error': 'Faltan campos requeridos: fecha, local'}), 400
    if not items:
        return jsonify({'error': 'Debes incluir al menos un producto'}), 400
    baja_grupo = int(_time_mod.time() * 1000)
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        for item in items:
            codigo = item.get('codigo', '').strip()
            nombre = item.get('nombre', '').strip()
            unidad = item.get('unidad', '').strip()
            cantidad = float(item.get('cantidad') or 0)
            costo_unitario = float(item.get('costo_unitario') or 0)
            costo_total = cantidad * costo_unitario
            cur.execute("""
                INSERT INTO goti.bajas_directas
                    (baja_grupo, fecha, local, codigo, nombre, unidad, cantidad, persona, motivo, documento, codigo_baja, costo_unitario, costo_total)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (baja_grupo, fecha, local, codigo, nombre, unidad, cantidad, '', motivo, documento or None, codigo_baja or None, costo_unitario, costo_total))
        for asig in asignaciones:
            persona = asig.get('persona', '').strip()
            monto = float(asig.get('monto') or 0)
            if persona and monto > 0:
                cur.execute("""
                    INSERT INTO goti.bajas_asignaciones
                        (baja_grupo, persona, monto, fecha, local, motivo)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (baja_grupo, persona, monto, fecha, local, motivo))
        conn.commit()
        return jsonify({'success': True, 'baja_grupo': baja_grupo})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/bajas/grupo/<int:baja_grupo>', methods=['DELETE'])
def eliminar_baja_grupo(baja_grupo):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.bajas_directas WHERE baja_grupo = %s", (baja_grupo,))
        cur.execute("DELETE FROM goti.bajas_asignaciones WHERE baja_grupo = %s", (baja_grupo,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


import threading
def _precargar_personas():
    for intento in range(6):
        token = _get_airtable_token()
        if not token:
            print(f'Pre-carga intento {intento+1}: AIRTABLE_TOKEN vacio, reintentando en 5s...')
            _time.sleep(5)
            continue
        try:
            _cargar_personas_airtable()
            print(f'Pre-carga personas OK (intento {intento+1}): {len(_personas_cache["datos"])} personas')
            return
        except Exception as e:
            print(f'Pre-carga intento {intento+1} error: {e}')
            _time.sleep(5)
    print('Pre-carga personas FALLO despues de 6 intentos')
threading.Thread(target=_precargar_personas, daemon=True).start()

# Inicializar tablas al arrancar
try:
    init_db()
except Exception as _e:
    print(f'Startup init_db error: {_e}')

# ==================== PANEL DE CONTROL ====================

@app.route('/api/panel/consultar', methods=['GET'])
def panel_consultar():
    """Consulta inventario por fecha y bodega opcional"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT local, codigo, nombre, unidad,
                   cantidad, cantidad_contada, cantidad_contada_2, costo_unitario
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s
        """
        params = [fecha]
        if bodega:
            query += ' AND local = %s'
            params.append(bodega)
        query += ' ORDER BY local, nombre'

        cur.execute(query, params)
        rows = cur.fetchall()

        return jsonify({
            'total': len(rows),
            'data': rows
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/panel/borrar-stock', methods=['POST'])
def panel_borrar_stock():
    """Pone cantidad=NULL para fecha/bodega. NO toca conteos."""
    data = request.get_json()
    fecha = data.get('fecha')
    bodega = data.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Contar afectados
        q_count = """
            SELECT COUNT(*) as cnt FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params = [fecha]
        if bodega:
            q_count += ' AND local = %s'
            params.append(bodega)

        cur.execute(q_count, params)
        count = cur.fetchone()['cnt']

        if count == 0:
            return jsonify({'affected': 0, 'message': 'No hay registros con stock para esa fecha'})

        # Ejecutar UPDATE
        q_update = """
            UPDATE goti.inventario_ciego_conteos
            SET cantidad = NULL
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params2 = [fecha]
        if bodega:
            q_update += ' AND local = %s'
            params2.append(bodega)

        cur.execute(q_update, params2)
        affected = cur.rowcount
        conn.commit()

        return jsonify({
            'affected': affected,
            'message': f'Stock borrado: {affected} registros actualizados'
        })
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/panel/contar-stock', methods=['GET'])
def panel_contar_stock():
    """Cuenta registros con stock para preview antes de borrar"""
    fecha = request.args.get('fecha')
    bodega = request.args.get('bodega', '')
    if not fecha:
        return jsonify({'error': 'Falta fecha'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT COUNT(*) as cnt FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND cantidad IS NOT NULL
        """
        params = [fecha]
        if bodega:
            query += ' AND local = %s'
            params.append(bodega)

        cur.execute(query, params)
        count = cur.fetchone()['cnt']

        return jsonify({'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ==================== ASIGNACION SEMANAL ====================

@app.route('/api/semanas', methods=['GET'])
def listar_semanas():
    """Lista semanas de inventario para una bodega"""
    local = request.args.get('local')
    if not local:
        return jsonify({'error': 'Falta parametro local'}), 400

    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT s.*,
                (SELECT COUNT(DISTINCT c.codigo)
                 FROM goti.inventario_ciego_conteos c
                 WHERE c.local = s.local
                   AND c.fecha BETWEEN s.fecha_inicio AND s.fecha_fin
                   AND (c.cantidad_contada IS NOT NULL OR c.cantidad_contada_2 IS NOT NULL)
                ) as total_productos,
                COALESCE((SELECT SUM(ap.monto)
                 FROM goti.asignacion_semanal a
                 JOIN goti.asignacion_semanal_personas ap ON ap.asignacion_semanal_id = a.id
                 WHERE a.semana_id = s.id
                ), 0) as total_asignado
            FROM goti.semanas_inventario s
            WHERE s.local = %s
        """
        params = [local]

        if fecha_desde:
            query += ' AND s.fecha_inicio >= %s'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND s.fecha_fin <= %s'
            params.append(fecha_hasta)

        query += ' ORDER BY s.fecha_inicio DESC'
        cur.execute(query, params)
        semanas = cur.fetchall()

        # Convert dates to strings
        for s in semanas:
            s['fecha_inicio'] = str(s['fecha_inicio'])
            s['fecha_fin'] = str(s['fecha_fin'])
            if s.get('cerrada_at'):
                s['cerrada_at'] = str(s['cerrada_at'])
            if s.get('created_at'):
                s['created_at'] = str(s['created_at'])

        return jsonify(semanas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/crear', methods=['POST'])
def crear_semana():
    """Crea o retorna una semana de inventario"""
    data = request.get_json()
    local = data.get('local')
    fecha_inicio = data.get('fecha_inicio')

    if not local or not fecha_inicio:
        return jsonify({'error': 'Faltan parametros local y fecha_inicio'}), 400

    from datetime import datetime, timedelta
    try:
        dt_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'fecha_inicio debe ser formato YYYY-MM-DD'}), 400

    # Validar que sea lunes (ISO weekday 1)
    if dt_inicio.isoweekday() != 1:
        return jsonify({'error': 'fecha_inicio debe ser un lunes'}), 400

    dt_fin = dt_inicio + timedelta(days=6)  # domingo

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Verificar si ya existe
        cur.execute("""
            SELECT * FROM goti.semanas_inventario
            WHERE fecha_inicio = %s AND local = %s
        """, (dt_inicio, local))
        existing = cur.fetchone()

        if existing:
            existing['fecha_inicio'] = str(existing['fecha_inicio'])
            existing['fecha_fin'] = str(existing['fecha_fin'])
            if existing.get('cerrada_at'):
                existing['cerrada_at'] = str(existing['cerrada_at'])
            if existing.get('created_at'):
                existing['created_at'] = str(existing['created_at'])
            return jsonify(existing)

        # Verificar que no haya otra semana abierta para este local
        cur.execute("""
            SELECT id, fecha_inicio, fecha_fin FROM goti.semanas_inventario
            WHERE local = %s AND estado = 'abierta'
        """, (local,))
        abierta = cur.fetchone()

        if abierta:
            return jsonify({
                'error': f'Ya existe una semana abierta para {local} ({abierta["fecha_inicio"]} - {abierta["fecha_fin"]}). Cierre primero antes de crear otra.'
            }), 409

        cur.execute("""
            INSERT INTO goti.semanas_inventario (fecha_inicio, fecha_fin, local)
            VALUES (%s, %s, %s)
            RETURNING *
        """, (dt_inicio, dt_fin, local))
        nueva = cur.fetchone()
        conn.commit()

        nueva['fecha_inicio'] = str(nueva['fecha_inicio'])
        nueva['fecha_fin'] = str(nueva['fecha_fin'])
        if nueva.get('created_at'):
            nueva['created_at'] = str(nueva['created_at'])

        return jsonify(nueva), 201
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/diferencias', methods=['GET'])
def diferencias_semana(semana_id):
    """Obtiene diferencias semanales de productos para una semana"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Obtener datos de la semana
        cur.execute("""
            SELECT * FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404

        fecha_inicio = semana['fecha_inicio']
        fecha_fin = semana['fecha_fin']
        local = semana['local']

        # Netear diferencias diarias por producto en la semana
        # Resta cantidad_justificada de cada dia (justificacion parcial)
        # dif_neta = dif_dia + cantidad_justificada (dif es negativa, justif reduce el faltante)
        cur.execute("""
            WITH diferencias_diarias AS (
                SELECT
                    codigo, nombre, COALESCE(unidad, '') as unidad, fecha,
                    cantidad as stock_sistema,
                    COALESCE(cantidad_contada_2, cantidad_contada) as contado,
                    COALESCE(cantidad_contada_2, cantidad_contada) - cantidad as dif_dia,
                    COALESCE(costo_unitario, 0) as costo_unitario,
                    COALESCE(corregido, FALSE) as corregido,
                    COALESCE(justificado, FALSE) as justificado,
                    COALESCE(cantidad_justificada, 0) as cant_justif
                FROM goti.inventario_ciego_conteos
                WHERE local = %s AND fecha BETWEEN %s AND %s
                  AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
            )
            SELECT
                codigo,
                nombre,
                unidad,
                SUM(
                    CASE
                        WHEN dif_dia < 0 THEN LEAST(dif_dia + cant_justif, 0)
                        WHEN dif_dia > 0 THEN GREATEST(dif_dia - cant_justif, 0)
                        ELSE 0
                    END
                ) as diferencia,
                AVG(costo_unitario) as costo_unitario,
                COUNT(*) as dias_contados,
                BOOL_AND(justificado) as justificado,
                SUM(cant_justif) as total_justificado,
                BOOL_OR(corregido) as tiene_correccion,
                json_agg(json_build_object(
                    'fecha', fecha,
                    'stock', stock_sistema,
                    'contado', contado,
                    'dif', dif_dia,
                    'corregido', corregido,
                    'justificado', justificado,
                    'cant_justif', cant_justif
                ) ORDER BY fecha) as detalle_diario
            FROM diferencias_diarias
            GROUP BY codigo, nombre, unidad
            HAVING SUM(CASE WHEN dif_dia < 0 THEN LEAST(dif_dia + cant_justif, 0) WHEN dif_dia > 0 THEN GREATEST(dif_dia - cant_justif, 0) ELSE 0 END) != 0
            ORDER BY nombre
        """, (local, fecha_inicio, fecha_fin))
        diferencias = cur.fetchall()

        # Serializar datos — costo incluye 20% por costos indirectos (no visible al usuario)
        FACTOR_COSTO_INDIRECTO = 1.20
        for d in diferencias:
            d['diferencia'] = float(d['diferencia']) if d['diferencia'] else 0
            costo_base = float(d['costo_unitario']) if d['costo_unitario'] else 0
            d['costo_unitario'] = round(costo_base * FACTOR_COSTO_INDIRECTO, 4)
            if d.get('detalle_diario'):
                for dd in d['detalle_diario']:
                    dd['fecha'] = str(dd['fecha'])
                    dd['stock'] = float(dd['stock']) if dd['stock'] else 0
                    dd['contado'] = float(dd['contado']) if dd['contado'] else 0
                    dd['dif'] = float(dd['dif']) if dd['dif'] else 0

        # Obtener asignaciones existentes para esta semana (con grupo_idx)
        cur.execute("""
            SELECT a.id, a.codigo, a.nombre, a.unidad, a.diferencia_semanal, a.costo_unitario,
                   COALESCE(a.grupo_idx, 0) as grupo_idx,
                   json_agg(json_build_object(
                       'id', ap.id,
                       'persona', ap.persona,
                       'cantidad', ap.cantidad,
                       'monto', ap.monto
                   )) FILTER (WHERE ap.id IS NOT NULL) as personas
            FROM goti.asignacion_semanal a
            LEFT JOIN goti.asignacion_semanal_personas ap
                ON ap.asignacion_semanal_id = a.id
            WHERE a.semana_id = %s
            GROUP BY a.id, a.codigo, a.nombre, a.unidad, a.diferencia_semanal, a.costo_unitario, a.grupo_idx
        """, (semana_id,))
        asignaciones = cur.fetchall()

        # Mapear asignaciones por codigo (lista, puede haber multiples por grupo)
        asig_map = {}
        for a in asignaciones:
            entry = {
                'id': a['id'],
                'nombre': a['nombre'],
                'unidad': a['unidad'],
                'diferencia_semanal': float(a['diferencia_semanal']) if a['diferencia_semanal'] else 0,
                'costo_unitario': float(a['costo_unitario']) if a['costo_unitario'] else 0,
                'grupo_idx': a['grupo_idx'],
                'personas': a['personas'] or []
            }
            if a['codigo'] not in asig_map:
                asig_map[a['codigo']] = []
            asig_map[a['codigo']].append(entry)

        # Combinar diferencias con asignaciones
        codigos_en_resultado = set()
        resultado = []
        for d in diferencias:
            item = dict(d)
            if d['codigo'] in asig_map:
                item['asignaciones'] = asig_map[d['codigo']]
            else:
                item['asignaciones'] = []
            resultado.append(item)
            codigos_en_resultado.add(d['codigo'])

        # Para semanas cerradas: incluir tambien productos asignados que ya no tienen diferencia neta
        if semana['estado'] == 'cerrada':
            for codigo, asig_list in asig_map.items():
                if codigo not in codigos_en_resultado and any(a['personas'] for a in asig_list):
                    first = asig_list[0]
                    resultado.append({
                        'codigo': codigo,
                        'nombre': first['nombre'],
                        'unidad': first['unidad'],
                        'diferencia': first['diferencia_semanal'],
                        'costo_unitario': first['costo_unitario'],
                        'dias_contados': 0,
                        'justificado': False,
                        'total_justificado': 0,
                        'tiene_correccion': False,
                        'detalle_diario': [],
                        'asignaciones': asig_list,
                    })

        semana_info = {
            'id': semana['id'],
            'fecha_inicio': str(semana['fecha_inicio']),
            'fecha_fin': str(semana['fecha_fin']),
            'local': semana['local'],
            'estado': semana['estado'],
            'cerrada_por': semana.get('cerrada_por'),
            'cerrada_at': str(semana['cerrada_at']) if semana.get('cerrada_at') else None,
        }

        return jsonify({
            'semana': semana_info,
            'diferencias': resultado
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/asignar', methods=['POST'])
def asignar_semana(semana_id):
    """Guarda asignaciones semanales de diferencias preservando estructura de grupos"""
    data = request.get_json()
    grupos = data.get('grupos', [])

    # Compatibilidad: si el frontend envia formato viejo {asignaciones:[...]},
    # convertir a formato nuevo {grupos:[...]} para no perder datos
    if not grupos and 'asignaciones' in data:
        asignaciones_viejas = data.get('asignaciones', [])
        if asignaciones_viejas:
            # Agrupar por set de personas (mismo comportamiento que tenia el frontend viejo)
            from collections import defaultdict
            grupos_temp = defaultdict(lambda: {'productos': [], 'personas_set': set()})
            for asig in asignaciones_viejas:
                personas = sorted([p.get('persona', '') for p in asig.get('personas', [])])
                key = '|'.join(personas)
                grupos_temp[key]['productos'].append({
                    'codigo': asig.get('codigo'),
                    'nombre': asig.get('nombre'),
                    'unidad': asig.get('unidad'),
                    'diferencia_semanal': asig.get('diferencia_semanal', 0),
                    'costo_unitario': asig.get('costo_unitario', 0),
                    'cantidad': sum(float(p.get('cantidad', 0)) for p in asig.get('personas', []))
                })
                for p in personas:
                    grupos_temp[key]['personas_set'].add(p)
            grupos = []
            for key, g in grupos_temp.items():
                grupos.append({
                    'productos': g['productos'],
                    'personas': sorted(g['personas_set'])
                })

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Verificar que la semana existe y esta abierta
        cur.execute("""
            SELECT * FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'abierta':
            return jsonify({'error': 'La semana esta cerrada, no se pueden modificar asignaciones'}), 400

        # Borrar asignaciones previas de esta semana
        cur.execute("""
            DELETE FROM goti.asignacion_semanal_personas
            WHERE asignacion_semanal_id IN (
                SELECT id FROM goti.asignacion_semanal WHERE semana_id = %s
            )
        """, (semana_id,))
        cur.execute("""
            DELETE FROM goti.asignacion_semanal WHERE semana_id = %s
        """, (semana_id,))

        # Insertar por grupo preservando la estructura original
        total_insertadas = 0
        for grupo_idx, grupo in enumerate(grupos):
            personas = grupo.get('personas', [])
            num_personas = len(personas)
            if num_personas == 0:
                continue

            for prod in grupo.get('productos', []):
                cantidad_total = float(prod.get('cantidad', 0))
                if cantidad_total <= 0:
                    continue
                costo = float(prod.get('costo_unitario', 0))

                cur.execute("""
                    INSERT INTO goti.asignacion_semanal
                        (semana_id, codigo, nombre, unidad, local, diferencia_semanal, costo_unitario, grupo_idx)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    semana_id,
                    prod.get('codigo'),
                    prod.get('nombre'),
                    prod.get('unidad'),
                    semana['local'],
                    prod.get('diferencia_semanal', 0),
                    costo,
                    grupo_idx
                ))
                asig_id = cur.fetchone()['id']

                # Dividir cantidad equitativamente entre personas del grupo
                cant_por_persona = cantidad_total / num_personas
                for persona_nombre in personas:
                    monto = cant_por_persona * costo
                    cur.execute("""
                        INSERT INTO goti.asignacion_semanal_personas
                            (asignacion_semanal_id, persona, cantidad, monto)
                        VALUES (%s, %s, %s, %s)
                    """, (asig_id, persona_nombre, round(cant_por_persona, 4), round(monto, 2)))

                total_insertadas += 1

        conn.commit()
        return jsonify({
            'ok': True,
            'message': f'{total_insertadas} asignaciones guardadas para semana {semana_id}'
        })
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/justificar-producto', methods=['POST'])
def justificar_producto_semanal(semana_id):
    """Justifica TODOS los dias de un producto en una semana (marca justificado=TRUE y cantidad_justificada=abs(dif))"""
    data = request.get_json()
    codigo = data.get('codigo')
    justificar = data.get('justificar', True)  # True=justificar, False=quitar justificacion

    if not codigo:
        return jsonify({'error': 'Falta codigo de producto'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT * FROM goti.semanas_inventario WHERE id = %s", (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404

        if justificar:
            # Justificar: marcar justificado=TRUE y cantidad_justificada=abs(diferencia) en todos los dias
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET justificado = TRUE,
                    cantidad_justificada = ABS(COALESCE(cantidad_contada_2, cantidad_contada) - cantidad)
                WHERE local = %s AND codigo = %s
                  AND fecha BETWEEN %s AND %s
                  AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
            """, (semana['local'], codigo, semana['fecha_inicio'], semana['fecha_fin']))
        else:
            # Quitar justificacion
            cur.execute("""
                UPDATE goti.inventario_ciego_conteos
                SET justificado = FALSE, cantidad_justificada = 0
                WHERE local = %s AND codigo = %s
                  AND fecha BETWEEN %s AND %s
            """, (semana['local'], codigo, semana['fecha_inicio'], semana['fecha_fin']))

        rows = cur.rowcount
        conn.commit()
        return jsonify({'ok': True, 'dias_actualizados': rows})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>/cerrar', methods=['POST'])
def cerrar_semana(semana_id):
    """Cierra una semana de inventario"""
    data = request.get_json() or {}
    cerrada_por = data.get('cerrada_por', 'sistema')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT estado FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'abierta':
            return jsonify({'error': 'La semana ya esta cerrada'}), 400

        cur.execute("""
            UPDATE goti.semanas_inventario
            SET estado = 'cerrada', cerrada_por = %s, cerrada_at = NOW()
            WHERE id = %s
            RETURNING *
        """, (cerrada_por, semana_id))
        updated = cur.fetchone()
        conn.commit()

        updated['fecha_inicio'] = str(updated['fecha_inicio'])
        updated['fecha_fin'] = str(updated['fecha_fin'])
        if updated.get('cerrada_at'):
            updated['cerrada_at'] = str(updated['cerrada_at'])
        if updated.get('created_at'):
            updated['created_at'] = str(updated['created_at'])

        return jsonify({'ok': True, 'semana': updated})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/<int:semana_id>', methods=['DELETE'])
def eliminar_semana(semana_id):
    """Elimina una semana (abierta o cerrada) y todas sus asignaciones (solo admin)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT estado FROM goti.semanas_inventario WHERE id = %s", (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404

        # Admin puede eliminar tanto abiertas como cerradas
        # Al eliminar, los productos asignados quedan sin responsable y deben ser reasignados

        # Eliminar asignaciones de personas
        cur.execute("""
            DELETE FROM goti.asignacion_semanal_personas
            WHERE asignacion_semanal_id IN (
                SELECT id FROM goti.asignacion_semanal WHERE semana_id = %s
            )
        """, (semana_id,))
        # Eliminar asignaciones
        cur.execute("DELETE FROM goti.asignacion_semanal WHERE semana_id = %s", (semana_id,))
        # Eliminar semana
        cur.execute("DELETE FROM goti.semanas_inventario WHERE id = %s", (semana_id,))
        conn.commit()

        return jsonify({'success': True, 'estado_previo': semana['estado']})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/semanas/<int:semana_id>/reabrir', methods=['POST'])
def reabrir_semana(semana_id):
    """Reabre una semana cerrada (solo admin)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT estado FROM goti.semanas_inventario WHERE id = %s
        """, (semana_id,))
        semana = cur.fetchone()
        if not semana:
            return jsonify({'error': 'Semana no encontrada'}), 404
        if semana['estado'] != 'cerrada':
            return jsonify({'error': 'La semana ya esta abierta'}), 400

        cur.execute("""
            UPDATE goti.semanas_inventario
            SET estado = 'abierta', cerrada_por = NULL, cerrada_at = NULL
            WHERE id = %s
            RETURNING *
        """, (semana_id,))
        updated = cur.fetchone()
        conn.commit()

        updated['fecha_inicio'] = str(updated['fecha_inicio'])
        updated['fecha_fin'] = str(updated['fecha_fin'])
        if updated.get('created_at'):
            updated['created_at'] = str(updated['created_at'])

        return jsonify({'ok': True, 'semana': updated})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/pendientes', methods=['GET'])
def semanas_pendientes():
    """Retorna semanas abiertas cuyo periodo ya termino (para recordatorios)"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            SELECT s.*
            FROM goti.semanas_inventario s
            WHERE s.estado = 'abierta'
              AND s.fecha_fin < CURRENT_DATE
            ORDER BY s.fecha_fin ASC
        """)
        semanas = cur.fetchall()

        for s in semanas:
            s['fecha_inicio'] = str(s['fecha_inicio'])
            s['fecha_fin'] = str(s['fecha_fin'])
            if s.get('created_at'):
                s['created_at'] = str(s['created_at'])

        return jsonify(semanas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/semanas/resumen-persona', methods=['GET'])
def resumen_persona_semanal():
    """Resumen de asignaciones por persona a traves de semanas cerradas"""
    local = request.args.get('local')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    if not local:
        return jsonify({'error': 'Falta parametro local'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        query = """
            SELECT ap.persona,
                   SUM(ap.cantidad) as total_cantidad,
                   SUM(ap.monto) as total_monto,
                   COUNT(DISTINCT a.semana_id) as semanas_count
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE s.local = %s AND s.estado = 'cerrada'
        """
        params = [local]

        if fecha_desde:
            query += ' AND s.fecha_inicio >= %s'
            params.append(fecha_desde)
        if fecha_hasta:
            query += ' AND s.fecha_fin <= %s'
            params.append(fecha_hasta)

        query += ' GROUP BY ap.persona ORDER BY total_monto DESC'
        cur.execute(query, params)
        resumen = cur.fetchall()

        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Cruce Operativo "CUADRAR" - Boton + Worker local
# ============================================================
# Flujo: Panel web -> POST /solicitar -> tarea pendiente
#        Worker PC FINANZAS -> GET /pendientes (cada 15s) -> toma tarea
#        Worker descarga Contifico, calcula cruce -> POST /resultado
#        Panel web -> GET /estado/<id> (polling) -> muestra resultado

# Token simple para autenticar al worker (env var)
WORKER_TOKEN = os.environ.get('CRUCE_WORKER_TOKEN', 'worker-foodix-2026-7K3xR9pL2qN8mZ4w')


def _archivar_version_cruce(cur, ejecucion_id, usuario):
    """Guarda una copia del cruce antes de que se pise. Devuelve su numero.

    Se llama justo antes de borrar el detalle. Si se llamara despues no habria
    nada que copiar, que es exactamente lo que venia pasando: cada cruce nuevo
    se llevaba por delante el anterior y no quedaba rastro de por donde
    empezaron ni de cuanto habian corregido.

    Devuelve None si no hay nada util que archivar: un cruce que fallo o que
    aun no termino no aporta informacion.
    """
    cur.execute("""
        SELECT bodega, fecha_toma, fecha_corte_contifico, estado, solicitado_por,
               solicitado_at, timestamp_cruce, total_productos_toma,
               total_productos_contifico, total_cruzados, total_con_diferencia,
               valor_total_dif
        FROM goti.cruce_operativo_ejecuciones WHERE id = %s
    """, (ejecucion_id,))
    e = cur.fetchone()
    if not e or e['estado'] != 'completado':
        return None

    cur.execute("""
        SELECT COALESCE(MAX(version), 0) + 1 AS siguiente
        FROM goti.cruce_operativo_versiones
        WHERE bodega = %s AND fecha_toma = %s
    """, (e['bodega'], e['fecha_toma']))
    version = cur.fetchone()['siguiente']

    cur.execute("SELECT COUNT(*) AS n FROM goti.cruce_operativo_detalle "
                "WHERE ejecucion_id = %s", (ejecucion_id,))
    filas_detalle = cur.fetchone()['n']

    cur.execute("""
        INSERT INTO goti.cruce_operativo_versiones
            (ejecucion_id, version, bodega, fecha_toma, fecha_corte_contifico,
             solicitado_por, solicitado_at, timestamp_cruce,
             total_productos_toma, total_productos_contifico, total_cruzados,
             total_filas_detalle, total_con_diferencia, valor_total_dif,
             archivado_por)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (ejecucion_id, version, e['bodega'], e['fecha_toma'],
          e['fecha_corte_contifico'], e['solicitado_por'], e['solicitado_at'],
          e['timestamp_cruce'], e['total_productos_toma'],
          e['total_productos_contifico'], e['total_cruzados'], filas_detalle,
          e['total_con_diferencia'], e['valor_total_dif'], usuario))
    version_id = cur.fetchone()['id']

    # Solo lo que DESCUADRA. Un producto que cuadra no aporta nada a la
    # pregunta que responde este historico, y son la mayoria: en Bodega
    # Principal, 65 de 210. El nombre y la categoria tampoco se copian: no
    # cambian entre versiones y ya viven en el catalogo.
    #
    # Que un codigo no aparezca en una version significa que ahi cuadraba, y
    # eso es suficiente para comparar: si estaba antes y ya no esta, se
    # arreglo; si no estaba y aparece, se rompio.
    cur.execute("""
        INSERT INTO goti.cruce_operativo_versiones_detalle
            (version_id, codigo, cantidad_toma, cantidad_sistema, diferencia,
             costo_unitario, valor_diferencia, factor)
        SELECT %s, codigo, cantidad_toma, cantidad_sistema, diferencia,
               costo_unitario, valor_diferencia, factor
        FROM goti.cruce_operativo_detalle
        WHERE ejecucion_id = %s AND COALESCE(diferencia, 0) <> 0
    """, (version_id, ejecucion_id))
    return version


@app.route('/api/cruce-op/solicitar', methods=['POST'])
def cruce_op_solicitar():
    """Llamado desde el panel cuando el usuario presiona CUADRAR.
    Crea una tarea pendiente que el worker tomara.
    Si ya existe completado, solo admin puede re-ejecutar.
    Si ya hay una pendiente o en proceso, devuelve esa misma."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha_toma = data.get('fecha_toma')
    fecha_corte = data.get('fecha_corte_contifico') or fecha_toma  # por defecto = fecha_toma
    usuario = data.get('usuario', 'panel')
    rol = data.get('rol', '')

    if bodega not in BODEGAS_OPERATIVAS:
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha_toma:
        return jsonify({'error': 'fecha_toma requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Ver si ya existe alguna ejecucion para esta bodega+fecha
        cur.execute("""
            SELECT id, estado FROM goti.cruce_operativo_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
            ORDER BY COALESCE(solicitado_at, timestamp_deteccion) DESC LIMIT 1
        """, (bodega, fecha_toma))
        existente = cur.fetchone()

        if existente and existente['estado'] in ('pendiente', 'en_proceso'):
            # Ya se esta procesando, devolver la misma
            return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})

        if existente and existente['estado'] == 'completado':
            # Cualquier usuario puede re-ejecutar un cruce completado.
            # Antes de borrar nada se guarda la version que se va a perder: es
            # el unico momento en que todavia existe.
            _archivar_version_cruce(cur, existente['id'], usuario)
            cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (existente['id'],))
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    fecha_corte_contifico=%s,
                    worker_lock=NULL, error_msg=NULL,
                    timestamp_descarga=NULL, timestamp_cruce=NULL,
                    total_productos_toma=NULL, total_productos_contifico=NULL,
                    total_cruzados=NULL, total_con_diferencia=NULL, valor_total_dif=NULL
                WHERE id = %s
            """, (usuario, fecha_corte, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        if existente:
            # Estado error: cualquiera puede reintentar
            cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (existente['id'],))
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    fecha_corte_contifico=%s,
                    worker_lock=NULL, error_msg=NULL,
                    timestamp_descarga=NULL, timestamp_cruce=NULL,
                    total_productos_toma=NULL, total_productos_contifico=NULL,
                    total_cruzados=NULL, total_con_diferencia=NULL, valor_total_dif=NULL
                WHERE id = %s
            """, (usuario, fecha_corte, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        # No existe: crear nueva
        cur.execute("""
            INSERT INTO goti.cruce_operativo_ejecuciones
            (bodega, fecha_toma, fecha_corte_contifico, estado, solicitado_por, solicitado_at)
            VALUES (%s, %s, %s, 'pendiente', %s, NOW())
            RETURNING id
        """, (bodega, fecha_toma, fecha_corte, usuario))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        print(f"Error en /api/cruce-op/solicitar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': 'Error interno del servidor', 'detalle': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/eliminar/<int:ejec_id>', methods=['DELETE'])
def cruce_op_eliminar(ejec_id):
    """Elimina una ejecucion y su detalle. Llamado desde el panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (ejec_id,))
        cur.execute("DELETE FROM goti.cruce_operativo_ejecuciones WHERE id = %s", (ejec_id,))
        conn.commit()
        return jsonify({'ok': True, 'eliminados': cur.rowcount})
    except Exception as e:
        print(f"Error en /api/cruce-op/eliminar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/cancelar/<int:ejec_id>', methods=['POST'])
def cruce_op_cancelar(ejec_id):
    """Cancela un cruce operativo desde el historial de tareas.

    Mismo criterio que en las cargas a locales: lo pendiente se para limpio y
    lo que el worker ya empezo se marca para que no se reintente, avisando de
    que puede terminar igual -el navegador que lo esta haciendo no se entera
    de que lo cancelaron-.

    Un cruce no crea documentos en Contifico, solo lee y compara, asi que
    cancelarlo a mitad no deja nada a medias alla: como mucho se pierde el
    resultado y hay que volver a pedirlo.
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, estado, bodega, fecha_toma
                       FROM goti.cruce_operativo_ejecuciones WHERE id = %s""", (ejec_id,))
        t = cur.fetchone()
        if not t:
            return jsonify({'error': 'no existe ese cruce'}), 404
        if t['estado'] in ('completado', 'cancelado'):
            return jsonify({'error': f"el cruce ya esta {t['estado']}, "
                                     f"no hay nada que cancelar"}), 409

        era = t['estado']
        cur.execute("""UPDATE goti.cruce_operativo_ejecuciones
                       SET estado = 'cancelado',
                           error_msg = %s,
                           timestamp_cruce = NOW()
                       WHERE id = %s""",
                    (f'cancelado desde el panel (estaba en {era})', ejec_id))
        conn.commit()

        aviso = None
        if era == 'en_proceso':
            aviso = ('El worker ya lo habia empezado. No se reintentara, pero '
                     'puede que termine de descargar antes de enterarse.')
        return jsonify({'ok': True, 'id': ejec_id, 'estado_anterior': era,
                        'aviso': aviso})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/versiones', methods=['GET'])
def cruce_op_versiones():
    """Historico de cruces de una bodega y fecha, del primero al ultimo.

    Incluye la version viva -la que esta ahora en la tabla de ejecuciones- al
    final de la lista, para poder ver el avance completo sin tener que juntar
    dos consultas.
    """
    bodega = request.args.get('bodega')
    fecha = request.args.get('fecha_toma')
    if not bodega or not fecha:
        return jsonify({'error': 'bodega y fecha_toma son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT version, ejecucion_id, solicitado_por, solicitado_at,
                   timestamp_cruce, total_productos_toma, total_cruzados,
                   total_con_diferencia, valor_total_dif, archivado_at
            FROM goti.cruce_operativo_versiones
            WHERE bodega = %s AND fecha_toma = %s
            ORDER BY version
        """, (bodega, fecha))
        filas = [dict(r) for r in cur.fetchall()]

        # La que esta viva ahora todavia no se ha archivado: se anade al final
        # con el numero que le tocaria.
        cur.execute("""
            SELECT id, estado, solicitado_por, solicitado_at, timestamp_cruce,
                   total_productos_toma, total_cruzados, total_con_diferencia,
                   valor_total_dif
            FROM goti.cruce_operativo_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
            ORDER BY id DESC LIMIT 1
        """, (bodega, fecha))
        viva = cur.fetchone()
        if viva and viva['estado'] == 'completado':
            filas.append({
                'version': len(filas) + 1,
                'ejecucion_id': viva['id'],
                'solicitado_por': viva['solicitado_por'],
                'solicitado_at': viva['solicitado_at'],
                'timestamp_cruce': viva['timestamp_cruce'],
                'total_productos_toma': viva['total_productos_toma'],
                'total_cruzados': viva['total_cruzados'],
                'total_con_diferencia': viva['total_con_diferencia'],
                'valor_total_dif': viva['valor_total_dif'],
                'archivado_at': None,
                'actual': True,
            })

        # Cuanto se corrigio de una version a la siguiente. Se calcula aqui y no
        # en el navegador para que el mismo numero salga igual en todas partes.
        anterior = None
        for f in filas:
            if anterior is not None:
                f['gano_productos'] = (anterior.get('total_con_diferencia') or 0) - \
                                      (f.get('total_con_diferencia') or 0)
                f['gano_valor'] = abs(float(anterior.get('valor_total_dif') or 0)) - \
                                  abs(float(f.get('valor_total_dif') or 0))
            anterior = f

        return jsonify({'bodega': bodega, 'fecha_toma': fecha,
                        'versiones': filas, 'total': len(filas)})
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/comparar', methods=['GET'])
def cruce_op_comparar():
    """Que cambio entre dos versiones, producto por producto.

    Separa lo arreglado de lo que se rompio por el camino, que es la parte que
    no se ve mirando solo el total: un descuadre puede bajar de 900 a 400 y
    haber tres productos nuevos descuadrados dentro.
    """
    bodega = request.args.get('bodega')
    fecha = request.args.get('fecha_toma')
    va = request.args.get('desde', type=int)
    vb = request.args.get('hasta', type=int)
    if not bodega or not fecha or va is None or vb is None:
        return jsonify({'error': 'bodega, fecha_toma, desde y hasta son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Los nombres no se guardan por version: no cambian y ocupaban mas que
        # todos los numeros juntos. Se sacan del cruce vivo, que es el que
        # tiene el catalogo al dia.
        cur.execute("""
            SELECT d.codigo, d.nombre
            FROM goti.cruce_operativo_detalle d
            JOIN goti.cruce_operativo_ejecuciones e ON e.id = d.ejecucion_id
            WHERE e.bodega = %s AND e.fecha_toma = %s
        """, (bodega, fecha))
        nombres = {r['codigo']: r['nombre'] for r in cur.fetchall()}

        def detalle_de(version):
            """Lo que descuadraba en una version, venga del historico o de la viva.

            Solo devuelve productos con diferencia. Que un codigo no este
            significa que ahi cuadraba, y de eso se deduce lo que se arreglo.
            """
            cur.execute("""
                SELECT id FROM goti.cruce_operativo_versiones
                WHERE bodega=%s AND fecha_toma=%s AND version=%s
            """, (bodega, fecha, version))
            v = cur.fetchone()
            if v:
                cur.execute("""
                    SELECT codigo, cantidad_toma, cantidad_sistema,
                           diferencia, valor_diferencia
                    FROM goti.cruce_operativo_versiones_detalle
                    WHERE version_id = %s
                """, (v['id'],))
                return {r['codigo']: dict(r) for r in cur.fetchall()}
            # No esta archivada: puede ser la viva, que si guarda todo el
            # detalle, asi que aqui si hay que filtrar lo que cuadra.
            cur.execute("""
                SELECT id FROM goti.cruce_operativo_ejecuciones
                WHERE bodega=%s AND fecha_toma=%s AND estado='completado'
                ORDER BY id DESC LIMIT 1
            """, (bodega, fecha))
            e = cur.fetchone()
            if not e:
                return {}
            cur.execute("""
                SELECT codigo, cantidad_toma, cantidad_sistema,
                       diferencia, valor_diferencia
                FROM goti.cruce_operativo_detalle
                WHERE ejecucion_id = %s AND COALESCE(diferencia, 0) <> 0
            """, (e['id'],))
            return {r['codigo']: dict(r) for r in cur.fetchall()}

        a, b = detalle_de(va), detalle_de(vb)
        if not a or not b:
            return jsonify({'error': 'alguna de las dos versiones no existe'}), 404

        arreglados, nuevos, siguen = [], [], []
        for cod in set(list(a.keys()) + list(b.keys())):
            da = float((a.get(cod) or {}).get('diferencia') or 0)
            db = float((b.get(cod) or {}).get('diferencia') or 0)
            fila = {
                'codigo': cod,
                'nombre': nombres.get(cod, cod),
                'diferencia_antes': da,
                'diferencia_ahora': db,
                'valor_antes': float((a.get(cod) or {}).get('valor_diferencia') or 0),
                'valor_ahora': float((b.get(cod) or {}).get('valor_diferencia') or 0),
            }
            if da != 0 and db == 0:
                arreglados.append(fila)
            elif da == 0 and db != 0:
                nuevos.append(fila)
            elif da != 0 and db != 0 and da != db:
                siguen.append(fila)
            elif da != 0 and da == db:
                siguen.append(fila)

        orden = lambda f: -abs(f['valor_antes'] or f['valor_ahora'])
        return jsonify({
            'bodega': bodega, 'fecha_toma': fecha, 'desde': va, 'hasta': vb,
            'arreglados': sorted(arreglados, key=orden),
            'nuevos': sorted(nuevos, key=lambda f: -abs(f['valor_ahora'])),
            'siguen_descuadrados': sorted(siguen, key=orden),
            'resumen': {
                'arreglados': len(arreglados),
                'nuevos': len(nuevos),
                'siguen': len(siguen),
            },
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/pendientes', methods=['GET'])
def cruce_op_pendientes():
    """Llamado por el worker. Devuelve tareas pendientes y las marca como en_proceso."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    worker_id = request.args.get('worker_id', 'pc-finanzas')
    if not worker_autorizado(worker_id):
        # Silencio en vez de error: el worker viejo reintentaria en bucle y
        # llenaria el log. Con la lista vacia simplemente no hace nada.
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Marca atomicamente las pendientes como en_proceso para este worker
        cur.execute("""
            UPDATE goti.cruce_operativo_ejecuciones
            SET estado = 'en_proceso',
                worker_lock = %s,
                timestamp_descarga = NOW()
            WHERE id IN (
                SELECT id FROM goti.cruce_operativo_ejecuciones
                WHERE estado = 'pendiente'
                ORDER BY solicitado_at ASC
                LIMIT 5
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha_toma, fecha_corte_contifico, solicitado_por, solicitado_at
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        result = [{
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'fecha_corte_contifico': r['fecha_corte_contifico'].isoformat() if r['fecha_corte_contifico'] else (r['fecha_toma'].isoformat() if r['fecha_toma'] else None),
            'solicitado_por': r['solicitado_por'],
        } for r in rows]
        return jsonify(result)
    except Exception as e:
        print(f"Error en /api/cruce-op/pendientes: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/resultado', methods=['POST'])
def cruce_op_resultado():
    """Llamado por el worker al terminar. Inserta detalle y marca completado/error."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')  # 'completado' o 'error'
    error_msg = data.get('error_msg')
    detalle = data.get('detalle', [])
    resumen = data.get('resumen', {})

    if not ejec_id:
        return jsonify({'error': 'id requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        if estado == 'error':
            cur.execute("""
                UPDATE goti.cruce_operativo_ejecuciones
                SET estado = 'error', error_msg = %s, timestamp_cruce = NOW()
                WHERE id = %s
            """, (error_msg, ejec_id))
            conn.commit()
            return jsonify({'ok': True})

        # Borrar detalle previo si existiera
        cur.execute("DELETE FROM goti.cruce_operativo_detalle WHERE ejecucion_id = %s", (ejec_id,))

        # Insertar detalle (batch insert para evitar timeout)
        if detalle:
            valores = []
            for d in detalle:
                valores.append((
                    ejec_id, d.get('codigo'), d.get('nombre'), d.get('categoria'),
                    d.get('unidad_destino'), d.get('unidad_toma'), d.get('factor'),
                    d.get('unidad_destino'), d.get('cantidad_toma'), d.get('cantidad_sistema'),
                    d.get('diferencia'), d.get('costo_unitario'), d.get('valor_diferencia'),
                    d.get('tipo_abc'), d.get('origen', 'cruce_operativo')
                ))
            cur.executemany("""
                INSERT INTO goti.cruce_operativo_detalle
                (ejecucion_id, codigo, nombre, categoria, unidad, unidad_toma, factor,
                 unidad_destino, cantidad_toma, cantidad_sistema, diferencia,
                 costo_unitario, valor_diferencia, tipo_abc, origen)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, valores)

        # Update ejecucion
        cur.execute("""
            UPDATE goti.cruce_operativo_ejecuciones
            SET estado = 'completado',
                total_productos_toma = %s,
                total_productos_contifico = %s,
                total_cruzados = %s,
                total_con_diferencia = %s,
                valor_total_dif = %s,
                timestamp_cruce = NOW()
            WHERE id = %s
        """, (
            resumen.get('total_productos_toma'),
            resumen.get('total_productos_contifico'),
            resumen.get('total_cruzados'),
            resumen.get('total_con_diferencia'),
            resumen.get('valor_total_dif'),
            ejec_id
        ))
        conn.commit()
        return jsonify({'ok': True, 'detalles_insertados': len(detalle)})
    except Exception as e:
        print(f"Error en /api/cruce-op/resultado: {e}")
        if conn:
            conn.rollback()
        return jsonify({'error': 'Error interno del servidor', 'detalle': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/estado/<int:ejec_id>', methods=['GET'])
def cruce_op_estado(ejec_id):
    """Polling desde el panel para saber estado de una ejecucion."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, bodega, fecha_toma, estado, solicitado_por, solicitado_at,
                   timestamp_descarga, timestamp_cruce, error_msg,
                   total_productos_toma, total_productos_contifico, total_cruzados,
                   total_con_diferencia, valor_total_dif
            FROM goti.cruce_operativo_ejecuciones WHERE id = %s
        """, (ejec_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'estado': r['estado'],
            'solicitado_por': r['solicitado_por'],
            'solicitado_at': r['solicitado_at'].isoformat() if r['solicitado_at'] else None,
            'timestamp_descarga': r['timestamp_descarga'].isoformat() if r['timestamp_descarga'] else None,
            'timestamp_cruce': r['timestamp_cruce'].isoformat() if r['timestamp_cruce'] else None,
            'error_msg': r['error_msg'],
            'total_productos_toma': r['total_productos_toma'],
            'total_productos_contifico': r['total_productos_contifico'],
            'total_cruzados': r['total_cruzados'],
            'total_con_diferencia': r['total_con_diferencia'],
            'valor_total_dif': float(r['valor_total_dif']) if r['valor_total_dif'] is not None else None,
        })
    except Exception as e:
        print(f"Error en /api/cruce-op/estado: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/cruce-op/fechas-disponibles', methods=['GET'])
def cruce_op_fechas():
    """Devuelve las fechas con toma fisica disponibles para una bodega."""
    bodega = request.args.get('bodega')
    tablas_centrales = {
        'bodega_principal': 'public.toma_bodega',
        'materia_prima':    'public.toma_materiaprima',
        'planta':           'public.toma_planta',
    }
    if bodega not in BODEGAS_OPERATIVAS:
        return jsonify({'error': 'bodega invalida'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if bodega in tablas_centrales:
            cur.execute(f"""
                SELECT fecha, COUNT(*) AS productos
                FROM {tablas_centrales[bodega]}
                WHERE fecha IS NOT NULL
                GROUP BY fecha ORDER BY fecha DESC
            """)
        else:
            cur.execute("""
                SELECT fecha, COUNT(*) AS productos
                FROM goti.inventario_ciego_conteos
                WHERE local = %s AND fecha IS NOT NULL
                GROUP BY fecha ORDER BY fecha DESC
            """, (bodega,))
        rows = cur.fetchall()
        return jsonify([{
            'fecha': r['fecha'].isoformat(),
            'productos': r['productos']
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/cruce-op/fechas-disponibles: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if conn:
            release_db(conn)


# ============================================================
# MODULO: Carga Toma Fisica a Contifico
# ============================================================
# Flujo: Panel web -> POST /solicitar -> tarea pendiente
#        Worker PC FINANZAS -> GET /pendientes-carga (cada 15s) -> toma tarea
#        Worker abre Contifico, llena formulario toma fisica -> POST /resultado-carga
#        Panel web -> GET /estado-carga/<id> (polling) -> muestra resultado
#        UNIQUE(bodega, fecha_toma) -> no permite cargar dos veces

@app.route('/api/carga-contifico/fechas-con-cruce', methods=['GET'])
def carga_contifico_fechas_con_cruce():
    """Devuelve las fechas que tienen cruce operativo completado para una bodega."""
    bodega = request.args.get('bodega')
    if bodega not in BODEGAS_OPERATIVAS:
        return jsonify({'error': 'bodega invalida'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT fecha_toma, total_cruzados, total_con_diferencia, valor_total_dif
            FROM goti.cruce_operativo_ejecuciones
            WHERE bodega = %s AND estado = 'completado'
            ORDER BY fecha_toma DESC
        """, (bodega,))
        rows = cur.fetchall()
        return jsonify([{
            'fecha': r['fecha_toma'].isoformat(),
            'cruzados': r['total_cruzados'],
            'con_dif': r['total_con_diferencia'],
            'valor_dif': float(r['valor_total_dif']) if r['valor_total_dif'] else 0,
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/carga-contifico/fechas-con-cruce: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/verificar', methods=['GET'])
def carga_contifico_verificar():
    """Verifica si ya existe una carga completada para bodega+fecha."""
    bodega = request.args.get('bodega')
    fecha = request.args.get('fecha')
    if not bodega or not fecha:
        return jsonify({'error': 'bodega y fecha requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, estado, solicitado_at, timestamp_fin, total_productos, productos_ok, productos_error
            FROM goti.carga_contifico_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
        """, (bodega, fecha))
        row = cur.fetchone()
        if not row:
            return jsonify({'existe': False, 'cargado': False})
        return jsonify({
            'existe': True,
            'cargado': row['estado'] == 'completado',
            'estado': row['estado'],
            'id': row['id'],
            'solicitado_at': row['solicitado_at'].isoformat() if row['solicitado_at'] else None,
            'timestamp_fin': row['timestamp_fin'].isoformat() if row['timestamp_fin'] else None,
            'total_productos': row['total_productos'],
            'productos_ok': row['productos_ok'],
            'productos_error': row['productos_error'],
        })
    except Exception as e:
        print(f"Error en /api/carga-contifico/verificar: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/solicitar', methods=['POST'])
def carga_contifico_solicitar():
    """Crea tarea de carga. Si ya esta completada, solo admin puede re-ejecutar."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha_toma = data.get('fecha_toma')
    usuario = data.get('usuario', 'panel')
    rol = data.get('rol', '')

    if bodega not in BODEGAS_OPERATIVAS:
        return jsonify({'error': 'bodega invalida'}), 400
    if not fecha_toma:
        return jsonify({'error': 'fecha_toma requerida'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, estado FROM goti.carga_contifico_ejecuciones
            WHERE bodega = %s AND fecha_toma = %s
        """, (bodega, fecha_toma))
        existente = cur.fetchone()

        if existente:
            if existente['estado'] == 'completado':
                if rol != 'admin':
                    return jsonify({'error': 'Ya fue cargado a Contifico. Solo el administrador puede re-ejecutar.', 'ya_cargado': True}), 409
                # Admin: resetear para re-ejecutar
                cur.execute("""
                    UPDATE goti.carga_contifico_ejecuciones
                    SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                        worker_lock=NULL, error_msg=NULL, timestamp_inicio=NULL, timestamp_fin=NULL,
                        total_productos=NULL, productos_ok=NULL, productos_error=NULL, productos_error_lista=NULL,
                        productos_sin_contar=NULL, total_en_toma=NULL
                    WHERE id = %s
                """, (usuario, existente['id']))
                conn.commit()
                return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})
            if existente['estado'] in ('pendiente', 'en_proceso'):
                return jsonify({'id': existente['id'], 'estado': existente['estado'], 'reused': True})
            # Estado error: resetear para reintentar
            cur.execute("""
                UPDATE goti.carga_contifico_ejecuciones
                SET estado='pendiente', solicitado_por=%s, solicitado_at=NOW(),
                    worker_lock=NULL, error_msg=NULL, timestamp_inicio=NULL, timestamp_fin=NULL,
                    total_productos=NULL, productos_ok=NULL, productos_error=NULL, productos_error_lista=NULL,
                        productos_sin_contar=NULL, total_en_toma=NULL
                WHERE id = %s
            """, (usuario, existente['id']))
            conn.commit()
            return jsonify({'id': existente['id'], 'estado': 'pendiente', 'reset': True})

        # No existe: crear nueva
        cur.execute("""
            INSERT INTO goti.carga_contifico_ejecuciones
            (bodega, fecha_toma, estado, solicitado_por, solicitado_at)
            VALUES (%s, %s, 'pendiente', %s, NOW())
            RETURNING id
        """, (bodega, fecha_toma, usuario))
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'id': new_id, 'estado': 'pendiente'})
    except Exception as e:
        print(f"Error en /api/carga-contifico/solicitar: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/cancelar/<int:ejec_id>', methods=['POST'])
def carga_contifico_cancelar(ejec_id):
    """Cancela una carga de ajuste a Contifico desde el historial de tareas.

    Esta es la que mas falta hacia poder parar: es la mas larga de las tres
    colas y, mientras corre, el worker no atiende nada mas.

    Si ya estaba en marcha se avisa, porque el navegador que la esta haciendo
    no se entera de la cancelacion: puede acabar creando el documento igual.
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, estado, bodega, fecha_toma
                       FROM goti.carga_contifico_ejecuciones WHERE id = %s""", (ejec_id,))
        t = cur.fetchone()
        if not t:
            return jsonify({'error': 'no existe esa carga'}), 404
        if t['estado'] in ('completado', 'cancelado'):
            return jsonify({'error': f"la carga ya esta {t['estado']}, "
                                     f"no hay nada que cancelar"}), 409

        era = t['estado']
        cur.execute("""UPDATE goti.carga_contifico_ejecuciones
                       SET estado = 'cancelado',
                           error_msg = %s,
                           timestamp_fin = NOW()
                       WHERE id = %s""",
                    (f'cancelada desde el panel (estaba en {era})', ejec_id))
        conn.commit()

        aviso = None
        if era == 'en_proceso':
            aviso = ('El worker ya la habia empezado. No se reintentara, pero '
                     'revisa en Contifico por si alcanzo a crear el documento.')
        return jsonify({'ok': True, 'id': ejec_id, 'estado_anterior': era,
                        'aviso': aviso})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/carga-contifico/pendientes', methods=['GET'])
def carga_contifico_pendientes():
    """Llamado por el worker. Devuelve tareas pendientes de carga y las marca en_proceso."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    worker_id = request.args.get('worker_id', 'pc-finanzas')
    if not worker_autorizado(worker_id):
        # Silencio en vez de error: el worker viejo reintentaria en bucle y
        # llenaria el log. Con la lista vacia simplemente no hace nada.
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_contifico_ejecuciones
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.carga_contifico_ejecuciones
                WHERE estado = 'pendiente'
                ORDER BY solicitado_at ASC
                LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha_toma, solicitado_por
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'tipo': 'carga_contifico',
        } for r in rows])
    except Exception as e:
        print(f"Error en /api/carga-contifico/pendientes: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/resultado', methods=['POST'])
def carga_contifico_resultado():
    """Llamado por el worker al terminar la carga."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401

    data = request.json or {}
    ejec_id = data.get('id')
    estado = data.get('estado', 'completado')
    error_msg = data.get('error_msg')

    if not ejec_id:
        return jsonify({'error': 'id requerido'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.carga_contifico_ejecuciones
            SET estado = %s, timestamp_fin = NOW(),
                total_productos = %s, productos_ok = %s,
                productos_error = %s, productos_error_lista = %s,
                productos_sin_contar = %s, total_en_toma = %s,
                error_msg = %s
            WHERE id = %s
        """, (
            estado,
            data.get('total_productos'),
            data.get('productos_ok'),
            data.get('productos_error'),
            data.get('productos_error_lista'),
            data.get('productos_sin_contar'),
            data.get('total_en_toma'),
            error_msg,
            ejec_id
        ))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        print(f"Error en /api/carga-contifico/resultado: {e}")
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/carga-contifico/estado/<int:ejec_id>', methods=['GET'])
def carga_contifico_estado(ejec_id):
    """Polling desde el panel."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, bodega, fecha_toma, estado, solicitado_at,
                   timestamp_inicio, timestamp_fin, error_msg,
                   total_productos, productos_ok, productos_error, productos_error_lista,
                   productos_sin_contar, total_en_toma
            FROM goti.carga_contifico_ejecuciones WHERE id = %s
        """, (ejec_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha_toma': r['fecha_toma'].isoformat() if r['fecha_toma'] else None,
            'estado': r['estado'],
            'solicitado_at': r['solicitado_at'].isoformat() if r['solicitado_at'] else None,
            'timestamp_inicio': r['timestamp_inicio'].isoformat() if r['timestamp_inicio'] else None,
            'timestamp_fin': r['timestamp_fin'].isoformat() if r['timestamp_fin'] else None,
            'error_msg': r['error_msg'],
            'total_productos': r['total_productos'],
            'productos_ok': r['productos_ok'],
            'productos_error': r['productos_error'],
            'productos_error_lista': r['productos_error_lista'],
            'productos_sin_contar': r['productos_sin_contar'],
            'total_en_toma': r['total_en_toma'],
        })
    except Exception as e:
        print(f"Error en /api/carga-contifico/estado: {e}")
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# ============================================================
# MODULO: Evaluacion Semanal por Local
# ============================================================

EVAL_LOCALES = [
    {'id': 'real_audiencia', 'nombre': 'Chios Real Audiencia'},
    {'id': 'floreana', 'nombre': 'Chios Floreana'},
    {'id': 'portugal', 'nombre': 'Chios Portugal'},
    {'id': 'santo_cachon_real', 'nombre': 'Santo Cachon Real'},
    {'id': 'santo_cachon_portugal', 'nombre': 'Santo Cachon Portugal'},
    {'id': 'simon_bolon', 'nombre': 'Simon Bolon'},
]

@app.route('/api/eval/categorias', methods=['GET'])
def eval_categorias():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, nombre, descripcion, orden, criterios FROM goti.eval_categorias WHERE activa = TRUE ORDER BY orden")
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/locales', methods=['GET'])
def eval_locales():
    return jsonify(EVAL_LOCALES)


@app.route('/api/eval/guardar', methods=['POST'])
def eval_guardar():
    """Guarda evaluacion semanal. Body: {local, semana_inicio, semana_fin, evaluaciones: [{categoria_id, puntaje, comentario}], evaluado_por}"""
    data = request.json or {}
    local = data.get('local')
    semana_inicio = data.get('semana_inicio')
    semana_fin = data.get('semana_fin')
    evaluaciones = data.get('evaluaciones', [])
    evaluado_por = data.get('evaluado_por', 'admin')

    if not local or not semana_inicio or not evaluaciones:
        return jsonify({'error': 'local, semana_inicio y evaluaciones requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        guardados = 0
        for ev in evaluaciones:
            cur.execute("""
                INSERT INTO goti.eval_semanal
                (local, semana_inicio, semana_fin, categoria_id, puntaje, comentario, evaluado_por, evaluado_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
                ON CONFLICT (local, semana_inicio, categoria_id)
                DO UPDATE SET puntaje = EXCLUDED.puntaje, comentario = EXCLUDED.comentario,
                              evaluado_por = EXCLUDED.evaluado_por, evaluado_at = NOW()
            """, (local, semana_inicio, semana_fin, ev['categoria_id'], ev['puntaje'], ev.get('comentario', ''), evaluado_por))
            guardados += 1
        conn.commit()
        return jsonify({'ok': True, 'guardados': guardados})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/semana', methods=['GET'])
def eval_semana():
    """Obtiene evaluaciones de una semana. Params: semana_inicio, local (opcional)"""
    semana = request.args.get('semana_inicio')
    local = request.args.get('local')
    if not semana:
        return jsonify({'error': 'semana_inicio requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if local:
            cur.execute("""
                SELECT e.id, e.local, e.categoria_id, c.nombre as categoria, e.puntaje, e.comentario, e.evaluado_por, e.evaluado_at
                FROM goti.eval_semanal e
                JOIN goti.eval_categorias c ON c.id = e.categoria_id
                WHERE e.semana_inicio = %s AND e.local = %s
                ORDER BY c.orden
            """, (semana, local))
        else:
            cur.execute("""
                SELECT e.id, e.local, e.categoria_id, c.nombre as categoria, e.puntaje, e.comentario, e.evaluado_por, e.evaluado_at
                FROM goti.eval_semanal e
                JOIN goti.eval_categorias c ON c.id = e.categoria_id
                WHERE e.semana_inicio = %s
                ORDER BY e.local, c.orden
            """, (semana,))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/ranking', methods=['GET'])
def eval_ranking():
    """Ranking de locales por promedio. Params: semana_inicio (opcional, default ultima disponible), ultimas_n (semanas a promediar, default 1)"""
    semana = request.args.get('semana_inicio')
    ultimas_n = int(request.args.get('ultimas_n', '1'))
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if semana:
            cur.execute("""
                SELECT local, ROUND(AVG(puntaje)::numeric, 2) as promedio, COUNT(DISTINCT categoria_id) as categorias_evaluadas
                FROM goti.eval_semanal
                WHERE semana_inicio = %s
                GROUP BY local ORDER BY promedio DESC
            """, (semana,))
        else:
            cur.execute("""
                SELECT local, ROUND(AVG(puntaje)::numeric, 2) as promedio, COUNT(DISTINCT semana_inicio) as semanas
                FROM goti.eval_semanal
                WHERE semana_inicio >= (
                    SELECT MAX(semana_inicio) - interval '%s weeks' FROM goti.eval_semanal
                )
                GROUP BY local ORDER BY promedio DESC
            """ % max(1, ultimas_n))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/tendencia', methods=['GET'])
def eval_tendencia():
    """Tendencia historica por local. Params: local (opcional), limite (semanas, default 12)"""
    local = request.args.get('local')
    limite = int(request.args.get('limite', '12'))
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if local:
            cur.execute("""
                SELECT semana_inicio, ROUND(AVG(puntaje)::numeric, 2) as promedio
                FROM goti.eval_semanal
                WHERE local = %s
                GROUP BY semana_inicio ORDER BY semana_inicio DESC LIMIT %s
            """, (local, limite))
        else:
            cur.execute("""
                SELECT local, semana_inicio, ROUND(AVG(puntaje)::numeric, 2) as promedio
                FROM goti.eval_semanal
                GROUP BY local, semana_inicio ORDER BY semana_inicio DESC, promedio DESC
                LIMIT %s
            """, (limite * 6,))
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/eval/semanas-disponibles', methods=['GET'])
def eval_semanas_disponibles():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT DISTINCT semana_inicio, semana_fin
            FROM goti.eval_semanal
            ORDER BY semana_inicio DESC LIMIT 52
        """)
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/evaluacion')
def evaluacion_page():
    return render_template('evaluacion.html')


# ============================================================
# MODULO: DEPOSITOS (lee desde AirTable)
# ============================================================
_AT_DEP_FB = _b64.b64decode('cGF0d1owSHBiRlQ5RkNoNWQuZDQwY2ExZTZlNGViYWRlZWE5ZjJmZGYyZTAwM2FhOGMxMGIyMjAzYzkxZjg2OTk1YmRiOTgyMjYwOTkzMzM3YQ==').decode()
AIRTABLE_DEPOSITOS_TOKEN = os.environ.get('AIRTABLE_DEPOSITOS_TOKEN', '') or _AT_DEP_FB
AIRTABLE_DEPOSITOS_BASE = 'apppZXgUChlBLbVpR'
AIRTABLE_DEPOSITOS_TABLE = 'tbldo5QTH6bBpgYbx'
AIRTABLE_TIENDAS_TABLE = 'tblxloBdnbdsGcuKR'

_tiendas_cache = {}
_tiendas_cache_ts = 0

def _at_headers():
    return {'Authorization': f'Bearer {AIRTABLE_DEPOSITOS_TOKEN}'}

def _cargar_tiendas():
    global _tiendas_cache, _tiendas_cache_ts
    import time as _t
    if _tiendas_cache and (_t.time() - _tiendas_cache_ts) < 600:
        return _tiendas_cache
    try:
        import requests as req
        r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_TIENDAS_TABLE}',
            headers=_at_headers(), params={'fields[]': ['Código', 'Marca']}, timeout=15)
        if r.status_code == 200:
            for rec in r.json().get('records', []):
                _tiendas_cache[rec['id']] = rec['fields'].get('Código', rec['id'])
            _tiendas_cache_ts = _t.time()
    except Exception as e:
        print(f'Error cargando tiendas: {e}')
    return _tiendas_cache

def _resolver_local(local_ids):
    if not local_ids:
        return 'Sin local'
    tiendas = _cargar_tiendas()
    nombres = [tiendas.get(lid, lid) for lid in local_ids]
    return ', '.join(nombres)


@app.route('/api/depositos/listar', methods=['GET'])
def depositos_listar():
    """Lista depositos desde AirTable con filtros."""
    import requests as req
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    estado = request.args.get('estado', '')
    cuadre = request.args.get('cuadre', '')

    try:
        # Construir formula de filtro
        filtros = []
        if fecha_desde:
            filtros.append(f"IS_AFTER({{Fecha}}, '{fecha_desde}')")
        if fecha_hasta:
            filtros.append(f"IS_BEFORE({{Fecha}}, DATEADD('{fecha_hasta}', 1, 'day'))")
        if estado:
            filtros.append(f"{{Estado}} = '{estado}'")
        if cuadre:
            filtros.append(f"{{Estado De Cuadre}} = '{cuadre}'")

        params = {
            'pageSize': 100,
            'sort[0][field]': 'Fecha',
            'sort[0][direction]': 'desc',
            'fields[]': ['Fecha', 'Local', 'Responsable De Caja', 'Monto Contado',
                         'Monto A Recibir', 'Diferencia Contado Vs. Recibido',
                         'Secuencia De Caja', 'Número De Depósitos', 'Estado',
                         'Estado De Cuadre', 'Observación', 'Evidencia', 'Evidencia Del Déposito',
                         'Fecha Creación', 'Correo (from Responsable De Caja)'],
        }
        if filtros:
            params['filterByFormula'] = 'AND(' + ','.join(filtros) + ')'

        all_records = []
        offset = None
        while True:
            if offset:
                params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}',
                headers=_at_headers(), params=params, timeout=20)
            if r.status_code != 200:
                return jsonify({'error': f'AirTable error: {r.status_code}'}), 500
            data = r.json()
            all_records.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset or len(all_records) >= 500:
                break

        # Resolver locales
        resultado = []
        for rec in all_records:
            f = rec['fields']
            evidencias = []
            for att in (f.get('Evidencia', []) + f.get('Evidencia Del Déposito', [])):
                if isinstance(att, dict):
                    thumb = att.get('thumbnails', {}).get('large', {}).get('url', '')
                    evidencias.append({'url': att.get('url', ''), 'thumb': thumb, 'filename': att.get('filename', '')})

            resultado.append({
                'id': rec['id'],
                'fecha': f.get('Fecha'),
                'local': _resolver_local(f.get('Local', [])),
                'monto_contado': f.get('Monto Contado', 0),
                'monto_recibir': f.get('Monto A Recibir', 0),
                'diferencia': f.get('Diferencia Contado Vs. Recibido', 0),
                'secuencia': f.get('Secuencia De Caja'),
                'num_depositos': f.get('Número De Depósitos'),
                'estado': f.get('Estado', ''),
                'cuadre': f.get('Estado De Cuadre', ''),
                'observacion': f.get('Observación', ''),
                'evidencias': evidencias,
                'fecha_creacion': f.get('Fecha Creación'),
                'responsable_email': (f.get('Correo (from Responsable De Caja)', [None]) or [None])[0],
            })

        return jsonify({'depositos': resultado, 'total': len(resultado)})
    except Exception as e:
        print(f'Error en depositos_listar: {e}')
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/resumen', methods=['GET'])
def depositos_resumen():
    """Resumen/KPIs de depositos."""
    import requests as req
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    try:
        filtros = []
        if fecha_desde:
            filtros.append(f"IS_AFTER({{Fecha}}, '{fecha_desde}')")
        if fecha_hasta:
            filtros.append(f"IS_BEFORE({{Fecha}}, DATEADD('{fecha_hasta}', 1, 'day'))")

        params = {
            'pageSize': 100,
            'fields[]': ['Fecha', 'Local', 'Monto Contado', 'Monto A Recibir',
                         'Diferencia Contado Vs. Recibido', 'Estado', 'Estado De Cuadre'],
        }
        if filtros:
            params['filterByFormula'] = 'AND(' + ','.join(filtros) + ')'

        all_records = []
        offset = None
        while True:
            if offset:
                params['offset'] = offset
            r = req.get(f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}',
                headers=_at_headers(), params=params, timeout=20)
            if r.status_code != 200:
                break
            data = r.json()
            all_records.extend(data.get('records', []))
            offset = data.get('offset')
            if not offset:
                break

        total_depositado = 0
        total_recibido = 0
        total_diferencia = 0
        descuadres = 0
        cuadran = 0
        por_local = {}
        pendientes = 0

        for rec in all_records:
            f = rec['fields']
            monto = f.get('Monto Contado', 0) or 0
            recibido = f.get('Monto A Recibir', 0) or 0
            dif = f.get('Diferencia Contado Vs. Recibido', 0) or 0
            total_depositado += monto
            total_recibido += recibido
            total_diferencia += abs(dif)

            if f.get('Estado De Cuadre') == 'Descuadra':
                descuadres += 1
            elif f.get('Estado De Cuadre') == 'Cuadra':
                cuadran += 1

            if f.get('Estado') not in ('Aprobado por Contabilidad',):
                pendientes += 1

            local = _resolver_local(f.get('Local', []))
            if local not in por_local:
                por_local[local] = {'monto': 0, 'depositos': 0, 'descuadres': 0}
            por_local[local]['monto'] += monto
            por_local[local]['depositos'] += 1
            if f.get('Estado De Cuadre') == 'Descuadra':
                por_local[local]['descuadres'] += 1

        return jsonify({
            'total_depositos': len(all_records),
            'total_depositado': round(total_depositado, 2),
            'total_recibido': round(total_recibido, 2),
            'total_diferencia': round(total_diferencia, 2),
            'cuadran': cuadran,
            'descuadres': descuadres,
            'pendientes': pendientes,
            'por_local': por_local,
        })
    except Exception as e:
        print(f'Error en depositos_resumen: {e}')
        return jsonify({'error': str(e)[:200]}), 500


@app.route('/api/depositos/aprobar', methods=['POST'])
def depositos_aprobar():
    """Aprueba un deposito en AirTable."""
    import requests as req
    data = request.json or {}
    record_id = data.get('id')
    if not record_id:
        return jsonify({'error': 'id requerido'}), 400

    try:
        r = req.patch(
            f'https://api.airtable.com/v0/{AIRTABLE_DEPOSITOS_BASE}/{AIRTABLE_DEPOSITOS_TABLE}/{record_id}',
            headers={**_at_headers(), 'Content-Type': 'application/json'},
            json={'fields': {
                'Estado': 'Aprobado por Contabilidad',
                'Fecha Aprobado Por Contabilidad': datetime.now().isoformat(),
            }},
            timeout=15,
        )
        if r.status_code == 200:
            return jsonify({'ok': True})
        return jsonify({'error': f'AirTable: {r.status_code}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500


# ==================== ADMIN USUARIOS ====================

SMTP_CONFIG = {
    'server': os.environ.get('SMTP_SERVER', 'smtp.gmail.com'),
    'port': int(os.environ.get('SMTP_PORT', '587')),
    'user': os.environ.get('SMTP_USER', 'ortiz.medranda@gmail.com'),
    'password': os.environ.get('SMTP_PASSWORD', 'cikp vxlq zlim dzzc'),
}
APP_URL = os.environ.get('APP_URL', 'https://inventario-ciego-5bdr.onrender.com')


def _enviar_email_invitacion(email_destino, nombre, username, token):
    """Envia email con link para establecer contrasena."""
    link = f"{APP_URL}/establecer-clave?token={token}"
    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:520px;margin:0 auto;padding:30px;background:#f8fafc;border-radius:12px;">
        <div style="text-align:center;margin-bottom:24px;">
            <h1 style="color:#123450;font-size:22px;margin:0;">FOODIX - Inventario</h1>
        </div>
        <div style="background:#fff;padding:28px;border-radius:10px;border:1px solid #e2e8f0;">
            <h2 style="color:#123450;font-size:18px;margin:0 0 12px;">Hola {nombre},</h2>
            <p style="color:#475569;font-size:14px;line-height:1.6;">
                Se ha creado tu cuenta en el sistema de inventario.
                Tu usuario es: <strong style="color:#123450;">{username}</strong>
            </p>
            <p style="color:#475569;font-size:14px;line-height:1.6;">
                Haz clic en el boton para establecer tu contrasena:
            </p>
            <div style="text-align:center;margin:24px 0;">
                <a href="{link}" style="background:#123450;color:#fff;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:600;font-size:14px;display:inline-block;">
                    Crear mi contrasena
                </a>
            </div>
            <p style="color:#94a3b8;font-size:12px;line-height:1.5;">
                Este enlace es valido por 48 horas. Si no puedes hacer clic, copia y pega esta URL en tu navegador:<br>
                <span style="color:#64748b;word-break:break-all;">{link}</span>
            </p>
        </div>
        <p style="color:#94a3b8;font-size:11px;text-align:center;margin-top:16px;">
            FOODIX S.A.S. — Sistema de Inventario Ciego
        </p>
    </div>
    """
    msg = MIMEMultipart('alternative')
    msg['Subject'] = f'FOODIX Inventario — Configura tu acceso'
    msg['From'] = f'FOODIX Inventario <{SMTP_CONFIG["user"]}>'
    msg['To'] = email_destino
    msg.attach(MIMEText(html, 'html'))

    server = smtplib.SMTP(SMTP_CONFIG['server'], SMTP_CONFIG['port'], timeout=15)
    server.starttls()
    server.login(SMTP_CONFIG['user'], SMTP_CONFIG['password'])
    server.sendmail(SMTP_CONFIG['user'], email_destino, msg.as_string())
    server.quit()


def _require_admin(data):
    """Valida que quien llama sea admin (username + password en el body)."""
    if not data:
        return None, jsonify({'error': 'Sin datos'}), 400
    admin_user = data.get('admin_user', '')
    admin_pass = data.get('admin_pass', '')
    if not admin_user or not admin_pass:
        return None, jsonify({'error': 'Credenciales de admin requeridas'}), 401
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""SELECT id FROM goti.usuarios
                       WHERE username = %s AND password = %s AND rol = 'admin' AND activo = TRUE""",
                    (admin_user, admin_pass))
        row = cur.fetchone()
        if not row:
            return None, jsonify({'error': 'No autorizado'}), 403
        return conn, None, None
    except Exception:
        release_db(conn)
        raise


@app.route('/api/admin/personas', methods=['GET'])
def admin_listar_personas():
    """Devuelve lista de personas activas desde AirTable con nombre y correo."""
    try:
        personas = _obtener_personas_con_correo()
        return jsonify(personas)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/admin/usuarios', methods=['GET'])
def admin_listar_usuarios():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT u.id, u.username, u.nombre, u.rol, u.activo, u.created_at, u.email,
                   COALESCE(array_agg(ub.bodega ORDER BY ub.bodega) FILTER (WHERE ub.bodega IS NOT NULL), '{}') AS bodegas
            FROM goti.usuarios u
            LEFT JOIN goti.usuario_bodegas ub ON ub.usuario_id = u.id
            GROUP BY u.id, u.username, u.nombre, u.rol, u.activo, u.created_at, u.email
            ORDER BY u.id
        """)
        usuarios = cur.fetchall()
        return jsonify(usuarios)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/usuarios', methods=['POST'])
def admin_crear_usuario():
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        username = data.get('username', '').strip().lower()
        nombre = data.get('nombre', '').strip()
        password = data.get('password', '').strip()
        email = data.get('email', '').strip().lower()
        rol = data.get('rol', 'subgerente')
        bodegas = data.get('bodegas', [])
        enviar_invitacion = data.get('enviar_invitacion', False)

        if not username or not nombre:
            return jsonify({'error': 'username y nombre son obligatorios'}), 400
        if enviar_invitacion and not email:
            return jsonify({'error': 'Email es obligatorio para enviar invitacion'}), 400
        if not enviar_invitacion and not password:
            return jsonify({'error': 'Debes asignar contrasena o enviar invitacion por email'}), 400

        cur.execute("SELECT id FROM goti.usuarios WHERE username = %s", (username,))
        if cur.fetchone():
            return jsonify({'error': f'El usuario "{username}" ya existe'}), 409

        # Generar token si enviar invitacion
        token = secrets.token_urlsafe(32) if enviar_invitacion else None
        token_expires = (datetime.utcnow() + timedelta(hours=48)).isoformat() if token else None
        pwd = password if password else '__pendiente__'

        cur.execute("""
            INSERT INTO goti.usuarios (username, password, nombre, rol, activo, email, invite_token, invite_token_expires)
            VALUES (%s, %s, %s, %s, TRUE, %s, %s, %s) RETURNING id
        """, (username, pwd, nombre, rol, email or None, token, token_expires))
        new_id = cur.fetchone()['id']

        for bod in bodegas:
            cur.execute("""INSERT INTO goti.usuario_bodegas (usuario_id, bodega)
                           VALUES (%s, %s) ON CONFLICT DO NOTHING""", (new_id, bod))

        conn.commit()

        # Enviar email
        msg_extra = ''
        if enviar_invitacion and token:
            try:
                _enviar_email_invitacion(email, nombre, username, token)
                msg_extra = f' — Invitacion enviada a {email}'
            except Exception as mail_err:
                msg_extra = f' — ERROR enviando email: {str(mail_err)[:100]}'

        return jsonify({'success': True, 'id': new_id, 'message': f'Usuario {username} creado{msg_extra}'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>', methods=['PUT'])
def admin_editar_usuario(uid):
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        username = data.get('username', '').strip().lower()
        nombre = data.get('nombre', '').strip()
        password = data.get('password', '').strip()
        rol = data.get('rol', 'subgerente')
        activo = data.get('activo', True)
        bodegas = data.get('bodegas', [])

        email = data.get('email', '').strip().lower() or None

        # Verificar que el nuevo username no exista en otro usuario
        if username:
            cur.execute("SELECT id FROM goti.usuarios WHERE username = %s AND id != %s", (username, uid))
            if cur.fetchone():
                return jsonify({'error': f'El usuario "{username}" ya existe'}), 409

        if password:
            cur.execute("""UPDATE goti.usuarios
                           SET username = %s, nombre = %s, password = %s, rol = %s, activo = %s, email = %s
                           WHERE id = %s""", (username, nombre, password, rol, activo, email, uid))
        else:
            cur.execute("""UPDATE goti.usuarios
                           SET username = %s, nombre = %s, rol = %s, activo = %s, email = %s
                           WHERE id = %s""", (username, nombre, rol, activo, email, uid))

        if cur.rowcount == 0:
            return jsonify({'error': 'Usuario no encontrado'}), 404

        cur.execute("DELETE FROM goti.usuario_bodegas WHERE usuario_id = %s", (uid,))
        for bod in bodegas:
            cur.execute("""INSERT INTO goti.usuario_bodegas (usuario_id, bodega)
                           VALUES (%s, %s) ON CONFLICT DO NOTHING""", (uid, bod))

        conn.commit()
        return jsonify({'success': True, 'message': 'Usuario actualizado'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>', methods=['DELETE'])
def admin_eliminar_usuario(uid):
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        cur.execute("SELECT username FROM goti.usuarios WHERE id = %s", (uid,))
        row = cur.fetchone()
        if row and row['username'] == 'admin':
            return jsonify({'error': 'No se puede eliminar al administrador principal'}), 403

        cur.execute("DELETE FROM goti.usuarios WHERE id = %s", (uid,))
        conn.commit()
        return jsonify({'success': True, 'message': 'Usuario eliminado'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/roles', methods=['GET'])
def admin_listar_roles():
    """Devuelve los modulos y permisos de cada rol."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT rol, modulo, puede_ver, puede_editar,
                       COALESCE(puede_eliminar, FALSE) as puede_eliminar
                       FROM goti.rol_modulos ORDER BY rol, modulo""")
        rows = cur.fetchall()
        result = {}
        for r in rows:
            result.setdefault(r['rol'], {})[r['modulo']] = {
                'ver': r['puede_ver'], 'editar': r['puede_editar'], 'eliminar': r['puede_eliminar']
            }
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/admin/roles', methods=['PUT'])
def admin_guardar_roles():
    """Guarda permisos de un rol. Body: { admin_user, admin_pass, rol, modulos: { modulo: {ver,editar,eliminar} } }"""
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        rol = data.get('rol', '').strip().lower()
        modulos = data.get('modulos', {})
        if rol not in ('subgerente', 'supervisor', 'gerente', 'admin'):
            return jsonify({'error': 'Rol invalido'}), 400
        cur.execute("DELETE FROM goti.rol_modulos WHERE rol = %s", (rol,))
        for mod, perms in modulos.items():
            ver = perms.get('ver', False)
            editar = perms.get('editar', False)
            eliminar = perms.get('eliminar', False)
            if ver or editar or eliminar:
                cur.execute("""INSERT INTO goti.rol_modulos (rol, modulo, puede_ver, puede_editar, puede_eliminar)
                               VALUES (%s, %s, %s, %s, %s) ON CONFLICT DO NOTHING""",
                            (rol, mod, ver, editar, eliminar))
        conn.commit()
        return jsonify({'success': True, 'message': f'Permisos de {rol} actualizados'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


@app.route('/api/admin/usuarios/<int:uid>/reenviar', methods=['POST'])
def admin_reenviar_invitacion(uid):
    """Genera nuevo token y reenvia email de invitacion."""
    data = request.json
    conn, err, code = _require_admin(data)
    if err:
        return err, code
    try:
        cur = conn.cursor()
        cur.execute("SELECT username, nombre, email FROM goti.usuarios WHERE id = %s", (uid,))
        user = cur.fetchone()
        if not user:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        if not user['email']:
            return jsonify({'error': 'El usuario no tiene email configurado'}), 400

        token = secrets.token_urlsafe(32)
        token_expires = (datetime.utcnow() + timedelta(hours=48)).isoformat()
        cur.execute("""UPDATE goti.usuarios
                       SET invite_token = %s, invite_token_expires = %s
                       WHERE id = %s""", (token, token_expires, uid))
        conn.commit()

        _enviar_email_invitacion(user['email'], user['nombre'], user['username'], token)
        return jsonify({'success': True, 'message': f'Invitacion reenviada a {user["email"]}'})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db(conn)


PAGINA_ESTABLECER_CLAVE = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Establecer Contrasena - FOODIX</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Inter', sans-serif; background: #F8FAFC; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
        .card { background: #fff; border-radius: 16px; box-shadow: 0 4px 24px rgba(15,23,42,0.08); max-width: 420px; width: 100%; padding: 40px 32px; }
        .logo { text-align: center; margin-bottom: 28px; }
        .logo h1 { color: #123450; font-size: 24px; }
        .logo p { color: #64748B; font-size: 13px; margin-top: 4px; }
        .info { background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 10px; padding: 14px 16px; margin-bottom: 20px; }
        .info p { color: #1E40AF; font-size: 13px; line-height: 1.5; }
        .info strong { color: #123450; }
        .form-group { margin-bottom: 16px; }
        .form-group label { display: block; font-size: 12px; font-weight: 600; color: #64748B; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 6px; }
        .form-group input { width: 100%; padding: 12px 14px; border: 1px solid #CBD5E1; border-radius: 10px; font-size: 14px; font-family: inherit; transition: border 0.2s; }
        .form-group input:focus { outline: none; border-color: #123450; box-shadow: 0 0 0 3px rgba(18,52,80,0.1); }
        .btn { width: 100%; padding: 14px; background: #123450; color: #fff; border: none; border-radius: 10px; font-size: 14px; font-weight: 600; cursor: pointer; transition: all 0.2s; }
        .btn:hover { background: #1a4a6e; }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .msg { text-align: center; padding: 12px; border-radius: 8px; margin-top: 16px; font-size: 13px; display: none; }
        .msg.ok { display: block; background: #D1FAE5; color: #065F46; }
        .msg.err { display: block; background: #FEE2E2; color: #991B1B; }
        .success-card { text-align: center; }
        .success-card .icon { font-size: 48px; margin-bottom: 16px; }
        .success-card a { display: inline-block; margin-top: 20px; padding: 12px 28px; background: #123450; color: #fff; border-radius: 10px; text-decoration: none; font-weight: 600; font-size: 14px; }
        .success-card a:hover { background: #1a4a6e; }
    </style>
</head>
<body>
    <div class="card" id="form-card">
        <div class="logo">
            <h1>FOODIX</h1>
            <p>Sistema de Inventario</p>
        </div>
        <div class="info">
            <p>Hola <strong>{{ nombre }}</strong>, establece la contrasena para tu usuario <strong>{{ username }}</strong></p>
        </div>
        <form id="set-pass-form" onsubmit="return guardar(event)">
            <div class="form-group">
                <label>Nueva contrasena</label>
                <input type="password" id="pass1" placeholder="Minimo 4 caracteres" required minlength="4">
            </div>
            <div class="form-group">
                <label>Confirmar contrasena</label>
                <input type="password" id="pass2" placeholder="Repite la contrasena" required minlength="4">
            </div>
            <button type="submit" class="btn" id="btn-guardar">Establecer contrasena</button>
        </form>
        <div id="msg" class="msg"></div>
    </div>
    <div class="card success-card" id="success-card" style="display:none;">
        <div class="icon">&#10004;</div>
        <h2 style="color:#065F46;font-size:20px;">Contrasena establecida</h2>
        <p style="color:#64748B;margin-top:8px;font-size:14px;">Ya puedes iniciar sesion con tu usuario y contrasena.</p>
        <a href="/">Ir al sistema</a>
    </div>
    <script>
    async function guardar(e) {
        e.preventDefault();
        const p1 = document.getElementById('pass1').value;
        const p2 = document.getElementById('pass2').value;
        const msg = document.getElementById('msg');
        if (p1 !== p2) { msg.className = 'msg err'; msg.textContent = 'Las contrasenas no coinciden'; return false; }
        document.getElementById('btn-guardar').disabled = true;
        try {
            const r = await fetch('/api/establecer-clave', {
                method: 'POST', headers: {'Content-Type':'application/json'},
                body: JSON.stringify({token: '{{ token }}', password: p1})
            });
            const d = await r.json();
            if (r.ok && d.success) {
                document.getElementById('form-card').style.display = 'none';
                document.getElementById('success-card').style.display = 'block';
            } else {
                msg.className = 'msg err'; msg.textContent = d.error || 'Error al guardar';
                document.getElementById('btn-guardar').disabled = false;
            }
        } catch(err) {
            msg.className = 'msg err'; msg.textContent = 'Error de conexion';
            document.getElementById('btn-guardar').disabled = false;
        }
        return false;
    }
    </script>
</body>
</html>
"""

PAGINA_TOKEN_INVALIDO = """
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Enlace invalido - FOODIX</title>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Inter', sans-serif; background: #F8FAFC; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
        .card { background: #fff; border-radius: 16px; box-shadow: 0 4px 24px rgba(15,23,42,0.08); max-width: 420px; width: 100%; padding: 40px 32px; text-align: center; }
        .icon { font-size: 48px; margin-bottom: 16px; }
        h2 { color: #991B1B; font-size: 20px; }
        p { color: #64748B; margin-top: 8px; font-size: 14px; }
    </style>
</head>
<body>
    <div class="card">
        <div class="icon">&#10060;</div>
        <h2>Enlace invalido o expirado</h2>
        <p>Pide al administrador que te reenvie la invitacion.</p>
    </div>
</body>
</html>
"""


@app.route('/api/establecer-clave', methods=['POST'])
def api_establecer_clave():
    """Endpoint para guardar la contrasena desde el formulario publico."""
    data = request.json or {}
    token = data.get('token', '')
    password = data.get('password', '')

    if not token or not password:
        return jsonify({'error': 'Token y contrasena requeridos'}), 400
    if len(password) < 4:
        return jsonify({'error': 'La contrasena debe tener al menos 4 caracteres'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, invite_token_expires FROM goti.usuarios
                       WHERE invite_token = %s AND activo = TRUE""", (token,))
        user = cur.fetchone()
        if not user:
            return jsonify({'error': 'Enlace invalido'}), 404
        if user['invite_token_expires'] and user['invite_token_expires'] < datetime.utcnow():
            return jsonify({'error': 'Enlace expirado. Pide al administrador que lo reenvie.'}), 410

        cur.execute("""UPDATE goti.usuarios
                       SET password = %s, invite_token = NULL, invite_token_expires = NULL
                       WHERE id = %s""", (password, user['id']))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            release_db(conn)


# ==================== DESCUENTOS NOMINA ====================

@app.route('/api/descuentos/reporte', methods=['GET'])
def descuentos_reporte():
    """Reporte de descuentos por persona en un rango de fechas (semanas cerradas)"""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local', '')
    solo_cerradas = request.args.get('solo_cerradas', '1')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        params = []
        where = ""
        if fecha_desde:
            where += " AND s.fecha_inicio >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            where += " AND s.fecha_fin <= %s"; params.append(fecha_hasta)
        if local:
            where += " AND s.local = %s"; params.append(local)
        if solo_cerradas == '1':
            where += " AND s.estado = 'cerrada'"

        # Detalle por persona, semana, local y producto
        cur.execute(f"""
            SELECT ap.persona,
                   s.fecha_inicio, s.fecha_fin, s.local,
                   a.codigo, a.nombre, a.unidad,
                   ap.cantidad, ap.monto,
                   a.costo_unitario, a.diferencia_semanal
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE 1=1 {where}
            ORDER BY ap.persona, s.fecha_inicio, s.local, a.nombre
        """, params)
        detalle = [dict(r) for r in cur.fetchall()]

        # Resumen por persona
        cur.execute(f"""
            SELECT ap.persona,
                   COUNT(DISTINCT s.id) as semanas,
                   COUNT(DISTINCT s.local) as locales,
                   COALESCE(SUM(ap.monto), 0) as total_monto
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE 1=1 {where}
            GROUP BY ap.persona
            ORDER BY total_monto DESC
        """, params)
        resumen = [dict(r) for r in cur.fetchall()]

        # Semanas incluidas
        cur.execute(f"""
            SELECT DISTINCT s.fecha_inicio, s.fecha_fin, s.local, s.estado
            FROM goti.semanas_inventario s
            JOIN goti.asignacion_semanal a ON a.semana_id = s.id
            JOIN goti.asignacion_semanal_personas ap ON ap.asignacion_semanal_id = a.id
            WHERE 1=1 {where}
            ORDER BY s.fecha_inicio, s.local
        """, params)
        semanas = [dict(r) for r in cur.fetchall()]

        return jsonify({
            'resumen': resumen,
            'detalle': detalle,
            'semanas': semanas,
            'total_personas': len(resumen),
            'total_descuento': sum(float(r['total_monto']) for r in resumen)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/descuentos/exportar-excel', methods=['GET'])
def descuentos_exportar_excel():
    """Exporta el reporte de descuentos a Excel"""
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local', '')
    solo_cerradas = request.args.get('solo_cerradas', '1')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        params = []
        where = ""
        if fecha_desde:
            where += " AND s.fecha_inicio >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            where += " AND s.fecha_fin <= %s"; params.append(fecha_hasta)
        if local:
            where += " AND s.local = %s"; params.append(local)
        if solo_cerradas == '1':
            where += " AND s.estado = 'cerrada'"

        # Resumen por persona
        cur.execute(f"""
            SELECT ap.persona,
                   COUNT(DISTINCT s.id) as semanas,
                   COALESCE(SUM(ap.monto), 0) as total_monto
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE 1=1 {where}
            GROUP BY ap.persona
            ORDER BY ap.persona
        """, params)
        resumen = cur.fetchall()

        # Detalle
        cur.execute(f"""
            SELECT ap.persona, s.fecha_inicio, s.fecha_fin, s.local,
                   a.codigo, a.nombre, ap.cantidad, ap.monto, a.costo_unitario
            FROM goti.asignacion_semanal_personas ap
            JOIN goti.asignacion_semanal a ON a.id = ap.asignacion_semanal_id
            JOIN goti.semanas_inventario s ON s.id = a.semana_id
            WHERE 1=1 {where}
            ORDER BY ap.persona, s.fecha_inicio, s.local
        """, params)
        detalle = cur.fetchall()

        wb = Workbook()
        # Hoja 1: Resumen por persona
        ws1 = wb.active
        ws1.title = "Resumen Descuentos"
        headers1 = ['Persona', 'Semanas', 'Total Descuento']
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="123450", end_color="123450", fill_type="solid")
        for col, h in enumerate(headers1, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
        total_gen = 0
        for i, r in enumerate(resumen, 2):
            ws1.cell(row=i, column=1, value=r['persona'])
            ws1.cell(row=i, column=2, value=int(r['semanas']))
            monto = float(r['total_monto'])
            ws1.cell(row=i, column=3, value=round(monto, 2)).number_format = '$#,##0.00'
            total_gen += monto
        # Fila total
        row_total = len(resumen) + 2
        ws1.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True, size=12)
        ws1.cell(row=row_total, column=3, value=round(total_gen, 2)).font = Font(bold=True, size=12)
        ws1.cell(row=row_total, column=3).number_format = '$#,##0.00'
        ws1.column_dimensions['A'].width = 35
        ws1.column_dimensions['B'].width = 12
        ws1.column_dimensions['C'].width = 18

        # Hoja 2: Detalle
        ws2 = wb.create_sheet("Detalle")
        headers2 = ['Persona', 'Semana Inicio', 'Semana Fin', 'Local', 'Codigo', 'Producto', 'Cantidad', 'Monto', 'Costo Unit.']
        for col, h in enumerate(headers2, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
        for i, r in enumerate(detalle, 2):
            ws2.cell(row=i, column=1, value=r['persona'])
            ws2.cell(row=i, column=2, value=str(r['fecha_inicio']))
            ws2.cell(row=i, column=3, value=str(r['fecha_fin']))
            ws2.cell(row=i, column=4, value=BODEGAS_NOMBRES.get(r['local'], r['local']))
            ws2.cell(row=i, column=5, value=r['codigo'])
            ws2.cell(row=i, column=6, value=r['nombre'])
            ws2.cell(row=i, column=7, value=float(r['cantidad'] or 0))
            ws2.cell(row=i, column=8, value=round(float(r['monto'] or 0), 2)).number_format = '$#,##0.00'
            ws2.cell(row=i, column=9, value=round(float(r['costo_unitario'] or 0), 4)).number_format = '$#,##0.0000'
        for col_letter in ['A','B','C','D','E','F','G','H','I']:
            ws2.column_dimensions[col_letter].width = 18 if col_letter in ['A','F'] else 14

        rango = f"{fecha_desde or 'inicio'}_a_{fecha_hasta or 'fin'}"
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'Descuentos_Nomina_{rango}.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


# ==================== CUADRES DE CAJA ====================

@app.route('/api/cuadres/listar', methods=['GET'])
def cuadres_listar():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local', '')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = "SELECT * FROM goti.cuadres_caja WHERE 1=1"
        params = []
        if fecha_desde:
            sql += " AND fecha >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND fecha <= %s"; params.append(fecha_hasta)
        if local:
            sql += " AND local = %s"; params.append(local)
        sql += " ORDER BY fecha DESC, local"
        cur.execute(sql, params)
        rows = cur.fetchall()
        return jsonify({'cuadres': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/cuadres/guardar', methods=['POST'])
def cuadres_guardar():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    if not fecha or not local:
        return jsonify({'error': 'Fecha y local requeridos'}), 400
    venta_sistema = float(data.get('venta_sistema', 0))
    efectivo_contado = float(data.get('efectivo_contado', 0))
    venta_tarjeta = float(data.get('venta_tarjeta', 0))
    venta_transferencia = float(data.get('venta_transferencia', 0))
    venta_plataformas = float(data.get('venta_plataformas', 0))
    otros_ingresos = float(data.get('otros_ingresos', 0))
    gastos_retiros = float(data.get('gastos_retiros', 0))
    efectivo_esperado = venta_sistema - venta_tarjeta - venta_transferencia - venta_plataformas + otros_ingresos - gastos_retiros
    diferencia = efectivo_contado - efectivo_esperado
    observacion = data.get('observacion', '')
    registrado_por = data.get('registrado_por', '')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.cuadres_caja (fecha, local, venta_sistema, efectivo_contado, venta_tarjeta,
                venta_transferencia, venta_plataformas, otros_ingresos, gastos_retiros,
                efectivo_esperado, diferencia, observacion, registrado_por)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (fecha, local) DO UPDATE SET
                venta_sistema=EXCLUDED.venta_sistema, efectivo_contado=EXCLUDED.efectivo_contado,
                venta_tarjeta=EXCLUDED.venta_tarjeta, venta_transferencia=EXCLUDED.venta_transferencia,
                venta_plataformas=EXCLUDED.venta_plataformas, otros_ingresos=EXCLUDED.otros_ingresos,
                gastos_retiros=EXCLUDED.gastos_retiros, efectivo_esperado=EXCLUDED.efectivo_esperado,
                diferencia=EXCLUDED.diferencia, observacion=EXCLUDED.observacion,
                registrado_por=EXCLUDED.registrado_por
            RETURNING id
        """, (fecha, local, venta_sistema, efectivo_contado, venta_tarjeta,
              venta_transferencia, venta_plataformas, otros_ingresos, gastos_retiros,
              efectivo_esperado, diferencia, observacion, registrado_por))
        row = cur.fetchone()
        conn.commit()
        return jsonify({'success': True, 'id': row['id']})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/cuadres/<int:cuadre_id>', methods=['DELETE'])
def cuadres_eliminar(cuadre_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.cuadres_caja WHERE id = %s", (cuadre_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/cuadres/resumen', methods=['GET'])
def cuadres_resumen():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        params = []
        where = ""
        if fecha_desde:
            where += " AND fecha >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            where += " AND fecha <= %s"; params.append(fecha_hasta)
        cur.execute(f"""
            SELECT COUNT(*) as total,
                   COALESCE(SUM(ABS(diferencia)),0) as total_diferencia,
                   COALESCE(AVG(diferencia),0) as avg_diferencia,
                   COALESCE(SUM(venta_sistema),0) as total_ventas,
                   COUNT(CASE WHEN ABS(diferencia) > 1 THEN 1 END) as con_descuadre
            FROM goti.cuadres_caja WHERE 1=1 {where}
        """, params)
        resumen = dict(cur.fetchone())
        cur.execute(f"""
            SELECT local, COUNT(*) as cuadres, COALESCE(SUM(diferencia),0) as diferencia_total,
                   COALESCE(SUM(ABS(diferencia)),0) as diferencia_abs,
                   COALESCE(AVG(diferencia),0) as diferencia_avg
            FROM goti.cuadres_caja WHERE 1=1 {where}
            GROUP BY local ORDER BY diferencia_abs DESC
        """, params)
        resumen['por_local'] = [dict(r) for r in cur.fetchall()]
        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


# ==================== DELIVERY / PLATAFORMAS ====================

@app.route('/api/delivery/listar', methods=['GET'])
def delivery_listar():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local', '')
    plataforma = request.args.get('plataforma', '')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = "SELECT * FROM goti.delivery_liquidaciones WHERE 1=1"
        params = []
        if fecha_desde:
            sql += " AND fecha >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND fecha <= %s"; params.append(fecha_hasta)
        if local:
            sql += " AND local = %s"; params.append(local)
        if plataforma:
            sql += " AND plataforma = %s"; params.append(plataforma)
        sql += " ORDER BY fecha DESC, local, plataforma"
        cur.execute(sql, params)
        rows = cur.fetchall()
        return jsonify({'liquidaciones': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/delivery/guardar', methods=['POST'])
def delivery_guardar():
    data = request.json
    fecha = data.get('fecha')
    local = data.get('local')
    plataforma = data.get('plataforma')
    if not fecha or not local or not plataforma:
        return jsonify({'error': 'Fecha, local y plataforma requeridos'}), 400
    venta_bruta = float(data.get('venta_bruta', 0))
    comision_pct = float(data.get('comision_pct', 0))
    comision_monto = float(data.get('comision_monto', 0))
    iva_comision = float(data.get('iva_comision', 0))
    propinas = float(data.get('propinas', 0))
    ajustes = float(data.get('ajustes', 0))
    neto_recibir = venta_bruta - comision_monto - iva_comision + propinas + ajustes
    depositado_real = float(data.get('depositado_real', 0))
    diferencia = depositado_real - neto_recibir
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.delivery_liquidaciones (fecha, local, plataforma, total_pedidos,
                venta_bruta, comision_pct, comision_monto, iva_comision, propinas, ajustes,
                neto_recibir, depositado_real, diferencia, referencia, observacion, registrado_por)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (fecha, local, plataforma, int(data.get('total_pedidos', 0)),
              venta_bruta, comision_pct, comision_monto, iva_comision, propinas, ajustes,
              neto_recibir, depositado_real, diferencia,
              data.get('referencia', ''), data.get('observacion', ''), data.get('registrado_por', '')))
        row = cur.fetchone()
        conn.commit()
        return jsonify({'success': True, 'id': row['id']})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/delivery/<int:liq_id>', methods=['DELETE'])
def delivery_eliminar(liq_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.delivery_liquidaciones WHERE id = %s", (liq_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/delivery/resumen', methods=['GET'])
def delivery_resumen():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        params = []
        where = ""
        if fecha_desde:
            where += " AND fecha >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            where += " AND fecha <= %s"; params.append(fecha_hasta)
        cur.execute(f"""
            SELECT COUNT(*) as total, COALESCE(SUM(venta_bruta),0) as total_ventas,
                   COALESCE(SUM(comision_monto),0) as total_comisiones,
                   COALESCE(SUM(neto_recibir),0) as total_neto,
                   COALESCE(SUM(depositado_real),0) as total_depositado,
                   COALESCE(SUM(ABS(diferencia)),0) as total_diferencia,
                   COALESCE(SUM(total_pedidos),0) as total_pedidos
            FROM goti.delivery_liquidaciones WHERE 1=1 {where}
        """, params)
        resumen = dict(cur.fetchone())
        cur.execute(f"""
            SELECT plataforma, COUNT(*) as liquidaciones, COALESCE(SUM(venta_bruta),0) as ventas,
                   COALESCE(SUM(comision_monto),0) as comisiones,
                   COALESCE(AVG(comision_pct),0) as comision_pct_avg,
                   COALESCE(SUM(ABS(diferencia)),0) as diferencia_abs,
                   COALESCE(SUM(total_pedidos),0) as pedidos
            FROM goti.delivery_liquidaciones WHERE 1=1 {where}
            GROUP BY plataforma ORDER BY ventas DESC
        """, params)
        resumen['por_plataforma'] = [dict(r) for r in cur.fetchall()]
        cur.execute(f"""
            SELECT local, COUNT(*) as liquidaciones, COALESCE(SUM(venta_bruta),0) as ventas,
                   COALESCE(SUM(ABS(diferencia)),0) as diferencia_abs
            FROM goti.delivery_liquidaciones WHERE 1=1 {where}
            GROUP BY local ORDER BY ventas DESC
        """, params)
        resumen['por_local'] = [dict(r) for r in cur.fetchall()]
        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


# ==================== REGISTRO DE FACTURAS ====================

@app.route('/api/facturas/listar', methods=['GET'])
def facturas_listar():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    local = request.args.get('local', '')
    categoria = request.args.get('categoria', '')
    estado_pago = request.args.get('estado_pago', '')
    proveedor = request.args.get('proveedor', '')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sql = "SELECT * FROM goti.facturas_registro WHERE 1=1"
        params = []
        if fecha_desde:
            sql += " AND fecha_emision >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            sql += " AND fecha_emision <= %s"; params.append(fecha_hasta)
        if local:
            sql += " AND local = %s"; params.append(local)
        if categoria:
            sql += " AND categoria = %s"; params.append(categoria)
        if estado_pago:
            sql += " AND estado_pago = %s"; params.append(estado_pago)
        if proveedor:
            sql += " AND proveedor ILIKE %s"; params.append(f'%{proveedor}%')
        sql += " ORDER BY fecha_emision DESC, local"
        cur.execute(sql, params)
        rows = cur.fetchall()
        return jsonify({'facturas': [dict(r) for r in rows]})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/facturas/guardar', methods=['POST'])
def facturas_guardar():
    data = request.json
    fecha = data.get('fecha_emision')
    local = data.get('local')
    proveedor = data.get('proveedor')
    if not fecha or not local or not proveedor:
        return jsonify({'error': 'Fecha, local y proveedor requeridos'}), 400
    subtotal_0 = float(data.get('subtotal_0', 0))
    subtotal_iva = float(data.get('subtotal_iva', 0))
    iva = float(data.get('iva', 0))
    total = subtotal_0 + subtotal_iva + iva
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.facturas_registro (fecha_emision, local, proveedor, ruc, numero_factura,
                autorizacion, subtotal_0, subtotal_iva, iva, total, categoria, forma_pago,
                estado_pago, observacion, registrado_por)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (fecha, local, proveedor, data.get('ruc',''), data.get('numero_factura',''),
              data.get('autorizacion',''), subtotal_0, subtotal_iva, iva, total,
              data.get('categoria','Otros'), data.get('forma_pago','Transferencia'),
              data.get('estado_pago','Pendiente'), data.get('observacion',''),
              data.get('registrado_por','')))
        row = cur.fetchone()
        conn.commit()
        return jsonify({'success': True, 'id': row['id']})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/facturas/<int:factura_id>', methods=['PUT'])
def facturas_actualizar(factura_id):
    data = request.json
    estado_pago = data.get('estado_pago')
    if not estado_pago:
        return jsonify({'error': 'estado_pago requerido'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE goti.facturas_registro SET estado_pago = %s WHERE id = %s", (estado_pago, factura_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/facturas/<int:factura_id>', methods=['DELETE'])
def facturas_eliminar(factura_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.facturas_registro WHERE id = %s", (factura_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)

@app.route('/api/facturas/resumen', methods=['GET'])
def facturas_resumen():
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        params = []
        where = ""
        if fecha_desde:
            where += " AND fecha_emision >= %s"; params.append(fecha_desde)
        if fecha_hasta:
            where += " AND fecha_emision <= %s"; params.append(fecha_hasta)
        cur.execute(f"""
            SELECT COUNT(*) as total, COALESCE(SUM(total),0) as total_facturado,
                   COALESCE(SUM(iva),0) as total_iva,
                   COUNT(CASE WHEN estado_pago = 'Pendiente' THEN 1 END) as pendientes,
                   COALESCE(SUM(CASE WHEN estado_pago = 'Pendiente' THEN total ELSE 0 END),0) as monto_pendiente
            FROM goti.facturas_registro WHERE 1=1 {where}
        """, params)
        resumen = dict(cur.fetchone())
        cur.execute(f"""
            SELECT categoria, COUNT(*) as facturas, COALESCE(SUM(total),0) as monto
            FROM goti.facturas_registro WHERE 1=1 {where}
            GROUP BY categoria ORDER BY monto DESC
        """, params)
        resumen['por_categoria'] = [dict(r) for r in cur.fetchall()]
        cur.execute(f"""
            SELECT local, COUNT(*) as facturas, COALESCE(SUM(total),0) as monto,
                   COUNT(CASE WHEN estado_pago = 'Pendiente' THEN 1 END) as pendientes
            FROM goti.facturas_registro WHERE 1=1 {where}
            GROUP BY local ORDER BY monto DESC
        """, params)
        resumen['por_local'] = [dict(r) for r in cur.fetchall()]
        return jsonify(resumen)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


# ==================== CONFIG PRODUCTOS POR MARCA ====================

@app.route('/api/admin/productos-marca', methods=['GET'])
def listar_productos_marca():
    """Lista productos configurados para una marca"""
    marca = request.args.get('marca', '')
    if not marca:
        return jsonify({'error': 'Marca requerida'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS goti.productos_por_marca (
                id SERIAL PRIMARY KEY,
                marca VARCHAR(50) NOT NULL,
                codigo VARCHAR(20) NOT NULL,
                nombre VARCHAR(100) NOT NULL,
                activo BOOLEAN DEFAULT TRUE,
                unidad VARCHAR(30) DEFAULT 'Unidad',
                equivalencia NUMERIC(12,4) DEFAULT 1,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(marca, codigo)
            )
        """)
        conn.commit()
        # Asegurar columnas nuevas existan
        cur.execute("ALTER TABLE goti.productos_por_marca ADD COLUMN IF NOT EXISTS unidad VARCHAR(30) DEFAULT 'Unidad'")
        cur.execute("ALTER TABLE goti.productos_por_marca ADD COLUMN IF NOT EXISTS equivalencia NUMERIC(12,4) DEFAULT 1")
        cur.execute("ALTER TABLE goti.productos_por_marca ADD COLUMN IF NOT EXISTS tipo_conteo VARCHAR(20) DEFAULT 'diario'")
        conn.commit()
        cur.execute("""
            SELECT id, marca, codigo, nombre, activo, unidad, equivalencia, tipo_conteo, created_at
            FROM goti.productos_por_marca
            WHERE marca = %s
            ORDER BY codigo
        """, (marca,))
        productos = [dict(r) for r in cur.fetchall()]
        return jsonify(productos)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca', methods=['POST'])
def agregar_producto_marca():
    """Agrega un producto a una marca"""
    data = request.json
    marca = data.get('marca', '').strip()
    codigo = data.get('codigo', '').strip().upper()
    nombre = data.get('nombre', '').strip().upper()
    unidad = data.get('unidad', 'Unidad').strip()
    equivalencia = data.get('equivalencia', 1)
    tipo_conteo = data.get('tipo_conteo', 'diario').strip()
    if tipo_conteo not in ('diario', 'cruce', 'ambos'):
        tipo_conteo = 'diario'
    if not marca or not codigo or not nombre:
        return jsonify({'error': 'marca, codigo y nombre son requeridos'}), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.productos_por_marca (marca, codigo, nombre, activo, unidad, equivalencia, tipo_conteo)
            VALUES (%s, %s, %s, TRUE, %s, %s, %s)
            ON CONFLICT (marca, codigo) DO UPDATE SET nombre = EXCLUDED.nombre, activo = TRUE, unidad = EXCLUDED.unidad, equivalencia = EXCLUDED.equivalencia, tipo_conteo = EXCLUDED.tipo_conteo
            RETURNING id, marca, codigo, nombre, activo, unidad, equivalencia, tipo_conteo
        """, (marca, codigo, nombre, unidad, equivalencia, tipo_conteo))
        conn.commit()
        prod = dict(cur.fetchone())
        return jsonify(prod)
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca/<int:prod_id>', methods=['DELETE'])
def eliminar_producto_marca(prod_id):
    """Elimina un producto de la configuracion"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM goti.productos_por_marca WHERE id = %s", (prod_id,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca/<int:prod_id>', methods=['PUT'])
def editar_producto_marca(prod_id):
    """Edita unidad y/o equivalencia de un producto"""
    data = request.json
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        sets = []
        vals = []
        if 'unidad' in data:
            sets.append("unidad = %s")
            vals.append(data['unidad'])
        if 'equivalencia' in data:
            sets.append("equivalencia = %s")
            vals.append(data['equivalencia'])
        if 'nombre' in data:
            sets.append("nombre = %s")
            vals.append(data['nombre'].strip().upper())
        if 'tipo_conteo' in data:
            tc = data['tipo_conteo'].strip()
            if tc not in ('diario', 'cruce', 'ambos'):
                tc = 'diario'
            sets.append("tipo_conteo = %s")
            vals.append(tc)
        if not sets:
            return jsonify({'error': 'Nada que actualizar'}), 400
        vals.append(prod_id)
        cur.execute(f"UPDATE goti.productos_por_marca SET {', '.join(sets)} WHERE id = %s RETURNING id, marca, codigo, nombre, activo, unidad, equivalencia, tipo_conteo", vals)
        conn.commit()
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Producto no encontrado'}), 404
        return jsonify(dict(row))
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca/toggle/<int:prod_id>', methods=['PUT'])
def toggle_producto_marca(prod_id):
    """Activa o desactiva un producto"""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.productos_por_marca SET activo = NOT activo WHERE id = %s
            RETURNING id, marca, codigo, nombre, activo, unidad, equivalencia, tipo_conteo
        """, (prod_id,))
        conn.commit()
        row = cur.fetchone()
        if not row:
            return jsonify({'error': 'Producto no encontrado'}), 404
        return jsonify(dict(row))
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca/carga-inicial', methods=['POST'])
def carga_inicial_productos():
    """Carga los productos hardcodeados o desde equivalencias_conteo para bodegas operativas"""
    PRODUCTOS_LOCALES = {
        'CHIOS': {
            'ALCH002': 'CLUB PEQ', 'ALMP001': 'PILSENER 600 ML', 'BEB019': 'AGUA DASANI',
            'BEB020': 'GUITIG', 'BEB021': 'COCA PEQ', 'BEB023': 'SPRITE PEQ',
            'BEB025': 'ZERO PEQ', 'BEB029': 'HATSU FRAMBUESA', 'BEB030': 'HATSU UVILLA',
            'CONG003': 'HELADO VAINILLA', 'CONGP001': 'PORTOBELLO PROCESADO',
            'CONGP006': 'PAN DE PORTOBELLO', 'CONGP007': 'PORTOBELLO CHEESE',
            'CONGP020': 'BOLITAS DE PAPA', 'CP001': 'ALAS CRUDAS PAQ',
            'CP002': 'POLLO ESTANDAR PAQ', 'CP003': 'HAMBURGUESA',
            'CP005': 'POLLO BOLDER PAQ', 'CP007': 'HAMBURGUESA SMASH',
            'DOG001': 'HELADO DE PERRO', 'DOG004': 'PUDIN POLLO PEQUENO',
            'DOMP003': 'GALLETA DE PERRO PAQ', 'DUL002': 'OREO', 'DUL013': 'KIT KAT CHOCOLATE',
            'EMB002': 'TOCINO', 'FRU012': 'FRUTILLA PORCIONADA PAQ', 'LAC004': 'HUEVOS',
            'LAC005': 'QUESO AMERICANO', 'LACP002': 'QUESO MOZZARELLA UNIDAD',
            'PASN008': 'PAN PRETZEL', 'PASN012': 'PAN DE PAPA',
        },
        'CACHON': {
            'ALCH001': 'PILSENER PEQ', 'ALCH002': 'CLUB PEQ', 'ALCH004': 'CORONA',
            'ALMP001': 'PILSENER 600 ML', 'BEB019': 'AGUA DASANI', 'BEB020': 'GUITIG',
            'BEB031': 'COCA VIDRIO', 'BEB032': 'SPRITE VIDRIO',
            'BEB035': 'FANTA VIDRIO', 'BEB040': 'FUZE TE VIDRIO', 'BEB048': 'ZERO VIDRIO',
            'BEB049': 'FRESA VIDRIO', 'CP012': 'CHULETA PAQ', 'CP013': 'FILETE DE PECHUGA',
            'CP014': 'PICAÑA A PAQ', 'CP015': 'PICAÑA C PAQ', 'CP016': 'LOMO FINO PAQ',
            'CP017': 'FILETE DE CARNE', 'CP018': 'COSTILLA PAQ', 'EMB013': 'CHISTORRA',
            'EMB012': 'MORCILLA', 'EMB010': 'CHORIZO CON ROMERO', 'EMB011': 'CHORIZO CON ALBAHACA',
            'EMB009': 'CHORIZO CHISTORRA', 'LAC004': 'HUEVOS', 'CP021': 'NEW YORK B PAQ',
            'CP022': 'RIBEYE B PAQ', 'CP023': 'NEW YORK C PAQ', 'CP024': 'RIBEYE C PAQ',
            'ZUM007': 'PULPA DE MARACUYA', 'ZUM008': 'PULPA DE MORA',
            'ZUM009': 'PULPA DE TOMATE DE ARBOL', 'ZUM010': 'PULPA DE NARANJILLA',
            'POST017': 'CHEESE CAKE', 'POST018': 'TRES LECHES',
        },
        'SIMON_BOLON': {
            'ACOM001': 'EMPANADA DE QUESO', 'ACOM002': 'EMPANADA DE CAMARON',
            'ACOM003': 'EMPANADA DE POLLO', 'ACOM005': 'CORVICHE DE ALBACORA',
            'ACOM006': 'MUCHIN', 'ALCH001': 'PILSENER PEQ', 'ALCH002': 'CLUB PEQ',
            'ALCH004': 'CORONA', 'BEB019': 'AGUA DASANI', 'BEB020': 'GUITIG',
            'BEB021': 'COCA PEQ', 'BEB023': 'SPRITE PEQ', 'BEB025': 'ZERO PEQ',
            'BEB031': 'COCA VIDRIO', 'BEB032': 'SPRITE VIDRIO', 'BEB035': 'FANTA VIDRIO',
            'BEB040': 'FUZE TE VIDRIO', 'BEB047': 'FANTA PEQ',
            'CP019': 'ESTOFADO DE CARNE PAQ', 'CP020': 'FRITADA PAQ',
            'DOG001': 'HELADO DE PERRO', 'DOG004': 'PUDIN POLLO PEQUENO',
            'FRU010': 'MORA', 'LAC004': 'HUEVOS',
            'MAR002': 'CAMARON', 'MAR003': 'ALBACORA', 'MAR005': 'CAMARON PAQ 110 GR',
            'MAR006': 'CAMARON PAQ 76 GR', 'PASN010': 'CHIFLES', 'ZUM003': 'ZUMO DE FRESA',
        },
    }
    # Mapeo de marca a bodega en equivalencias_conteo
    BODEGAS_OPERATIVAS_MAP = {
        'BODEGA_PRINCIPAL': 'bodega_principal',
        'MATERIA_PRIMA': 'materia_prima',
        'PLANTA': 'planta',
    }

    marca = request.json.get('marca', '') if request.json else ''
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        total = 0

        # Si es bodega operativa, cargar desde tablas de toma fisica usando INSERT directo
        if marca in BODEGAS_OPERATIVAS_MAP:
            TABLAS_TOMA = {
                'BODEGA_PRINCIPAL': 'public.toma_bodega',
                'MATERIA_PRIMA': 'public.toma_materiaprima',
                'PLANTA': 'public.toma_planta',
            }
            tabla = TABLAS_TOMA.get(marca)
            if tabla:
                # NO se borra nada. Antes habia aqui un DELETE de toda la marca
                # y despues este INSERT rehacia las equivalencias con la regla
                # "1000 si dice kg, 1 para el resto". Un clic destruia meses de
                # correcciones: asi es como HERR005 paso de 30 a 1.
                #
                # El ON CONFLICT de abajo hace lo correcto: refresca nombre y
                # unidad -lo unico que esta carga sabe- y deja la equivalencia
                # como este. La regla por defecto solo se aplica a productos
                # NUEVOS, que es para lo que sirve.

                # INSERT directo desde la tabla de toma (evita problemas de fetch)
                # Usamos subquery con ROW_NUMBER para deduplicar por codigo
                cur.execute(f"""
                    INSERT INTO goti.productos_por_marca (marca, codigo, nombre, activo, unidad, equivalencia)
                    SELECT
                        %s as marca,
                        sub.codigo,
                        sub.producto,
                        TRUE as activo,
                        COALESCE(sub.unidad, 'Unidad') as unidad,
                        CASE WHEN LOWER(COALESCE(sub.unidad,'')) LIKE '%%kg%%' THEN 1000 ELSE 1 END as equivalencia
                    FROM (
                        SELECT codigo, producto, unidad,
                               ROW_NUMBER() OVER (PARTITION BY codigo ORDER BY fecha DESC) as rn
                        FROM {tabla}
                        WHERE fecha >= CURRENT_DATE - INTERVAL '90 days'
                          AND codigo IS NOT NULL AND codigo != ''
                    ) sub
                    WHERE sub.rn = 1
                    ON CONFLICT (marca, codigo) DO UPDATE
                    SET nombre = EXCLUDED.nombre, unidad = EXCLUDED.unidad
                """, (marca,))
                total = cur.rowcount
                # Cuantas equivalencias se respetaron, para que la respuesta lo
                # diga y nadie se quede con la duda de si se perdio algo.
                cur.execute("""
                    SELECT COUNT(*) AS n FROM goti.productos_por_marca
                    WHERE marca = %s AND COALESCE(equivalencia, 1) NOT IN (1, 1000)
                """, (marca,))
                respetadas = cur.fetchone()['n']
        else:
            # Cargar desde lista hardcodeada para locales
            marcas_cargar = [marca] if marca else list(PRODUCTOS_LOCALES.keys())
            for m in marcas_cargar:
                if m not in PRODUCTOS_LOCALES:
                    continue
                for codigo, nombre in PRODUCTOS_LOCALES[m].items():
                    cur.execute("""
                        INSERT INTO goti.productos_por_marca (marca, codigo, nombre, activo)
                        VALUES (%s, %s, %s, TRUE)
                        ON CONFLICT (marca, codigo) DO NOTHING
                    """, (m, codigo, nombre))
                    total += 1

        conn.commit()
        salida = {'ok': True, 'insertados': total}
        if marca in BODEGAS_OPERATIVAS_MAP:
            salida['equivalencias_respetadas'] = respetadas
            salida['aviso'] = (f'{respetadas} equivalencias corregidas a mano se '
                               f'conservaron. Esta carga solo actualiza nombre y '
                               f'unidad, y agrega productos nuevos.')
        return jsonify(salida)
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/admin/productos-marca/fix-equivalencias-kg', methods=['POST'])
def fix_equivalencias_kg():
    """Pone equivalencia=1000 a los productos en Kg que AUN no tengan factor propio.

    Ya no pisa lo corregido a mano. La regla "kilos a gramos son 1000" parece
    obvia pero no siempre es cierta: la propia lista de excepciones escritas en
    el codigo -DETERGENTE, JACK DANIEL- se descubrio a base de encontrar
    productos que no encajaban, y hubo mas despues. Un factor que alguien puso
    mirando el producto vale mas que una regla general.

    Por eso solo toca los que estan en el valor por defecto (1 o sin dato), y
    exige confirmar: ?confirmar=si.
    """
    if request.args.get('confirmar') != 'si':
        return jsonify({
            'error': 'Hace falta confirmar. Esta accion cambia equivalencias en '
                     'las tres bodegas operativas.',
            'como': 'repetir la llamada con ?confirmar=si',
        }), 400
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Solo los que siguen en el valor por defecto. Lo que alguien ya
        # corrigio se queda como esta.
        cur.execute("""
            UPDATE goti.productos_por_marca
            SET equivalencia = 1000
            WHERE marca IN ('BODEGA_PRINCIPAL', 'MATERIA_PRIMA', 'PLANTA')
              AND (LOWER(unidad) LIKE '%kg%' OR LOWER(unidad) LIKE '%kilogramo%')
              AND COALESCE(equivalencia, 1) = 1
              AND UPPER(nombre) NOT LIKE '%DETERGENTE%'
              AND UPPER(nombre) NOT LIKE '%JACK DANIEL%'
        """)
        actualizados = cur.rowcount
        conn.commit()
        return jsonify({'ok': True, 'actualizados': actualizados,
                        'aviso': 'Solo se tocaron los que estaban en el valor por '
                                 'defecto. Lo corregido a mano no se modifico.'})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: release_db(conn)


# ============================================================
# MODULO FLUJO DE CAJA
# ============================================================

def fc_get_movimientos_db():
    """Conexion a BD movimientos con reintentos"""
    import time
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Intentar pool primero
            try:
                conn = _get_movimientos_pool().getconn()
                conn.cursor().execute("SELECT 1")
                conn.rollback()
                return conn
            except Exception:
                pass

            # Fallback a conexion directa
            conn = psycopg2.connect(
                host=os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
                database='movimientos',
                user=os.environ.get('DB_USER', 'adminChios'),
                password=os.environ.get('DB_PASSWORD', 'Burger2023'),
                port=os.environ.get('DB_PORT', '5432'),
                sslmode='require',
                connect_timeout=15
            )
            return conn
        except Exception as e:
            print(f"fc_get_movimientos_db intento {attempt+1}/{max_retries}: {e}")
            if attempt < max_retries - 1:
                time.sleep(1)

    # Ultimo intento
    return psycopg2.connect(
        host=os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
        database='movimientos',
        user=os.environ.get('DB_USER', 'adminChios'),
        password=os.environ.get('DB_PASSWORD', 'Burger2023'),
        port=os.environ.get('DB_PORT', '5432'),
        sslmode='require',
        connect_timeout=20
    )

def fc_release_movimientos_db(conn):
    """Libera conexion de movimientos al pool"""
    try:
        if conn and not conn.closed:
            _get_movimientos_pool().putconn(conn)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

def fc_dia_deposito_tc(fecha_venta):
    """Calcula dia de deposito TC (2 dias habiles)"""
    dia = fecha_venta.weekday()
    if dia == 0: return fecha_venta + timedelta(days=2)  # Lun->Mie
    elif dia == 1: return fecha_venta + timedelta(days=2)  # Mar->Jue
    elif dia == 2: return fecha_venta + timedelta(days=2)  # Mie->Vie
    elif dia == 3: return fecha_venta + timedelta(days=4)  # Jue->Lun
    elif dia == 4: return fecha_venta + timedelta(days=3)  # Vie->Lun
    elif dia == 5: return fecha_venta + timedelta(days=3)  # Sab->Mar
    else: return fecha_venta + timedelta(days=2)  # Dom->Mar

def fc_dia_deposito_efectivo_g1(fecha_cobro):
    """Grupo 1 efectivo: libran Dom/Lun"""
    dia = fecha_cobro.weekday()
    if dia == 0: return fecha_cobro + timedelta(days=1)  # Lun->Mar
    elif dia == 1: return fecha_cobro + timedelta(days=1)  # Mar->Mie
    elif dia == 2: return fecha_cobro + timedelta(days=1)
    elif dia == 3: return fecha_cobro + timedelta(days=1)
    elif dia == 4: return fecha_cobro + timedelta(days=1)
    elif dia == 5: return fecha_cobro + timedelta(days=3)  # Sab->Mar
    else: return fecha_cobro + timedelta(days=2)  # Dom->Mar

def fc_dia_deposito_efectivo_g2(fecha_cobro):
    """Grupo 2 efectivo: libran Lun/Mar"""
    dia = fecha_cobro.weekday()
    if dia == 0: return fecha_cobro + timedelta(days=2)  # Lun->Mie
    elif dia == 1: return fecha_cobro + timedelta(days=1)  # Mar->Mie
    elif dia == 2: return fecha_cobro + timedelta(days=1)
    elif dia == 3: return fecha_cobro + timedelta(days=1)
    elif dia == 4: return fecha_cobro + timedelta(days=1)
    elif dia == 5: return fecha_cobro + timedelta(days=4)  # Sab->Mie
    else: return fecha_cobro + timedelta(days=3)  # Dom->Mie

def fc_dia_deposito_deuna(fecha_cobro):
    """DEUNA deposita en Pichincha dia siguiente (o martes si fin de semana)"""
    dia = fecha_cobro.weekday()
    if dia <= 3: return fecha_cobro + timedelta(days=1)  # Lun-Jue -> dia siguiente
    elif dia == 4: return fecha_cobro + timedelta(days=3)  # Vie->Lun
    elif dia == 5: return fecha_cobro + timedelta(days=3)  # Sab->Mar
    else: return fecha_cobro + timedelta(days=2)  # Dom->Mar

GRUPO1_EFECTIVO = ['REAL', 'FLOREANA', 'PORTUGAL', 'SANTO CACHON PORTUGAL']
GRUPO2_EFECTIVO = ['SIMON BOLON', 'SANTO CACHON REAL']

@app.route('/api/flujo-caja/datos', methods=['GET'])
def flujo_caja_datos():
    """Endpoint para obtener datos de flujo de caja"""
    conn = None
    try:
        # Leer parametros de query
        fecha_param = request.args.get('fecha', '')
        num_semanas = int(request.args.get('semanas', '5'))

        conn = fc_get_movimientos_db()
        cur = conn.cursor()

        # Determinar fecha de inicio (lunes)
        if fecha_param:
            try:
                lunes = datetime.strptime(fecha_param, '%Y-%m-%d').date()
                # Asegurar que sea lunes
                if lunes.weekday() != 0:
                    lunes = lunes - timedelta(days=lunes.weekday())
            except:
                hoy = datetime.now(TZ_ECUADOR).date()
                lunes = hoy - timedelta(days=hoy.weekday())
        else:
            hoy = datetime.now(TZ_ECUADOR).date()
            lunes = hoy - timedelta(days=hoy.weekday())

        fecha_inicio_datos = lunes - timedelta(days=7)

        # Generar semanas
        semanas = []
        fecha_fin_proyeccion = lunes + timedelta(weeks=num_semanas)
        for i in range(num_semanas):
            inicio = lunes + timedelta(weeks=i)
            fin = inicio + timedelta(days=6)
            num = inicio.isocalendar()[1]
            dias = [str(inicio + timedelta(days=j)) for j in range(7)]
            semanas.append({'num': num, 'inicio': str(inicio), 'fin': str(fin), 'dias': dias})

        hoy = datetime.now(TZ_ECUADOR).date()

        # ============ PROMEDIOS HISTORICOS (ultimas 8 semanas) ============
        fecha_hist_inicio = hoy - timedelta(weeks=8)

        # Promedio TC por dia de semana
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago ILIKE '%%tarjeta%%'
                AND fecha >= %s AND fecha < %s
            )
            SELECT EXTRACT(DOW FROM fecha) as dia_semana, AVG(total) as promedio
            FROM (SELECT fecha, SUM(valor) as total FROM unicos GROUP BY fecha) sub
            GROUP BY dia_semana
        ''', (fecha_hist_inicio, hoy))
        promedios_tc = {int(r[0]): float(r[1]) for r in cur.fetchall()}

        # Promedio Efectivo por dia de semana
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago = 'Efectivo'
                AND fecha >= %s AND fecha < %s
            )
            SELECT EXTRACT(DOW FROM fecha) as dia_semana, AVG(total) as promedio
            FROM (SELECT fecha, SUM(valor) as total FROM unicos GROUP BY fecha) sub
            GROUP BY dia_semana
        ''', (fecha_hist_inicio, hoy))
        promedios_efectivo = {int(r[0]): float(r[1]) for r in cur.fetchall()}

        # Promedio DEUNA por dia de semana
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago = 'Transferencia'
                AND cta_afectada = 'BANCO DEUNA' AND fecha >= %s AND fecha < %s
            )
            SELECT EXTRACT(DOW FROM fecha) as dia_semana, AVG(total) as promedio
            FROM (SELECT fecha, SUM(valor) as total FROM unicos GROUP BY fecha) sub
            GROUP BY dia_semana
        ''', (fecha_hist_inicio, hoy))
        promedios_deuna = {int(r[0]): float(r[1]) for r in cur.fetchall()}

        # ============ GENERAR VENTAS PROYECTADAS PARA DIAS SIN DATOS ============
        ventas_tc_proyectadas = {}
        ventas_efectivo_proyectadas = {}
        ventas_deuna_proyectadas = {}

        # Para cada dia desde inicio de datos hasta fin de proyeccion
        # Incluimos dias pasados recientes para que tengan proyecciones si no hay datos
        dia_actual = fecha_inicio_datos
        while dia_actual < fecha_fin_proyeccion:
            dia_str = str(dia_actual)
            dow = dia_actual.weekday()
            # PostgreSQL DOW: 0=domingo, 1=lunes... Python weekday: 0=lunes, 1=martes...
            dow_pg = (dow + 1) % 7  # Convertir a formato PostgreSQL

            # Proyectar todos los dias (pasados y futuros) - luego se filtran los que tienen datos reales
            if dow_pg in promedios_tc:
                ventas_tc_proyectadas[dia_str] = promedios_tc[dow_pg]
            if dow_pg in promedios_efectivo:
                ventas_efectivo_proyectadas[dia_str] = promedios_efectivo[dow_pg]
            if dow_pg in promedios_deuna:
                ventas_deuna_proyectadas[dia_str] = promedios_deuna[dow_pg]

            dia_actual += timedelta(days=1)

        # Obtener ventas TC
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago ILIKE '%%tarjeta%%' AND fecha >= %s
            )
            SELECT fecha, SUM(valor) as total FROM unicos GROUP BY fecha ORDER BY fecha
        ''', (fecha_inicio_datos,))
        ventas_tc = {str(r[0]): float(r[1]) for r in cur.fetchall()}

        # Agregar ventas proyectadas para dias sin datos
        for fecha_str, total in ventas_tc_proyectadas.items():
            if fecha_str not in ventas_tc:
                ventas_tc[fecha_str] = total

        # Calcular depositos TC (86% neto)
        depositos_tc = {}
        for fecha_str, total in ventas_tc.items():
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            fecha_dep = fc_dia_deposito_tc(fecha)
            dep_str = str(fecha_dep)
            if dep_str not in depositos_tc:
                depositos_tc[dep_str] = {'bruto': 0, 'neto': 0}
            depositos_tc[dep_str]['bruto'] += total
            depositos_tc[dep_str]['neto'] += total * 0.86

        # Obtener cobros efectivo
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, centro_costo, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago = 'Efectivo' AND fecha >= %s
            )
            SELECT fecha, centro_costo, SUM(valor) as total FROM unicos GROUP BY fecha, centro_costo ORDER BY fecha
        ''', (fecha_inicio_datos,))
        cobros_efectivo = {}
        for r in cur.fetchall():
            fecha = str(r[0])
            local = r[1]
            total = float(r[2])
            if fecha not in cobros_efectivo:
                cobros_efectivo[fecha] = {'G1': 0, 'G2': 0}
            if local in GRUPO1_EFECTIVO:
                cobros_efectivo[fecha]['G1'] += total
            elif local in GRUPO2_EFECTIVO:
                cobros_efectivo[fecha]['G2'] += total

        # Agregar efectivo proyectado (dividir 50/50 entre G1 y G2)
        for fecha_str, total in ventas_efectivo_proyectadas.items():
            if fecha_str not in cobros_efectivo:
                cobros_efectivo[fecha_str] = {'G1': total * 0.5, 'G2': total * 0.5}

        # Calcular depositos efectivo
        depositos_efectivo = {}
        for fecha_str, grupos in cobros_efectivo.items():
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            if grupos['G1'] > 0:
                dep = str(fc_dia_deposito_efectivo_g1(fecha))
                if dep not in depositos_efectivo:
                    depositos_efectivo[dep] = {'total': 0}
                depositos_efectivo[dep]['total'] += grupos['G1']
            if grupos['G2'] > 0:
                dep = str(fc_dia_deposito_efectivo_g2(fecha))
                if dep not in depositos_efectivo:
                    depositos_efectivo[dep] = {'total': 0}
                depositos_efectivo[dep]['total'] += grupos['G2']

        # Obtener cobros DEUNA
        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI' AND forma_cobro_pago = 'Transferencia'
                AND cta_afectada = 'BANCO DEUNA' AND fecha >= %s
            )
            SELECT fecha, SUM(valor) as total FROM unicos GROUP BY fecha ORDER BY fecha
        ''', (fecha_inicio_datos,))
        cobros_deuna = {str(r[0]): float(r[1]) for r in cur.fetchall()}

        # Agregar DEUNA proyectado
        for fecha_str, total in ventas_deuna_proyectadas.items():
            if fecha_str not in cobros_deuna:
                cobros_deuna[fecha_str] = total

        # Calcular depositos DEUNA
        depositos_deuna = {}
        for fecha_str, total in cobros_deuna.items():
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            dep = str(fc_dia_deposito_deuna(fecha))
            if dep not in depositos_deuna:
                depositos_deuna[dep] = {'total': 0}
            depositos_deuna[dep]['total'] += total

        # Calcular totales por semana
        totales_produbanco = {}
        totales_pichincha = {}
        for sem in semanas:
            total_prod = 0
            total_pich = 0
            for dia in sem['dias']:
                total_prod += depositos_tc.get(dia, {}).get('neto', 0)
                total_prod += depositos_efectivo.get(dia, {}).get('total', 0)
                total_pich += depositos_deuna.get(dia, {}).get('total', 0)
            totales_produbanco[sem['num']] = total_prod
            totales_pichincha[sem['num']] = total_pich

        return jsonify({
            'ok': True,
            'semanas': semanas,
            'depositos_tc': depositos_tc,
            'depositos_efectivo': depositos_efectivo,
            'depositos_deuna': depositos_deuna,
            'totales_produbanco': totales_produbanco,
            'totales_pichincha': totales_pichincha
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/guardar', methods=['POST'])
def flujo_caja_guardar():
    """Guardar datos de flujo de caja trabajados"""
    conn = None
    try:
        # Asegurar que las nuevas columnas existan
        conn_temp = fc_get_movimientos_db()
        cur_temp = conn_temp.cursor()
        cur_temp.execute('''
            ALTER TABLE flujo_caja_guardado
            ADD COLUMN IF NOT EXISTS saldo_produbanco NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS saldo_pichincha NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS plataformas JSONB DEFAULT '{}'::jsonb
        ''')
        conn_temp.commit()
        fc_release_movimientos_db(conn_temp)

        data = request.get_json()
        fecha_semana = data.get('fecha_semana')
        semana_num = data.get('semana_num')
        # Soportar tanto formato viejo (saldo_inicial) como nuevo (saldo_produbanco/pichincha)
        saldo_produbanco = data.get('saldo_produbanco', data.get('saldo_inicial', 0))
        saldo_pichincha = data.get('saldo_pichincha', 0)
        ajustes_tc = json.dumps(data.get('ajustes_tc', {}))
        ajustes_efectivo = json.dumps(data.get('ajustes_efectivo', {}))
        ajustes_deuna = json.dumps(data.get('ajustes_deuna', {}))
        traspasos = json.dumps(data.get('traspasos', {}))
        plataformas = json.dumps(data.get('plataformas', {}))
        egresos_dict = data.get('egresos', {})
        egresos = json.dumps(egresos_dict)
        usuario = data.get('usuario', 'admin')

        # PROTECCION contra perdida de datos: si los egresos entrantes no traen
        # ningun valor/saldo/dias/deuda, NO sobrescribir egresos ya guardados.
        egresos_entrantes_vacios = not any(
            (it.get('valores') or it.get('saldo') or it.get('dias') or it.get('deuda'))
            for items in egresos_dict.values() for it in items
        )

        conn = fc_get_movimientos_db()
        cur = conn.cursor()

        # Upsert: insertar o actualizar si ya existe.
        # Si los egresos entrantes estan vacios y la fila existente tiene datos,
        # se conservan los egresos existentes.
        cur.execute('''
            INSERT INTO flujo_caja_guardado
                (fecha_semana, semana_num, saldo_produbanco, saldo_pichincha, ajustes_tc, ajustes_efectivo,
                 ajustes_deuna, traspasos, plataformas, egresos, created_by, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (fecha_semana) DO UPDATE SET
                semana_num = EXCLUDED.semana_num,
                saldo_produbanco = EXCLUDED.saldo_produbanco,
                saldo_pichincha = EXCLUDED.saldo_pichincha,
                ajustes_tc = EXCLUDED.ajustes_tc,
                ajustes_efectivo = EXCLUDED.ajustes_efectivo,
                ajustes_deuna = EXCLUDED.ajustes_deuna,
                traspasos = EXCLUDED.traspasos,
                plataformas = EXCLUDED.plataformas,
                egresos = CASE
                    WHEN %s AND COALESCE(flujo_caja_guardado.egresos::text, '{}') NOT IN ('{}', 'null')
                    THEN flujo_caja_guardado.egresos
                    ELSE EXCLUDED.egresos
                END,
                updated_at = NOW()
            RETURNING id
        ''', (fecha_semana, semana_num, saldo_produbanco, saldo_pichincha, ajustes_tc, ajustes_efectivo,
              ajustes_deuna, traspasos, plataformas, egresos, usuario, egresos_entrantes_vacios))

        row_id = cur.fetchone()[0]
        conn.commit()

        return jsonify({'ok': True, 'id': row_id, 'mensaje': f'Semana {semana_num} guardada'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/cargar-guardado', methods=['GET'])
def flujo_caja_cargar_guardado():
    """Cargar datos guardados de flujo de caja para las semanas solicitadas"""
    conn = None
    try:
        fechas = request.args.get('fechas', '')  # Comma-separated list of dates
        if not fechas:
            return jsonify({'ok': True, 'guardados': {}})

        lista_fechas = [f.strip() for f in fechas.split(',')]

        conn = fc_get_movimientos_db()
        cur = conn.cursor()

        # Asegurar que las nuevas columnas existan
        cur.execute('''
            ALTER TABLE flujo_caja_guardado
            ADD COLUMN IF NOT EXISTS saldo_produbanco NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS saldo_pichincha NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS plataformas JSONB DEFAULT '{}'::jsonb
        ''')
        conn.commit()

        cur.execute('''
            SELECT fecha_semana, semana_num,
                   COALESCE(saldo_produbanco, saldo_inicial, 0) as saldo_produbanco,
                   COALESCE(saldo_pichincha, 0) as saldo_pichincha,
                   ajustes_tc, ajustes_efectivo, ajustes_deuna, traspasos,
                   COALESCE(plataformas, '{}'::jsonb) as plataformas,
                   egresos, updated_at
            FROM flujo_caja_guardado
            WHERE fecha_semana = ANY(%s::date[])
            ORDER BY updated_at ASC NULLS FIRST
        ''', (lista_fechas,))

        guardados = {}
        for row in cur.fetchall():
            fecha_str = str(row[0])
            guardados[fecha_str] = {
                'semana_num': row[1],
                'saldo_produbanco': float(row[2]) if row[2] else 0,
                'saldo_pichincha': float(row[3]) if row[3] else 0,
                'ajustes_tc': row[4] or {},
                'ajustes_efectivo': row[5] or {},
                'ajustes_deuna': row[6] or {},
                'traspasos': row[7] or {},
                'plataformas': row[8] or {},
                'egresos': row[9] or {},
                'updated_at': row[10].isoformat() if row[10] else None
            }

        return jsonify({'ok': True, 'guardados': guardados})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/proveedores', methods=['GET'])
def flujo_caja_proveedores_listar():
    """Listar proveedores del catalogo"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS fc_proveedores (
                id SERIAL PRIMARY KEY,
                nombre TEXT UNIQUE NOT NULL,
                nombre_comercial TEXT DEFAULT '',
                criticidad TEXT DEFAULT 'BAJO',
                dias_credito INTEGER DEFAULT 0,
                dia_despacho TEXT DEFAULT '',
                productos_servicios TEXT DEFAULT '',
                observaciones TEXT DEFAULT '',
                ruc TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        ''')
        # La tabla ya existia sin estas columnas en las BD desplegadas
        cur.execute("""ALTER TABLE fc_proveedores
            ADD COLUMN IF NOT EXISTS ruc TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS telefono TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS tipo_proveedor TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS apertura TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS credito_mixto_monto NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS credito_mixto_dias INTEGER DEFAULT 0""")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_fc_proveedores_ruc ON fc_proveedores(ruc) WHERE ruc <> ''")
        conn.commit()
        cur.execute('''SELECT id, nombre, nombre_comercial, criticidad, dias_credito, dia_despacho,
                              productos_servicios, observaciones, ruc, telefono, tipo_proveedor, apertura,
                              COALESCE(credito_mixto_monto, 0), COALESCE(credito_mixto_dias, 0)
                       FROM fc_proveedores ORDER BY nombre''')
        proveedores = []
        for r in cur.fetchall():
            proveedores.append({
                'id': r[0], 'nombre': r[1], 'nombre_comercial': r[2] or '',
                'criticidad': r[3] or 'BAJO', 'dias_credito': r[4] or 0,
                'dia_despacho': r[5] or '', 'productos_servicios': r[6] or '',
                'observaciones': r[7] or '', 'ruc': r[8] or '',
                'telefono': r[9] or '', 'tipo_proveedor': r[10] or '', 'apertura': r[11] or '',
                'credito_mixto_monto': float(r[12] or 0), 'credito_mixto_dias': r[13] or 0
            })
        return jsonify({'ok': True, 'proveedores': proveedores})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/proveedores', methods=['POST'])
def flujo_caja_proveedores_guardar():
    """Crear o actualizar proveedor"""
    conn = None
    try:
        data = request.get_json()
        nombre = data.get('nombre', '').strip()
        if not nombre:
            return jsonify({'error': 'Nombre requerido'}), 400
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("""ALTER TABLE fc_proveedores
            ADD COLUMN IF NOT EXISTS ruc TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS telefono TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS tipo_proveedor TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS apertura TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS credito_mixto_monto NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS credito_mixto_dias INTEGER DEFAULT 0""")
        cur.execute('''
            INSERT INTO fc_proveedores (nombre, nombre_comercial, criticidad, dias_credito, dia_despacho, productos_servicios, observaciones, ruc, telefono, tipo_proveedor, apertura, credito_mixto_monto, credito_mixto_dias)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (nombre) DO UPDATE SET
                nombre_comercial = EXCLUDED.nombre_comercial,
                criticidad = EXCLUDED.criticidad,
                dias_credito = EXCLUDED.dias_credito,
                dia_despacho = EXCLUDED.dia_despacho,
                productos_servicios = EXCLUDED.productos_servicios,
                observaciones = EXCLUDED.observaciones,
                ruc = EXCLUDED.ruc,
                telefono = EXCLUDED.telefono,
                tipo_proveedor = EXCLUDED.tipo_proveedor,
                apertura = EXCLUDED.apertura,
                credito_mixto_monto = EXCLUDED.credito_mixto_monto,
                credito_mixto_dias = EXCLUDED.credito_mixto_dias,
                updated_at = NOW()
            RETURNING id
        ''', (nombre, (data.get('nombre_comercial') or '').strip() or nombre, data.get('criticidad', 'BAJO'),
              data.get('dias_credito', 0), data.get('dia_despacho', ''),
              data.get('productos_servicios', ''), data.get('observaciones', ''),
              re.sub(r'[^0-9]', '', str(data.get('ruc') or '')),
              (data.get('telefono') or '').strip(),
              (data.get('tipo_proveedor') or '').strip().upper(),
              (data.get('apertura') or '').strip().upper(),
              _fc_num(data.get('credito_mixto_monto')),
              int(_fc_num(data.get('credito_mixto_dias')))))
        prov_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': prov_id})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/proveedores/bulk', methods=['POST'])
def flujo_caja_proveedores_bulk():
    """Guardar multiples proveedores de una vez"""
    conn = None
    try:
        data = request.get_json()
        proveedores = data.get('proveedores', [])
        if not proveedores:
            return jsonify({'error': 'Sin proveedores'}), 400
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute('''CREATE TABLE IF NOT EXISTS fc_proveedores (
            id SERIAL PRIMARY KEY, nombre TEXT UNIQUE NOT NULL,
            nombre_comercial TEXT DEFAULT '', criticidad TEXT DEFAULT 'BAJO',
            dias_credito INTEGER DEFAULT 0, dia_despacho TEXT DEFAULT '',
            productos_servicios TEXT DEFAULT '', observaciones TEXT DEFAULT '',
            ruc TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW(), updated_at TIMESTAMP DEFAULT NOW())''')
        cur.execute("""ALTER TABLE fc_proveedores
            ADD COLUMN IF NOT EXISTS ruc TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS telefono TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS tipo_proveedor TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS apertura TEXT DEFAULT '',
            ADD COLUMN IF NOT EXISTS credito_mixto_monto NUMERIC(14,2) DEFAULT 0,
            ADD COLUMN IF NOT EXISTS credito_mixto_dias INTEGER DEFAULT 0""")
        # Un solo INSERT con todas las filas. Antes era un execute() por proveedor
        # dentro del loop: con 217 proveedores eran 217 viajes a Azure y se sentia.
        filas = []
        vistos = set()
        for p in proveedores:
            nombre = (p.get('nombre') or '').strip()
            if not nombre: continue
            # execute_values manda todo en un statement: dos filas con el mismo
            # nombre chocarian entre si ("ON CONFLICT DO UPDATE command cannot
            # affect row a second time"). Gana la ultima, como en el loop viejo.
            clave = nombre.upper()
            if clave in vistos:
                filas = [f for f in filas if f[0].upper() != clave]
            vistos.add(clave)
            filas.append((
                nombre,
                (p.get('nombre_comercial') or '').strip() or nombre,
                p.get('criticidad') or 'BAJO',
                p.get('dias_credito') or 0,
                p.get('dia_despacho') or '',
                p.get('productos_servicios') or '',
                p.get('observaciones') or '',
                re.sub(r'[^0-9]', '', str(p.get('ruc') or '')),
                (p.get('telefono') or '').strip(),
                (p.get('tipo_proveedor') or '').strip().upper(),
                (p.get('apertura') or '').strip().upper(),
                _fc_num(p.get('credito_mixto_monto')),
                int(_fc_num(p.get('credito_mixto_dias'))),
            ))
        if not filas:
            return jsonify({'error': 'Sin proveedores con nombre valido'}), 400
        execute_values(cur, '''
            INSERT INTO fc_proveedores (nombre, nombre_comercial, criticidad, dias_credito, dia_despacho, productos_servicios, observaciones, ruc, telefono, tipo_proveedor, apertura, credito_mixto_monto, credito_mixto_dias)
            VALUES %s
            ON CONFLICT (nombre) DO UPDATE SET
                nombre_comercial = COALESCE(NULLIF(EXCLUDED.nombre_comercial, ''), fc_proveedores.nombre_comercial),
                criticidad = CASE WHEN EXCLUDED.criticidad != 'BAJO' THEN EXCLUDED.criticidad ELSE fc_proveedores.criticidad END,
                dias_credito = CASE WHEN EXCLUDED.dias_credito > 0 THEN EXCLUDED.dias_credito ELSE fc_proveedores.dias_credito END,
                dia_despacho = COALESCE(NULLIF(EXCLUDED.dia_despacho, ''), fc_proveedores.dia_despacho),
                productos_servicios = COALESCE(NULLIF(EXCLUDED.productos_servicios, ''), fc_proveedores.productos_servicios),
                observaciones = COALESCE(NULLIF(EXCLUDED.observaciones, ''), fc_proveedores.observaciones),
                ruc = COALESCE(NULLIF(EXCLUDED.ruc, ''), fc_proveedores.ruc),
                telefono = COALESCE(NULLIF(EXCLUDED.telefono, ''), fc_proveedores.telefono),
                tipo_proveedor = COALESCE(NULLIF(EXCLUDED.tipo_proveedor, ''), fc_proveedores.tipo_proveedor),
                apertura = COALESCE(NULLIF(EXCLUDED.apertura, ''), fc_proveedores.apertura),
                credito_mixto_monto = CASE WHEN EXCLUDED.credito_mixto_monto > 0 THEN EXCLUDED.credito_mixto_monto ELSE fc_proveedores.credito_mixto_monto END,
                credito_mixto_dias = CASE WHEN EXCLUDED.credito_mixto_dias > 0 THEN EXCLUDED.credito_mixto_dias ELSE fc_proveedores.credito_mixto_dias END,
                updated_at = NOW()
        ''', filas, page_size=250)
        guardados = len(filas)
        conn.commit()
        return jsonify({'ok': True, 'guardados': guardados})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/proveedores/<int:prov_id>', methods=['DELETE'])
def flujo_caja_proveedores_eliminar(prov_id):
    """Eliminar proveedor del catalogo"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute('DELETE FROM fc_proveedores WHERE id = %s RETURNING nombre', (prov_id,))
        row = cur.fetchone()
        conn.commit()
        if row: return jsonify({'ok': True, 'eliminado': row[0]})
        return jsonify({'error': 'No encontrado'}), 404
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def _fc_num(valor):
    """Numero tolerante: el front manda strings, vacios y hasta '1.200,00'."""
    if valor is None:
        return 0
    if isinstance(valor, (int, float)):
        return float(valor)
    limpio = str(valor).strip().replace('$', '').replace(' ', '')
    if not limpio:
        return 0
    # Si trae los dos separadores, la coma es de miles; si solo trae coma, es decimal
    if ',' in limpio and '.' in limpio:
        limpio = limpio.replace(',', '')
    elif ',' in limpio:
        limpio = limpio.replace(',', '.')
    try:
        return float(limpio)
    except ValueError:
        return 0


def _fc_crear_tabla_recurrentes(cur):
    cur.execute('''
        CREATE TABLE IF NOT EXISTS fc_pagos_recurrentes (
            id SERIAL PRIMARY KEY,
            nombre TEXT NOT NULL,
            grupo TEXT DEFAULT 'pagos-fijos',
            monto NUMERIC(14,2) DEFAULT 0,
            frecuencia TEXT DEFAULT 'mensual',
            dia_mes INTEGER DEFAULT 1,
            dia_semana INTEGER DEFAULT 0,
            banco TEXT DEFAULT 'produbanco',
            activo BOOLEAN DEFAULT TRUE,
            observaciones TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    # Vigencia: desde cuando rige, hasta cuando (NULL = indefinido) y cuotas pactadas
    cur.execute("ALTER TABLE fc_pagos_recurrentes ADD COLUMN IF NOT EXISTS fecha_inicio DATE")
    cur.execute("ALTER TABLE fc_pagos_recurrentes ADD COLUMN IF NOT EXISTS fecha_fin DATE")
    cur.execute("ALTER TABLE fc_pagos_recurrentes ADD COLUMN IF NOT EXISTS total_cuotas INTEGER DEFAULT 0")
    # RUC del beneficiario: un pago fijo a nombre de una persona (p.ej. un arriendo)
    # no se puede validar solo por el nombre. Con el RUC se amarra a fc_proveedores.
    cur.execute("ALTER TABLE fc_pagos_recurrentes ADD COLUMN IF NOT EXISTS ruc TEXT DEFAULT ''")


@app.route('/api/flujo-caja/recurrentes', methods=['GET'])
def flujo_caja_recurrentes_listar():
    """Listar pagos recurrentes"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_recurrentes(cur)
        conn.commit()
        cur.execute('''SELECT id, nombre, grupo, monto, frecuencia, dia_mes, dia_semana, banco,
                       activo, observaciones, fecha_inicio, fecha_fin, total_cuotas,
                       COALESCE(ruc, '')
                       FROM fc_pagos_recurrentes ORDER BY grupo, nombre''')
        pagos = []
        for r in cur.fetchall():
            pagos.append({
                'id': r[0], 'nombre': r[1], 'grupo': r[2], 'monto': float(r[3] or 0),
                'frecuencia': r[4] or 'mensual', 'dia_mes': r[5] or 1,
                'dia_semana': r[6] or 0, 'banco': r[7] or 'produbanco',
                'activo': r[8], 'observaciones': r[9] or '',
                'fecha_inicio': r[10].isoformat() if r[10] else '',
                'fecha_fin': r[11].isoformat() if r[11] else '',
                'total_cuotas': r[12] or 0,
                'ruc': r[13] or ''
            })
        return jsonify({'ok': True, 'pagos': pagos})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/recurrentes', methods=['POST'])
def flujo_caja_recurrentes_guardar():
    """Crear o actualizar pago recurrente"""
    conn = None
    try:
        data = request.get_json()
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_recurrentes(cur)
        pago_id = data.get('id', 0)
        f_ini = (data.get('fecha_inicio') or '').strip() or None
        f_fin = (data.get('fecha_fin') or '').strip() or None
        cuotas = int(data.get('total_cuotas') or 0)
        # Solo digitos: el RUC es la llave contra fc_proveedores y Contifico
        ruc = ''.join(c for c in str(data.get('ruc') or '') if c.isdigit())
        if pago_id and pago_id > 0:
            cur.execute('''
                UPDATE fc_pagos_recurrentes SET nombre=%s, grupo=%s, monto=%s, frecuencia=%s,
                    dia_mes=%s, dia_semana=%s, banco=%s, activo=%s, observaciones=%s,
                    fecha_inicio=%s, fecha_fin=%s, total_cuotas=%s, ruc=%s, updated_at=NOW()
                WHERE id=%s RETURNING id
            ''', (data.get('nombre',''), data.get('grupo','pagos-fijos'), data.get('monto',0),
                  data.get('frecuencia','mensual'), data.get('dia_mes',1), data.get('dia_semana',0),
                  data.get('banco','produbanco'), data.get('activo',True), data.get('observaciones',''),
                  f_ini, f_fin, cuotas, ruc, pago_id))
        else:
            cur.execute('''
                INSERT INTO fc_pagos_recurrentes (nombre, grupo, monto, frecuencia, dia_mes, dia_semana,
                    banco, activo, observaciones, fecha_inicio, fecha_fin, total_cuotas, ruc)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
            ''', (data.get('nombre',''), data.get('grupo','pagos-fijos'), data.get('monto',0),
                  data.get('frecuencia','mensual'), data.get('dia_mes',1), data.get('dia_semana',0),
                  data.get('banco','produbanco'), data.get('activo',True), data.get('observaciones',''),
                  f_ini, f_fin, cuotas, ruc))
        pago_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': pago_id})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/recurrentes/<int:pago_id>', methods=['DELETE'])
def flujo_caja_recurrentes_eliminar(pago_id):
    """Eliminar pago recurrente"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute('DELETE FROM fc_pagos_recurrentes WHERE id = %s RETURNING nombre', (pago_id,))
        row = cur.fetchone()
        conn.commit()
        if row: return jsonify({'ok': True, 'eliminado': row[0]})
        return jsonify({'error': 'No encontrado'}), 404
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def _fc_crear_tabla_eliminados(cur):
    cur.execute('''CREATE TABLE IF NOT EXISTS fc_egresos_eliminados (
        grupo TEXT NOT NULL,
        nombre TEXT NOT NULL,
        eliminado_desde DATE NOT NULL,
        eliminado_por TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY (grupo, nombre))''')


@app.route('/api/flujo-caja/egresos-eliminados', methods=['GET'])
def flujo_caja_eliminados_get():
    """Items de egreso dados de baja con su fecha de vigencia"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_eliminados(cur)
        conn.commit()
        cur.execute('SELECT grupo, nombre, eliminado_desde FROM fc_egresos_eliminados')
        eliminados = [{'grupo': r[0], 'nombre': r[1], 'eliminado_desde': r[2].isoformat()} for r in cur.fetchall()]
        return jsonify({'ok': True, 'eliminados': eliminados})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/egresos-eliminados', methods=['POST'])
def flujo_caja_eliminados_marcar():
    """Marcar item como eliminado desde una semana (conserva el historico anterior)"""
    conn = None
    try:
        data = request.get_json()
        grupo = (data.get('grupo') or '').strip()
        nombre = (data.get('nombre') or '').strip()
        desde = (data.get('eliminado_desde') or '').strip()
        if not grupo or not nombre or not desde:
            return jsonify({'error': 'grupo, nombre y eliminado_desde son requeridos'}), 400
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_eliminados(cur)
        # El frontend compara nombres con trim().toUpperCase(), asi que una baja
        # escrita con otra capitalizacion dejaria dos filas y ganaria la primera.
        # Se limpian las variantes antes de insertar la exacta.
        cur.execute('''DELETE FROM fc_egresos_eliminados
            WHERE UPPER(TRIM(grupo)) = UPPER(TRIM(%s))
              AND UPPER(TRIM(nombre)) = UPPER(TRIM(%s))
              AND nombre <> %s''', (grupo, nombre, nombre))
        cur.execute('''INSERT INTO fc_egresos_eliminados (grupo, nombre, eliminado_desde, eliminado_por)
            VALUES (%s, %s, %s, %s) ON CONFLICT (grupo, nombre) DO UPDATE SET
            eliminado_desde = EXCLUDED.eliminado_desde, created_at = NOW()''',
            (grupo, nombre, desde, data.get('usuario', '')))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/egresos-eliminados/reactivar', methods=['POST'])
def flujo_caja_eliminados_reactivar():
    """Quitar la baja de un item (vuelve a aparecer en todas las semanas)"""
    conn = None
    try:
        data = request.get_json()
        grupo = (data.get('grupo') or '').strip()
        nombre = (data.get('nombre') or '').strip()
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_eliminados(cur)
        # Case/espacio-insensible: el nombre puede volver escrito distinto y aun
        # asi tiene que reactivar la misma baja (igual que compara el frontend).
        cur.execute('''DELETE FROM fc_egresos_eliminados
            WHERE UPPER(TRIM(grupo)) = UPPER(TRIM(%s))
              AND UPPER(TRIM(nombre)) = UPPER(TRIM(%s))''', (grupo, nombre))
        conn.commit()
        return jsonify({'ok': True, 'reactivados': cur.rowcount})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def _fc_crear_tabla_ahorro(cur):
    cur.execute('''CREATE TABLE IF NOT EXISTS fc_ahorro_deuda (
        id SMALLINT PRIMARY KEY DEFAULT 1,
        ahorro_semanal NUMERIC(14,2) DEFAULT 0,
        aportes_extra JSONB DEFAULT '{}'::jsonb,
        updated_at TIMESTAMP DEFAULT NOW())''')


@app.route('/api/flujo-caja/ahorro-deuda', methods=['GET'])
def flujo_caja_ahorro_deuda_get():
    """Config de ahorro semanal destinado a pago de deuda vencida"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_ahorro(cur)
        conn.commit()
        cur.execute('SELECT ahorro_semanal, aportes_extra FROM fc_ahorro_deuda WHERE id = 1')
        row = cur.fetchone()
        if row:
            return jsonify({'ok': True, 'ahorro_semanal': float(row[0] or 0), 'aportes_extra': row[1] or {}})
        return jsonify({'ok': True, 'ahorro_semanal': 0, 'aportes_extra': {}})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/ahorro-deuda', methods=['POST'])
def flujo_caja_ahorro_deuda_guardar():
    """Guardar config de ahorro semanal para deuda"""
    conn = None
    try:
        data = request.get_json()
        ahorro = float(data.get('ahorro_semanal', 0) or 0)
        aportes = data.get('aportes_extra', {}) or {}
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_ahorro(cur)
        cur.execute('''INSERT INTO fc_ahorro_deuda (id, ahorro_semanal, aportes_extra)
            VALUES (1, %s, %s::jsonb) ON CONFLICT (id) DO UPDATE SET
            ahorro_semanal = EXCLUDED.ahorro_semanal,
            aportes_extra = EXCLUDED.aportes_extra,
            updated_at = NOW()''', (ahorro, json.dumps(aportes)))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def _fc_crear_tabla_cartera(cur):
    """Cartera por pagar de CADA semana. El XLS que se carga es el de esa semana:
    el proveedor que no viene en el archivo de la semana no debe aparecer en ella."""
    cur.execute('''CREATE TABLE IF NOT EXISTS fc_cartera_semana (
        semana_inicio DATE NOT NULL,
        proveedor TEXT NOT NULL,
        ruc TEXT DEFAULT '',
        saldo NUMERIC(14,2) DEFAULT 0,
        facturas INTEGER DEFAULT 0,
        detalle JSONB DEFAULT '[]'::jsonb,
        cargado_at TIMESTAMP DEFAULT NOW(),
        PRIMARY KEY (semana_inicio, proveedor))''')
    # El detalle de facturas vive aqui, UNA sola vez por semana y proveedor, en vez
    # de repetirse dentro de los egresos de cada semana visible
    cur.execute("ALTER TABLE fc_cartera_semana ADD COLUMN IF NOT EXISTS detalle JSONB DEFAULT '[]'::jsonb")


@app.route('/api/flujo-caja/cartera-semana', methods=['GET'])
def flujo_caja_cartera_semana_get():
    """Proveedores de la cartera de las semanas pedidas: ?fechas=2026-08-17,2026-08-24"""
    conn = None
    try:
        fechas = [f.strip() for f in (request.args.get('fechas') or '').split(',') if f.strip()]
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_cartera(cur)
        conn.commit()
        if not fechas:
            return jsonify({'ok': True, 'semanas': {}})
        cur.execute('''SELECT semana_inicio, proveedor, ruc, saldo, facturas, detalle
                       FROM fc_cartera_semana WHERE semana_inicio = ANY(%s::date[])
                       ORDER BY semana_inicio, proveedor''', (fechas,))
        semanas = {}
        for r in cur.fetchall():
            semanas.setdefault(r[0].isoformat(), []).append({
                'proveedor': r[1], 'ruc': r[2] or '',
                'saldo': float(r[3] or 0), 'facturas': r[4] or 0,
                'detalle': r[5] or []})
        return jsonify({'ok': True, 'semanas': semanas})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/cartera-semana', methods=['POST'])
def flujo_caja_cartera_semana_guardar():
    """Registra la cartera de una semana. REEMPLAZA la que hubiera: el archivo manda."""
    conn = None
    try:
        data = request.get_json()
        semana = (data.get('semana_inicio') or '').strip()
        proveedores = data.get('proveedores') or []
        if not semana:
            return jsonify({'error': 'semana_inicio requerida'}), 400
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_cartera(cur)
        cur.execute('DELETE FROM fc_cartera_semana WHERE semana_inicio = %s', (semana,))
        borradas = cur.rowcount
        filas, vistos = [], set()
        for p in proveedores:
            nombre = (p.get('proveedor') or '').strip()
            if not nombre or nombre.upper() in vistos:
                continue
            vistos.add(nombre.upper())
            detalle = p.get('detalle') or []
            filas.append((semana, nombre,
                          re.sub(r'[^0-9]', '', str(p.get('ruc') or '')),
                          float(p.get('saldo') or 0),
                          int(p.get('facturas') or len(detalle)),
                          json.dumps(detalle)))
        if filas:
            execute_values(cur, '''INSERT INTO fc_cartera_semana
                (semana_inicio, proveedor, ruc, saldo, facturas, detalle)
                VALUES %s''', filas, page_size=500,
                template='(%s, %s, %s, %s, %s, %s::jsonb)')
        conn.commit()
        return jsonify({'ok': True, 'guardados': len(filas), 'reemplazados': borradas})
    except Exception as e:
        if conn: conn.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/cartera-semana/detalle', methods=['POST'])
def flujo_caja_cartera_detalle_guardar():
    """Actualiza SOLO el detalle de facturas de proveedores de una semana.

    Se usa al guardar el flujo: el usuario pudo cambiar fecha de pago, abono o el
    vencimiento de una factura y eso tiene que sobrevivir sin volver a cargar el XLS.
    No borra la cartera: solo toca las filas que se mandan.
    """
    conn = None
    try:
        data = request.get_json()
        semana = (data.get('semana_inicio') or '').strip()
        proveedores = data.get('proveedores') or []
        if not semana:
            return jsonify({'error': 'semana_inicio requerida'}), 400
        if not proveedores:
            return jsonify({'ok': True, 'actualizados': 0})
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_cartera(cur)
        filas = []
        for p in proveedores:
            nombre = (p.get('proveedor') or '').strip()
            if not nombre:
                continue
            detalle = p.get('detalle') or []
            filas.append((semana, nombre, json.dumps(detalle), len(detalle)))
        if filas:
            execute_values(cur, """
                UPDATE fc_cartera_semana AS c
                SET detalle = d.detalle::jsonb, facturas = d.n
                FROM (VALUES %s) AS d(semana, proveedor, detalle, n)
                WHERE c.semana_inicio = d.semana::date
                  AND UPPER(TRIM(c.proveedor)) = UPPER(TRIM(d.proveedor))
            """, filas, page_size=300)
        actualizados = cur.rowcount
        conn.commit()
        return jsonify({'ok': True, 'actualizados': actualizados})
    except Exception as e:
        if conn: conn.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def _fc_crear_tabla_liquidez(cur):
    cur.execute('''CREATE TABLE IF NOT EXISTS fc_config_liquidez (
        id SMALLINT PRIMARY KEY DEFAULT 1,
        minimo_produbanco NUMERIC(14,2) DEFAULT 0,
        minimo_pichincha NUMERIC(14,2) DEFAULT 0,
        semanas_cobertura NUMERIC(5,2) DEFAULT 2,
        updated_at TIMESTAMP DEFAULT NOW())''')
    # Acumulado del fondo ANEFI. NO es caja disponible: es plata apartada que genera
    # intereses, asi que no entra en saldos, cobertura ni alertas. Vive aqui para
    # poder mostrarla al lado de los bancos sin contaminar el flujo.
    cur.execute("ALTER TABLE fc_config_liquidez ADD COLUMN IF NOT EXISTS saldo_anefi NUMERIC(14,2) DEFAULT 0")
    # Fecha a la que corresponde ese saldo (el corte de la cartola). Sin esto no se
    # puede saber si un interes ya venia incluido en el saldo o hay que sumarlo.
    cur.execute("ALTER TABLE fc_config_liquidez ADD COLUMN IF NOT EXISTS saldo_anefi_fecha DATE")


def _fc_crear_tabla_anefi(cur):
    """Intereses y ajustes del fondo. NO son movimientos de caja: no salen de ningun
    banco, son ganancia dentro del fondo. Solo suman al acumulado."""
    cur.execute('''CREATE TABLE IF NOT EXISTS fc_anefi_movimientos (
        id SERIAL PRIMARY KEY,
        fecha DATE NOT NULL,
        monto NUMERIC(14,2) NOT NULL,
        concepto TEXT DEFAULT '',
        created_at TIMESTAMP DEFAULT NOW())''')
    cur.execute("CREATE INDEX IF NOT EXISTS idx_fc_anefi_fecha ON fc_anefi_movimientos(fecha)")


@app.route('/api/flujo-caja/config-liquidez', methods=['GET'])
def flujo_caja_config_liquidez_get():
    """Saldo minimo por banco para las alertas de liquidez"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_liquidez(cur)
        conn.commit()
        cur.execute('''SELECT minimo_produbanco, minimo_pichincha, semanas_cobertura,
                              COALESCE(saldo_anefi, 0), saldo_anefi_fecha
                       FROM fc_config_liquidez WHERE id = 1''')
        row = cur.fetchone()
        if row:
            return jsonify({'ok': True, 'minimo_produbanco': float(row[0] or 0),
                            'minimo_pichincha': float(row[1] or 0),
                            'semanas_cobertura': float(row[2] or 2),
                            'saldo_anefi': float(row[3] or 0),
                            'saldo_anefi_fecha': row[4].isoformat() if row[4] else ''})
        return jsonify({'ok': True, 'minimo_produbanco': 0, 'minimo_pichincha': 0,
                        'semanas_cobertura': 2, 'saldo_anefi': 0, 'saldo_anefi_fecha': ''})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/config-liquidez', methods=['POST'])
def flujo_caja_config_liquidez_guardar():
    """Guardar los minimos por banco"""
    conn = None
    try:
        data = request.get_json()
        min_pro = float(data.get('minimo_produbanco', 0) or 0)
        min_pich = float(data.get('minimo_pichincha', 0) or 0)
        semanas = float(data.get('semanas_cobertura', 2) or 2)
        anefi = _fc_num(data.get('saldo_anefi', 0))
        anefi_fecha = (data.get('saldo_anefi_fecha') or '').strip() or None
        if min_pro < 0 or min_pich < 0 or semanas < 0:
            return jsonify({'error': 'Los valores no pueden ser negativos'}), 400
        if anefi < 0:
            return jsonify({'error': 'El acumulado de ANEFI no puede ser negativo'}), 400
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_liquidez(cur)
        cur.execute('''INSERT INTO fc_config_liquidez (id, minimo_produbanco, minimo_pichincha, semanas_cobertura, saldo_anefi, saldo_anefi_fecha)
            VALUES (1, %s, %s, %s, %s, %s) ON CONFLICT (id) DO UPDATE SET
            minimo_produbanco = EXCLUDED.minimo_produbanco,
            minimo_pichincha = EXCLUDED.minimo_pichincha,
            semanas_cobertura = EXCLUDED.semanas_cobertura,
            saldo_anefi = EXCLUDED.saldo_anefi,
            saldo_anefi_fecha = EXCLUDED.saldo_anefi_fecha,
            updated_at = NOW()''', (min_pro, min_pich, semanas, anefi, anefi_fecha))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/anefi-resumen', methods=['GET'])
def flujo_caja_anefi_resumen():
    """Estado del fondo ANEFI para mostrarlo fuera del flujo de caja.

    El acumulado NO es caja disponible: es plata apartada que genera intereses.
    Los aportes salen del banco (por eso siguen siendo egreso en el flujo), pero
    aqui se suman al fondo. Solo cuenta lo POSTERIOR a la fecha de corte del
    saldo: lo anterior ya venia dentro de la cartola.
    """
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_liquidez(cur)
        _fc_crear_tabla_anefi(cur)
        conn.commit()

        cur.execute("SELECT COALESCE(saldo_anefi, 0), saldo_anefi_fecha FROM fc_config_liquidez WHERE id = 1")
        row = cur.fetchone()
        saldo = float(row[0] or 0) if row else 0.0
        corte = row[1].isoformat() if (row and row[1]) else ''

        # Intereses y ajustes registrados despues del corte
        if corte:
            cur.execute("SELECT COALESCE(SUM(monto), 0) FROM fc_anefi_movimientos WHERE fecha > %s", (corte,))
        else:
            cur.execute("SELECT COALESCE(SUM(monto), 0) FROM fc_anefi_movimientos")
        intereses = float(cur.fetchone()[0] or 0)

        # Aportes y rescates guardados en el flujo (grupo de inversiones)
        cur.execute("SELECT egresos FROM flujo_caja_guardado WHERE egresos IS NOT NULL")
        aportes = 0.0
        detalle_aportes = []
        for (eg,) in cur.fetchall():
            if isinstance(eg, str):
                eg = json.loads(eg)
            for item in (eg or {}).get('inversiones', []) or []:
                for dia, monto in (item.get('valores') or {}).items():
                    if corte and dia <= corte:
                        continue
                    val = float(monto or 0)
                    if not val:
                        continue
                    aportes += val
                    detalle_aportes.append({'fecha': dia, 'monto': val,
                                            'concepto': item.get('nombre', 'ANEFI')})
        detalle_aportes.sort(key=lambda x: x['fecha'])

        return jsonify({'ok': True, 'saldo': saldo, 'fecha_corte': corte,
                        'intereses': intereses, 'aportes': aportes,
                        'total': saldo + intereses + aportes,
                        'detalle_aportes': detalle_aportes})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/anefi-movimientos', methods=['GET'])
def flujo_caja_anefi_listar():
    """Intereses y ajustes registrados del fondo ANEFI"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_anefi(cur)
        conn.commit()
        cur.execute('''SELECT id, fecha, monto, concepto FROM fc_anefi_movimientos
                       ORDER BY fecha DESC, id DESC''')
        movs = [{'id': r[0], 'fecha': r[1].isoformat(), 'monto': float(r[2] or 0),
                 'concepto': r[3] or ''} for r in cur.fetchall()]
        return jsonify({'ok': True, 'movimientos': movs})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/anefi-movimientos', methods=['POST'])
def flujo_caja_anefi_guardar():
    """Registrar un interes o ajuste del fondo"""
    conn = None
    try:
        data = request.get_json()
        fecha = (data.get('fecha') or '').strip()
        if not fecha:
            return jsonify({'error': 'La fecha es obligatoria'}), 400
        monto = _fc_num(data.get('monto'))
        if abs(monto) < 0.005:
            return jsonify({'error': 'El monto no puede ser cero'}), 400
        concepto = (data.get('concepto') or '').strip()[:200]
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        _fc_crear_tabla_anefi(cur)
        cur.execute('''INSERT INTO fc_anefi_movimientos (fecha, monto, concepto)
                       VALUES (%s, %s, %s) RETURNING id''', (fecha, monto, concepto))
        mov_id = cur.fetchone()[0]
        conn.commit()
        return jsonify({'ok': True, 'id': mov_id})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/anefi-movimientos/<int:mov_id>', methods=['DELETE'])
def flujo_caja_anefi_eliminar(mov_id):
    """Borrar un interes o ajuste mal registrado"""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute('DELETE FROM fc_anefi_movimientos WHERE id = %s RETURNING concepto', (mov_id,))
        row = cur.fetchone()
        conn.commit()
        if row:
            return jsonify({'ok': True})
        return jsonify({'error': 'No encontrado'}), 404
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/ventas-por-local', methods=['GET'])
def flujo_caja_ventas_por_local():
    """Ventas mensuales por centro de costo para grafico"""
    conn = None
    try:
        meses = int(request.args.get('meses', '12'))
        conn = fc_get_movimientos_db()
        cur = conn.cursor()

        cur.execute('''
            WITH unicos AS (
                SELECT DISTINCT ON (documento_id, fecha, valor) fecha, centro_costo, valor
                FROM contifico_cobrospagos
                WHERE tipo_registro = 'CLI'
                AND fecha >= CURRENT_DATE - (%s || ' months')::interval
                AND centro_costo IS NOT NULL AND centro_costo != ''
            )
            SELECT TO_CHAR(fecha, 'YYYY-MM') as mes,
                   centro_costo,
                   SUM(valor) as total
            FROM unicos
            GROUP BY mes, centro_costo
            ORDER BY mes, centro_costo
        ''', (meses,))

        datos = {}
        locales_set = set()
        for r in cur.fetchall():
            mes = r[0]
            local = r[1]
            total = float(r[2])
            if mes not in datos:
                datos[mes] = {}
            datos[mes][local] = total
            locales_set.add(local)

        return jsonify({
            'ok': True,
            'datos': datos,
            'locales': sorted(list(locales_set)),
            'meses': sorted(datos.keys())
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/flujo-caja/facturas-proveedor', methods=['GET'])
def flujo_caja_facturas_proveedor():
    """Buscar facturas de compras pendientes de un proveedor"""
    conn = None
    try:
        nombre = request.args.get('nombre', '')
        if not nombre or len(nombre) < 3:
            return jsonify({'ok': False, 'error': 'Nombre muy corto (min 3 caracteres)'})

        conn = fc_get_movimientos_db()
        cur = conn.cursor()

        # Buscar en fact_detallada_compras (facturas de compra de proveedores)
        cur.execute('''
            SELECT DISTINCT numero_documento, fecha_emision::date, total,
                   razon_social, autorizacion
            FROM fact_detallada_compras
            WHERE razon_social ILIKE %s
            AND fecha_emision >= CURRENT_DATE - INTERVAL '6 months'
            ORDER BY fecha_emision DESC
            LIMIT 100
        ''', (f'%{nombre}%',))

        facturas = []
        for r in cur.fetchall():
            facturas.append({
                'num': r[0] or '',
                'fecha': str(r[1]) if r[1] else '',
                'monto': float(r[2]) if r[2] else 0,
                'proveedor': r[3] or '',
                'autorizacion': r[4] or '',
                'vencimiento': ''
            })

        return jsonify({'ok': True, 'facturas': facturas, 'total': len(facturas)})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


# =====================================================
# VOUCHER SCANNER - Endpoint comparacion Azure
# =====================================================
@app.route('/api/vouchers/cobros-tarjeta', methods=['GET'])
def vs_cobros_tarjeta():
    """Obtiene cobros con tarjeta de Azure para comparar con vouchers Supabase"""
    fecha = request.args.get('fecha')
    local = request.args.get('local')  # Filtro opcional por local
    if not fecha:
        return jsonify({'error': 'Fecha requerida'}), 400

    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Mapeo de IDs Supabase a nombres de local en Azure
        mapeo_locales = {
            'vj6e98Ao7UrAEdWB': 'PORTUGAL',
            'gDGe76ymMCn21dn2': 'REAL',
            'pXKdwmJWZcgk3agW': 'SANTO CACHON PORTUGAL',
            '4pzb864o4uBWpeEw': 'SANTO CACHON REAL',
            'rQzaOyDP4cElmdp4': 'SIMON BOLON'
        }
        local_azure = mapeo_locales.get(local) if local else None

        # Cobros con tarjeta - detalle individual
        if local_azure:
            cur.execute('''
                SELECT
                    codigo_comprobante as factura,
                    centro_costo as local,
                    valor,
                    persona,
                    lote
                FROM contifico_cobrospagos
                WHERE fecha = %s AND forma_cobro_pago = 'Tarjeta Credito' AND centro_costo = %s
                ORDER BY lote, codigo_comprobante
            ''', (fecha, local_azure))
        else:
            cur.execute('''
                SELECT
                    codigo_comprobante as factura,
                    centro_costo as local,
                    valor,
                    persona,
                    lote
                FROM contifico_cobrospagos
                WHERE fecha = %s AND forma_cobro_pago = 'Tarjeta Credito'
                ORDER BY centro_costo, lote, codigo_comprobante
            ''', (fecha,))

        registros = []
        total_monto = 0
        for row in cur.fetchall():
            factura = row['factura'] or ''
            local = row['local'] or '(sin local)'
            valor = float(row['valor'] or 0)
            registros.append({
                'factura': factura,
                'local': local,
                'valor': valor,
                'persona': row['persona'] or '',
                'lote': row['lote'] or ''
            })
            total_monto += valor

        # Obtener lotes únicos con totales
        lotes_unicos = {}
        for r in registros:
            lote = r['lote'] or ''
            if lote:
                if lote not in lotes_unicos:
                    lotes_unicos[lote] = {'lote': lote, 'local': r['local'], 'txn': 0, 'total': 0}
                lotes_unicos[lote]['txn'] += 1
                lotes_unicos[lote]['total'] += r['valor']

        return jsonify({
            'ok': True,
            'fecha': fecha,
            'registros': registros,
            'lotes': list(lotes_unicos.values()),
            'total_transacciones': len(registros),
            'total_monto': total_monto
        })

    except Exception as e:
        print(f"Error en /api/vouchers/cobros-tarjeta: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/vouchers/cobros-deuna', methods=['GET'])
def vs_cobros_deuna():
    """Obtiene cobros DEUNA de Azure (contifico_cobrospagos) para comparar con vouchers Supabase"""
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha:
        return jsonify({'error': 'Fecha requerida'}), 400

    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Mapeo de IDs Supabase a nombres de local en Azure
        mapeo_locales = {
            'vj6e98Ao7UrAEdWB': 'PORTUGAL',
            'gDGe76ymMCn21dn2': 'REAL',
            'pXKdwmJWZcgk3agW': 'SANTO CACHON PORTUGAL',
            '4pzb864o4uBWpeEw': 'SANTO CACHON REAL',
            'rQzaOyDP4cElmdp4': 'SIMON BOLON'
        }
        local_azure = mapeo_locales.get(local) if local else None

        # Consultar contifico_cobrospagos (DEUNA = Transferencia + BANCO DEUNA)
        if local_azure:
            cur.execute('''
                SELECT
                    codigo_comprobante as factura,
                    centro_costo as local,
                    valor,
                    persona
                FROM contifico_cobrospagos
                WHERE fecha = %s
                  AND forma_cobro_pago = 'Transferencia'
                  AND cta_afectada = 'BANCO DEUNA'
                  AND centro_costo = %s
                ORDER BY codigo_comprobante
            ''', (fecha, local_azure))
        else:
            cur.execute('''
                SELECT
                    codigo_comprobante as factura,
                    centro_costo as local,
                    valor,
                    persona
                FROM contifico_cobrospagos
                WHERE fecha = %s
                  AND forma_cobro_pago = 'Transferencia'
                  AND cta_afectada = 'BANCO DEUNA'
                ORDER BY centro_costo, codigo_comprobante
            ''', (fecha,))

        registros = []
        total_monto = 0
        for row in cur.fetchall():
            factura = row['factura'] or ''
            local_name = row['local'] or '(sin local)'
            valor = float(row['valor'] or 0)
            registros.append({
                'factura': factura,
                'local': local_name,
                'valor': valor,
                'persona': row['persona'] or ''
            })
            total_monto += valor

        return jsonify({
            'ok': True,
            'fecha': fecha,
            'registros': registros,
            'total_transacciones': len(registros),
            'total_monto': total_monto
        })

    except Exception as e:
        print(f"Error en /api/vouchers/cobros-deuna: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/vouchers/cortesias', methods=['GET'])
def vs_cortesias():
    """Obtiene cortesias (DNAs) de una fecha para control"""
    fecha = request.args.get('fecha')
    local = request.args.get('local')
    if not fecha:
        return jsonify({'error': 'Fecha requerida'}), 400

    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        mapeo_locales = {
            'vj6e98Ao7UrAEdWB': 'PORTUGAL',
            'gDGe76ymMCn21dn2': 'REAL',
            'pXKdwmJWZcgk3agW': 'SANTO CACHON PORTUGAL',
            '4pzb864o4uBWpeEw': 'SANTO CACHON REAL',
            'rQzaOyDP4cElmdp4': 'SIMON BOLON',
            'isla_floreana': 'FLOREANA'
        }
        local_azure = mapeo_locales.get(local) if local else None

        if local_azure:
            cur.execute('''
                SELECT num_documento, persona, vendedor, centro_costo, nota, total
                FROM fact_detallada
                WHERE fecha = %s
                  AND tipo_documento = 'DNA'
                  AND centro_costo = %s
                ORDER BY num_documento
            ''', (fecha, local_azure))
        else:
            cur.execute('''
                SELECT num_documento, persona, vendedor, centro_costo, nota, total
                FROM fact_detallada
                WHERE fecha = %s
                  AND tipo_documento = 'DNA'
                ORDER BY centro_costo, num_documento
            ''', (fecha,))

        registros = []
        total_monto = 0
        for row in cur.fetchall():
            monto = float(row['total'] or 0)
            registros.append({
                'dna': row['num_documento'] or '',
                'cliente': row['persona'] or '',
                'vendedor': row['vendedor'] or '',
                'local': row['centro_costo'] or '',
                'nota': row['nota'] or '',
                'monto': monto
            })
            total_monto += monto

        return jsonify({
            'ok': True,
            'fecha': fecha,
            'registros': registros,
            'total_cortesias': len(registros),
            'total_monto': total_monto
        })

    except Exception as e:
        print(f"Error en /api/vouchers/cortesias: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


# ============================================================
# INVENTARIO LOCALES (actualizar cantidad / toma fisica)
# ============================================================
# NOTA: WORKER_TOKEN ya definido arriba (linea ~4159)

@app.route('/api/inventario-locales/solicitar', methods=['POST'])
def inventario_locales_solicitar():
    """Admin solicita tarea de actualizar cantidad o toma fisica."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha = data.get('fecha')
    accion = data.get('accion')
    usuario = data.get('usuario', 'admin')

    if not bodega or not fecha or not accion:
        return jsonify({'error': 'bodega, fecha y accion son requeridos'}), 400

    if accion not in ('actualizar_cantidad', 'toma_fisica'):
        return jsonify({'error': 'accion debe ser actualizar_cantidad o toma_fisica'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.tareas_inventario_locales (bodega, fecha, accion, solicitado_por)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (bodega, fecha, accion, usuario))
        tarea_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'ok': True, 'id': tarea_id})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/inventario-locales/pendientes', methods=['GET'])
def inventario_locales_pendientes():
    """Worker toma tareas pendientes."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    worker_id = request.args.get('worker_id', 'pc-finanzas')
    if not worker_autorizado(worker_id):
        # Silencio en vez de error: el worker viejo reintentaria en bucle y
        # llenaria el log. Con la lista vacia simplemente no hace nada.
        return jsonify([])
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.tareas_inventario_locales
            SET estado = 'en_proceso', worker_lock = %s, timestamp_inicio = NOW()
            WHERE id IN (
                SELECT id FROM goti.tareas_inventario_locales
                WHERE estado = 'pendiente' ORDER BY solicitado_at ASC LIMIT 1
                FOR UPDATE SKIP LOCKED
            )
            RETURNING id, bodega, fecha, accion
        """, (worker_id,))
        rows = cur.fetchall()
        conn.commit()
        return jsonify([{
            'id': r['id'],
            'bodega': r['bodega'],
            'fecha': r['fecha'].isoformat() if r['fecha'] else None,
            'accion': r['accion'],
        } for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/inventario-locales/resultado', methods=['POST'])
def inventario_locales_resultado():
    """Worker reporta resultado."""
    token = request.headers.get('X-Worker-Token')
    if token != WORKER_TOKEN:
        return jsonify({'error': 'unauthorized'}), 401
    data = request.json or {}
    ejec_id = data.get('id')
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            UPDATE goti.tareas_inventario_locales
            SET estado = %s, timestamp_fin = NOW(), total_productos = %s,
                url_contifico = %s, error_msg = %s
            WHERE id = %s
        """, (
            data.get('estado', 'completado'),
            data.get('total_productos'),
            data.get('url_contifico'),
            data.get('error_msg'),
            ejec_id
        ))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/inventario-locales/historial', methods=['GET'])
def inventario_locales_historial():
    """Lista historial de tareas."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        # Las dos colas en una sola lista. 'origen' dice de cual viene cada
        # fila: hace falta para cancelar, porque los id se repiten entre tablas
        # (hay una carga #77 y un cruce #77, y no son la misma cosa).
        cur.execute("""
            SELECT 'carga' AS origen,
                   id, bodega, fecha, accion, estado, solicitado_por, solicitado_at,
                   timestamp_inicio, timestamp_fin, total_productos,
                   url_contifico, error_msg
            FROM goti.tareas_inventario_locales

            UNION ALL

            SELECT 'cruce' AS origen,
                   id, bodega, fecha_toma AS fecha, 'cruce_operativo' AS accion,
                   estado, solicitado_por, solicitado_at,
                   timestamp_descarga AS timestamp_inicio,
                   timestamp_cruce   AS timestamp_fin,
                   total_productos_toma AS total_productos,
                   NULL AS url_contifico, error_msg
            FROM goti.cruce_operativo_ejecuciones

            UNION ALL

            SELECT 'ajuste' AS origen,
                   id, bodega, fecha_toma AS fecha, 'carga_ajuste' AS accion,
                   estado, solicitado_por, solicitado_at,
                   timestamp_inicio, timestamp_fin,
                   COALESCE(productos_ok, total_productos) AS total_productos,
                   NULL AS url_contifico, error_msg
            FROM goti.carga_contifico_ejecuciones

            ORDER BY solicitado_at DESC
            LIMIT 100
        """)
        return jsonify(cur.fetchall())
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/inventario-locales/estado/<int:tarea_id>', methods=['GET'])
def inventario_locales_estado(tarea_id):
    """Estado de una tarea especifica."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM goti.tareas_inventario_locales WHERE id = %s", (tarea_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify(dict(r))
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


@app.route('/api/inventario-locales/cancelar/<int:tarea_id>', methods=['POST'])
def inventario_locales_cancelar(tarea_id):
    """Cancela una tarea de carga a locales.

    Se podia crear tareas pero no pararlas: si alguien pedia una por error o
    con la fecha equivocada, solo quedaba esperar a que el worker la hiciera.

    Que se puede cancelar y que no:
      - 'pendiente'  : se cancela limpio, el worker ya no la toma.
      - 'en_proceso' : el worker YA la esta haciendo. Se marca cancelada para
                       que no se reintente, pero puede que el documento acabe
                       creandose igual en Contifico: el navegador no se entera.
                       Por eso se avisa en la respuesta en vez de decir que se
                       paro, que seria mentira.
      - terminadas   : no se tocan. Cancelar algo ya hecho no deshace nada.
    """
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, estado, bodega, fecha, accion
                       FROM goti.tareas_inventario_locales WHERE id = %s""", (tarea_id,))
        t = cur.fetchone()
        if not t:
            return jsonify({'error': 'no existe esa tarea'}), 404
        if t['estado'] in ('completado', 'cancelado'):
            return jsonify({'error': f"la tarea ya esta {t['estado']}, no hay nada que cancelar"}), 409

        era = t['estado']
        cur.execute("""UPDATE goti.tareas_inventario_locales
                       SET estado = 'cancelado',
                           error_msg = %s,
                           timestamp_fin = NOW()
                       WHERE id = %s""",
                    (f'cancelada desde el panel (estaba en {era})', tarea_id))
        conn.commit()

        aviso = None
        if era == 'en_proceso':
            aviso = ('El worker ya la habia empezado. No se reintentara, pero '
                     'revisa en Contifico por si alcanzo a crear el documento.')
        return jsonify({'ok': True, 'id': tarea_id, 'estado_anterior': era,
                        'aviso': aviso})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/inventario-locales/borrar', methods=['POST'])
def inventario_locales_borrar():
    """Borra datos de inventario (cantidad + conteos) para un local y fecha."""
    data = request.json or {}
    bodega = data.get('bodega')
    fecha = data.get('fecha')
    usuario = data.get('usuario', 'admin')

    if not bodega or not fecha:
        return jsonify({'error': 'bodega y fecha son requeridos'}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        # Borrar registros
        cur.execute("""
            DELETE FROM goti.inventario_ciego_conteos
            WHERE local = %s AND fecha = %s::date
        """, (bodega, fecha))
        borrados = cur.rowcount
        conn.commit()

        return jsonify({
            'ok': True,
            'borrados': borrados,
            'mensaje': f'Borrados {borrados} registros de {bodega} fecha {fecha}'
        })
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: release_db(conn)


# ================================================================
# MODULO NOMINA Y TTHH - Endpoints API
# BD: movimientos (Azure PostgreSQL) - Schema: nomina_*
# ================================================================

_nomina_pool = None

def _get_nomina_pool():
    """Pool dedicado para nomina con RealDictCursor"""
    global _nomina_pool
    if _nomina_pool is None:
        _nomina_pool = SimpleConnectionPool(
            minconn=1, maxconn=3,
            host=os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
            database='movimientos',
            user=os.environ.get('DB_USER', 'adminChios'),
            password=os.environ.get('DB_PASSWORD', 'Burger2023'),
            port=os.environ.get('DB_PORT', '5432'),
            sslmode='require',
            connect_timeout=10,
            cursor_factory=RealDictCursor
        )
    return _nomina_pool

def _get_mov_conn():
    """Conexion a BD movimientos para modulo nomina (con RealDictCursor)"""
    conn = _get_nomina_pool().getconn()
    try:
        conn.cursor().execute("SELECT 1")
        conn.rollback()
    except Exception:
        try:
            _get_nomina_pool().putconn(conn, close=True)
        except Exception:
            try: conn.close()
            except Exception: pass
        conn = psycopg2.connect(
            host=os.environ.get('DB_HOST', 'chiosburguer.postgres.database.azure.com'),
            database='movimientos',
            user=os.environ.get('DB_USER', 'adminChios'),
            password=os.environ.get('DB_PASSWORD', 'Burger2023'),
            port=os.environ.get('DB_PORT', '5432'),
            sslmode='require',
            cursor_factory=RealDictCursor
        )
    return conn

def _release_mov(conn):
    try:
        if conn and not conn.closed:
            _get_nomina_pool().putconn(conn)
    except Exception:
        try: conn.close()
        except Exception: pass

def _init_nomina_schema():
    """Crea tablas nomina_* en BD movimientos si no existen"""
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS nomina_empleados (
                id SERIAL PRIMARY KEY,
                tipo_documento TEXT DEFAULT 'cedula',
                cedula TEXT NOT NULL UNIQUE,
                nombre_completo TEXT,
                primer_nombre TEXT DEFAULT '',
                segundo_nombre TEXT,
                apellido_paterno TEXT DEFAULT '',
                apellido_materno TEXT,
                email TEXT,
                fecha_nacimiento DATE,
                genero TEXT,
                nacionalidad TEXT DEFAULT 'Ecuatoriano',
                estado_civil TEXT,
                direccion TEXT,
                celular TEXT,
                empresa TEXT DEFAULT 'FOODIX SAS',
                marca TEXT,
                tienda TEXT,
                area TEXT,
                cargo_texto TEXT,
                cargo_jerarquico_texto TEXT,
                estado TEXT DEFAULT 'Activo',
                fecha_ingreso DATE,
                fecha_salida DATE,
                tipo_contrato TEXT,
                etapa TEXT,
                jornada TEXT DEFAULT 'Completa',
                horas_mes NUMERIC(6,2) DEFAULT 240,
                salario NUMERIC(10,2) NOT NULL DEFAULT 0,
                descuento_por_cargo NUMERIC(10,2) DEFAULT 0,
                bono_cumpleanos NUMERIC(10,2) DEFAULT 0,
                decimos TEXT DEFAULT 'Mensualizado',
                forma_pago TEXT DEFAULT 'Transferencia',
                banco TEXT,
                tipo_cuenta TEXT,
                numero_cuenta TEXT,
                motivo_salida TEXT,
                emergencia1_nombre TEXT,
                emergencia1_telefono TEXT,
                emergencia1_relacion TEXT,
                emergencia2_nombre TEXT,
                emergencia2_telefono TEXT,
                emergencia2_relacion TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS nomina_rubros (
                id SERIAL PRIMARY KEY,
                codigo TEXT NOT NULL UNIQUE,
                nombre TEXT NOT NULL,
                tipo TEXT NOT NULL,
                es_predeterminado BOOLEAN DEFAULT FALSE,
                porcentaje NUMERIC(8,4),
                valor_fijo NUMERIC(10,2),
                aplica_iess BOOLEAN DEFAULT FALSE,
                aplica_ir BOOLEAN DEFAULT FALSE,
                activo BOOLEAN DEFAULT TRUE,
                orden INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS nomina_liquidacion_grupo (
                id SERIAL PRIMARY KEY,
                tipo TEXT NOT NULL,
                periodo TEXT NOT NULL,
                fecha_liquidacion DATE NOT NULL,
                estado TEXT DEFAULT 'Borrador',
                total_ingresos NUMERIC(12,2) DEFAULT 0,
                total_descuentos NUMERIC(12,2) DEFAULT 0,
                total_pagar NUMERIC(12,2) DEFAULT 0,
                total_prestaciones NUMERIC(12,2) DEFAULT 0,
                total_costo_empleado NUMERIC(12,2) DEFAULT 0,
                num_empleados INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS nomina_liquidacion_detalle (
                id SERIAL PRIMARY KEY,
                liquidacion_grupo_id INTEGER NOT NULL REFERENCES nomina_liquidacion_grupo(id) ON DELETE CASCADE,
                empleado_id INTEGER NOT NULL REFERENCES nomina_empleados(id),
                nombre_completo TEXT,
                sueldo_base NUMERIC(10,2) NOT NULL,
                dias_trabajados NUMERIC(5,1) DEFAULT 30,
                total_ingresos NUMERIC(10,2) DEFAULT 0,
                total_descuentos NUMERIC(10,2) DEFAULT 0,
                liquido_pagar NUMERIC(10,2) DEFAULT 0,
                total_prestaciones NUMERIC(10,2) DEFAULT 0,
                costo_empleado NUMERIC(10,2) DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS nomina_liquidacion_rubro (
                id SERIAL PRIMARY KEY,
                liquidacion_detalle_id INTEGER NOT NULL REFERENCES nomina_liquidacion_detalle(id) ON DELETE CASCADE,
                concepto TEXT NOT NULL,
                tipo TEXT NOT NULL,
                cantidad NUMERIC(8,2) DEFAULT 1,
                valor NUMERIC(10,2) DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS nomina_centros_costo (
                id SERIAL PRIMARY KEY,
                codigo TEXT NOT NULL UNIQUE,
                nombre TEXT NOT NULL,
                activo BOOLEAN DEFAULT TRUE
            );

            CREATE TABLE IF NOT EXISTS nomina_tipos_contrato (
                id SERIAL PRIMARY KEY,
                nombre TEXT NOT NULL,
                dias_duracion INTEGER,
                activo BOOLEAN DEFAULT TRUE
            );

            CREATE TABLE IF NOT EXISTS nomina_config_cargo (
                id SERIAL PRIMARY KEY,
                cargo_jerarquico TEXT NOT NULL UNIQUE,
                descuento_por_cargo NUMERIC(10,2) DEFAULT 0,
                bono_cumpleanos NUMERIC(10,2) DEFAULT 0,
                activo BOOLEAN DEFAULT TRUE
            );
        """)

        # Insertar rubros predeterminados si no existen
        cur.execute("SELECT COUNT(*) as c FROM nomina_rubros")
        if cur.fetchone()['c'] == 0:
            rubros = [
                ('ING001', 'Sueldo Base', 'ingreso', None, True, True, 1),
                ('ING004', 'Comisiones', 'ingreso', None, True, True, 4),
                ('ING009', 'Vacaciones Pagadas', 'ingreso', None, True, True, 9),
                ('DES001', 'IESS Personal', 'descuento', 9.45, False, False, 1),
                ('DES003', 'Prestamo IESS Quirografario', 'descuento', None, False, False, 3),
                ('DES007', 'Multas', 'descuento', None, False, False, 7),
                ('PRE001', 'IESS Patronal', 'prestacion', 12.15, False, False, 1),
                ('PRE002', 'IECE+SECAP', 'prestacion', 1.0, False, False, 2),
                ('PRE004', 'Provision Vacaciones', 'prestacion', None, False, False, 4),
                ('PRE005', 'Provision Decimotercero', 'prestacion', None, False, False, 5),
                ('PRE006', 'Provision Decimocuarto', 'prestacion', None, False, False, 6),
            ]
            for r in rubros:
                cur.execute("""
                    INSERT INTO nomina_rubros (codigo, nombre, tipo, porcentaje, aplica_iess, aplica_ir, orden)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, r)

        # Insertar config cargos jerarquicos si no existen
        cur.execute("SELECT COUNT(*) as c FROM nomina_config_cargo")
        if cur.fetchone()['c'] == 0:
            cargos = [
                ('Director General', 100, 30), ('Jefe de Area', 35, 30),
                ('Gerente de Tienda', 25, 30), ('Subgerente de Tienda', 20, 20),
                ('Analista', 20, 20), ('Supervisor', 20, 20),
                ('Operario de Produccion', 15, 15), ('Polifuncional', 15, 15),
                ('Auxiliar', 15, 15), ('Asistente', 15, 15),
                ('Pasante / Practicante', 10, 10),
            ]
            for c in cargos:
                cur.execute("""
                    INSERT INTO nomina_config_cargo (cargo_jerarquico, descuento_por_cargo, bono_cumpleanos)
                    VALUES (%s, %s, %s)
                """, c)

        # Insertar tipos contrato si no existen
        cur.execute("SELECT COUNT(*) as c FROM nomina_tipos_contrato")
        if cur.fetchone()['c'] == 0:
            tipos = [
                'Contrato Indefinido con Periodo de Prueba', 'Contrato Sector Turistico',
                'Contrato Pasantias', 'Contrato Practicas', 'Contrato por Servicios Profesionales',
                'Contrato Emergente', 'Contrato Funciones de Confianza',
                'Contrato Jornada Parcial Permanente', 'Contrato Artesanal',
            ]
            for t in tipos:
                cur.execute("INSERT INTO nomina_tipos_contrato (nombre) VALUES (%s)", (t,))

        conn.commit()
        print('[NOMINA] Schema inicializado correctamente')
    except Exception as e:
        if conn: conn.rollback()
        print(f'[NOMINA] Error inicializando schema: {e}')
    finally:
        if conn: _release_mov(conn)

# Inicializar schema al arrancar
_init_nomina_schema()

# ---------- CONSTANTES NOMINA ECUADOR 2026 ----------
NOM_SBU = 482.0
NOM_IESS_PERSONAL = 0.0945
NOM_IESS_PATRONAL = 0.1215
NOM_IECE_SECAP = 0.01
NOM_FR_RATE = 1/12  # 8.33%
NOM_ANTICIPO_PCT = 0.40

def _nom_round(n):
    """Redondeo bancario a 2 decimales"""
    return round(float(n or 0) + 1e-9, 2)

# ---------- CATALOGOS ----------
@app.route('/api/nomina/catalogos', methods=['GET'])
def nomina_catalogos():
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("SELECT nombre FROM nomina_tipos_contrato WHERE activo = TRUE ORDER BY nombre")
        tipos = [r['nombre'] for r in cur.fetchall()]
        cur.execute("SELECT cargo_jerarquico, descuento_por_cargo, bono_cumpleanos FROM nomina_config_cargo WHERE activo = TRUE ORDER BY cargo_jerarquico")
        cargos = cur.fetchall()
        return jsonify({
            'tipos_contrato': tipos,
            'cargos_jerarquicos': [c['cargo_jerarquico'] for c in cargos],
            'config_cargos': {c['cargo_jerarquico']: {'descuento': float(c['descuento_por_cargo']), 'bono': float(c['bono_cumpleanos'])} for c in cargos},
            'tiendas': ['Chios Real Audiencia','Chios Portugal','Chios Floreana','Santo Cachon Real Audiencia','Santo Cachon Portugal','Simon Bolon','Planta','Oficinas'],
            'areas': ['Operaciones','Produccion','Administracion','Contabilidad','Marketing','Talento Humano'],
            'marcas': ['Chios','Santo Cachon','Simon Bolon','FOODIX'],
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

# ---------- DASHBOARD ----------
@app.route('/api/nomina/dashboard', methods=['GET'])
def nomina_dashboard():
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) as c FROM nomina_empleados WHERE estado = 'Activo'")
        activos = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM nomina_empleados WHERE estado != 'Activo'")
        inactivos = cur.fetchone()['c']
        cur.execute("SELECT COALESCE(SUM(salario),0) as s, COALESCE(AVG(salario),0) as a FROM nomina_empleados WHERE estado = 'Activo'")
        sal = cur.fetchone()
        # Costo mensual estimado: salario + IESS patronal + IECE/SECAP + provisiones
        costo_base = float(sal['s'])
        costo_mensual = _nom_round(costo_base * (1 + NOM_IESS_PATRONAL + NOM_IECE_SECAP + NOM_FR_RATE + 1/12 + NOM_SBU/(12*costo_base) if costo_base > 0 else 0))
        if costo_base == 0:
            costo_mensual = 0

        cur.execute("SELECT COALESCE(area, 'Sin area') as area, COUNT(*) as cantidad FROM nomina_empleados WHERE estado = 'Activo' GROUP BY area ORDER BY cantidad DESC")
        por_area = cur.fetchall()
        cur.execute("SELECT COALESCE(tienda, 'Sin tienda') as tienda, COUNT(*) as cantidad FROM nomina_empleados WHERE estado = 'Activo' GROUP BY tienda ORDER BY cantidad DESC")
        por_tienda = cur.fetchall()
        cur.execute("SELECT COALESCE(tipo_contrato, 'Sin tipo') as tipo_contrato, COUNT(*) as cantidad FROM nomina_empleados WHERE estado = 'Activo' GROUP BY tipo_contrato ORDER BY cantidad DESC")
        por_contrato = cur.fetchall()

        return jsonify({
            'activos': activos, 'inactivos': inactivos,
            'costo_mensual': costo_mensual,
            'promedio_salario': float(sal['a']),
            'por_area': [dict(r) for r in por_area],
            'por_tienda': [dict(r) for r in por_tienda],
            'por_contrato': [dict(r) for r in por_contrato],
        })
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

# ---------- EMPLEADOS CRUD ----------
@app.route('/api/nomina/empleados', methods=['GET'])
def nomina_empleados_lista():
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        filtro = request.args.get('filtro', 'Activo')
        buscar = request.args.get('buscar', '').strip()
        where = []
        params = []
        if filtro:
            where.append("estado = %s")
            params.append(filtro)
        if buscar:
            where.append("(nombre_completo ILIKE %s OR cedula ILIKE %s OR cargo_texto ILIKE %s)")
            like = f'%{buscar}%'
            params.extend([like, like, like])
        w = ('WHERE ' + ' AND '.join(where)) if where else ''
        cur.execute(f"""
            SELECT id, cedula, nombre_completo, cargo_texto, tienda, salario, fecha_ingreso, estado, area, marca
            FROM nomina_empleados {w}
            ORDER BY nombre_completo
        """, params)
        return jsonify({'empleados': [dict(r) for r in cur.fetchall()]})
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/empleados/<int:emp_id>', methods=['GET'])
def nomina_empleado_detalle(emp_id):
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM nomina_empleados WHERE id = %s", (emp_id,))
        emp = cur.fetchone()
        if not emp:
            return jsonify({'error': 'Empleado no encontrado'}), 404
        return jsonify(dict(emp))
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/empleados', methods=['POST'])
def nomina_empleado_crear():
    conn = None
    try:
        data = request.get_json()
        conn = _get_mov_conn()
        cur = conn.cursor()
        campos = ['cedula','nombre_completo','email','celular','fecha_nacimiento','genero','estado_civil',
            'nacionalidad','direccion','empresa','marca','tienda','area','cargo_texto','cargo_jerarquico_texto',
            'fecha_ingreso','tipo_contrato','estado','salario','descuento_por_cargo','bono_cumpleanos',
            'jornada','horas_mes','decimos','forma_pago','banco','tipo_cuenta','numero_cuenta',
            'emergencia1_nombre','emergencia1_telefono','emergencia1_relacion',
            'emergencia2_nombre','emergencia2_telefono','emergencia2_relacion']
        vals = []
        placeholders = []
        cols = []
        for c in campos:
            v = data.get(c)
            if v is not None and v != '':
                cols.append(c)
                placeholders.append('%s')
                if c in ('salario','descuento_por_cargo','bono_cumpleanos','horas_mes'):
                    vals.append(float(v) if v else 0)
                else:
                    vals.append(v)
        cur.execute(f"""
            INSERT INTO nomina_empleados ({','.join(cols)})
            VALUES ({','.join(placeholders)})
            RETURNING id
        """, vals)
        new_id = cur.fetchone()['id']
        conn.commit()
        return jsonify({'ok': True, 'id': new_id})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/empleados/<int:emp_id>', methods=['PUT'])
def nomina_empleado_editar(emp_id):
    conn = None
    try:
        data = request.get_json()
        conn = _get_mov_conn()
        cur = conn.cursor()
        campos = ['cedula','nombre_completo','email','celular','fecha_nacimiento','genero','estado_civil',
            'nacionalidad','direccion','empresa','marca','tienda','area','cargo_texto','cargo_jerarquico_texto',
            'fecha_ingreso','tipo_contrato','estado','salario','descuento_por_cargo','bono_cumpleanos',
            'jornada','horas_mes','decimos','forma_pago','banco','tipo_cuenta','numero_cuenta',
            'emergencia1_nombre','emergencia1_telefono','emergencia1_relacion',
            'emergencia2_nombre','emergencia2_telefono','emergencia2_relacion']
        sets = []
        vals = []
        for c in campos:
            if c in data:
                v = data[c]
                if c in ('salario','descuento_por_cargo','bono_cumpleanos','horas_mes'):
                    v = float(v) if v else 0
                elif v == '':
                    v = None
                sets.append(f"{c} = %s")
                vals.append(v)
        if not sets:
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        sets.append("updated_at = NOW()")
        vals.append(emp_id)
        cur.execute(f"UPDATE nomina_empleados SET {','.join(sets)} WHERE id = %s", vals)
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/empleados/<int:emp_id>', methods=['DELETE'])
def nomina_empleado_eliminar(emp_id):
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("DELETE FROM nomina_empleados WHERE id = %s", (emp_id,))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

# ---------- NOMINAS - PROCESAMIENTO ----------
@app.route('/api/nomina/nominas', methods=['GET'])
def nomina_listar():
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, tipo, periodo, fecha_liquidacion, estado,
                   total_ingresos, total_descuentos, total_pagar, num_empleados, created_at
            FROM nomina_liquidacion_grupo
            ORDER BY periodo DESC, created_at DESC
        """)
        return jsonify({'nominas': [dict(r) for r in cur.fetchall()]})
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/nominas/procesar', methods=['POST'])
def nomina_procesar():
    """Calcula nomina para todos los empleados activos del periodo"""
    conn = None
    try:
        data = request.get_json()
        periodo = data.get('periodo')
        tipo = data.get('tipo', 'Mensual')
        if not periodo:
            return jsonify({'error': 'Periodo requerido'}), 400

        conn = _get_mov_conn()
        cur = conn.cursor()

        # Obtener empleados activos
        cur.execute("""
            SELECT id, nombre_completo, salario, descuento_por_cargo, bono_cumpleanos,
                   fecha_ingreso, cargo_jerarquico_texto, jornada, horas_mes, decimos
            FROM nomina_empleados WHERE estado = 'Activo' AND salario > 0
            ORDER BY nombre_completo
        """)
        empleados = cur.fetchall()
        if not empleados:
            return jsonify({'error': 'No hay empleados activos con salario'}), 400

        # Crear grupo de liquidacion
        cur.execute("""
            INSERT INTO nomina_liquidacion_grupo (tipo, periodo, fecha_liquidacion)
            VALUES (%s, %s, NOW()::date) RETURNING id
        """, (tipo, periodo))
        grupo_id = cur.fetchone()['id']

        total_ing = 0
        total_des = 0
        total_pagar = 0
        total_prest = 0
        total_costo = 0

        for emp in empleados:
            salario = float(emp['salario'])
            dias = 30
            sueldo_devengado = _nom_round(salario * dias / 30)

            if tipo == 'Anticipo':
                # Anticipo: 40% del devengado
                liquido = _nom_round(sueldo_devengado * NOM_ANTICIPO_PCT)
                ing = liquido
                des = 0
                prest = 0
                costo = ing
                rubros_data = [('Anticipo Quincena', 'ingreso', 1, liquido)]
            else:
                # Calculo mensual completo
                desc_cargo = float(emp['descuento_por_cargo'] or 0)
                bono_cumple = float(emp['bono_cumpleanos'] or 0)

                # Ingresos
                ing_sueldo = sueldo_devengado
                ing_total = ing_sueldo + bono_cumple

                # Descuentos
                base_iess = ing_sueldo  # Solo sueldo para IESS
                iess_personal = _nom_round(base_iess * NOM_IESS_PERSONAL)
                des_total = iess_personal + desc_cargo

                # Liquido
                liquido = _nom_round(ing_total - des_total)

                # Prestaciones (costo empleador)
                iess_patronal = _nom_round(base_iess * NOM_IESS_PATRONAL)
                iece_secap = _nom_round(base_iess * NOM_IECE_SECAP)
                prov_vacaciones = _nom_round(base_iess / 24)
                prov_decimo3 = _nom_round(base_iess / 12)
                prov_decimo4 = _nom_round(NOM_SBU / 12)
                fondos_reserva = _nom_round(base_iess * NOM_FR_RATE)

                prest = _nom_round(iess_patronal + iece_secap + prov_vacaciones + prov_decimo3 + prov_decimo4 + fondos_reserva)
                ing = ing_total
                des = des_total
                costo = _nom_round(ing + prest)

                rubros_data = [
                    ('Sueldo Base', 'ingreso', dias, ing_sueldo),
                    ('Bono Cumpleanos', 'ingreso', 1, bono_cumple),
                    ('IESS Personal 9.45%', 'descuento', 1, iess_personal),
                    ('Descuento por Cargo', 'descuento', 1, desc_cargo),
                    ('IESS Patronal 12.15%', 'prestacion', 1, iess_patronal),
                    ('IECE+SECAP 1%', 'prestacion', 1, iece_secap),
                    ('Prov. Vacaciones', 'prestacion', 1, prov_vacaciones),
                    ('Prov. Decimotercero', 'prestacion', 1, prov_decimo3),
                    ('Prov. Decimocuarto', 'prestacion', 1, prov_decimo4),
                    ('Fondos de Reserva', 'prestacion', 1, fondos_reserva),
                ]

            # Insertar detalle
            cur.execute("""
                INSERT INTO nomina_liquidacion_detalle
                    (liquidacion_grupo_id, empleado_id, nombre_completo, sueldo_base, dias_trabajados,
                     total_ingresos, total_descuentos, liquido_pagar, total_prestaciones, costo_empleado)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
            """, (grupo_id, emp['id'], emp['nombre_completo'], salario, dias, ing, des, liquido, prest, costo))
            det_id = cur.fetchone()['id']

            # Insertar rubros
            for concepto, rtipo, cant, val in rubros_data:
                if val != 0:
                    cur.execute("""
                        INSERT INTO nomina_liquidacion_rubro (liquidacion_detalle_id, concepto, tipo, cantidad, valor)
                        VALUES (%s,%s,%s,%s,%s)
                    """, (det_id, concepto, rtipo, cant, val))

            total_ing += ing
            total_des += des
            total_pagar += liquido
            total_prest += prest
            total_costo += costo

        # Actualizar totales del grupo
        cur.execute("""
            UPDATE nomina_liquidacion_grupo
            SET total_ingresos=%s, total_descuentos=%s, total_pagar=%s,
                total_prestaciones=%s, total_costo_empleado=%s, num_empleados=%s
            WHERE id=%s
        """, (_nom_round(total_ing), _nom_round(total_des), _nom_round(total_pagar),
              _nom_round(total_prest), _nom_round(total_costo), len(empleados), grupo_id))

        conn.commit()
        return jsonify({'ok': True, 'grupo_id': grupo_id, 'num_empleados': len(empleados),
                        'total_pagar': _nom_round(total_pagar)})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/nominas/<int:grupo_id>', methods=['GET'])
def nomina_detalle(grupo_id):
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("SELECT * FROM nomina_liquidacion_grupo WHERE id = %s", (grupo_id,))
        grupo = cur.fetchone()
        if not grupo:
            return jsonify({'error': 'Nomina no encontrada'}), 404
        cur.execute("""
            SELECT d.*, e.cedula, e.cargo_texto, e.tienda
            FROM nomina_liquidacion_detalle d
            JOIN nomina_empleados e ON e.id = d.empleado_id
            WHERE d.liquidacion_grupo_id = %s
            ORDER BY d.nombre_completo
        """, (grupo_id,))
        detalles = []
        for row in cur.fetchall():
            d = dict(row)
            cur.execute("""
                SELECT concepto, tipo, cantidad, valor
                FROM nomina_liquidacion_rubro WHERE liquidacion_detalle_id = %s ORDER BY id
            """, (d['id'],))
            d['rubros'] = [dict(r) for r in cur.fetchall()]
            detalles.append(d)
        return jsonify({'grupo': dict(grupo), 'detalles': detalles})
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)

@app.route('/api/nomina/nominas/<int:grupo_id>/aprobar', methods=['POST'])
def nomina_aprobar(grupo_id):
    conn = None
    try:
        conn = _get_mov_conn()
        cur = conn.cursor()
        cur.execute("UPDATE nomina_liquidacion_grupo SET estado = 'Aprobado' WHERE id = %s AND estado = 'Borrador'", (grupo_id,))
        if cur.rowcount == 0:
            return jsonify({'error': 'Solo se pueden aprobar nominas en estado Borrador'}), 400
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn: _release_mov(conn)



# ============================================================
# TELEGRAM: destinatarios de las notificaciones
# ============================================================
# Antes los chat_id vivian escritos en notificar_telegram.py: dar de alta o de
# baja a alguien obligaba a editar el codigo y desplegar. Ahora estan en
# goti.telegram_destinatarios y se administran desde el panel.
#
# Quien le escribe /start al bot entra solo como 'pendiente' (lo hace el cron
# registrar_telegram.py); desde aqui se le asignan bodegas y pasa a 'asignado'.

BODEGAS_TELEGRAM = [
    'TODAS',
    'REAL', 'FLOREANA', 'PORTUGAL',
    'SANTO CACHON REAL', 'SANTO CACHON PORTUGAL', 'SIMON BOLON',
    'BODEGA PRINCIPAL', 'BODEGA MATERIA PRIMA', 'PLANTA DE PRODUCCION', 'BODEGA PULMON',
]
OPERACIONES_TELEGRAM = [
    'TODAS', 'Baja', 'Ingreso Extraordinario', 'Traslado',
    'Conteo', 'Produccion', 'Toma Fisica', 'Cruce Operativo',
]


@app.route('/api/telegram/opciones', methods=['GET'])
def telegram_opciones():
    """Listas para armar los desplegables del panel."""
    return jsonify({'bodegas': BODEGAS_TELEGRAM, 'operaciones': OPERACIONES_TELEGRAM})


@app.route('/api/telegram/destinatarios', methods=['GET'])
def telegram_listar():
    """Lista los destinatarios. Los 'pendiente' salen primero: son los que
    acaban de activar el bot y esperan que se les asigne bodega."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, chat_id, nombre, username, bodegas, operaciones,
                   avisos, activo, estado, notas, creado_at
            FROM goti.telegram_destinatarios
            ORDER BY (estado = 'pendiente') DESC, nombre NULLS LAST, chat_id
        """)
        filas = []
        for r in cur.fetchall():
            filas.append({
                'id': r['id'],
                'chat_id': str(r['chat_id']),          # str: JS pierde precision en enteros grandes
                'nombre': r['nombre'],
                'username': r['username'],
                'bodegas': r['bodegas'] or [],
                'operaciones': r['operaciones'] or [],
                'avisos': r['avisos'],
                'activo': r['activo'],
                'estado': r['estado'],
                'notas': r['notas'],
                'creado_at': r['creado_at'].isoformat() if r['creado_at'] else None,
            })
        pendientes = sum(1 for f in filas if f['estado'] == 'pendiente')
        return jsonify({'destinatarios': filas, 'total': len(filas),
                        'pendientes': pendientes})
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/telegram/destinatarios', methods=['POST'])
def telegram_guardar():
    """Crea o actualiza un destinatario.

    Si se le asignan bodegas, pasa automaticamente de 'pendiente' a 'asignado':
    asi no hay que acordarse de cambiar el estado a mano.
    """
    data = request.json or {}
    chat_id = str(data.get('chat_id', '')).strip()
    if not chat_id.lstrip('-').isdigit():
        return jsonify({'error': 'chat_id debe ser numerico'}), 400

    bodegas = [b for b in (data.get('bodegas') or []) if b]
    operaciones = [o for o in (data.get('operaciones') or []) if o] or ['TODAS']
    avisos = (data.get('avisos') or 'ambos').lower()
    if avisos not in ('ambos', 'exito', 'error'):
        return jsonify({'error': "avisos debe ser ambos, exito o error"}), 400
    activo = bool(data.get('activo', True))
    estado = 'asignado' if bodegas else 'pendiente'

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO goti.telegram_destinatarios
                (chat_id, nombre, username, bodegas, operaciones, avisos,
                 activo, estado, notas, actualizado_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s, NOW())
            ON CONFLICT (chat_id) DO UPDATE SET
                nombre = EXCLUDED.nombre,
                username = COALESCE(EXCLUDED.username, goti.telegram_destinatarios.username),
                bodegas = EXCLUDED.bodegas,
                operaciones = EXCLUDED.operaciones,
                avisos = EXCLUDED.avisos,
                activo = EXCLUDED.activo,
                estado = EXCLUDED.estado,
                notas = EXCLUDED.notas,
                actualizado_at = NOW()
            RETURNING id, estado
        """, (int(chat_id), (data.get('nombre') or '')[:120] or None,
              (data.get('username') or '')[:80] or None,
              bodegas, operaciones, avisos, activo, estado,
              (data.get('notas') or '')[:500] or None))
        r = cur.fetchone()
        conn.commit()
        return jsonify({'ok': True, 'id': r['id'], 'estado': r['estado']})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/telegram/destinatarios/<int:dest_id>', methods=['DELETE'])
def telegram_borrar(dest_id):
    """Da de baja. Por defecto solo DESACTIVA (activo=false) para conservar el
    historial; con ?definitivo=1 borra la fila."""
    definitivo = request.args.get('definitivo') == '1'
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        if definitivo:
            cur.execute("DELETE FROM goti.telegram_destinatarios WHERE id = %s", (dest_id,))
            accion = 'eliminado'
        else:
            cur.execute("""UPDATE goti.telegram_destinatarios
                           SET activo = FALSE, actualizado_at = NOW() WHERE id = %s""", (dest_id,))
            accion = 'desactivado'
        conn.commit()
        if cur.rowcount == 0:
            return jsonify({'error': 'no encontrado'}), 404
        return jsonify({'ok': True, 'accion': accion})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)


@app.route('/api/telegram/probar/<int:dest_id>', methods=['POST'])
def telegram_probar(dest_id):
    """Manda un mensaje de prueba para confirmar que el numero recibe."""
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT chat_id, nombre, bodegas FROM goti.telegram_destinatarios
                       WHERE id = %s""", (dest_id,))
        r = cur.fetchone()
        if not r:
            return jsonify({'error': 'no encontrado'}), 404
        token = os.environ.get('TELEGRAM_TOKEN', '')
        if not token:
            return jsonify({'error': 'falta TELEGRAM_TOKEN en el servidor'}), 500
        bods = ', '.join(r['bodegas']) if r['bodegas'] else 'ninguna'
        texto = ('🔔 <b>Mensaje de prueba</b>\n\n'
                 f'Si lees esto, tus avisos estan llegando bien.\n'
                 f'Bodegas asignadas: <b>{bods}</b>')
        resp = requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                             json={'chat_id': r['chat_id'], 'text': texto,
                                   'parse_mode': 'HTML'}, timeout=15)
        if resp.ok:
            return jsonify({'ok': True})
        return jsonify({'error': f'Telegram respondio {resp.status_code}: {resp.text[:150]}'}), 502
    except Exception as e:
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)



@app.route('/api/telegram/sincronizar-nombres', methods=['POST'])
def telegram_sincronizar_nombres():
    """Completa los nombres preguntandoselos a Telegram.

    Los 14 destinatarios que venian escritos en notificar_telegram.py se
    migraron solo con el numero: en la tabla salian todos como '-' y no habia
    forma de saber de quien era cada chat. getChat devuelve nombre y usuario de
    cualquier chat que el bot conozca, asi que se rellenan solos en vez de
    teclearlos uno por uno.

    Solo toca las filas sin nombre; lo que ya se escribio a mano no se pisa.
    """
    token = os.environ.get('TELEGRAM_TOKEN', '')
    if not token:
        return jsonify({'error': 'falta TELEGRAM_TOKEN en el servidor'}), 500

    conn = None
    actualizados = sin_datos = 0
    motivo = ''          # primer fallo, para poder decir POR QUE no se pudo
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""SELECT id, chat_id FROM goti.telegram_destinatarios
                       WHERE nombre IS NULL OR nombre = ''""")
        pendientes = cur.fetchall()

        for r in pendientes:
            try:
                resp = requests.get(f'https://api.telegram.org/bot{token}/getChat',
                                    params={'chat_id': r['chat_id']}, timeout=12)
                d = resp.json() if resp.ok else {}
            except Exception as e:
                d = {}
                if not motivo:
                    motivo = f'no se pudo llamar a Telegram: {str(e)[:120]}'
            chat = (d.get('result') or {}) if d.get('ok') else {}
            nombre = (f"{chat.get('first_name', '')} {chat.get('last_name', '')}".strip()
                      or chat.get('title') or '')
            if not nombre:
                # Sin esto el panel decia "sin datos" y no habia forma de saber si
                # Telegram no conocia el chat o si la llamada ni siquiera salio.
                if not motivo:
                    motivo = (d.get('description')
                              or f'Telegram respondio sin nombre: {str(d)[:120]}')
                sin_datos += 1
                continue
            cur.execute("""UPDATE goti.telegram_destinatarios
                           SET nombre = %s,
                               username = COALESCE(%s, username),
                               actualizado_at = NOW()
                           WHERE id = %s""",
                        (nombre[:120], (chat.get('username') or None), r['id']))
            actualizados += 1

        conn.commit()
        return jsonify({'ok': True, 'actualizados': actualizados,
                        'sin_datos': sin_datos, 'revisados': len(pendientes),
                        'motivo': motivo})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({'error': str(e)[:200]}), 500
    finally:
        if conn:
            release_db(conn)

# =====================================================
# MODULO CONTROL DE COSTOS
# Lee de costos_resumen_diario: ya viene deduplicado y valorizado
# como cantidad * costo_promedio (nunca valor_total, que es el total
# del documento repetido en cada linea).
# =====================================================

CO_UMBRAL_PRECIO = 8.0      # % de variacion para alertar
CO_MIN_CONSUMO = 100.0      # $ de consumo en 30 dias para que el producto cuente
CO_SEMANAS_PATRON = 6       # semanas hacia atras para el patron de consumo

# Un producto tiene un costo tipico. El dia que Contifico devuelve un costo
# muchas veces mayor, esa linea no es consumo: es un dato roto, y hay que
# sacarlo de las cuentas antes de que arrastre todo lo demas.
#
# El caso que dejo el modulo inservible: el 19 de julio el CHICHARRON salio a
# 14,97 dolares el gramo. Un solo producto, un solo dia, 236.721 dolares: el
# 28% de los 839.406 de cuatro meses. Y como el patron de consumo se calculaba
# con el PROMEDIO de los ultimos domingos, ese dia envenenaba la referencia
# durante seis semanas -promedio 29.844 contra una mediana de 286-, asi que
# todos los domingos siguientes aparecian como caidas del 99%.
#
# El costo se disparo porque el stock de CHICHARRON estaba en menos 294 kg, y
# el promedio ponderado de Contifico se vuelve absurdo cuando el stock es
# negativo. Eso venia del bot de produccion, que registraba mil veces menos
# producto terminado del que se hacia. Corregido el 26-ago-2026.
CO_FACTOR_COSTO_RARO = 5.0   # veces por encima de su costo habitual
CO_MIN_VALOR_ROTO = 20.0     # $ de la linea, para no perseguir centavos

# Una linea sana es la que NO tiene el costo disparado respecto al del propio
# producto. Se usa en todas las cuentas del modulo, para que ninguna cifra
# quede contaminada por un costo imposible.
CO_SQL_TIPICO_PROD = """
    tipico_prod AS (
        SELECT codigo_prod,
               percentile_cont(0.5) WITHIN GROUP (ORDER BY costo_unitario) AS costo_med
        FROM costos_resumen_diario
        WHERE tipo = 'EGR' AND costo_unitario > 0
          AND fecha > %(f)s - 120 AND fecha <= %(f)s
        GROUP BY codigo_prod
    )
"""

CO_SQL_SANO = """
    sano AS (
        SELECT c.*
        FROM costos_resumen_diario c
        LEFT JOIN tipico_prod t ON t.codigo_prod = c.codigo_prod
        WHERE c.tipo = 'EGR'
          AND c.fecha > %(f)s - (%(sem)s * 7) AND c.fecha <= %(f)s
          AND NOT (COALESCE(t.costo_med, 0) > 0
                   AND c.costo_unitario > t.costo_med * %(fac)s
                   AND c.valor >= %(minroto)s)
    )
"""


def co_dia(valor, por_defecto):
    """Una fecha del querystring, o la de por defecto si no viene o viene mal."""
    if valor:
        try:
            return datetime.strptime(valor, '%Y-%m-%d').date()
        except ValueError:
            pass
    return por_defecto


CO_DIAS_MADUREZ = 3   # dias que tarda un dia en terminar de cargarse


def co_ultimo_dia_completo():
    """Un dia no queda completo el mismo dia: la sincronizacion sigue trayendo
    movimientos suyos durante varios dias. Comparar un dia a medias contra dias
    enteros marcaria caidas de consumo que no existen, por eso el analisis se
    para sobre el ultimo dia ya maduro."""
    return datetime.now(TZ_ECUADOR).date() - timedelta(days=CO_DIAS_MADUREZ)


def co_completitud(cur, fecha):
    """Que tan cargado esta un dia respecto a lo tipico de ese dia de semana.
    Devuelve (porcentaje, lineas_dia, lineas_tipicas) o (None, ...) si no hay base."""
    cur.execute("""
        WITH dia AS (
            SELECT coalesce(sum(lineas), 0) AS n
            FROM costos_resumen_diario WHERE fecha = %(f)s AND tipo = 'EGR'
        ),
        tipico AS (
            SELECT percentile_cont(0.5) WITHIN GROUP (ORDER BY n) AS n
            FROM (
                SELECT fecha, sum(lineas) AS n
                FROM costos_resumen_diario
                WHERE tipo = 'EGR'
                  AND fecha < %(f)s - %(mad)s
                  AND fecha >= %(f)s - 60
                  AND extract(dow FROM fecha) = extract(dow FROM %(f)s::date)
                GROUP BY fecha
            ) t
        )
        SELECT dia.n, tipico.n FROM dia, tipico
    """, {'f': fecha, 'mad': CO_DIAS_MADUREZ})
    r = cur.fetchone()
    if not r or not r[1]:
        return None, (r[0] if r else 0), 0
    return round(100.0 * float(r[0]) / float(r[1]), 1), int(r[0]), int(float(r[1]))


@app.route('/api/costos/alertas', methods=['GET'])
def costos_alertas():
    """Todo lo que hay que mirar de un dia, en el orden en que hay que mirarlo.

    Devuelve cuatro cosas, y ese es el orden que importa:

      1. rotos    - lineas con un costo imposible. Van primero porque mientras
                    esten ahi, ninguna otra cifra del dia es de fiar.
      2. resumen  - cuanto se consumio el dia y cuanto se suele consumir un dia
                    como ese. Una sola cifra para saber si el dia fue normal.
      3. precios  - productos que cambiaron de costo, con lo que eso cuesta al
                    mes si se queda asi.
      4. consumos - categorias que se salieron de su patron.

    Las cuentas 2, 3 y 4 se hacen SOLO sobre lineas sanas.
    """
    conn = None
    try:
        fecha = co_dia(request.args.get('fecha'), co_ultimo_dia_completo())
        umbral = float(request.args.get('umbral', CO_UMBRAL_PRECIO))
        min_consumo = float(request.args.get('min_consumo', CO_MIN_CONSUMO))

        p = {'f': fecha, 'u': umbral, 'm': min_consumo,
             'sem': CO_SEMANAS_PATRON, 'fac': CO_FACTOR_COSTO_RARO,
             'minroto': CO_MIN_VALOR_ROTO,
             'mindif': float(request.args.get('min_dif', 50))}

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '90s'")

        # --- 1. Lineas con un costo imposible -------------------------------
        cur.execute("WITH " + CO_SQL_TIPICO_PROD + """
            SELECT c.codigo_prod, max(c.nombre_prod), max(c.categoria), max(c.bodega),
                   sum(c.cantidad), sum(c.valor),
                   sum(c.valor) / NULLIF(sum(c.cantidad), 0) AS costo_dia,
                   max(t.costo_med) AS costo_med
            FROM costos_resumen_diario c
            JOIN tipico_prod t ON t.codigo_prod = c.codigo_prod
            WHERE c.fecha = %(f)s AND c.tipo = 'EGR' AND c.cantidad > 0
              AND t.costo_med > 0 AND c.valor >= %(minroto)s
              AND c.costo_unitario > t.costo_med * %(fac)s
            GROUP BY c.codigo_prod
            ORDER BY sum(c.valor) DESC LIMIT 20
        """, p)
        rotos = []
        for r in cur.fetchall():
            med = float(r[7] or 0)
            cd = float(r[6] or 0)
            rotos.append({
                'codigo_prod': r[0], 'nombre_prod': r[1], 'categoria': r[2],
                'bodega': r[3], 'cantidad': float(r[4] or 0), 'valor': float(r[5] or 0),
                'costo_dia': cd, 'costo_tipico': med,
                'veces': round(cd / med, 1) if med else None,
            })

        # --- 2. El dia contra un dia como el ---------------------------------
        cur.execute("WITH " + CO_SQL_TIPICO_PROD + "," + CO_SQL_SANO + """,
            hist AS (
                SELECT fecha, sum(valor) AS v FROM sano
                WHERE fecha < %(f)s
                  AND extract(dow FROM fecha) = extract(dow FROM %(f)s::date)
                GROUP BY fecha
            )
            SELECT (SELECT coalesce(sum(valor), 0) FROM sano WHERE fecha = %(f)s),
                   percentile_cont(0.5) WITHIN GROUP (ORDER BY v),
                   count(*)
            FROM hist
        """, p)
        r = cur.fetchone()
        valor_dia = float(r[0] or 0)
        valor_tipico = float(r[1] or 0)
        resumen = {
            'valor_dia': valor_dia,
            'valor_tipico': valor_tipico,
            'muestras': int(r[2] or 0),
            'diferencia': valor_dia - valor_tipico,
            'desvio': ((valor_dia - valor_tipico) / valor_tipico * 100) if valor_tipico else None,
            'dia_semana': ['lunes', 'martes', 'miercoles', 'jueves', 'viernes',
                           'sabado', 'domingo'][fecha.weekday()],
        }

        # --- 3. Precios que se movieron --------------------------------------
        cur.execute("WITH " + CO_SQL_TIPICO_PROD + """,
            hoy AS (
                SELECT c.codigo_prod, max(c.nombre_prod) AS nombre_prod,
                       max(c.categoria) AS categoria, max(c.bodega) AS bodega,
                       sum(c.valor) / NULLIF(sum(c.cantidad), 0) AS costo_hoy,
                       sum(c.cantidad) AS cantidad_hoy
                FROM costos_resumen_diario c
                LEFT JOIN tipico_prod t ON t.codigo_prod = c.codigo_prod
                WHERE c.fecha = %(f)s AND c.tipo = 'EGR' AND c.costo_unitario > 0
                  -- sin nombre en el catalogo no hay sobre que actuar
                  AND c.codigo_prod <> 'SIN_CODIGO' AND c.nombre_prod <> '(SIN NOMBRE)'
                  -- lo que tiene el costo disparado ya salio arriba, como dato roto
                  AND NOT (COALESCE(t.costo_med, 0) > 0
                           AND c.costo_unitario > t.costo_med * %(fac)s
                           AND c.valor >= %(minroto)s)
                GROUP BY c.codigo_prod
            ),
            base AS (
                SELECT codigo_prod,
                       percentile_cont(0.5) WITHIN GROUP (ORDER BY costo_unitario) AS costo_base,
                       sum(valor)    AS valor_30d,
                       sum(cantidad) AS cantidad_30d
                FROM costos_resumen_diario
                WHERE fecha >= %(f)s - 30 AND fecha < %(f)s
                  AND tipo = 'EGR' AND costo_unitario > 0
                GROUP BY codigo_prod
            )
            SELECT h.codigo_prod, h.nombre_prod, h.categoria, h.bodega,
                   h.costo_hoy, b.costo_base,
                   (h.costo_hoy - b.costo_base) / b.costo_base * 100 AS variacion,
                   b.cantidad_30d,
                   (h.costo_hoy - b.costo_base) * b.cantidad_30d AS impacto_mes,
                   b.valor_30d
            FROM hoy h
            JOIN base b ON b.codigo_prod = h.codigo_prod
            WHERE b.costo_base > 0
              AND abs((h.costo_hoy - b.costo_base) / b.costo_base * 100) >= %(u)s
              AND b.valor_30d >= %(m)s
            ORDER BY abs((h.costo_hoy - b.costo_base) * b.cantidad_30d) DESC
            LIMIT 60
        """, p)

        precios = [{
            'codigo_prod': r[0], 'nombre_prod': r[1], 'categoria': r[2], 'bodega': r[3],
            'costo_hoy': float(r[4] or 0), 'costo_base': float(r[5] or 0),
            'variacion': float(r[6] or 0), 'cantidad_30d': float(r[7] or 0),
            'impacto_mes': float(r[8] or 0), 'valor_30d': float(r[9] or 0),
        } for r in cur.fetchall()]

        # --- 4. Categorias fuera de su patron --------------------------------
        # La referencia es la MEDIANA de los ultimos dias iguales, no el
        # promedio: con el promedio, un solo dia raro deja la referencia
        # inservible durante seis semanas.
        cur.execute("WITH " + CO_SQL_TIPICO_PROD + "," + CO_SQL_SANO + """,
            dia AS (
                SELECT bodega, categoria, sum(valor) AS valor_dia
                FROM sano WHERE fecha = %(f)s GROUP BY bodega, categoria
            ),
            historico AS (
                SELECT bodega, categoria, fecha, sum(valor) AS valor
                FROM sano
                WHERE fecha < %(f)s
                  AND extract(dow FROM fecha) = extract(dow FROM %(f)s::date)
                GROUP BY bodega, categoria, fecha
            ),
            patron AS (
                SELECT bodega, categoria,
                       percentile_cont(0.5) WITHIN GROUP (ORDER BY valor) AS tipico,
                       count(*) AS muestras
                FROM historico GROUP BY bodega, categoria
            )
            SELECT d.bodega, d.categoria, d.valor_dia, p.tipico, p.muestras,
                   (d.valor_dia - p.tipico) / NULLIF(p.tipico, 0) * 100 AS desvio,
                   d.valor_dia - p.tipico AS diferencia
            FROM dia d
            JOIN patron p ON p.bodega = d.bodega AND p.categoria = d.categoria
            WHERE p.muestras >= 3 AND p.tipico > 20
              AND abs((d.valor_dia - p.tipico) / NULLIF(p.tipico, 0) * 100) >= 30
              -- ademas del desvio relativo, una diferencia en dolares que valga
              -- la pena mirar: una categoria chica se dispara con 15 dolares
              AND abs(d.valor_dia - p.tipico) >= %(mindif)s
            ORDER BY abs(d.valor_dia - p.tipico) DESC
            LIMIT 40
        """, p)

        consumos = [{
            'bodega': r[0], 'categoria': r[1], 'valor_dia': float(r[2] or 0),
            'promedio': float(r[3] or 0), 'muestras': r[4],
            'desvio': float(r[5] or 0), 'diferencia': float(r[6] or 0),
        } for r in cur.fetchall()]

        # --- Calidad del dato del dia ----------------------------------------
        cur.execute("""
            SELECT count(*) FILTER (WHERE sin_costo > 0) AS filas_sin_costo,
                   count(*) AS filas,
                   coalesce(sum(valor), 0) AS valor_total,
                   count(*) FILTER (WHERE nombre_prod = '(SIN NOMBRE)') AS sin_catalogo
            FROM costos_resumen_diario
            WHERE fecha = %s AND tipo = 'EGR'
        """, (fecha,))
        q = cur.fetchone()
        pct_completo, lineas_dia, lineas_tipicas = co_completitud(cur, fecha)

        return jsonify({
            'ok': True,
            'fecha': fecha.isoformat(),
            'dias_madurez': CO_DIAS_MADUREZ,
            'umbral': umbral,
            'semanas_patron': CO_SEMANAS_PATRON,
            'resumen': resumen,
            'rotos': rotos,
            'precios': precios,
            'consumos': consumos,
            'calidad': {
                'filas_sin_costo': q[0] or 0,
                'filas': q[1] or 0,
                'valor_total': float(q[2] or 0),
                'sin_catalogo': q[3] or 0,
                'completitud': pct_completo,
                'lineas_dia': lineas_dia,
                'lineas_tipicas': lineas_tipicas,
            },
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/resumen', methods=['GET'])
def costos_resumen():
    """Consumo por categoria o por producto en un rango. nivel=categoria|producto"""
    conn = None
    try:
        hasta = co_dia(request.args.get('hasta'), datetime.now(TZ_ECUADOR).date())
        desde = co_dia(request.args.get('desde'), hasta - timedelta(days=30))
        nivel = request.args.get('nivel', 'categoria')
        bodega = (request.args.get('bodega') or '').strip()
        categoria = (request.args.get('categoria') or '').strip()

        filtros = ["fecha >= %(d)s", "fecha <= %(h)s", "tipo = 'EGR'"]
        params = {'d': desde, 'h': hasta}
        if bodega:
            filtros.append("bodega = %(b)s"); params['b'] = bodega
        if categoria:
            filtros.append("categoria = %(c)s"); params['c'] = categoria
        where = ' AND '.join(filtros)

        # Periodo anterior de igual duracion, para comparar
        dias = (hasta - desde).days + 1
        params['pd'] = desde - timedelta(days=dias)
        params['ph'] = desde - timedelta(days=1)
        where_prev = where.replace('%(d)s', '%(pd)s').replace('%(h)s', '%(ph)s')

        if nivel == 'producto':
            campos = "categoria, codigo_prod, max(nombre_prod) AS nombre_prod"
            grupo = "categoria, codigo_prod"
            llave = "categoria || '|' || codigo_prod"
        else:
            campos = "categoria"
            grupo = "categoria"
            llave = "categoria"

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '90s'")
        cur.execute(f"""
            WITH actual AS (
                SELECT {campos}, sum(valor) AS valor, sum(cantidad) AS cantidad,
                       sum(sin_costo) AS sin_costo,
                       {llave} AS llave
                FROM costos_resumen_diario WHERE {where} GROUP BY {grupo}
            ),
            previo AS (
                SELECT {llave} AS llave, sum(valor) AS valor_prev
                FROM costos_resumen_diario WHERE {where_prev} GROUP BY {grupo}
            )
            SELECT a.*, COALESCE(p.valor_prev, 0) AS valor_prev
            FROM actual a LEFT JOIN previo p ON p.llave = a.llave
            ORDER BY a.valor DESC LIMIT 300
        """, params)

        filas = []
        for r in cur.fetchall():
            if nivel == 'producto':
                cat, cod, nom, valor, cant, sin_c, _llave, prev = r
                item = {'categoria': cat, 'codigo_prod': cod, 'nombre_prod': nom}
            else:
                cat, valor, cant, sin_c, _llave, prev = r
                item = {'categoria': cat}
            valor = float(valor or 0); prev = float(prev or 0)
            item.update({
                'valor': valor, 'cantidad': float(cant or 0),
                'sin_costo': int(sin_c or 0), 'valor_prev': prev,
                'variacion': ((valor - prev) / prev * 100) if prev else None,
            })
            filas.append(item)

        return jsonify({'ok': True, 'nivel': nivel, 'desde': desde.isoformat(),
                        'hasta': hasta.isoformat(), 'filas': filas,
                        'total': sum(f['valor'] for f in filas)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/producto', methods=['GET'])
def costos_producto():
    """Historia diaria de precio y consumo de un producto."""
    conn = None
    try:
        codigo = (request.args.get('codigo') or '').strip()
        if not codigo:
            return jsonify({'error': 'codigo requerido'}), 400
        dias = min(int(request.args.get('dias', 90)), 365)
        hasta = datetime.now(TZ_ECUADOR).date()
        desde = hasta - timedelta(days=dias)

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '60s'")
        cur.execute("""
            SELECT fecha,
                   sum(valor) / NULLIF(sum(cantidad), 0) AS costo,
                   sum(cantidad) AS cantidad, sum(valor) AS valor
            FROM costos_resumen_diario
            WHERE codigo_prod = %s AND tipo = 'EGR'
              AND fecha >= %s AND fecha <= %s
            GROUP BY fecha ORDER BY fecha
        """, (codigo, desde, hasta))
        serie = [{'fecha': r[0].isoformat(),
                  'costo': float(r[1]) if r[1] is not None else None,
                  'cantidad': float(r[2] or 0), 'valor': float(r[3] or 0)}
                 for r in cur.fetchall()]

        cur.execute("""
            SELECT max(nombre_prod), max(categoria)
            FROM costos_resumen_diario WHERE codigo_prod = %s
        """, (codigo,))
        info = cur.fetchone()

        return jsonify({'ok': True, 'codigo': codigo,
                        'nombre': info[0] if info else None,
                        'categoria': info[1] if info else None,
                        'serie': serie})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/filtros', methods=['GET'])
def costos_filtros():
    """Bodegas y categorias disponibles, y hasta que dia hay resumen."""
    conn = None
    try:
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '60s'")
        cur.execute("""
            SELECT DISTINCT bodega FROM costos_resumen_diario
            WHERE fecha >= CURRENT_DATE - 60 ORDER BY 1
        """)
        bodegas = [r[0] for r in cur.fetchall()]
        cur.execute("""
            SELECT DISTINCT categoria FROM costos_resumen_diario
            WHERE fecha >= CURRENT_DATE - 60 ORDER BY 1
        """)
        categorias = [r[0] for r in cur.fetchall()]
        cur.execute("SELECT min(fecha), max(fecha), count(*) FROM costos_resumen_diario")
        rango = cur.fetchone()
        return jsonify({'ok': True, 'bodegas': bodegas, 'categorias': categorias,
                        'desde': rango[0].isoformat() if rango[0] else None,
                        'hasta': rango[1].isoformat() if rango[1] else None,
                        'filas': rango[2]})
    except Exception as e:
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/refrescar', methods=['POST'])
def costos_refrescar():
    """Recalcula el resumen de los ultimos N dias (por defecto 3)."""
    conn = None
    try:
        dias = min(int((request.get_json(silent=True) or {}).get('dias', 3)), 40)
        from costos_resumen import asegurar_tabla, refrescar_rango
        hasta = datetime.now(TZ_ECUADOR).date() + timedelta(days=1)
        desde = hasta - timedelta(days=dias)

        conn = fc_get_movimientos_db()
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '240s'")
        asegurar_tabla(cur)
        filas = refrescar_rango(cur, desde, hasta)
        return jsonify({'ok': True, 'desde': desde.isoformat(),
                        'hasta': hasta.isoformat(), 'filas': filas})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn:
            try: conn.autocommit = False
            except Exception: pass
            fc_release_movimientos_db(conn)


# ============================================================
# MODULO: Precios de Compra
# ============================================================
# Cada barra de esta pantalla es un INGRESO -un ING de
# contifico_movimientos-, no una linea de factura. Es la diferencia que
# importa: la factura dice lo que se pidio, el ingreso dice lo que entro a la
# bodega, y son dos cosas distintas. De las papas francesas entraron 140
# ingresos en 90 dias y solo 84 traen una factura enlazada.
#
# El precio de cada ingreso sale de total_movimiento partido para la cantidad.
# No de precio, que viene en cero en 140 de 140. Para CONG001 eso da 1,25 el
# kilo, y los ingresos a 12,50 son los que hay que revisar.
#
# La unidad se resuelve sola y sin adivinar: un movimiento siempre viene en la
# unidad con la que Contifico lleva el producto. Si es Gramos, se muestra por
# kilo. Con las facturas esto era imposible, porque cada proveedor factura como
# quiere -el mismo producto en kilos y en gramos-.
#
# El proveedor se saca del numero de factura que alguien escribio en la
# descripcion del ingreso. Cuando no hay numero, la mercaderia entro sin
# factura enlazada y no se sabe de quien vino: es la serie mas grande de todas,
# el 55% del dinero, y es el punto ciego mas grande que tiene esto.

CP_DIAS = 90                  # ventana por defecto
CP_FACTOR_SOSPECHOSO = 5.0    # veces la mediana para no creerse un precio
CP_SIN_FACTURA = 'Sin factura (ingreso manual)'

# El numero de factura tal como lo escriben: 001-010-000000276
CP_RE_FACTURA = r'([0-9]{3})-([0-9]{3})-0*([0-9]{4,})'

# Los ingresos de la ventana, con su precio unitario y su numero de factura si
# es que lo trae escrito.
CP_SQL_ING = """
    ing AS (
        SELECT m.producto_id, m.fecha, m.codigo, m.cantidad,
               m.total_movimiento AS valor,
               m.total_movimiento / NULLIF(m.cantidad, 0) AS pu,
               m.bodega_destino_id, m.descripcion,
               substring(m.descripcion from %(re)s) IS NOT NULL AS con_factura
        FROM contifico_movimientos m
        WHERE m.tipo = 'ING'
          AND m.fecha >= CURRENT_DATE - %(d)s
          AND m.cantidad > 0 AND m.total_movimiento > 0
    ),
    med AS (
        SELECT producto_id,
               percentile_cont(0.5) WITHIN GROUP (ORDER BY pu) AS m
        FROM ing GROUP BY producto_id
    ),
    marc AS (
        -- Ni por arriba ni por abajo: un precio cinco veces la mediana no se
        -- cree, y uno cinco veces por debajo tampoco. Sin el lado bajo, el
        -- "mejor precio" lo fija un dato roto y todo lo demas sale mal.
        SELECT i.*, (d.m > 0 AND (i.pu > d.m * %(fac)s OR i.pu < d.m / %(fac)s)) AS rara
        FROM ing i JOIN med d ON d.producto_id = i.producto_id
    )
"""


def cp_unidad(unidad):
    """Como mostrar los precios de un producto.

    Un movimiento siempre viene en la unidad de Contifico, asi que aqui no hay
    nada que adivinar: gramos se muestran por kilo, mililitros por litro.
    """
    u = (unidad or '').strip().upper()
    if u == 'GRAMOS':
        return 'kg', 1000.0
    if u == 'MILILITROS':
        return 'L', 1000.0
    return (unidad or 'unidad').lower(), 1.0


def cp_proveedores_por_factura(cur, dias):
    """De que proveedor es cada numero de factura."""
    cur.execute("""
        SELECT num_documento, max(persona) FROM fact_detallada_compras
        WHERE fecha >= CURRENT_DATE - %s - 45 AND persona IS NOT NULL
        GROUP BY num_documento
    """, (dias,))
    import re as _re
    rx = _re.compile(CP_RE_FACTURA)
    mapa = {}
    for doc, per in cur.fetchall():
        m = rx.search(doc or '')
        if m:
            mapa['%s-%s-%s' % (m.group(1), m.group(2), m.group(3))] = per
    return mapa


@app.route('/api/costos/compras', methods=['GET'])
def costos_compras():
    """Donde se esta pagando de mas, ordenado por cuanto."""
    conn = None
    try:
        dias = min(int(request.args.get('dias', CP_DIAS)), 365)
        p = {'d': dias, 'fac': CP_FACTOR_SOSPECHOSO, 're': CP_RE_FACTURA}

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '120s'")
        cur.execute("WITH " + CP_SQL_ING + """,
            mejor AS (
                SELECT producto_id, min(pu) AS mp
                FROM marc WHERE NOT rara GROUP BY producto_id
            ),
            tot AS (
                SELECT producto_id, count(*) AS n, sum(cantidad) AS q,
                       sum(valor) AS g, count(*) FILTER (WHERE rara) AS nr,
                       sum(valor) FILTER (WHERE NOT con_factura) AS g_sin,
                       count(DISTINCT date_trunc('day', fecha)) AS ndias
                FROM marc GROUP BY producto_id
            ),
            ultima AS (
                SELECT DISTINCT ON (producto_id) producto_id, fecha, pu
                FROM marc WHERE NOT rara ORDER BY producto_id, fecha DESC, valor DESC
            ),
            demas AS (
                SELECT m.producto_id, sum((m.pu - j.mp) * m.cantidad) AS dm
                FROM marc m JOIN mejor j ON j.producto_id = m.producto_id
                WHERE NOT m.rara GROUP BY m.producto_id
            )
            SELECT pr.codigo, pr.nombre, coalesce(u.nombre, ''),
                   t.n, t.q, t.g, t.nr, coalesce(t.g_sin, 0),
                   j.mp, ul.pu, ul.fecha, coalesce(d.dm, 0)
            FROM tot t
            JOIN contifico_productos pr ON pr.id = t.producto_id
            LEFT JOIN contifico_unidades u ON u.id = pr.unidad_id
            JOIN mejor j  ON j.producto_id = t.producto_id
            JOIN ultima ul ON ul.producto_id = t.producto_id
            LEFT JOIN demas d ON d.producto_id = t.producto_id
            WHERE t.n >= 2
            ORDER BY coalesce(d.dm, 0) DESC
            LIMIT 150
        """, p)

        filas = []
        for r in cur.fetchall():
            uni, factor = cp_unidad(r[2])
            mejor = float(r[8] or 0) * factor
            ultimo = float(r[9] or 0) * factor
            gasto = float(r[5] or 0)
            filas.append({
                'codigo': r[0], 'nombre': r[1], 'unidad': uni,
                'ingresos': r[3], 'cantidad': float(r[4] or 0) / factor,
                'gasto': gasto, 'sospechosos': r[6],
                'pct_sin_factura': (float(r[7] or 0) / gasto * 100) if gasto else 0,
                'mejor_precio': mejor, 'ultimo_precio': ultimo,
                'ultima_fecha': r[10].isoformat() if r[10] else None,
                'pagado_de_mas': float(r[11] or 0),
                'vs_mejor': ((ultimo - mejor) / mejor * 100) if mejor else None,
            })

        return jsonify({'ok': True, 'dias': dias, 'filas': filas,
                        'pagado_de_mas': sum(f['pagado_de_mas'] for f in filas)})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/compras/producto', methods=['GET'])
def costos_compras_producto():
    """Un producto: cada ingreso, de quien vino y a que precio."""
    conn = None
    try:
        codigo = (request.args.get('codigo') or '').strip().upper()
        if not codigo:
            return jsonify({'error': 'codigo requerido'}), 400
        dias = min(int(request.args.get('dias', CP_DIAS)), 365)

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '120s'")

        cur.execute("""
            SELECT p.id, p.nombre, coalesce(u.nombre, '')
            FROM contifico_productos p
            LEFT JOIN contifico_unidades u ON u.id = p.unidad_id
            WHERE p.codigo = %s""", (codigo,))
        info = cur.fetchone()
        if not info:
            return jsonify({'error': 'producto %s no esta en el catalogo' % codigo}), 404
        pid, nombre, unidad = info
        uni, factor = cp_unidad(unidad)

        cur.execute("""
            SELECT m.fecha, m.codigo, m.cantidad, m.total_movimiento,
                   coalesce(b.nombre, ''), m.descripcion
            FROM contifico_movimientos m
            LEFT JOIN contifico_bodegas b ON b.id = m.bodega_destino_id
            WHERE m.producto_id = %s AND m.tipo = 'ING'
              AND m.fecha >= CURRENT_DATE - %s
              AND m.cantidad > 0 AND m.total_movimiento > 0
            ORDER BY m.fecha, m.codigo""", (pid, dias))
        crudos = cur.fetchall()

        porfactura = cp_proveedores_por_factura(cur, dias)
        cur.execute("""SELECT DISTINCT persona FROM fact_detallada_compras
                       WHERE persona IS NOT NULL AND length(persona) > 8""")
        personas = [x[0] for x in cur.fetchall()]

        import re as _re
        rx = _re.compile(CP_RE_FACTURA)
        ingresos = []
        for f, cod, q, tot, bod, desc in crudos:
            q = float(q or 0)
            precio = (float(tot or 0) / q * factor) if q else 0.0
            factura, prov = None, None
            m = rx.search(desc or '')
            if m:
                factura = '%s-%s-%s' % (m.group(1), m.group(2), m.group(3))
                prov = porfactura.get(factura)
            if not prov:
                # A veces el nombre del proveedor esta escrito y el numero no
                d = (desc or '').upper()
                for x in personas:
                    if x.upper()[:16] in d:
                        prov = x
                        break
            ingresos.append({
                'fecha': f.isoformat(), 'ingreso': cod,
                'cantidad': q / factor, 'valor': float(tot or 0),
                'precio': precio, 'bodega': bod,
                'factura': factura, 'proveedor': prov or CP_SIN_FACTURA,
                'sin_factura': not bool(m),
            })

        precios = sorted(x['precio'] for x in ingresos if x['precio'] > 0)
        mediana = precios[len(precios) // 2] if precios else 0
        for x in ingresos:
            x['sospechosa'] = bool(mediana and (x['precio'] > mediana * CP_FACTOR_SOSPECHOSO
                                                or x['precio'] < mediana / CP_FACTOR_SOSPECHOSO))

        creibles = [x for x in ingresos if not x['sospechosa']]
        mejor = min((x['precio'] for x in creibles), default=0.0)
        mejor_x = next((x for x in creibles if x['precio'] == mejor), None)
        ultima = creibles[-1] if creibles else None
        ultima_real = ingresos[-1] if ingresos else None
        pagado_de_mas = sum((x['precio'] - mejor) * x['cantidad'] for x in creibles)

        series = {}
        for x in ingresos:
            s = series.setdefault(x['proveedor'], {
                'proveedor': x['proveedor'], 'compras': 0, 'cantidad': 0.0,
                'gasto': 0.0, 'minimo': None, 'ultimo': None, 'ultima_fecha': None,
                'sin_factura': x['proveedor'] == CP_SIN_FACTURA})
            s['compras'] += 1
            s['cantidad'] += x['cantidad']
            s['gasto'] += x['valor']
            if not x['sospechosa']:
                s['minimo'] = x['precio'] if s['minimo'] is None else min(s['minimo'], x['precio'])
            if s['ultima_fecha'] is None or x['fecha'] >= s['ultima_fecha']:
                s['ultima_fecha'], s['ultimo'] = x['fecha'], x['precio']
        gasto = sum(s['gasto'] for s in series.values())
        proveedores = sorted(series.values(), key=lambda x: -x['gasto'])
        for s in proveedores:
            s['promedio'] = s['gasto'] / s['cantidad'] if s['cantidad'] else 0
            s['parte'] = (s['gasto'] / gasto * 100) if gasto else 0
            s['mejor'] = s['minimo'] is not None and mejor and abs(s['minimo'] - mejor) < 1e-9

        # A que costo salio a ventas: si el costo de salida se envenena, el
        # producto ensucia todos los informes aunque se haya comprado bien.
        cur.execute("""
            SELECT min(costo_unitario), max(costo_unitario),
                   percentile_cont(0.5) WITHIN GROUP (ORDER BY costo_unitario),
                   count(*)
            FROM costos_resumen_diario
            WHERE codigo_prod = %s AND tipo = 'EGR' AND costo_unitario > 0
              AND fecha >= CURRENT_DATE - %s""", (codigo, dias))
        c = cur.fetchone()
        salida = None
        if c and c[3]:
            cmed = float(c[2] or 0)
            cmax = float(c[1] or 0)
            # El costo de salida se juzga contra lo que costo comprar. Medirlo
            # contra si mismo no sirve: si esta mal todos los dias, la mediana
            # tambien esta mal y no salta nada.
            salida = {
                'minimo': float(c[0] or 0) * factor, 'maximo': cmax * factor,
                'tipico': cmed * factor, 'dias': c[3],
                'veces_compra': (cmed * factor / mejor) if mejor else None,
                'envenenado': bool(mejor and cmed * factor > mejor * 3),
            }

        return jsonify({
            'ok': True, 'codigo': codigo, 'nombre': nombre, 'unidad': uni,
            'dias': dias, 'ingresos': ingresos, 'proveedores': proveedores,
            'salida': salida,
            'resumen': {
                'n': len(ingresos),
                'n_sospechosas': sum(1 for x in ingresos if x['sospechosa']),
                'n_series': len(proveedores),
                'cantidad': sum(x['cantidad'] for x in ingresos),
                'gasto': gasto,
                'pct_sin_factura': (sum(s['gasto'] for s in proveedores if s['sin_factura'])
                                    / gasto * 100) if gasto else 0,
                'mejor_precio': mejor,
                'mejor_fecha': mejor_x['fecha'] if mejor_x else None,
                'mejor_proveedor': mejor_x['proveedor'] if mejor_x else None,
                'ultimo_precio': ultima['precio'] if ultima else None,
                'ultima_fecha': ultima['fecha'] if ultima else None,
                'ultimo_proveedor': ultima['proveedor'] if ultima else None,
                'ultimo_sospechoso': bool(ultima_real['sospechosa']) if ultima_real else False,
                'sospechoso_precio': (ultima_real['precio']
                                      if ultima_real and ultima_real['sospechosa'] else None),
                'sospechoso_fecha': (ultima_real['fecha']
                                     if ultima_real and ultima_real['sospechosa'] else None),
                'vs_mejor': (((ultima['precio'] - mejor) / mejor * 100)
                             if ultima and mejor else None),
                'pagado_de_mas': pagado_de_mas,
            },
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


# ============================================================
# MODULO: Tablero de fallos
# ============================================================
# Lo primero que tiene que ver un gerente no es una tabla de precios: es que
# esta roto, cuanto cuesta y que hacer. Las tablas van despues.
#
# Cada revision devuelve lo mismo -cuanta plata, cuantos productos, quienes
# son los peores y que hacer- para que la pantalla las pinte todas igual y se
# puedan ordenar por dinero.

TB_DIAS = 90


@app.route('/api/costos/tablero', methods=['GET'])
def costos_tablero():
    """Las seis cosas que pueden estar mal, con su costo."""
    conn = None
    try:
        dias = min(int(request.args.get('dias', TB_DIAS)), 365)
        p = {'d': dias, 'fac': CP_FACTOR_SOSPECHOSO, 're': CP_RE_FACTURA}
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '150s'")
        fallos = []

        # ---- 1. Mercaderia que entro sin factura -------------------------
        # No se sabe de quien vino ni si el precio fue el pactado. Es el punto
        # ciego mas grande: sobre esa plata no se puede negociar nada.
        cur.execute("WITH " + CP_SQL_ING + """
            SELECT coalesce(sum(valor) FILTER (WHERE NOT con_factura), 0),
                   coalesce(sum(valor), 0),
                   count(*) FILTER (WHERE NOT con_factura),
                   count(DISTINCT producto_id) FILTER (WHERE NOT con_factura)
            FROM marc
        """, p)
        sf, tot, nsf, psf = cur.fetchone()
        sf, tot = float(sf or 0), float(tot or 0)
        cur.execute("WITH " + CP_SQL_ING + """
            SELECT pr.codigo, pr.nombre, sum(m.valor), count(*)
            FROM marc m JOIN contifico_productos pr ON pr.id = m.producto_id
            WHERE NOT m.con_factura
            GROUP BY pr.codigo, pr.nombre ORDER BY sum(m.valor) DESC LIMIT 12
        """, p)
        fallos.append({
            'id': 'sin_factura', 'titulo': 'Mercaderia que entro sin factura',
            'valor': sf, 'gravedad': 'alta' if tot and sf / tot > 0.4 else 'media',
            'cifra': '%.0f%% de todo lo que entro' % (100 * sf / tot if tot else 0),
            'resumen': ('%s de %s en %d dias entraron sin un numero de factura en el ingreso, '
                        'repartidos en %d ingresos de %d productos.'
                        % (fc_money(sf), fc_money(tot), dias, nsf or 0, psf or 0)),
            'porque': ('Sin factura enlazada no se sabe de que proveedor vino ni si se pago el '
                       'precio pactado. Sobre esa plata no se puede negociar ni reclamar.'),
            'hacer': 'Exigir que el ingreso lleve el numero de factura. Es un campo, no un proceso nuevo.',
            'detalle': [{'codigo': r[0], 'nombre': r[1], 'valor': float(r[2] or 0), 'n': r[3]}
                        for r in cur.fetchall()],
            'cols': ['Producto', 'Ingresos', 'Entro sin factura'],
        })

        # ---- 2. Consumo que no se puede valorizar ------------------------
        # Si Contifico no le pone costo, el producto no existe para ningun
        # informe: no se puede saber cuanto costo lo que se consumio.
        cur.execute("""
            SELECT count(*) FILTER (WHERE coalesce(valor, 0) = 0), count(*),
                   count(DISTINCT codigo_prod) FILTER (WHERE coalesce(valor, 0) = 0),
                   count(DISTINCT codigo_prod)
            FROM costos_resumen_diario
            WHERE tipo = 'EGR' AND fecha >= CURRENT_DATE - %s
        """, (dias,))
        lc, lt, pc, pt = cur.fetchone()
        cur.execute("""
            SELECT codigo_prod, max(nombre_prod), max(categoria),
                   count(DISTINCT fecha), sum(cantidad)
            FROM costos_resumen_diario
            WHERE tipo = 'EGR' AND fecha >= CURRENT_DATE - %s
            GROUP BY codigo_prod HAVING sum(coalesce(valor, 0)) = 0
            ORDER BY count(DISTINCT fecha) DESC, sum(cantidad) DESC LIMIT 12
        """, (dias,))
        fallos.append({
            'id': 'sin_costo', 'titulo': 'Productos que se consumen sin costo',
            'valor': None, 'gravedad': 'alta' if pt and pc / pt > 0.25 else 'media',
            'cifra': '%d de %d productos' % (pc or 0, pt or 0),
            'resumen': ('%d productos se movieron en %d dias y Contifico no les asigna costo, '
                        'asi que su consumo vale cero. Son %d de %d lineas de movimiento.'
                        % (pc or 0, dias, lc or 0, lt or 0)),
            'porque': ('Un producto sin costo es invisible: no aparece en variaciones, no suma al '
                       'consumo y no se puede saber si se gasto de mas. Son casi todos los '
                       'procesados en bodega -salsas, zumos, carnicos-.'),
            'hacer': 'Ponerles costo en Contifico, o costearlos desde la formula de produccion.',
            'detalle': [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2],
                         'n': r[3], 'cantidad': float(r[4] or 0)} for r in cur.fetchall()],
            'cols': ['Producto', 'Dias con movimiento', 'Cantidad sin valorizar'],
        })

        # ---- 3. Costo de salida envenenado -------------------------------
        # Se compra bien y se descarga la venta a un costo disparatado. Ahi el
        # margen de ese producto y todo informe que lo incluya estan mal.
        cur.execute("WITH " + CP_SQL_ING + """,
            mejor AS (SELECT producto_id, min(pu) mp FROM marc WHERE NOT rara GROUP BY producto_id),
            compra AS (
                SELECT pr.codigo, m.mp,
                       CASE WHEN upper(coalesce(u.nombre, '')) IN ('GRAMOS', 'MILILITROS')
                            THEN 1000.0 ELSE 1.0 END AS factor
                FROM mejor m
                JOIN contifico_productos pr ON pr.id = m.producto_id
                LEFT JOIN contifico_unidades u ON u.id = pr.unidad_id
            ),
            venta AS (
                SELECT codigo_prod,
                       percentile_cont(0.5) WITHIN GROUP (ORDER BY costo_unitario) cu,
                       sum(valor) gastado, max(nombre_prod) nombre
                FROM costos_resumen_diario
                WHERE tipo = 'EGR' AND costo_unitario > 0 AND fecha >= CURRENT_DATE - %(d)s
                GROUP BY codigo_prod
            )
            SELECT v.codigo_prod, v.nombre, v.cu * c.factor, c.mp * c.factor,
                   (v.cu * c.factor) / NULLIF(c.mp * c.factor, 0), v.gastado
            FROM venta v JOIN compra c ON c.codigo = v.codigo_prod
            WHERE c.mp > 0 AND v.cu * c.factor > c.mp * c.factor * 3
            ORDER BY v.gastado DESC LIMIT 12
        """, p)
        env = [{'codigo': r[0], 'nombre': r[1], 'salida': float(r[2] or 0),
                'compra': float(r[3] or 0), 'veces': float(r[4] or 0),
                'valor': float(r[5] or 0)} for r in cur.fetchall()]
        fallos.append({
            'id': 'envenenado', 'titulo': 'Sale a ventas mucho mas caro de lo que entra',
            'valor': sum(x['valor'] for x in env),
            'gravedad': 'alta' if env else 'ok',
            'cifra': '%d productos' % len(env),
            'resumen': ('%d productos se descargan a ventas a un costo que no se parece al que se '
                        'compraron. Se valorizo %s de consumo con esos costos.'
                        % (len(env), fc_money(sum(x['valor'] for x in env)))),
            'porque': ('El costo de salida lo calcula Contifico como promedio ponderado, y deja de '
                       'tener sentido cuando el stock queda en negativo. Con eso mal, el margen '
                       'del producto y cualquier informe que lo incluya estan mal.'),
            'hacer': 'Revisar el stock de esos productos antes que el costo: casi siempre esta negativo.',
            'detalle': env,
            'cols': ['Producto', 'Entra a', 'Sale a', 'Veces'],
        })

        # ---- 4. Precios imposibles en el ingreso -------------------------
        cur.execute("WITH " + CP_SQL_ING + """
            SELECT pr.codigo, pr.nombre, count(*), sum(m.valor),
                   max(m.pu), min(d.m)
            FROM marc m
            JOIN contifico_productos pr ON pr.id = m.producto_id
            JOIN med d ON d.producto_id = m.producto_id
            WHERE m.rara
            GROUP BY pr.codigo, pr.nombre
            ORDER BY sum(m.valor) DESC LIMIT 12
        """, p)
        raros = [{'codigo': r[0], 'nombre': r[1], 'n': r[2], 'valor': float(r[3] or 0),
                  'precio': float(r[4] or 0), 'normal': float(r[5] or 0)}
                 for r in cur.fetchall()]
        cur.execute("WITH " + CP_SQL_ING + """
            SELECT count(*), coalesce(sum(valor), 0), count(DISTINCT producto_id)
            FROM marc WHERE rara""", p)
        nr, vr, prr = cur.fetchone()
        fallos.append({
            'id': 'precio_raro', 'titulo': 'Ingresos con un precio que no puede ser',
            'valor': float(vr or 0), 'gravedad': 'alta' if (nr or 0) > 50 else 'media',
            'cifra': '%d ingresos' % (nr or 0),
            'resumen': ('%d ingresos de %d productos entraron a un precio muy distinto al habitual '
                        'de ese mismo producto, por %s.' % (nr or 0, prr or 0, fc_money(vr))),
            'porque': ('Un precio fuera de escala casi siempre es la unidad mal puesta: gramos '
                       'donde van kilos, o al reves. Ensucia el costo del producto y arrastra '
                       'los informes de las semanas siguientes.'),
            'hacer': 'Corregir esos ingresos en Contifico. Estan listados con su numero.',
            'detalle': raros,
            'cols': ['Producto', 'Ingresos', 'Precio puesto', 'Precio normal'],
        })

        # ---- 5. Se paga de mas ------------------------------------------
        cur.execute("WITH " + CP_SQL_ING + """,
            mejor AS (SELECT producto_id, min(pu) mp FROM marc WHERE NOT rara GROUP BY producto_id),
            demas AS (
                SELECT m.producto_id, sum((m.pu - j.mp) * m.cantidad) dm, count(*) n
                FROM marc m JOIN mejor j ON j.producto_id = m.producto_id
                WHERE NOT m.rara GROUP BY m.producto_id
            )
            SELECT pr.codigo, pr.nombre, d.dm, d.n
            FROM demas d JOIN contifico_productos pr ON pr.id = d.producto_id
            WHERE d.dm > 0 ORDER BY d.dm DESC LIMIT 12
        """, p)
        caros = [{'codigo': r[0], 'nombre': r[1], 'valor': float(r[2] or 0), 'n': r[3]}
                 for r in cur.fetchall()]
        cur.execute("WITH " + CP_SQL_ING + """,
            mejor AS (SELECT producto_id, min(pu) mp FROM marc WHERE NOT rara GROUP BY producto_id)
            SELECT coalesce(sum((m.pu - j.mp) * m.cantidad), 0)
            FROM marc m JOIN mejor j ON j.producto_id = m.producto_id WHERE NOT m.rara""", p)
        dm = float(cur.fetchone()[0] or 0)
        fallos.append({
            'id': 'pagado_de_mas', 'titulo': 'Se pago mas que el mejor precio del mismo producto',
            'valor': dm, 'gravedad': 'media',
            'cifra': fc_money(dm),
            'resumen': ('Comprando siempre al mejor precio que ya se consiguio para cada producto, '
                        'se habrian ahorrado %s en %d dias.' % (fc_money(dm), dias)),
            'porque': ('El mismo producto entra a precios distintos segun el proveedor y el dia. '
                       'La diferencia contra el mejor precio conseguido es lo que se dejo sobre la mesa.'),
            'hacer': 'Mirar los de arriba: ahi esta concentrada casi toda la diferencia.',
            'detalle': caros,
            'cols': ['Producto', 'Ingresos', 'Se pago de mas'],
        })

        fallos.sort(key=lambda x: -(x['valor'] or 0))
        return jsonify({'ok': True, 'dias': dias, 'fallos': fallos})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


def fc_money(v):
    """Un numero de dinero, corto y legible."""
    try:
        v = float(v or 0)
    except (TypeError, ValueError):
        return '$0'
    return '$%s' % format(int(round(v)), ',d').replace(',', '.')


# ============================================================
# MODULO: Costos - ventanas envenenadas
# ============================================================
# Contifico no reprocesa hacia atras. Si un producto entra con el costo mal,
# todo lo que sale o se traslada mientras ese costo esta vigente se va con el
# costo malo, y corregir el ingreso despues no repara nada. Por eso no alcanza
# con mirar el ingreso ni con mirar la salida: hay que mirar la linea de tiempo
# del producto y encontrar las ventanas en las que el costo estuvo roto.
#
# Todo sale de contifico_movimientos y de un solo campo:
#
#     costo unitario = total_movimiento / cantidad
#
# total_movimiento esta en el 99% de los movimientos; costo_promedio solo en el
# 60%, y por leerlo a el parecia que 168 productos no tenian costo cuando si lo
# tienen.
#
# El ajuste de costo (AJU) queda fuera a proposito: corrige el costo hacia
# adelante, no mueve mercaderia, y meterlo en la banda distorsiona lo que se
# considera normal.

CV_DIAS = 180        # cuanto historial se mira
CV_VENTANA = 90      # sobre cuantos dias se calcula lo normal
CV_SIGMAS = 2.0      # cuando se considera que el costo se salio
CV_MIN_MUESTRAS = 8  # dias con movimiento antes de creerle a la desviacion

# La sincronizacion vuelve a insertar el mismo documento con un movimiento_id
# nuevo en cada corrida. Sin deduplicar, un producto aparece con varias veces
# el consumo que tuvo. Verificado con EGR 202608124294, que estaba 3 veces.
CV_DEDUP = ("m.codigo, m.producto_id, m.cantidad, m.total_movimiento, "
            "m.bodega_origen_id, m.fecha")

CV_SQL = """
WITH dedup AS (
    SELECT DISTINCT ON (%(k)s)
           m.fecha, m.tipo, m.producto_id, m.cantidad, m.total_movimiento
    FROM contifico_movimientos m
    WHERE m.tipo IN ('ING', 'EGR', 'TRA')
      AND m.fecha >= CURRENT_DATE - %%(dias)s
      AND m.cantidad > 0 AND m.total_movimiento > 0
      %(filtro)s
    ORDER BY %(k)s, m.id
),
dia AS (
    SELECT producto_id, fecha,
           -- Dos costos, no uno. El de salida manda porque es el que se lleva
           -- el dinero; el de entrada se vigila aparte porque es la causa.
           sum(total_movimiento) FILTER (WHERE tipo IN ('EGR', 'TRA'))
             / NULLIF(sum(cantidad) FILTER (WHERE tipo IN ('EGR', 'TRA')), 0) AS costo,
           sum(total_movimiento) FILTER (WHERE tipo = 'ING')
             / NULLIF(sum(cantidad) FILTER (WHERE tipo = 'ING'), 0)           AS costo_entrada,
           coalesce(sum(cantidad)         FILTER (WHERE tipo = 'EGR'), 0) AS sal,
           coalesce(sum(total_movimiento) FILTER (WHERE tipo = 'EGR'), 0) AS sal_val,
           coalesce(sum(cantidad)         FILTER (WHERE tipo = 'TRA'), 0) AS tras,
           coalesce(sum(total_movimiento) FILTER (WHERE tipo = 'TRA'), 0) AS tras_val,
           coalesce(sum(cantidad)         FILTER (WHERE tipo = 'ING'), 0) AS ent,
           coalesce(sum(total_movimiento) FILTER (WHERE tipo = 'ING'), 0) AS ent_val
    FROM dedup GROUP BY producto_id, fecha
),
-- Sin salida no hay costo de salida, y ese dia no entra en la banda: si
-- entrara, un dia que solo tuvo ingreso movería la referencia sin motivo.
con_salida AS (
    SELECT * FROM dia WHERE costo IS NOT NULL
),
banda AS (
    SELECT d.*,
           avg(costo)         OVER v AS media,
           stddev_samp(costo) OVER v AS desv,
           count(*)           OVER v AS muestras
    FROM con_salida d
    WINDOW v AS (PARTITION BY producto_id ORDER BY fecha
                 RANGE BETWEEN INTERVAL '%(vent)s days' PRECEDING
                           AND INTERVAL '1 day' PRECEDING)
)
SELECT producto_id, fecha, costo, costo_entrada, media, desv, muestras,
       sal, sal_val, tras, tras_val, ent, ent_val,
       CASE WHEN desv > 0 AND muestras >= %(min)s
            THEN (costo - media) / desv END AS sigmas
FROM banda ORDER BY producto_id, fecha
"""


def cv_sql(filtro=''):
    return CV_SQL % {'k': CV_DEDUP, 'filtro': filtro,
                     'vent': CV_VENTANA, 'min': CV_MIN_MUESTRAS}


def cv_ventanas(filas, sigmas=CV_SIGMAS):
    """Agrupa los dias fuera de banda en ventanas, y calcula el daño.

    Una ventana empieza el primer dia que el costo se sale y termina cuando
    vuelve dentro. El daño es lo que salio y se traslado durante la ventana,
    valorado contra el costo que ese producto deberia haber tenido.
    """
    ventanas, abierta = [], None
    for f in filas:
        costo = float(f['costo'] or 0)
        sig = f['sigmas']
        fuera = sig is not None and abs(float(sig)) > sigmas
        if fuera and abierta is None:
            abierta = {'desde': f['fecha'], 'hasta': f['fecha'],
                       'normal': float(f['media'] or 0), 'pico': costo,
                       'sal': 0.0, 'sal_val': 0.0, 'tras': 0.0, 'tras_val': 0.0,
                       'dias': 0, 'sigmas': float(sig)}
        if abierta is None:
            continue
        abierta['hasta'] = f['fecha']
        abierta['dias'] += 1
        if abs(costo - abierta['normal']) > abs(abierta['pico'] - abierta['normal']):
            abierta['pico'] = costo
        for k in ('sal', 'sal_val', 'tras', 'tras_val'):
            abierta[k] += float(f[k] or 0)
        if sig is not None and abs(float(sig)) > abs(abierta['sigmas']):
            abierta['sigmas'] = float(sig)
        if not fuera:
            ventanas.append(abierta)
            abierta = None
    if abierta:
        ventanas.append(abierta)

    for v in ventanas:
        # Lo que decide la gravedad es cuanto se movio el COSTO UNITARIO.
        v['veces'] = (v['pico'] / v['normal']) if v['normal'] else None
        v['variacion'] = ((v['pico'] - v['normal']) / v['normal'] * 100) if v['normal'] else None
        # El dinero queda como contexto: dice sobre cuanto aplica la desviacion,
        # pero no decide nada.
        v['movido'] = v['sal'] + v['tras']
        v['valorado'] = v['sal_val'] + v['tras_val']
    return ventanas


def cv_catalogo(cur):
    """codigo, nombre y categoria de cada producto."""
    cur.execute("""
        SELECT p.id, p.codigo, p.nombre, coalesce(c.nombre, 'SIN CATEGORIA'),
               coalesce(u.nombre, '')
        FROM contifico_productos p
        LEFT JOIN contifico_categorias c ON c.id = p.categoria_id
        LEFT JOIN contifico_unidades u ON u.id = p.unidad_id
    """)
    return {r[0]: {'codigo': r[1], 'nombre': r[2], 'categoria': r[3], 'unidad': r[4]}
            for r in cur.fetchall()}


@app.route('/api/costos/ventanas', methods=['GET'])
def costos_ventanas():
    """Todas las ventanas en las que el costo de un producto estuvo roto."""
    conn = None
    try:
        dias = min(int(request.args.get('dias', CV_DIAS)), 400)
        sigmas = float(request.args.get('sigmas', CV_SIGMAS))
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '240s'")
        cat = cv_catalogo(cur)

        cur.execute(cv_sql(), {'dias': dias})
        cols = [d[0] for d in cur.description]
        por_producto = {}
        for r in cur.fetchall():
            f = dict(zip(cols, r))
            por_producto.setdefault(f['producto_id'], []).append(f)

        productos, todas, vistos = [], [], set()
        for pid, filas in por_producto.items():
            info = cat.get(pid)
            if not info or not info['codigo']:
                continue
            vs = [v for v in cv_ventanas(filas, sigmas) if v['movido'] > 0]
            if not vs:
                continue
            for v in vs:
                v.update({'codigo': info['codigo'], 'nombre': info['nombre'],
                          'categoria': info['categoria'], 'unidad': info['unidad'],
                          'desde': v['desde'].isoformat(), 'hasta': v['hasta'].isoformat()})
            todas.extend(vs)
            vistos.add(pid)
            peor = max(vs, key=lambda x: abs(x['sigmas']))
            productos.append({
                'codigo': info['codigo'], 'nombre': info['nombre'],
                'categoria': info['categoria'], 'unidad': info['unidad'],
                'ventanas': len(vs), 'peor': peor,
                'sigmas': peor['sigmas'], 'veces': peor['veces'],
                'movido': sum(v['movido'] for v in vs),
                'valorado': sum(v['valorado'] for v in vs),
                'dias_con_movimiento': len(filas),
            })

        productos.sort(key=lambda x: -abs(x['sigmas']))
        todas.sort(key=lambda x: -abs(x['sigmas']))

        # Cuantos productos tuvieron el costo roto cada dia. Si el problema
        # fuera azaroso las barras serian parejas; si se amontonan, hay un
        # proceso detras.
        pordia = {}
        for pid, filas in por_producto.items():
            if pid not in vistos:
                continue
            for f in filas:
                if f['sigmas'] is None or abs(float(f['sigmas'])) <= sigmas:
                    continue
                d = pordia.setdefault(f['fecha'].isoformat(), {'fecha': f['fecha'].isoformat(),
                                                               'productos': 0, 'arriba': 0, 'abajo': 0})
                d['productos'] += 1
                if float(f['sigmas']) > 0:
                    d['arriba'] += 1
                else:
                    d['abajo'] += 1
        por_dia = sorted(pordia.values(), key=lambda x: x['fecha'])

        porcat = {}
        for p in productos:
            c = porcat.setdefault(p['categoria'], {'categoria': p['categoria'],
                                                   'productos': 0, 'ventanas': 0,
                                                   'peor_sigmas': 0.0})
            c['productos'] += 1
            c['ventanas'] += p['ventanas']
            if abs(p['sigmas']) > abs(c['peor_sigmas']):
                c['peor_sigmas'] = p['sigmas']
        categorias = sorted(porcat.values(), key=lambda x: -x['productos'])

        return jsonify({
            'ok': True, 'dias': dias, 'sigmas': sigmas, 'ventana_base': CV_VENTANA,
            'n_productos': len(productos), 'n_ventanas': len(todas),
            'por_dia': por_dia,
            'productos': productos[:120],
            'ventanas': todas[:120],
            'categorias': categorias,
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/linea', methods=['GET'])
def costos_linea():
    """La linea de tiempo de un producto: costo diario, banda normal y ventanas."""
    conn = None
    try:
        codigo = (request.args.get('codigo') or '').strip().upper()
        if not codigo:
            return jsonify({'error': 'codigo requerido'}), 400
        dias = min(int(request.args.get('dias', CV_DIAS)), 400)
        sigmas = float(request.args.get('sigmas', CV_SIGMAS))

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '120s'")
        cur.execute("""
            SELECT p.id, p.nombre, coalesce(c.nombre, 'SIN CATEGORIA'), coalesce(u.nombre, '')
            FROM contifico_productos p
            LEFT JOIN contifico_categorias c ON c.id = p.categoria_id
            LEFT JOIN contifico_unidades u ON u.id = p.unidad_id
            WHERE p.codigo = %s""", (codigo,))
        info = cur.fetchone()
        if not info:
            return jsonify({'error': 'no existe el producto %s' % codigo}), 404
        pid, nombre, categoria, unidad = info

        cur.execute(cv_sql(" AND m.producto_id = %(pid)s "), {'dias': dias, 'pid': pid})
        cols = [d[0] for d in cur.description]
        filas = [dict(zip(cols, r)) for r in cur.fetchall()]
        ventanas = cv_ventanas(filas, sigmas)

        serie = [{
            'fecha': f['fecha'].isoformat(),
            'costo': float(f['costo'] or 0),
            'costo_entrada': float(f['costo_entrada']) if f['costo_entrada'] is not None else None,
            'media': float(f['media']) if f['media'] is not None else None,
            'desv': float(f['desv']) if f['desv'] is not None else None,
            'sigmas': float(f['sigmas']) if f['sigmas'] is not None else None,
            'sal': float(f['sal'] or 0), 'sal_val': float(f['sal_val'] or 0),
            'tras': float(f['tras'] or 0), 'tras_val': float(f['tras_val'] or 0),
            'ent': float(f['ent'] or 0), 'ent_val': float(f['ent_val'] or 0),
            'fuera': f['sigmas'] is not None and abs(float(f['sigmas'])) > sigmas,
        } for f in filas]

        for v in ventanas:
            v['desde'] = v['desde'].isoformat()
            v['hasta'] = v['hasta'].isoformat()

        return jsonify({
            'ok': True, 'codigo': codigo, 'nombre': nombre,
            'categoria': categoria, 'unidad': unidad, 'dias': dias,
            'serie': serie,
            'ventanas': sorted([v for v in ventanas if v['movido'] > 0],
                               key=lambda x: -abs(x['sigmas'])),
            'sin_movimiento': len([v for v in ventanas if v['movido'] <= 0]),
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


CN_RECIENTE = 30      # dias del periodo nuevo
CN_ANTERIOR = 90      # hasta donde se mira el periodo viejo
CN_FACTOR = 2.0       # al doble o a la mitad, cambio de escala


@app.route('/api/costos/niveles', methods=['GET'])
def costos_niveles():
    """Productos cuyo costo de salida cambio de escala y se quedo asi.

    Es el punto ciego de la banda movil: lo que sube y vuelve lo agarra la
    banda; lo que sube y se queda lo absorbe y deja de avisar. Aqui no hay
    media movil, se comparan dos periodos.
    """
    conn = None
    try:
        factor = float(request.args.get('factor', CN_FACTOR))
        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '180s'")
        cur.execute("""
            WITH dedup AS (
                SELECT DISTINCT ON (%s)
                       m.fecha, m.tipo, m.producto_id, m.cantidad, m.total_movimiento
                FROM contifico_movimientos m
                WHERE m.tipo IN ('EGR', 'TRA')
                  AND m.fecha >= CURRENT_DATE - %%(ant)s
                  AND m.cantidad > 0 AND m.total_movimiento > 0
                ORDER BY %s, m.id
            ),
            dia AS (
                SELECT producto_id, fecha,
                       sum(total_movimiento) / NULLIF(sum(cantidad), 0) AS costo,
                       sum(cantidad) AS cant, sum(total_movimiento) AS valor
                FROM dedup GROUP BY producto_id, fecha
            ),
            comp AS (
                SELECT producto_id,
                       percentile_cont(0.5) WITHIN GROUP (ORDER BY costo)
                         FILTER (WHERE fecha >= CURRENT_DATE - %%(rec)s) AS ahora,
                       percentile_cont(0.5) WITHIN GROUP (ORDER BY costo)
                         FILTER (WHERE fecha < CURRENT_DATE - %%(rec)s) AS antes,
                       count(*) FILTER (WHERE fecha >= CURRENT_DATE - %%(rec)s) AS n_ahora,
                       count(*) FILTER (WHERE fecha < CURRENT_DATE - %%(rec)s) AS n_antes,
                       sum(cant)  FILTER (WHERE fecha >= CURRENT_DATE - %%(rec)s) AS cant_ahora,
                       sum(valor) FILTER (WHERE fecha >= CURRENT_DATE - %%(rec)s) AS valor_ahora
                FROM dia GROUP BY producto_id
            )
            SELECT p.codigo, p.nombre, coalesce(c.nombre, 'SIN CATEGORIA'),
                   coalesce(u.nombre, ''),
                   x.antes, x.ahora, x.n_antes, x.n_ahora,
                   x.cant_ahora, x.valor_ahora,
                   x.ahora / NULLIF(x.antes, 0) AS razon
            FROM comp x
            JOIN contifico_productos p ON p.id = x.producto_id
            LEFT JOIN contifico_categorias c ON c.id = p.categoria_id
            LEFT JOIN contifico_unidades u ON u.id = p.unidad_id
            WHERE x.antes > 0 AND x.ahora > 0
              AND x.n_antes >= 5 AND x.n_ahora >= 3
              AND (x.ahora > x.antes * %%(f)s OR x.ahora < x.antes / %%(f)s)
            ORDER BY greatest(x.ahora / NULLIF(x.antes, 0),
                              x.antes / NULLIF(x.ahora, 0)) DESC
            LIMIT 80
        """ % (CV_DEDUP, CV_DEDUP),
            {'rec': CN_RECIENTE, 'ant': CN_ANTERIOR, 'f': factor})

        filas = []
        for r in cur.fetchall():
            antes, ahora = float(r[4] or 0), float(r[5] or 0)
            cant = float(r[8] or 0)
            valor = float(r[9] or 0)
            # Lo que costo el consumo de los ultimos 30 dias contra lo que
            # habria costado al nivel anterior.
            debio = cant * antes
            filas.append({
                'codigo': r[0], 'nombre': r[1], 'categoria': r[2], 'unidad': r[3],
                'antes': antes, 'ahora': ahora,
                'dias_antes': r[6], 'dias_ahora': r[7],
                'cantidad': cant, 'valor': valor, 'debio': debio,
                'diferencia': valor - debio,
                'razon': float(r[10] or 0),
                'sentido': 'subio' if ahora > antes else 'bajo',
            })
        return jsonify({'ok': True, 'factor': factor,
                        'reciente': CN_RECIENTE, 'anterior': CN_ANTERIOR,
                        'filas': filas})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


@app.route('/api/costos/movimientos', methods=['GET'])
def costos_movimientos():
    """Los movimientos de un producto en un rango, uno por uno.

    Es lo que hace falta para poder actuar: no basta con saber que el 30 de
    abril el costo se rompio, hay que saber cual fue el documento, de que
    bodega salio y a que costo, para poder ir a Contifico a mirarlo.
    """
    conn = None
    try:
        codigo = (request.args.get('codigo') or '').strip().upper()
        desde = request.args.get('desde')
        hasta = request.args.get('hasta') or desde
        if not codigo or not desde:
            return jsonify({'error': 'codigo y desde requeridos'}), 400

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '90s'")
        cur.execute("""
            SELECT p.id, p.nombre, coalesce(u.nombre, '')
            FROM contifico_productos p
            LEFT JOIN contifico_unidades u ON u.id = p.unidad_id
            WHERE p.codigo = %s""", (codigo,))
        info = cur.fetchone()
        if not info:
            return jsonify({'error': 'no existe %s' % codigo}), 404
        pid, nombre, unidad = info

        cur.execute("""
            SELECT DISTINCT ON (m.codigo, m.cantidad, m.total_movimiento,
                                m.bodega_origen_id, m.fecha)
                   m.fecha, m.tipo, m.codigo, m.cantidad, m.total_movimiento,
                   m.costo_promedio,
                   coalesce(bo.nombre, ''), coalesce(bd.nombre, ''),
                   m.descripcion, m.estado,
                   -- Cuantas lineas tiene el documento. Si tiene mas de una,
                   -- total_movimiento es el total de todas y no sirve para
                   -- sacar el costo de esta.
                   (SELECT count(*) FROM contifico_movimientos d
                     WHERE d.codigo = m.codigo AND d.fecha = m.fecha) AS lineas_doc
            FROM contifico_movimientos m
            LEFT JOIN contifico_bodegas bo ON bo.id = m.bodega_origen_id
            LEFT JOIN contifico_bodegas bd ON bd.id = m.bodega_destino_id
            WHERE m.producto_id = %(pid)s
              AND m.tipo IN ('ING', 'EGR', 'TRA')
              AND m.fecha >= %(d)s AND m.fecha <= %(h)s
              AND m.cantidad > 0
            ORDER BY m.codigo, m.cantidad, m.total_movimiento,
                     m.bodega_origen_id, m.fecha, m.id
        """, {'pid': pid, 'd': desde, 'h': hasta})

        movs = []
        for r in cur.fetchall():
            lineas = int(r[10] or 1)
            cant = float(r[3] or 0)
            # El costo unitario solo se puede sacar del total cuando el
            # documento trae una sola linea. Si trae varias, se usa el
            # costo_promedio -que viene redondeado a centavos- y si tampoco
            # esta, se dice que no se sabe, en vez de inventarlo.
            if lineas == 1 and cant:
                costo, origen_costo = float(r[4] or 0) / cant, 'documento'
            elif r[5] and float(r[5]) > 0:
                costo, origen_costo = float(r[5]), 'costo promedio'
            else:
                costo, origen_costo = None, None
            movs.append({
                'fecha': r[0].isoformat(), 'tipo': r[1], 'documento': r[2],
                'cantidad': cant, 'total_documento': float(r[4] or 0),
                'costo': costo, 'origen_costo': origen_costo,
                'lineas_doc': lineas,
                'origen': r[6], 'destino': r[7],
                'descripcion': ' · '.join((r[8] or '').split(chr(10)))[:120],
                'estado': r[9],
            })
        movs.sort(key=lambda x: (x['fecha'], -(x['costo'] or 0)))

        return jsonify({'ok': True, 'codigo': codigo, 'nombre': nombre,
                        'unidad': unidad, 'desde': desde, 'hasta': hasta,
                        'movimientos': movs})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)[:300]}), 500
    finally:
        if conn: fc_release_movimientos_db(conn)


# =====================================================================
#  PANEL DE COSTOS  -  un solo endpoint para todo el tablero
# =====================================================================
# Todo sale de costos_diario, que trae los movimientos valorizados contra el
# costo real (costo_vigente_producto) y con los traslados por partida doble.
#
# Un solo endpoint, y no uno por bloque, para que el filtro mueva el tablero
# entero de una vez. Antes cada bloque consultaba por su lado y podian acabar
# mostrando periodos distintos sin que se notara.
#
# Los cuatro filtros -desde, hasta, bodega, categoria- se aplican a todo.

PN_DIAS = 90

# --- desfases de costo ------------------------------------------------
# Un desfase es un tramo en el que el producto costo algo distinto de lo que
# suele costar. La referencia no es un promedio -que se ensucia con el propio
# desfase- sino el costo que MAS DIAS estuvo vigente en el periodo: lo que ese
# producto vale normalmente.
#
# El daño es la cantidad que se movio durante el tramo, valorada contra la
# diferencia. Contifico no reprocesa hacia atras: lo que salio mientras el
# costo estuvo mal, salio mal, y ya no se corrige solo.
PN_DESVIO = 0.15     # 15% de separacion para considerarlo desfase

PN_SQL_DESFASES = """
WITH lim AS (SELECT %(d)s::date AS d, %(h)s::date AS h),
tramo AS (
    SELECT v.producto_id, v.codigo_prod, v.costo, v.fuente, v.confianza,
           greatest(v.desde, lim.d) AS ini,
           least(v.hasta, lim.h)   AS fin
    FROM costo_vigente_producto v, lim
    WHERE v.desde <= lim.h AND v.hasta >= lim.d AND v.costo > 0
),
condias AS (SELECT *, (fin - ini + 1) AS dias FROM tramo),
-- Lo que ese producto vale normalmente: la MEDIANA de los dias, no el
-- tramo mas largo. Con "el mas largo" el filete de carne tomaba 3,25 como
-- referencia -su tramo mas duradero- y marcaba como desfase los diez tramos
-- normales de 0,87 a 1,68. La mediana ponderada por dias parte la serie por
-- la mitad y no se deja arrastrar por un solo tramo, dure lo que dure.
acumulado AS (
    SELECT producto_id, costo, dias,
           sum(dias) OVER (PARTITION BY producto_id ORDER BY costo,
                           costo ROWS UNBOUNDED PRECEDING) AS acum,
           sum(dias) OVER (PARTITION BY producto_id)        AS total
    FROM condias
),
refer AS (
    SELECT DISTINCT ON (producto_id) producto_id, costo AS ref
    FROM acumulado WHERE acum >= total / 2.0
    ORDER BY producto_id, costo
),
fuera AS (
    SELECT c.*, r.ref
    FROM condias c JOIN refer r USING (producto_id)
    WHERE r.ref > 0
      AND (c.costo / r.ref > 1 + %(desvio)s OR r.ref / c.costo > 1 + %(desvio)s)
),
movido AS (
    SELECT f.codigo_prod, f.ini,
           sum(d.cantidad) AS qty, sum(d.valor) AS val,
           max(d.nombre_prod) AS nombre, max(d.categoria) AS categoria,
           max(d.unidad) AS unidad, count(DISTINCT d.fecha) AS dias_mov
    FROM fuera f
    JOIN costos_diario d ON d.codigo_prod = f.codigo_prod
     AND d.tipo IN ('EGR', 'TRA_SALE')
     AND d.fecha BETWEEN f.ini AND f.fin
     ___FILTRO___
    GROUP BY 1, 2
)
SELECT f.codigo_prod, m.nombre, m.categoria, m.unidad,
       f.ini, f.fin, f.dias, f.ref, f.costo,
       (f.costo / f.ref - 1) * 100 AS desvio,
       coalesce(m.qty, 0), coalesce(m.dias_mov, 0),
       (f.costo - f.ref) * coalesce(m.qty, 0) AS dano,
       f.fuente, f.confianza
FROM fuera f JOIN movido m ON m.codigo_prod = f.codigo_prod AND m.ini = f.ini
ORDER BY abs((f.costo - f.ref) * coalesce(m.qty, 0)) DESC
LIMIT 40
"""




def pn_filtro(bodega, categoria, producto='', centro='', alias='d'):
    """El filtro comun, como SQL y parametros. Vacio si no filtran nada."""
    sql, p = '', {}
    if bodega:
        sql += ' AND %s.bodega = %%(bodega)s' % alias
        p['bodega'] = bodega
    if categoria:
        sql += ' AND %s.categoria = %%(categoria)s' % alias
        p['categoria'] = categoria
    if producto:
        sql += ' AND %s.codigo_prod = %%(producto)s' % alias
        p['producto'] = producto
    if centro:
        sql += ' AND %s.centro_costo = %%(centro)s' % alias
        p['centro'] = centro
    return sql, p


@app.route('/api/costos/panel', methods=['GET'])
def costos_panel():
    conn = None
    try:
        hasta = co_dia(request.args.get('hasta'), co_ultimo_dia_completo())
        desde = co_dia(request.args.get('desde'), hasta - timedelta(days=PN_DIAS))
        if desde > hasta:
            desde, hasta = hasta, desde
        bodega = (request.args.get('bodega') or '').strip()
        categoria = (request.args.get('categoria') or '').strip()
        producto = (request.args.get('producto') or '').strip().upper()

        centro = (request.args.get('centro') or '').strip()

        f, fp = pn_filtro(bodega, categoria, producto, centro)
        p = {'d': desde, 'h': hasta}
        p.update(fp)

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '120s'")

        # --- que se puede filtrar --------------------------------------
        cur.execute("""SELECT DISTINCT bodega FROM costos_diario
                       WHERE bodega <> '(SIN BODEGA)' ORDER BY 1""")
        bodegas = [r[0] for r in cur.fetchall()]
        cur.execute("""SELECT DISTINCT categoria FROM costos_diario
                       WHERE categoria <> '(SIN CATEGORIA)' ORDER BY 1""")
        categorias = [r[0] for r in cur.fetchall()]
        cur.execute("""SELECT DISTINCT centro_costo FROM costos_diario
                       WHERE centro_costo IS NOT NULL ORDER BY 1""")
        centros = [r[0] for r in cur.fetchall()]
        # La lista completa alimenta el buscador del filtro. Son unos 450
        # productos: cabe de sobra en la respuesta y evita un viaje aparte.
        cur.execute("""SELECT codigo_prod, max(nombre_prod), max(categoria)
                       FROM costos_diario WHERE codigo_prod <> 'SIN_CODIGO'
                       GROUP BY 1 ORDER BY 2""")
        productos = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2]}
                     for r in cur.fetchall()]
        cur.execute("SELECT min(fecha), max(fecha), max(actualizado_en) FROM costos_diario")
        rmin, rmax, actualizado = cur.fetchone()

        # --- lo que se consumio ----------------------------------------
        cur.execute("""
            SELECT coalesce(sum(d.valor), 0), count(DISTINCT d.codigo_prod),
                   coalesce(sum(d.cantidad) FILTER (WHERE d.sin_costo > 0), 0),
                   count(DISTINCT d.codigo_prod) FILTER (WHERE d.sin_costo > 0)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f, p)
        consumo, n_prod, cant_sin, prod_sin = cur.fetchone()

        # --- los cambios de costo, con lo que costaron -----------------
        # Un cambio de costo no es noticia por si mismo. Lo es por la cantidad
        # que salio despues, valorada a la diferencia: eso es plata que se fue
        # al costo nuevo. Por ahi se ordena la lista.
        cur.execute("""
            WITH tr AS (
                SELECT producto_id, codigo_prod, desde, hasta, costo, fuente,
                       confianza, lag(costo) OVER w AS antes
                FROM costo_vigente_producto
                WINDOW w AS (PARTITION BY producto_id ORDER BY desde)
            ),
            camb AS (
                SELECT * FROM tr
                WHERE antes IS NOT NULL AND antes > 0 AND costo > 0
                  AND desde BETWEEN %(d)s AND %(h)s
            ),
            cons AS (
                SELECT c.codigo_prod, c.desde, sum(d.cantidad) AS qty,
                       max(d.nombre_prod) AS nombre, max(d.categoria) AS categoria,
                       max(d.unidad) AS unidad
                FROM camb c
                JOIN costos_diario d ON d.codigo_prod = c.codigo_prod
                 AND d.tipo = 'EGR'
                 AND d.fecha BETWEEN greatest(c.desde, %(d)s) AND least(c.hasta, %(h)s)
                 """ + f + """
                GROUP BY 1, 2
            )
            SELECT c.codigo_prod, o.nombre, o.categoria, o.unidad, c.desde,
                   c.antes, c.costo, coalesce(o.qty, 0),
                   (c.costo - c.antes) * coalesce(o.qty, 0) AS impacto,
                   c.fuente, c.confianza
            -- INNER, no LEFT: un cambio de costo entra en la lista solo si
            -- ese producto salio de verdad dentro de lo que se esta mirando.
            -- Con LEFT, filtrar por bodega no quitaba ningun cambio: seguian
            -- los 60 en pantalla, solo que con cantidad cero.
            FROM camb c JOIN cons o
              ON o.codigo_prod = c.codigo_prod AND o.desde = c.desde
            ORDER BY abs((c.costo - c.antes) * coalesce(o.qty, 0)) DESC
            LIMIT 60""", p)
        cambios = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2], 'unidad': r[3],
                    'fecha': r[4].isoformat(), 'antes': float(r[5]), 'ahora': float(r[6]),
                    'pct': (float(r[6]) / float(r[5]) - 1) * 100 if r[5] else None,
                    'cantidad': float(r[7] or 0), 'impacto': float(r[8] or 0),
                    'fuente': r[9], 'confianza': r[10]} for r in cur.fetchall()]
        impacto = sum(c['impacto'] for c in cambios)

        # --- los desfases de costo -------------------------------------
        p['desvio'] = PN_DESVIO
        cur.execute(PN_SQL_DESFASES.replace('___FILTRO___', f), p)
        desfases = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2], 'unidad': r[3],
                     'desde': r[4].isoformat(), 'hasta': r[5].isoformat(), 'dias': r[6],
                     'normal': float(r[7]), 'costo': float(r[8]), 'desvio': float(r[9]),
                     'cantidad': float(r[10] or 0), 'dias_mov': r[11],
                     'dano': float(r[12] or 0), 'fuente': r[13], 'confianza': r[14]}
                    for r in cur.fetchall()]
        dano_total = sum(abs(x['dano']) for x in desfases)

        # --- consumo dia a dia -----------------------------------------
        # Con un producto elegido la serie lleva ademas su costo unitario de
        # cada dia: es la linea que muestra como vario el costo.
        cur.execute("""
            SELECT d.fecha, sum(d.valor), count(DISTINCT d.codigo_prod),
                   CASE WHEN sum(d.cantidad) <> 0
                        THEN sum(d.valor) / sum(d.cantidad) END
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            GROUP BY 1 ORDER BY 1""", p)
        serie = [{'fecha': r[0].isoformat(), 'valor': float(r[1] or 0), 'prods': r[2],
                  'costo': float(r[3]) if r[3] is not None else None}
                 for r in cur.fetchall()]

        # --- por bodega y por categoria --------------------------------
        cur.execute("""
            SELECT d.bodega, sum(d.valor), count(DISTINCT d.codigo_prod)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            GROUP BY 1 ORDER BY 2 DESC""", p)
        por_bodega = [{'nombre': r[0], 'valor': float(r[1] or 0), 'prods': r[2]}
                      for r in cur.fetchall()]

        # El centro de costo agrupa bodegas -Principal y Pulmon son uno solo-,
        # asi que no es la misma vista con otro nombre.
        cur.execute("""
            SELECT d.centro_costo, sum(d.valor), count(DISTINCT d.codigo_prod)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            GROUP BY 1 ORDER BY 2 DESC""", p)
        por_centro = [{'nombre': r[0], 'valor': float(r[1] or 0), 'prods': r[2]}
                      for r in cur.fetchall()]

        cur.execute("""
            SELECT d.categoria, sum(d.valor), count(DISTINCT d.codigo_prod)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            GROUP BY 1 ORDER BY 2 DESC LIMIT 14""", p)
        por_categoria = [{'nombre': r[0], 'valor': float(r[1] or 0), 'prods': r[2]}
                         for r in cur.fetchall()]

        # --- en que se va la plata -------------------------------------
        cur.execute("""
            SELECT d.codigo_prod, max(d.nombre_prod), max(d.categoria), max(d.unidad),
                   sum(d.valor), sum(d.cantidad), min(d.confianza)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            GROUP BY 1 ORDER BY 5 DESC LIMIT 25""", p)
        top = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2], 'unidad': r[3],
                'valor': float(r[4] or 0), 'cantidad': float(r[5] or 0),
                'confianza': r[6]} for r in cur.fetchall()]

        # --- lo que se mueve sin costo ---------------------------------
        cur.execute("""
            SELECT d.codigo_prod, max(d.nombre_prod), max(d.categoria), max(d.unidad),
                   sum(d.cantidad), count(DISTINCT d.fecha)
            FROM costos_diario d
            WHERE d.tipo = 'EGR' AND d.fecha BETWEEN %(d)s AND %(h)s
              AND d.sin_costo > 0""" + f + """
            GROUP BY 1 ORDER BY 6 DESC, 5 DESC LIMIT 25""", p)
        sin_costo = [{'codigo': r[0], 'nombre': r[1], 'categoria': r[2], 'unidad': r[3],
                      'cantidad': float(r[4] or 0), 'dias': r[5]} for r in cur.fetchall()]

        # Con un producto elegido, su historia de costo es lo primero que se
        # quiere ver, asi que va en la misma respuesta y no en un segundo viaje.
        tramos = []
        if producto:
            cur.execute("""SELECT desde, hasta, costo, fuente, confianza
                           FROM costo_vigente_producto
                           WHERE codigo_prod = %s ORDER BY desde""", (producto,))
            tramos = [{'desde': r[0].isoformat(), 'hasta': r[1].isoformat(),
                       'costo': float(r[2]), 'fuente': r[3], 'confianza': r[4]}
                      for r in cur.fetchall()]

        cur.close()
        return jsonify({
            'ok': True,
            'desde': desde.isoformat(), 'hasta': hasta.isoformat(),
            'bodega': bodega, 'categoria': categoria, 'producto': producto,
            'centro': centro,
            'tramos': tramos,
            'filtros': {'bodegas': bodegas, 'categorias': categorias,
                        'productos': productos, 'centros': centros,
                        'min': rmin.isoformat() if rmin else None,
                        'max': rmax.isoformat() if rmax else None,
                        'actualizado': actualizado.isoformat() if actualizado else None},
            'kpi': {'consumo': float(consumo or 0), 'productos': n_prod or 0,
                    'cambios': len(cambios), 'impacto': impacto,
                    'desfases': len(desfases), 'dano': dano_total,
                    'prod_desfase': len(set(x['codigo'] for x in desfases)),
                    'sin_costo_prod': prod_sin or 0,
                    'sin_costo_cant': float(cant_sin or 0)},
            'serie': serie, 'bodegas': por_bodega, 'centros': por_centro,
            'categorias': por_categoria,
            'top': top, 'cambios': cambios, 'sin_costo': sin_costo,
            'desfases': desfases, 'desvio': PN_DESVIO,
        })
    except Exception as e:
        app.logger.exception('costos_panel')
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/costos/panel/movimientos', methods=['GET'])
def costos_panel_movimientos():
    """Lo que abre el popup: los dias, los documentos y los tramos de costo."""
    conn = None
    try:
        codigo = (request.args.get('codigo') or '').strip().upper()
        if not codigo:
            return jsonify({'ok': False, 'error': 'falta el codigo'}), 400
        hasta = co_dia(request.args.get('hasta'), co_ultimo_dia_completo())
        desde = co_dia(request.args.get('desde'), hasta - timedelta(days=PN_DIAS))
        bodega = (request.args.get('bodega') or '').strip()
        f, fp = pn_filtro(bodega, '')  # el codigo ya viene aparte
        p = {'c': codigo, 'd': desde, 'h': hasta}
        p.update(fp)

        conn = fc_get_movimientos_db()
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '60s'")
        cur.execute("""
            SELECT d.fecha, d.bodega, d.tipo, d.cantidad, d.costo_unitario,
                   d.valor, d.docs, d.confianza, d.sin_costo, d.unidad, d.nombre_prod
            FROM costos_diario d
            WHERE d.codigo_prod = %(c)s AND d.fecha BETWEEN %(d)s AND %(h)s""" + f + """
            ORDER BY d.fecha DESC, d.valor DESC LIMIT 300""", p)
        filas = [{'fecha': r[0].isoformat(), 'bodega': r[1], 'tipo': r[2],
                  'cantidad': float(r[3] or 0),
                  'costo': float(r[4]) if r[4] is not None else None,
                  'valor': float(r[5] or 0), 'docs': r[6], 'confianza': r[7],
                  'sin_costo': r[8], 'unidad': r[9], 'nombre': r[10]}
                 for r in cur.fetchall()]

        cur.execute("""SELECT desde, hasta, costo, fuente, confianza
                       FROM costo_vigente_producto
                       WHERE codigo_prod = %s ORDER BY desde""", (codigo,))
        tramos = [{'desde': r[0].isoformat(), 'hasta': r[1].isoformat(),
                   'costo': float(r[2]), 'fuente': r[3], 'confianza': r[4]}
                  for r in cur.fetchall()]
        cur.close()
        return jsonify({'ok': True, 'codigo': codigo, 'filas': filas, 'tramos': tramos,
                        'nombre': filas[0]['nombre'] if filas else codigo})
    except Exception as e:
        app.logger.exception('costos_panel_movimientos')
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if conn:
            fc_release_movimientos_db(conn)


@app.route('/api/costos/panel/refrescar', methods=['POST'])
def costos_panel_refrescar():
    """Rehace el costo vigente y revaloriza los ultimos dias.

    Hay que dispararlo despues de que la sincronizacion de movimientos
    termine. Sin esto el tablero se queda con la ultima foto: el resumen
    anterior llevaba parado desde el 26-ago porque nadie llamaba a su
    refresco.

    Los dias van hacia atras porque un dia no queda completo el mismo dia:
    la sincronizacion sigue trayendo movimientos suyos durante un par de
    dias mas.
    """
    conn = None
    try:
        cuerpo = request.get_json(silent=True) or {}
        dias = min(int(cuerpo.get('dias', 7)), 60)
        rehacer = bool(cuerpo.get('rehacer_costos', True))

        import costo_vigente
        import costos_diario as cd

        hasta = datetime.now(TZ_ECUADOR).date() + timedelta(days=1)
        desde = hasta - timedelta(days=dias)

        conn = fc_get_movimientos_db()
        conn.autocommit = True
        cur = conn.cursor()
        cur.execute("SET statement_timeout = '600s'")

        pasos = []
        if rehacer:
            # El costo vigente se rearma entero: son tramos encadenados, y
            # recalcular solo el ultimo trozo dejaria el enlace roto.
            tramos = costo_vigente.construir(
                cur, costo_vigente.DESDE_DEFECTO, log=lambda s: pasos.append(s.strip()))
        else:
            tramos = None

        filas = cd.refrescar(cur, desde, hasta, log=None)
        return jsonify({'ok': True, 'desde': desde.isoformat(),
                        'hasta': (hasta - timedelta(days=1)).isoformat(),
                        'tramos_costo': tramos, 'filas': filas, 'pasos': pasos})
    except Exception as e:
        app.logger.exception('costos_panel_refrescar')
        return jsonify({'ok': False, 'error': str(e)[:300]}), 500
    finally:
        if conn:
            try:
                conn.autocommit = False
            except Exception:
                pass
            fc_release_movimientos_db(conn)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
