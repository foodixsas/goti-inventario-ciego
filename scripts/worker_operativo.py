"""
Worker local del Cruce Operativo - PC FINANZAS
================================================
Loop infinito que cada 15 segundos consulta tareas pendientes en el backend
de Render. Cuando hay una, descarga saldos de Contifico via Selenium,
calcula el cruce contra la toma fisica usando equivalencias_conteo,
y sube el resultado al backend.

Ejecutar:
    pythonw cruce_worker.py        (background)
    python  cruce_worker.py        (consola visible para debug)

Tarea programada Windows para arrancar al boot.
"""
import os, sys, time, glob, json, re, traceback, subprocess, threading
from datetime import date, datetime, timedelta, timezone
from concurrent.futures import ThreadPoolExecutor, as_completed

# Forzar UTF-8 en stdout
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

import requests
import psycopg2
from psycopg2.extras import RealDictCursor
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service

# ============ CACHE CHROMEDRIVER (evita descargar cada vez) ============
_CACHED_DRIVER_PATH = None

def get_chromedriver_path():
    """Obtiene el path del chromedriver, cacheando para no descargar cada vez."""
    global _CACHED_DRIVER_PATH
    if _CACHED_DRIVER_PATH and os.path.exists(_CACHED_DRIVER_PATH):
        return _CACHED_DRIVER_PATH
    del_sistema = os.environ.get('CHROMEDRIVER', '/usr/bin/chromedriver')
    if os.path.exists(del_sistema):
        _CACHED_DRIVER_PATH = del_sistema
    else:
        # Import perezoso: webdriver_manager no se instala en la imagen Docker,
        # donde el chromedriver ya viene con chromium.
        from webdriver_manager.chrome import ChromeDriverManager
        _CACHED_DRIVER_PATH = ChromeDriverManager().install()
    return _CACHED_DRIVER_PATH

def matar_procesos_chrome():
    """Mata procesos chrome/chromedriver huerfanos (solo Windows).

    En el contenedor de Render no aplica: cada corrida arranca limpia y
    'taskkill' no existe en Linux.
    """
    if os.name != 'nt':
        return
    try:
        subprocess.run(['taskkill', '/F', '/IM', 'chromedriver.exe'],
                      capture_output=True, timeout=5)
    except Exception:
        pass
    try:
        # Solo matar chrome.exe si no hay ventanas visibles del usuario
        result = subprocess.run(['tasklist', '/FI', 'IMAGENAME eq chrome.exe'],
                               capture_output=True, text=True, timeout=5)
        if 'chrome.exe' in result.stdout:
            # Contar procesos - si hay muchos, probablemente son zombies
            count = result.stdout.count('chrome.exe')
            if count > 10:  # Muchos procesos = zombies
                subprocess.run(['taskkill', '/F', '/IM', 'chrome.exe'],
                              capture_output=True, timeout=10)
    except Exception:
        pass

# ============ CONFIGURACION ============
BACKEND_URL  = os.environ.get('CRUCE_BACKEND', 'https://inventario-ciego-5bdr.onrender.com')
WORKER_TOKEN = os.environ.get('CRUCE_WORKER_TOKEN', 'worker-foodix-2026-7K3xR9pL2qN8mZ4w')
WORKER_ID    = os.environ.get('CRUCE_WORKER_ID', 'pc-finanzas')
POLL_SEC     = int(os.environ.get('CRUCE_POLL_SEC', '15'))

# Procesamiento paralelo
MAX_PARALLEL_WORKERS = int(os.environ.get('CRUCE_MAX_PARALLEL', '2'))  # Max 2 Chrome simultaneos
_lock_log = threading.Lock()  # Para logs thread-safe

DB_CONFIG = {
    'host': os.environ.get('DB_HOST', ''),
    'port': int(os.environ.get('DB_PORT', '5432')),
    'dbname': os.environ.get('DB_NAME', 'InventariosLocales'),
    'user': os.environ.get('DB_USER', ''),
    'password': os.environ.get('DB_PASSWORD', ''),
    'sslmode': 'require',
}

CONTIFICO = {
    'login_url':   'https://base.contifico.com/sistema/accounts/login/',
    'reporte_url': 'https://1793168604001.contifico.com/sistema/reportes/saldos_inventario/',
    'usuario':     os.environ.get('CONTIFICO_WEB_USUARIO', ''),
    'password':    os.environ.get('CONTIFICO_WEB_PASSWORD', ''),
}

# Mapeo bodega_id (BD) -> nombre Contifico exacto + tabla toma fisica + marca para equivalencias
BODEGAS = {
    'bodega_principal': {
        'contifico': 'BODEGA PRINCIPAL',
        'tabla_toma': 'public.toma_bodega',
        'marca': 'BODEGA_PRINCIPAL',
    },
    'materia_prima': {
        'contifico': 'BODEGA MATERIA PRIMA',
        'tabla_toma': 'public.toma_materiaprima',
        'marca': 'MATERIA_PRIMA',
    },
    'planta': {
        'contifico': 'BODEGA PLANTA DE PRODUCCION',
        'tabla_toma': 'public.toma_planta',
        'marca': 'PLANTA',
    },
}

DOWNLOAD_DIR = os.environ.get('DOWNLOAD_DIR') or os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'descargas')
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# ============ BODEGAS LOCALES (ventas) ============
BODEGAS_LOCALES = {
    'real_audiencia': {'contifico': 'BODEGA CHIOS REAL', 'tipo': 'CHIOS'},
    'floreana': {'contifico': 'BODEGA CHIOS FLOREANA', 'tipo': 'CHIOS'},
    'portugal': {'contifico': 'BODEGA CHIOS PORTUGAL', 'tipo': 'CHIOS'},
    'santo_cachon_real': {'contifico': 'BODEGA SANTO CACHON REAL', 'tipo': 'CACHON'},
    'santo_cachon_portugal': {'contifico': 'BODEGA SANTO CACHON PORTUGAL', 'tipo': 'CACHON'},
    'simon_bolon': {'contifico': 'BODEGA SIMON BOLON', 'tipo': 'SIMON_BOLON'},
}

# ============ LOG ============
def log(msg, level='INFO', worker_id=None):
    with _lock_log:
        prefix = f"[W{worker_id}] " if worker_id else ""
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [{level}] {prefix}{msg}", flush=True)

# ============ HTTP HELPERS ============
def get_pendientes():
    try:
        r = requests.get(
            f'{BACKEND_URL}/api/cruce-op/pendientes',
            headers={'X-Worker-Token': WORKER_TOKEN},
            params={'worker_id': WORKER_ID},
            timeout=20,
        )
        if r.status_code == 200:
            return r.json()
        log(f'pendientes -> HTTP {r.status_code}: {r.text[:200]}', 'WARN')
    except Exception as e:
        log(f'pendientes -> EXCEPTION: {e}', 'WARN')
    return []

def post_resultado(payload):
    max_reintentos = 3
    for intento in range(max_reintentos):
        try:
            r = requests.post(
                f'{BACKEND_URL}/api/cruce-op/resultado',
                headers={'X-Worker-Token': WORKER_TOKEN, 'Content-Type': 'application/json'},
                data=json.dumps(payload, default=str),
                timeout=90,
            )
            if r.status_code == 200:
                return True
            if r.status_code in (502, 503, 504) and intento < max_reintentos - 1:
                log(f'resultado -> HTTP {r.status_code}, reintento {intento+2}/{max_reintentos}...', 'WARN')
                time.sleep(5 * (intento + 1))
                continue
            log(f'resultado -> HTTP {r.status_code}: {r.text[:200]}', 'ERROR')
        except Exception as e:
            if intento < max_reintentos - 1:
                log(f'resultado -> EXCEPTION {e}, reintento {intento+2}/{max_reintentos}...', 'WARN')
                time.sleep(5 * (intento + 1))
                continue
            log(f'resultado -> EXCEPTION: {e}', 'ERROR')
    return False

def get_pendientes_carga():
    """Poll para tareas de carga a Contifico."""
    try:
        r = requests.get(
            f'{BACKEND_URL}/api/carga-contifico/pendientes',
            headers={'X-Worker-Token': WORKER_TOKEN},
            params={'worker_id': WORKER_ID},
            timeout=20,
        )
        if r.status_code == 200:
            return r.json()
        log(f'pendientes_carga -> HTTP {r.status_code}: {r.text[:200]}', 'WARN')
    except Exception as e:
        log(f'pendientes_carga -> EXCEPTION: {e}', 'WARN')
    return []

def post_resultado_carga(payload):
    """Envia resultado de carga a Contifico al backend."""
    try:
        r = requests.post(
            f'{BACKEND_URL}/api/carga-contifico/resultado',
            headers={'X-Worker-Token': WORKER_TOKEN, 'Content-Type': 'application/json'},
            data=json.dumps(payload, default=str),
            timeout=60,
        )
        if r.status_code == 200:
            return True
        log(f'resultado_carga -> HTTP {r.status_code}: {r.text[:200]}', 'ERROR')
    except Exception as e:
        log(f'resultado_carga -> EXCEPTION: {e}', 'ERROR')
    return False

# ============ BD HELPERS ============
def db_query(sql, params=None):
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(sql, params or ())
        return cur.fetchall()
    finally:
        conn.close()

def cargar_equivalencias(bodega):
    """Lee equivalencias desde productos_por_marca (módulo Config Productos del sistema)."""
    # Determinar la marca según el tipo de bodega
    if bodega in BODEGAS:
        marca = BODEGAS[bodega].get('marca', bodega.upper())
    elif bodega in BODEGAS_LOCALES:
        marca = BODEGAS_LOCALES[bodega].get('tipo', 'CHIOS')
    else:
        marca = bodega.upper()

    rows = db_query(
        """SELECT codigo, nombre as producto, unidad as unidad_toma,
                  COALESCE(equivalencia, 1) as factor, 'Unidad' as unidad_destino
           FROM goti.productos_por_marca
           WHERE marca = %s AND activo = TRUE""",
        (marca,)
    )
    log(f'    Equivalencias cargadas de productos_por_marca (marca={marca}): {len(rows)}')
    return {r['codigo']: r for r in rows}

def cargar_toma_fisica(tabla, fecha):
    rows = db_query(
        f"SELECT codigo, producto, total, unidad, categoria, \"Tipo A,B o C\" AS tipo_abc "
        f"FROM {tabla} WHERE fecha = %s",
        (fecha,)
    )
    return {r['codigo']: r for r in rows}

# ============ SELENIUM CONTIFICO ============
# Configuracion de timeouts (en segundos) - REDUCIDOS para evitar cuelgues
TIMEOUT_PAGE_LOAD = 90   # antes: 300
TIMEOUT_SCRIPT = 180     # aumentado para exportarExcel()
TIMEOUT_DESCARGA = 120   # antes: 300
TIMEOUT_WAIT = 45        # antes: 120

# Timeout de la conexion HTTP entre Selenium y chromedriver. urllib3 lo pone en
# 120s y ese corte no lo cambia ningun set_*_timeout: la tarea 249 murio con
# "HTTPConnectionPool ... Read timed out (read timeout=120)" mientras Contifico
# seguia generando el Excel. Tiene que quedar POR ENCIMA de TIMEOUT_SCRIPT para
# que el que corte sea Selenium, que da un error legible, y no la conexion.
TIMEOUT_CONEXION = int(os.environ.get('TIMEOUT_CONEXION', '300'))

# Modo headless (sin ventana) - mas rapido y menos recursos
USE_HEADLESS = os.environ.get('CRUCE_HEADLESS', '1') == '1'

def make_chrome(parallel_id=None):
    """Crea instancia de Chrome. Si parallel_id se especifica, usa carpeta de descarga separada."""
    # Matar procesos zombie solo si no es paralelo (evita matar otros workers)
    if parallel_id is None:
        matar_procesos_chrome()

    # Carpeta de descarga: separada por worker si es paralelo
    if parallel_id is not None:
        download_dir = os.path.join(DOWNLOAD_DIR, f'worker_{parallel_id}')
        os.makedirs(download_dir, exist_ok=True)
    else:
        download_dir = DOWNLOAD_DIR

    opts = webdriver.ChromeOptions()
    opts.add_experimental_option('prefs', {
        'download.default_directory': download_dir,
        'download.prompt_for_download': False,
        'profile.default_content_settings.popups': 0,
        'download.directory_upgrade': True,
        'safebrowsing.enabled': True,
    })

    # Opciones para estabilidad
    opts.add_argument('--no-sandbox')
    opts.add_argument('--disable-dev-shm-usage')
    opts.add_argument('--disable-gpu')
    opts.add_argument('--disable-extensions')
    opts.add_argument('--disable-infobars')
    opts.add_argument('--window-size=1400,900')

    # Headless mode (configurable via env var CRUCE_HEADLESS=0 para desactivar)
    if USE_HEADLESS:
        opts.add_argument('--headless=new')
        log(f'Chrome en modo HEADLESS (sin ventana)', worker_id=parallel_id)

    # En la imagen Docker el binario es chromium, no chrome
    binario = os.environ.get('CHROME_BIN')
    if binario:
        opts.binary_location = binario

    # Usar chromedriver cacheado
    driver_path = get_chromedriver_path()
    driver = webdriver.Chrome(service=Service(driver_path), options=opts)

    # Timeouts reducidos
    driver.set_page_load_timeout(TIMEOUT_PAGE_LOAD)
    driver.set_script_timeout(TIMEOUT_SCRIPT)
    # El corte de 120s que mato la tarea 249 vive aqui, en el cliente HTTP que
    # habla con chromedriver, y no lo toca ningun set_*_timeout. En selenium
    # 4.40 hay que escribirlo en la instancia: RemoteConnection.set_timeout()
    # lanza AttributeError y se queda en nada.
    try:
        driver.command_executor.client_config.timeout = TIMEOUT_CONEXION
    except Exception as e:
        log(f'no se pudo fijar el timeout de conexion: {str(e)[:80]}', 'WARN')
    driver.implicitly_wait(5)

    # Guardar la carpeta de descarga en el driver para uso posterior
    driver._download_dir = download_dir

    return driver

def login_contifico(driver):
    driver.get(CONTIFICO['login_url'])
    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, 'username'))).send_keys(CONTIFICO['usuario'])
    driver.find_element(By.NAME, 'password').send_keys(CONTIFICO['password'] + Keys.RETURN)
    # Esperar a que cargue el dashboard (antes: sleep(4))
    try:
        WebDriverWait(driver, 15).until(
            lambda d: 'login' not in d.current_url.lower() or len(d.find_elements(By.CSS_SELECTOR, '.dashboard, .main-content, #content')) > 0
        )
    except Exception:
        time.sleep(2)  # Fallback minimo

def _set_fecha_contifico(driver, fecha_dmY):
    """Setea la fecha disparando eventos para que el framework la reconozca."""
    driver.execute_script("""
        var el = document.getElementById('id_fecha_corte');
        if (!el) return;
        var nativeSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        nativeSetter.call(el, arguments[0]);
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
    """, fecha_dmY)

def _esperar_archivo_xls(antes, download_dir, timeout=None):
    """Espera hasta `timeout` seg a que aparezca un .xls* nuevo en download_dir."""
    if timeout is None:
        timeout = TIMEOUT_DESCARGA
    log(f'    Esperando descarga (max {timeout}s)...')
    for i in range(timeout):
        time.sleep(1)
        todos = set(glob.glob(os.path.join(download_dir, '*.xls*')))
        nuevos = todos - antes
        finales = [f for f in nuevos if not f.endswith('.crdownload')]
        if finales:
            log(f'    Archivo descargado: {os.path.basename(finales[0])}')
            return finales[0]
        # Log progreso cada 15 segundos
        if (i + 1) % 15 == 0:
            en_progreso = [f for f in nuevos if f.endswith('.crdownload')]
            log(f'    Esperando... {i+1}/{timeout}s | en_progreso={len(en_progreso)} | total_xls={len(todos)}')
    log(f'    TIMEOUT: No se descargo archivo en {timeout}s', 'ERROR')
    return None

def descargar_saldos(driver, nombre_bodega_contifico, fecha_iso):
    """Descarga el Excel de saldos para una bodega y una fecha (YYYY-MM-DD).
    Devuelve la ruta al archivo Excel descargado."""
    # Usar carpeta de descarga del driver (puede ser específica por worker)
    download_dir = getattr(driver, '_download_dir', DOWNLOAD_DIR)
    fecha_dmY = datetime.strptime(fecha_iso, '%Y-%m-%d').strftime('%d/%m/%Y')

    # exportarExcel() tarda y dispara una navegacion. El corte a los 90s venia
    # del page_load_timeout, NO del de script: hay que subir LOS DOS. Va aqui
    # dentro y no en cada llamador, porque asi lo heredan los tres caminos
    # (cruce, conteo operativo y toma fisica). Antes solo estaba parcheado en la
    # funcion de verificacion y el conteo real seguia cayendose.
    driver.set_page_load_timeout(TIMEOUT_SCRIPT)
    driver.set_script_timeout(TIMEOUT_SCRIPT)
    antes = set(glob.glob(os.path.join(download_dir, '*.xls*')))

    driver.get(CONTIFICO['reporte_url'])
    campo = WebDriverWait(driver, TIMEOUT_WAIT).until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, 'input.object-description[data_id="id_bodega_id"]')))

    # Fecha — con dispatchEvent para que el framework la reconozca
    _set_fecha_contifico(driver, fecha_dmY)

    # Seleccionar bodega (sleeps reducidos)
    campo.click()
    campo.clear()
    campo.send_keys(nombre_bodega_contifico)
    time.sleep(1.5)  # antes: 3
    campo.send_keys(Keys.DOWN)
    campo.send_keys(Keys.ENTER)
    time.sleep(1)  # antes: 2

    # Re-set fecha (a veces se resetea al elegir bodega)
    _set_fecha_contifico(driver, fecha_dmY)

    # Generar reporte via form submit (el boton btn-primary es invisible en Contifico)
    try:
        form = driver.find_element(By.TAG_NAME, 'form')
        driver.execute_script('arguments[0].submit();', form)
        log('    Form de reporte enviado')
    except Exception as e:
        log(f'    WARN: no se pudo enviar form: {e}', 'WARN')

    # Esperar a que la tabla cargue filas (timeout reducido)
    try:
        WebDriverWait(driver, TIMEOUT_WAIT).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR, 'table tbody tr')) > 0)
        time.sleep(1)  # antes: 2
    except Exception:
        log(f'    WARN: tabla sin filas tras {TIMEOUT_WAIT}s, intentando exportar', 'WARN')
        time.sleep(1)

    n_filas = len(driver.find_elements(By.CSS_SELECTOR, 'table tbody tr'))
    log(f'    Tabla tiene {n_filas} filas')

    # Exportar Excel
    log('    Llamando exportarExcel()...')
    log(f'    Carpeta descargas: {download_dir}')
    log(f'    Archivos antes: {len(antes)}')
    try:
        driver.execute_script('exportarExcel();')
        log('    exportarExcel() ejecutado, esperando archivo...')
    except Exception as e:
        log(f'    ERROR en exportarExcel(): {e}', 'ERROR')
        raise
    archivo = _esperar_archivo_xls(antes, download_dir)
    if archivo:
        return archivo

    # No se logro descargar — guardar screenshot para diagnostico
    ss = os.path.join(download_dir, f'debug_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png')
    try:
        driver.save_screenshot(ss)
        log(f'    Screenshot de fallo guardado: {ss}', 'ERROR')
    except Exception:
        pass
    raise TimeoutError('No se descargo el archivo Excel de Contifico')

def parsear_saldos(ruta_excel, nombre_bodega_esperado):
    """Lee el Excel descargado y devuelve dict {codigo: {nombre, unidad, stock, costo}}"""
    # Validar header bodega
    df_h = pd.read_excel(ruta_excel, header=None, nrows=5)
    bod = ''
    for _, row in df_h.iterrows():
        for v in row.values:
            s = str(v).strip()
            if s.lower().startswith('bodega:'):
                bod = s.split(':', 1)[1].strip()
    if bod and bod.upper() != nombre_bodega_esperado.upper():
        raise ValueError(f'Excel de bodega incorrecta: esperado "{nombre_bodega_esperado}", descargo "{bod}"')

    df = pd.read_excel(ruta_excel, header=5)
    df.columns = [str(c).strip() for c in df.columns]
    log(f'    Columnas Excel: {list(df.columns)}')

    def norm(s):
        # quita acentos y pasa a minusculas
        import unicodedata
        return ''.join(c for c in unicodedata.normalize('NFD', str(s))
                       if unicodedata.category(c) != 'Mn').lower().strip()

    col_cod = next((c for c in df.columns
                    if norm(c).startswith('cod') and 'cat' not in norm(c) and 'barr' not in norm(c)), None)
    col_uni = next((c for c in df.columns if 'unid' in norm(c)), None)
    col_stock = next((c for c in df.columns if norm(c) == 'stock'), None)
    col_costo = next((c for c in df.columns if 'costo' in norm(c)), None)
    col_nom = next((c for c in df.columns if norm(c) == 'nombre'), None)

    if not col_cod or not col_uni or not col_stock:
        raise ValueError(f'No se encontraron columnas requeridas en el Excel. '
                         f'cod={col_cod} uni={col_uni} stock={col_stock}. '
                         f'Columnas disponibles: {list(df.columns)}')

    out = {}
    for _, r in df.iterrows():
        cod = str(r[col_cod]).strip()
        if not cod or cod.lower() == 'nan':
            continue
        try:
            stock = float(r[col_stock]) if pd.notna(r[col_stock]) else 0.0
        except Exception:
            stock = 0.0
        try:
            costo = float(r[col_costo]) if (col_costo and pd.notna(r[col_costo])) else 0.0
        except Exception:
            costo = 0.0
        out[cod] = {
            'nombre': str(r[col_nom]) if col_nom else '',
            'unidad': str(r[col_uni]).strip(),
            'stock': stock,
            'costo': costo,
        }
    return out

# ============ CRUCE ============
def calcular_cruce(toma, equivs, contifico):
    """Cruza toma fisica vs contifico aplicando equivalencias."""
    detalle = []
    sin_equivalencia = []
    stock_negativo = []
    for cod, t in toma.items():
        eq = equivs.get(cod)
        cont = contifico.get(cod)
        # Si no aparece en Contifico = stock 0 (Contifico oculta los stock=0)
        stock_c = float(cont['stock']) if cont else 0.0
        costo = float(cont['costo']) if cont else 0.0
        unidad_c = cont['unidad'] if cont else None
        nombre = (cont['nombre'] if cont else None) or t['producto']

        if eq:
            factor = float(eq['factor'])
            unidad_toma = eq['unidad_toma']
            unidad_destino = eq['unidad_destino']
            cant_conv = float(t['total'] or 0) * factor
        else:
            # No hay equivalencia cargada: la cantidad va tal cual. Se anota
            # para avisar, porque la diferencia que salga aqui no es real.
            factor = None
            unidad_toma = (t['unidad'] or '').strip()
            unidad_destino = unidad_c or unidad_toma
            cant_conv = float(t['total'] or 0)
            sin_equivalencia.append((cod, nombre))

        # Un saldo negativo no existe fisicamente: o falta registrar ingresos o
        # hay movimientos mal cargados. Ademas envenena el cruce, porque la
        # diferencia contra un negativo suma en vez de restar.
        if stock_c < 0:
            stock_negativo.append((cod, nombre, stock_c, unidad_c, stock_c * costo))

        diferencia = cant_conv - stock_c
        valor_dif = diferencia * costo

        detalle.append({
            'codigo': cod,
            'nombre': nombre,
            'categoria': t.get('categoria'),
            'unidad_toma': unidad_toma,
            'factor': factor,
            'unidad_destino': unidad_destino,
            'cantidad_toma': cant_conv,
            'cantidad_sistema': stock_c,
            'diferencia': diferencia,
            'costo_unitario': costo,
            'valor_diferencia': valor_dif,
            'tipo_abc': t.get('tipo_abc'),
            'origen': 'cruce_operativo',
        })

    # Resumen
    cruzados = len(detalle)
    con_dif = sum(1 for d in detalle if abs(d['diferencia']) >= 0.01)
    valor_total = sum(d['valor_diferencia'] or 0 for d in detalle if abs(d['diferencia']) >= 0.01)
    resumen = {
        'total_productos_toma': len(toma),
        'total_productos_contifico': len(contifico),
        'total_cruzados': cruzados,
        'total_con_diferencia': con_dif,
        'valor_total_dif': round(valor_total, 2),
        'sin_equivalencia': sin_equivalencia,
        # Por valor y no por cantidad: las unidades se mezclan (lo que va en
        # gramos siempre daria el numero mas grande) y lo que interesa
        # primero es el negativo que mas plata representa.
        'stock_negativo': sorted(stock_negativo, key=lambda x: x[4]),
    }
    return detalle, resumen

# ============ PROCESAR UNA TAREA ============
# Retorna (driver_sigue_vivo_bool, ok_bool)
def avisar_sin_equivalencia(proceso, bodega, faltantes):
    """Avisa por Telegram de los productos que salieron sin equivalencia.

    Sin este aviso el hueco es invisible. La conversion no falla con un error:
    se saltea, la cantidad viaja en la unidad de la toma y la fila cuadra
    contra un Contifico que habla en otra unidad. El descuadre que aparece se
    lee como faltante de mercaderia y no lo es.

    Paso con SORBETES EXTRAGRUESOS el 31-ago-2026: 7 paquetes contra 700
    unidades en Contifico, -34,28 USD que no existian. El producto estaba en la
    toma fisica pero nadie lo habia dado de alta en Config Productos. Un
    producto nuevo tiene que gritar, no pasar de largo.
    """
    if not faltantes:
        return
    nombre = BODEGA_TELEGRAM.get(bodega, bodega.upper())
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from notificar_telegram import notificar_error

        lineas = ['{} - {}'.format(c, n) for c, n in faltantes[:15]]
        if len(faltantes) > 15:
            lineas.append('... y {} mas'.format(len(faltantes) - 15))
        detalle = '{}: {} producto(s) sin equivalencia{}{}'.format(
            proceso, len(faltantes), chr(10), chr(10).join(lineas))
        notificar_error(
            'Equivalencias', nombre, detalle,
            'La cantidad va SIN convertir: la diferencia que salga en estos '
            'productos no es real. Cargar el factor en Config Productos y '
            'volver a correr.')
        log('  aviso enviado: {} producto(s) sin equivalencia'.format(len(faltantes)), 'WARN')
    except Exception as e:
        log('  aviso de equivalencias fallo: {}'.format(str(e)[:120]), 'WARN')


def avisar_toma_incompleta(bodega, fecha_toma, cargados, sin_contar, total):
    """Avisa que parte de la toma llego en blanco y subio en cero.

    Sube toda la toma, la misma lista que lee el cruce, y lo que nadie conto va
    en cero: si no se conto, el inventario de ese producto queda en cero. El
    aviso no es para frenar nada, es para que quede escrito cuales fueron.
    Poner en cero el saldo de 193 productos es un hecho grande y no puede pasar
    sin registro.

    La fila existe en la toma, con su producto y su usuario; lo que no tiene es
    nada tecleado en 'cantidades'.
    """
    if not sin_contar:
        return
    nombre = BODEGA_TELEGRAM.get(bodega, bodega.upper())
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from notificar_telegram import destinatarios_para, enviar_mensaje

        try:
            f = datetime.strptime(str(fecha_toma)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            f = str(fecha_toma)

        pct = (100.0 * cargados / total) if total else 0
        lineas = ['• <b>{}</b> {}'.format(c, (p or '')[:40]) for c, p in sin_contar[:15]]
        if len(sin_contar) > 15:
            lineas.append('... y {} mas'.format(len(sin_contar) - 15))

        mensaje = (
            '⚠️ <b>TOMA FISICA CON PRODUCTOS EN CERO</b>\n'
            '📍 {}\n'
            'Toma del {}\n\n'
            'Se cargaron <b>{}</b> de <b>{}</b> productos ({:.0f}%).\n'
            '<b>{}</b> llegaron sin contar y se subieron <b>EN CERO</b>:\n\n'
            '{}\n\n'
            'El saldo de estos productos en Contifico queda en cero. Si alguno '
            'tenia existencia y no se conto, hay que volver a contarlo y cargar '
            'la toma de nuevo.'
        ).format(nombre, f, cargados, total, pct, len(sin_contar), chr(10).join(lineas))

        enviados = destinatarios_para(nombre, 'Toma fisica', 'error')
        for chat_id in enviados:
            enviar_mensaje(chat_id, mensaje)
        log('  aviso enviado: {} de {} productos en cero -> {} chats'.format(
            len(sin_contar), total, len(enviados)), 'WARN')
    except Exception as e:
        log('  aviso de productos en cero fallo: {}'.format(str(e)[:120]), 'WARN')


def avisar_stock_negativo(bodega, fecha_toma, negativos):
    """Avisa, marcado como importante, de los productos con saldo negativo.

    Un stock negativo no existe: nadie tiene menos veinte kilos en una percha.
    Significa que Contifico descargo consumo que nunca se ingreso, o que hay
    movimientos cargados contra el producto equivocado. Va aparte del aviso
    normal del cruce y con otro encabezado porque no es el resultado del dia:
    es un dato de Contifico que hay que ir a corregir a mano, y hasta que se
    corrija la diferencia de ese producto no significa nada.
    """
    if not negativos:
        return
    nombre = BODEGA_TELEGRAM.get(bodega, bodega.upper())
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from notificar_telegram import destinatarios_para, enviar_mensaje

        try:
            f = datetime.strptime(str(fecha_toma)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            f = str(fecha_toma)

        lineas = []
        for cod, prod, stock, unidad, valor in negativos[:20]:
            lineas.append('• <b>{}</b> {}: {:,.2f} {} ({:+,.2f} USD)'.format(
                cod, (prod or '')[:38], stock, (unidad or '').strip(), valor))
        if len(negativos) > 20:
            lineas.append('... y {} mas'.format(len(negativos) - 20))

        total = sum(n[4] for n in negativos)
        mensaje = (
            '‼️ <b>IMPORTANTE — STOCK NEGATIVO EN CONTIFICO</b>\n'
            '📍 {}\n'
            'Cruce de la toma del {}\n\n'
            '<b>{}</b> producto(s) con saldo negativo, {:+,.2f} USD en total:\n\n'
            '{}\n\n'
            '⚠️ Un saldo negativo no existe fisicamente: falta registrar '
            'ingresos o hay movimientos cargados contra el producto '
            'equivocado. Mientras siga negativo, la diferencia del cruce en '
            'esos productos no significa nada.'
        ).format(nombre, f, len(negativos), total, chr(10).join(lineas))

        enviados = destinatarios_para(nombre, 'Stock negativo', 'error')
        for chat_id in enviados:
            enviar_mensaje(chat_id, mensaje)
        log('  aviso enviado: {} producto(s) con stock negativo -> {} chats'.format(
            len(negativos), len(enviados)), 'WARN')
    except Exception as e:
        log('  aviso de stock negativo fallo: {}'.format(str(e)[:120]), 'WARN')


def avisar_cruce_operativo(bodega, fecha_toma, resumen, error=None):
    """Avisa por Telegram del resultado del cruce operativo.

    Con su propio try, igual que el aviso de la toma fisica: que falle el
    mensaje no puede dar por fallido un cruce que si se calculo y ya quedo
    guardado.

    El importe se manda con su signo y con su lectura en palabras: un numero
    suelto no dice si sobra o falta mercaderia, y esa es justo la pregunta que
    se hace quien lo recibe.
    """
    nombre = BODEGA_TELEGRAM.get(bodega, bodega.upper())
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from notificar_telegram import notificar_exito, notificar_error

        try:
            f = datetime.strptime(str(fecha_toma)[:10], '%Y-%m-%d').strftime('%d/%m/%Y')
        except Exception:
            f = str(fecha_toma)

        if error:
            notificar_error('Cruce Operativo', nombre,
                            'Toma del {}'.format(f), error[:200])
            return

        valor = float(resumen.get('valor_total_dif') or 0)
        if valor > 0:
            lectura = 'sobra mercaderia frente a Contifico'
        elif valor < 0:
            lectura = 'falta mercaderia frente a Contifico'
        else:
            lectura = 'sin diferencia de valor'

        partes = [
            'Toma del {}'.format(f),
            '{} productos cruzados'.format(resumen.get('total_cruzados') or 0),
            '{} con diferencia'.format(resumen.get('total_con_diferencia') or 0),
            'Valor de la diferencia: {:+,.2f} USD ({})'.format(valor, lectura),
        ]
        notificar_exito('Cruce Operativo', nombre, chr(10).join(partes), '')
    except Exception as e:
        log('  aviso por Telegram fallo: {}'.format(str(e)[:120]), 'WARN')


def procesar_tarea(tarea, driver):
    ejec_id = tarea['id']
    bodega = tarea['bodega']
    fecha_toma = tarea['fecha_toma']
    fecha_corte = tarea.get('fecha_corte_contifico') or fecha_toma
    log(f'>>> Procesando ejec_id={ejec_id} bodega={bodega} toma={fecha_toma} corte_contifico={fecha_corte}')

    cfg = BODEGAS.get(bodega)
    if not cfg:
        post_resultado({'id': ejec_id, 'estado': 'error', 'error_msg': f'bodega desconocida: {bodega}'})
        return True, False

    driver_ok = True
    try:
        log('  - Cargando equivalencias y toma fisica desde BD...')
        equivs = cargar_equivalencias(bodega)
        toma = cargar_toma_fisica(cfg['tabla_toma'], fecha_toma)
        log(f'    equivalencias={len(equivs)}  toma={len(toma)}')

        if not toma:
            post_resultado({'id': ejec_id, 'estado': 'error',
                            'error_msg': f'No hay toma fisica para {bodega} en {fecha_toma}'})
            return True, False

        log(f'  - Descargando saldos de Contifico al corte {fecha_corte}...')
        archivo = descargar_saldos(driver, cfg['contifico'], fecha_corte)
        log(f'    archivo: {os.path.basename(archivo)}')

        contifico = parsear_saldos(archivo, cfg['contifico'])
        log(f'    productos en Contifico: {len(contifico)}')

        log('  - Calculando cruce...')
        detalle, resumen = calcular_cruce(toma, equivs, contifico)
        log(f'    cruzados={resumen["total_cruzados"]}  con_dif={resumen["total_con_diferencia"]}  valor_total={resumen["valor_total_dif"]}')
        faltantes = resumen.get('sin_equivalencia') or []
        if faltantes:
            log(f'    SIN EQUIVALENCIA: {len(faltantes)} -> '
                + ', '.join(c for c, _ in faltantes[:10]), 'WARN')
            avisar_sin_equivalencia('Cruce operativo', bodega, faltantes)

        negativos = resumen.get('stock_negativo') or []
        if negativos:
            log(f'    STOCK NEGATIVO: {len(negativos)} -> '
                + ', '.join(n[0] for n in negativos[:10]), 'WARN')
            avisar_stock_negativo(bodega, fecha_toma, negativos)

        log('  - Subiendo resultado al backend...')
        ok = post_resultado({
            'id': ejec_id,
            'estado': 'completado',
            'detalle': detalle,
            'resumen': resumen,
        })
        log(f'    upload {"OK" if ok else "FALLO"}')
        avisar_cruce_operativo(bodega, fecha_toma, resumen)
        return True, ok
    except Exception as e:
        tb = traceback.format_exc()
        log(f'ERROR procesando {ejec_id}: {e}\n{tb}', 'ERROR')
        post_resultado({'id': ejec_id, 'estado': 'error', 'error_msg': str(e)[:500]})
        avisar_cruce_operativo(bodega, fecha_toma, {}, error=str(e))
        msg = str(e).lower()
        # Si el error es del driver/sesion, marcamos driver como muerto para reiniciar
        if any(k in msg for k in ('invalid session', 'disconnected', 'not connected',
                                   'no such window', 'connection refused', 'target closed',
                                   'chrome not reachable')):
            driver_ok = False
        return driver_ok, False

# ============ CARGA TOMA FISICA A CONTIFICO ============
CONTIFICO_TOMA_URL = 'https://1793168604001.contifico.com/sistema/inventario/tomafisica/registrar/'

BODEGAS_CARGA = {
    'bodega_principal': 'BODEGA PRINCIPAL',
    'materia_prima':    'BODEGA MATERIA PRIMA',
    'planta':           'BODEGA PLANTA DE PRODUCCION',
}

def cargar_toma_para_contifico(tabla, fecha, equivs):
    """Lee la toma fisica y aplica equivalencias para obtener cantidades en unidad Contifico.

    Entra todo lo que se conto, incluido lo contado en CERO: si alguien miro el
    estante y escribio 0, Contifico tiene que bajar ese producto a cero. Antes
    el filtro era 'total > 0' y esos ceros se perdian -unos 36 por semana solo
    en Materia Prima-, con el agravante de que el cruce si los lee: marcaba la
    diferencia y la carga no la corregia nunca.

    No entra lo que nadie conto. Se distinguen por el texto que tecleo la
    persona, guardado en 'cantidades': '+0' lleva un digito y es un conteo; '+'
    y la cadena vacia no. La diferencia importa mucho: una toma fisica FIJA el
    saldo, asi que subir en cero un producto que nadie miro seria declararlo
    agotado. En Planta serian 197 productos de golpe.
    """
    # Entra TODA la toma, la misma lista que lee el cruce. Lo que nadie conto
    # sube en cero.
    #
    # Antes se filtraba y solo subia lo contado, para no declarar agotado lo que
    # nadie miro. En Planta eso dejaba fuera 193 de 224 productos y partia el
    # inventario en dos: el cruce medido contra la toma completa y Contifico
    # ajustado solo en una fraccion. Medido sobre la toma del 31-ago, de los 177
    # que quedaban afuera 130 ya estaban en cero, 17 estaban en NEGATIVO -y
    # subirlos en cero los corrige- y 30 tenian saldo, 2.157,49 USD.
    #
    # CARGA_SOLO_CONTADOS=1 vuelve al camino viejo si hiciera falta.
    solo_contados = os.environ.get('CARGA_SOLO_CONTADOS', '0') == '1'
    rows = db_query(
        f"SELECT codigo, producto, total, unidad, cantidades FROM {tabla} "
        f"WHERE fecha = %s",
        (fecha,)
    )
    productos = []
    sin_equivalencia = []
    sin_contar = []
    for r in rows:
        cod = r['codigo']
        total = float(r['total'] or 0)
        if not (total > 0 or (total == 0 and re.search(r'[0-9]', r['cantidades'] or ''))):
            # Se anota igual: subir 193 productos en cero es un hecho que tiene
            # que avisarse, aunque ahora si se suban.
            sin_contar.append((cod, r['producto']))
            if solo_contados:
                continue
        eq = equivs.get(cod)
        if eq:
            cantidad = total * float(eq['factor'])
        else:
            # Aqui el hueco es mas caro que en el cruce: esta cantidad se
            # ESCRIBE en Contifico y una toma fisica fija el saldo. Subir
            # 7 donde van 700 declara un faltante que nadie tuvo.
            cantidad = total
            sin_equivalencia.append((cod, r['producto']))
        productos.append({
            'codigo': cod,
            'nombre': r['producto'],
            'cantidad': cantidad,
        })
    # len(rows) aparte: con el modo por defecto coincide con len(productos),
    # pero con CARGA_SOLO_CONTADOS=1 no, y el panel necesita el total real.
    return productos, sin_equivalencia, sin_contar, len(rows)


def cerrar_modales(driver):
    """Cierra cualquier modal bootbox o dialog que este bloqueando la pagina."""
    try:
        modales = driver.find_elements(By.CSS_SELECTOR, '.bootbox.modal.in, .modal.in')
        for modal in modales:
            if modal.is_displayed():
                btns_close = modal.find_elements(By.CSS_SELECTOR, '.close, .bootbox-close-button')
                for btn in btns_close:
                    try:
                        btn.click()
                        time.sleep(0.5)
                        return True
                    except:
                        pass
                driver.execute_script("$('.bootbox.modal').modal('hide'); $('.modal-backdrop').remove();")
                time.sleep(0.5)
                return True
    except:
        pass
    try:
        driver.execute_script("""
            document.querySelectorAll('.bootbox.modal, .modal-backdrop').forEach(function(el){
                el.style.display='none'; el.classList.remove('in');
            });
            document.body.classList.remove('modal-open');
            document.body.style.paddingRight='';
        """)
    except:
        pass
    return False


def js_set_value_and_trigger(driver, element_id, value):
    """Establece un valor via JavaScript y dispara eventos de cambio."""
    driver.execute_script(f"""
        var el = document.getElementById('{element_id}');
        if (el) {{
            el.value = '{value}';
            el.dispatchEvent(new Event('input', {{bubbles: true}}));
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
            el.dispatchEvent(new Event('keyup', {{bubbles: true}}));
            if (typeof calcularSubtotal === 'function') {{ calcularSubtotal(el); }}
        }}
    """)


def registrar_toma_por_archivo(driver, bodega_contifico, productos, fecha_form):
    """Sube la toma fisica como Excel por la pestana 'Carga Masiva'.

    Mismo formulario que la carga fila a fila, misma URL: lo unico que cambia
    es que en vez de teclear cada producto se adjunta un archivo. La diferencia
    de tiempo es de una hora a dos minutos.

    Devuelve (productos_ok, errores) para poder sustituir a registrar_toma_bodega
    sin tocar a quien la llama. Lanza excepcion si Contifico no devuelve el
    numero de documento o si el movimiento no queda GENERADO: sin Generar el
    inventario no se ajusta, y dar por buena una toma que no ajusto seria peor
    que fallar.
    """
    import openpyxl
    wait = WebDriverWait(driver, 60)

    # 1. Excel con el formato de la plantilla oficial
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    seguro = re.sub(r'[^A-Za-z0-9]+', '_', bodega_contifico).strip('_').lower()
    ruta = os.path.join(DOWNLOAD_DIR,
                        'tf_%s_%s.xlsx' % (seguro, fecha_form.replace('/', '-')))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Producto (Código)', 'Cantidad'])
    for p in productos:
        cant = float(p['cantidad'])
        # El codigo va como TEXTO: la plantilla lo exige.
        ws.cell(row=ws.max_row + 1, column=1, value=str(p['codigo'])).number_format = '@'
        ws.cell(row=ws.max_row, column=2,
                value=int(cant) if cant == int(cant) else cant)
    wb.save(ruta)
    log(f'    Excel generado: {os.path.basename(ruta)} ({len(productos)} productos)')

    # 2. Cabecera
    log(f'    Navegando a formulario de toma fisica...')
    driver.get(CONTIFICO_TOMA_URL)
    wait.until(EC.presence_of_element_located((By.ID, 'id_fecha')))
    time.sleep(2)

    log(f'    Configurando fecha: {fecha_form}')
    campo_fecha = driver.find_element(By.ID, 'id_fecha')
    driver.execute_script("arguments[0].value = '';", campo_fecha)
    campo_fecha.click()
    campo_fecha.clear()
    campo_fecha.send_keys(fecha_form)
    campo_fecha.send_keys(Keys.ESCAPE)
    time.sleep(1)

    log(f'    Seleccionando bodega: {bodega_contifico}')
    campo_bodega = driver.find_element(
        By.CSS_SELECTOR, 'input.object-description[data_id="id_bodega"]')
    campo_bodega.click()
    campo_bodega.clear()
    campo_bodega.send_keys(bodega_contifico)
    time.sleep(2)
    campo_bodega.send_keys(Keys.DOWN)
    campo_bodega.send_keys(Keys.ENTER)
    time.sleep(2)

    if not (campo_bodega.get_attribute('value') or '').strip():
        raise Exception('Bodega NO seleccionada: el autocompletado dejo el campo vacio')

    try:
        campo_desc = driver.find_element(By.ID, 'id_descripcion')
        driver.execute_script("arguments[0].focus();", campo_desc)
        campo_desc.clear()
        campo_desc.send_keys(f'AJUSTE DE INVENTARIO {bodega_contifico} - {fecha_form}')
    except Exception:
        pass      # la descripcion no es obligatoria

    # 3. Adjuntar y registrar
    inp = driver.find_element(By.ID, 'id_archivo')
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    inp.send_keys(ruta)
    time.sleep(2)
    log('    Archivo adjuntado, guardando...')
    driver.execute_script("registrarMovimiento();")
    time.sleep(9)

    # 4. Sin numero TFI no hay documento
    cuerpo = driver.find_element(By.TAG_NAME, 'body').text
    m = re.search(r'TFI\s+\d+', cuerpo)
    if not m:
        resumen = ' | '.join(l.strip() for l in cuerpo.splitlines()[:12] if l.strip())
        raise Exception(f'Contifico no devolvio numero TFI. Pantalla: {resumen[:300]}')
    num_doc = m.group(0)
    url_doc = driver.current_url
    log(f'    Registrado: {num_doc}  (Pendiente)')

    # 5. Generar: es el paso que ajusta el inventario
    driver.execute_script("mostrar_generar_movimientos();")
    time.sleep(3)
    try:
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.ID, 'btndlgContinuar'))).click()
    except Exception:
        driver.execute_script("generarMovimiento();")
    time.sleep(8)

    # 6. Confirmar releyendo el documento
    driver.get(url_doc)
    time.sleep(3)
    cuerpo = driver.find_element(By.TAG_NAME, 'body').text
    if 'Generado' not in cuerpo:
        raise Exception(f'{num_doc} se registro pero NO quedo Generado '
                        f'(no ajusta inventario). Revisar en Contifico.')
    log(f'    {num_doc} GENERADO: el inventario quedo ajustado')
    return len(productos), []


def registrar_toma_bodega(driver, bodega_contifico, productos, fecha_form):
    """Registra la toma fisica de UNA bodega en Contifico via Selenium.
    Retorna (productos_ok, lista_errores)."""
    wait = WebDriverWait(driver, 30)

    log(f'    Navegando a formulario de toma fisica...')
    driver.get(CONTIFICO_TOMA_URL)
    time.sleep(3)
    wait.until(EC.presence_of_element_located((By.ID, 'id_fecha')))
    time.sleep(2)

    # 1. FECHA
    log(f'    Configurando fecha: {fecha_form}')
    campo_fecha = driver.find_element(By.ID, 'id_fecha')
    driver.execute_script("arguments[0].value = '';", campo_fecha)
    campo_fecha.click()
    campo_fecha.clear()
    campo_fecha.send_keys(fecha_form)
    campo_fecha.send_keys(Keys.ESCAPE)
    time.sleep(0.5)

    # 2. BODEGA (autocomplete)
    log(f'    Seleccionando bodega: {bodega_contifico}')
    campo_bodega = driver.find_element(By.CSS_SELECTOR, 'input.object-description[data_id="id_bodega"]')
    campo_bodega.click()
    campo_bodega.clear()
    time.sleep(0.5)
    campo_bodega.send_keys(bodega_contifico)
    time.sleep(2)

    try:
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.ui-autocomplete .ui-menu-item')))
        time.sleep(0.5)
        driver.find_element(By.CSS_SELECTOR, '.ui-autocomplete .ui-menu-item:first-child').click()
        time.sleep(2)
    except:
        time.sleep(2)
        campo_bodega.send_keys(Keys.DOWN)
        time.sleep(1)
        campo_bodega.send_keys(Keys.ENTER)
        time.sleep(2)

    valor_bodega = campo_bodega.get_attribute('value') or ''
    if not valor_bodega.strip():
        raise Exception(f'Bodega NO seleccionada - campo vacio despues del autocomplete')
    log(f'    Bodega confirmada: {valor_bodega}')

    # Verificar hidden id_bodega
    try:
        hidden_bodega = driver.find_element(By.ID, 'id_bodega').get_attribute('value') or ''
        if not hidden_bodega.strip():
            campo_bodega.click()
            time.sleep(0.5)
            campo_bodega.send_keys(Keys.DOWN)
            time.sleep(0.5)
            campo_bodega.send_keys(Keys.ENTER)
            time.sleep(2)
    except:
        pass

    # Esperar formulario de detalle
    log('    Esperando formulario de detalle...')
    for _ in range(10):
        try:
            primer_campo = driver.find_element(By.CSS_SELECTOR, 'input.object-description[data_id="id_detalle_1-producto"]')
            if primer_campo.is_displayed() and primer_campo.is_enabled():
                break
        except:
            pass
        time.sleep(1)
    else:
        try:
            driver.execute_script('movimiento.agregarDetalle();')
            time.sleep(2)
        except:
            pass
    time.sleep(1)

    # 3. DESCRIPCION
    cerrar_modales(driver)
    time.sleep(1)
    descripcion = f'TOMA FISICA {bodega_contifico} - {fecha_form}'
    campo_desc = driver.find_element(By.ID, 'id_descripcion')
    try:
        campo_desc.click()
    except:
        cerrar_modales(driver)
        time.sleep(1)
        driver.execute_script("arguments[0].focus(); arguments[0].click();", campo_desc)
    campo_desc.clear()
    campo_desc.send_keys(descripcion)
    time.sleep(0.5)

    # 4. AGREGAR PRODUCTOS
    log(f'    Agregando {len(productos)} productos...')
    productos_ok = 0
    productos_error = []
    fila_actual = 1

    for i, prod in enumerate(productos):
        codigo = prod['codigo']
        cantidad = prod['cantidad']
        nombre = prod['nombre']

        try:
            cerrar_modales(driver)
            time.sleep(0.3)

            selector_prod = f'input.object-description[data_id="id_detalle_{fila_actual}-producto"]'
            try:
                campo_producto = driver.find_element(By.CSS_SELECTOR, selector_prod)
                if not campo_producto.is_displayed():
                    raise Exception('Campo no visible')
            except:
                driver.execute_script('movimiento.agregarDetalle();')
                time.sleep(1)
                cerrar_modales(driver)
                all_prods = driver.find_elements(By.CSS_SELECTOR, 'input.object-description[data_id*="detalle_"]')
                visible_prods = [p for p in all_prods if p.is_displayed() and 'producto' in (p.get_attribute('data_id') or '')]
                if visible_prods:
                    campo_producto = visible_prods[-1]
                    data_id = campo_producto.get_attribute('data_id')
                    fila_actual = int(data_id.split('detalle_')[1].split('-')[0])
                else:
                    log(f'      [{i+1}/{len(productos)}] ERROR {codigo}: No hay campo de producto disponible', 'ERROR')
                    productos_error.append(codigo)
                    continue

            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo_producto)
            time.sleep(0.3)

            try:
                campo_producto.click()
            except:
                driver.execute_script("arguments[0].focus(); arguments[0].click();", campo_producto)
            campo_producto.clear()
            time.sleep(0.2)
            campo_producto.send_keys(codigo)
            time.sleep(1.5)

            campo_producto.send_keys(Keys.DOWN)
            time.sleep(0.3)
            campo_producto.send_keys(Keys.ENTER)
            time.sleep(1.5)

            cerrar_modales(driver)
            time.sleep(0.3)

            # Llenar cantidad_registrada
            id_cantidad = f'id_detalle_{fila_actual}-cantidad_registrada'
            campo_cantidad = driver.find_element(By.ID, id_cantidad)

            cant_str = str(int(cantidad)) if cantidad == int(cantidad) else str(cantidad)

            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo_cantidad)
                time.sleep(0.2)
                driver.execute_script("arguments[0].value = '';", campo_cantidad)
                campo_cantidad.click()
                campo_cantidad.clear()
                campo_cantidad.send_keys(cant_str)
                campo_cantidad.send_keys(Keys.TAB)
            except:
                cerrar_modales(driver)
                js_set_value_and_trigger(driver, id_cantidad, cant_str)

            time.sleep(0.5)
            fila_actual += 1
            productos_ok += 1

            if (i + 1) % 10 == 0 or (i + 1) == len(productos):
                log(f'      [{i+1}/{len(productos)}] {codigo} - {nombre}: {cantidad}')

        except Exception as e:
            err_msg = str(e).split('\n')[0][:80]
            log(f'      [{i+1}/{len(productos)}] ERROR {codigo} ({nombre}): {err_msg}', 'ERROR')
            productos_error.append(codigo)
            cerrar_modales(driver)
            try:
                driver.execute_script('movimiento.agregarDetalle();')
                time.sleep(0.5)
                cerrar_modales(driver)
                all_prods = driver.find_elements(By.CSS_SELECTOR, 'input.object-description[data_id*="detalle_"]')
                visible_prods = [p for p in all_prods if p.is_displayed() and 'producto' in (p.get_attribute('data_id') or '')]
                if visible_prods:
                    data_id = visible_prods[-1].get_attribute('data_id')
                    fila_actual = int(data_id.split('detalle_')[1].split('-')[0])
            except:
                fila_actual += 1

    log(f'    Productos cargados: {productos_ok}/{len(productos)}')
    if productos_error:
        log(f'    Productos con error: {", ".join(productos_error)}')

    # 5. GUARDAR
    log(f'    Guardando toma fisica...')
    cerrar_modales(driver)
    time.sleep(2)
    try:
        btn_guardar = driver.find_element(By.ID, 'btn-guardar')
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn_guardar)
        time.sleep(0.5)
        try:
            btn_guardar.click()
        except:
            driver.execute_script("document.getElementById('btn-guardar').click();")
        log('    -> Click en Guardar OK')
        time.sleep(5)
    except Exception as e:
        log(f'    -> Error al guardar: {e}', 'ERROR')
        try:
            driver.execute_script("document.getElementById('btn-guardar').click();")
            time.sleep(5)
        except:
            log('    -> No se pudo guardar', 'ERROR')

    cerrar_modales(driver)
    time.sleep(1)

    # 6. GENERAR
    log(f'    Generando movimiento...')
    time.sleep(2)
    try:
        btn_generar = driver.find_element(By.ID, 'btn-generar')
        try:
            btn_generar.click()
        except:
            driver.execute_script("document.getElementById('btn-generar').click();")
        log('    -> Click en Generar OK')
        time.sleep(3)
    except Exception as e:
        log(f'    -> Error al generar: {e}', 'ERROR')

    # 7. CONFIRMAR
    log(f'    Confirmando generacion de movimiento...')
    time.sleep(2)
    try:
        btn_continuar = wait.until(EC.element_to_be_clickable((By.ID, 'btndlgContinuar')))
        btn_continuar.click()
        log('    -> Click en Continuar OK')
        time.sleep(5)
    except:
        try:
            driver.execute_script('generarMovimiento();')
            log('    -> generarMovimiento() ejecutado via JS')
            time.sleep(5)
        except Exception as e:
            log(f'    -> No se pudo confirmar: {e}', 'ERROR')

    url_final = driver.current_url
    log(f'    URL final: {url_final}')

    return productos_ok, productos_error


def procesar_tarea_carga(tarea, driver):
    """Procesa una tarea de carga de toma fisica a Contifico.
    Retorna (driver_sigue_vivo, ok)."""
    ejec_id = tarea['id']
    bodega = tarea['bodega']
    fecha_toma = tarea['fecha_toma']
    fecha_form = datetime.strptime(fecha_toma, '%Y-%m-%d').strftime('%d/%m/%Y')

    bodega_contifico = BODEGAS_CARGA.get(bodega)
    if not bodega_contifico:
        post_resultado_carga({'id': ejec_id, 'estado': 'error', 'error_msg': f'bodega desconocida: {bodega}'})
        return True, False

    cfg = BODEGAS.get(bodega)
    if not cfg:
        post_resultado_carga({'id': ejec_id, 'estado': 'error', 'error_msg': f'config bodega no encontrada: {bodega}'})
        return True, False

    log(f'>>> CARGA CONTIFICO ejec_id={ejec_id} bodega={bodega} fecha={fecha_toma}')
    driver_ok = True
    try:
        log('  - Cargando equivalencias y toma fisica...')
        equivs = cargar_equivalencias(bodega)
        productos, faltantes, sin_contar, total_toma = cargar_toma_para_contifico(
            cfg['tabla_toma'], fecha_toma, equivs)
        log(f'    equivalencias={len(equivs)}  productos_a_cargar={len(productos)}'
            f'  sin_contar={len(sin_contar)}  total_en_la_toma={total_toma}')
        if sin_contar:
            log(f'    EN CERO: {len(sin_contar)} de {total_toma} productos llegaron '
                f'en blanco desde el conteo y suben en cero', 'WARN')
            avisar_toma_incompleta(bodega, fecha_toma, len(productos), sin_contar, total_toma)
        if faltantes:
            log(f'    SIN EQUIVALENCIA: {len(faltantes)} -> '
                + ', '.join(c for c, _ in faltantes[:10]), 'WARN')
            avisar_sin_equivalencia('Carga de toma fisica a Contifico', bodega, faltantes)

        if not productos:
            post_resultado_carga({'id': ejec_id, 'estado': 'error',
                                  'error_msg': f'No hay toma fisica para {bodega} en {fecha_toma}'})
            return True, False

        log(f'  - Registrando en Contifico: {bodega_contifico} ({len(productos)} productos)...')
        # Por archivo. Fila a fila tardaba cerca de una hora con 111 productos y
        # dejaba al worker bloqueado, con el resto de la cola esperando.
        # CARGA_FILA_A_FILA=1 vuelve al camino viejo si hiciera falta.
        if os.environ.get('CARGA_FILA_A_FILA', '0') == '1':
            ok, errores = registrar_toma_bodega(driver, bodega_contifico,
                                                productos, fecha_form)
        else:
            ok, errores = registrar_toma_por_archivo(driver, bodega_contifico,
                                                     productos, fecha_form)

        post_resultado_carga({
            'id': ejec_id,
            'estado': 'completado',
            'total_productos': len(productos),
            'productos_ok': ok,
            'productos_error': len(errores),
            'productos_error_lista': ', '.join(errores) if errores else None,
            # Lo que la toma tenia y no se conto. Va al resultado para que el
            # panel pueda decir '31 de 224' en vez de '31/31 OK'.
            'productos_sin_contar': len(sin_contar),
            'total_en_toma': total_toma,
        })
        log(f'  - CARGA COMPLETADA: {ok}/{len(productos)} productos OK'
            f'  (de {total_toma} en la toma; {len(sin_contar)} sin contar)')
        return True, True
    except Exception as e:
        tb = traceback.format_exc()
        log(f'ERROR procesando carga {ejec_id}: {e}\n{tb}', 'ERROR')
        post_resultado_carga({'id': ejec_id, 'estado': 'error', 'error_msg': str(e)[:500]})
        msg = str(e).lower()
        if any(k in msg for k in ('invalid session', 'disconnected', 'not connected',
                                   'no such window', 'connection refused', 'target closed',
                                   'chrome not reachable')):
            driver_ok = False
        return driver_ok, False


# ============ CONTEO OPERATIVO (descarga Contifico + seleccion productos) ============
# Los productos se consultan al backend via /api/conteo-op/productos-semana
# BP: 14 fijos (tipo_conteo='fijo') | MP: 10 aleatorios semanales | Planta: 10 aleatorios semanales

AIRTABLE_TOKEN = os.environ.get('AIRTABLE_TOKEN_CONTEO', '')
AIRTABLE_BASE = 'app5zYXr1GmF2bmVF'
AIRTABLE_TABLE = 'tbl8hyvwwfSnrspAt'

AIRTABLE_CAMPO_BODEGA = {
    'bodega_principal': 'Conteo Bodega Principal',
    'materia_prima': 'Conteo Bodega Materia Prima',
    'planta': 'Conteo Planta Producción',
}

def cargar_productos_airtable(bodega):
    """Carga productos de AirTable filtrados por bodega, retorna dict {codigo: tipo_abc}."""
    campo_conteo = AIRTABLE_CAMPO_BODEGA.get(bodega)
    if not campo_conteo:
        return {}
    all_records = []
    offset = None
    while True:
        params = {
            'pageSize': 100,
            'fields[]': ['Código', 'Tipo A,B o C', campo_conteo, 'Estado'],
        }
        if offset:
            params['offset'] = offset
        r = requests.get(
            f'https://api.airtable.com/v0/{AIRTABLE_BASE}/{AIRTABLE_TABLE}',
            headers={'Authorization': f'Bearer {AIRTABLE_TOKEN}'},
            params=params, timeout=15,
        )
        data = r.json()
        all_records.extend(data.get('records', []))
        offset = data.get('offset')
        if not offset:
            break

    resultado = {}
    for rec in all_records:
        f = rec.get('fields', {})
        cod = f.get('Código', '')
        if not cod:
            continue
        if f.get(campo_conteo) != 'Sí':
            continue
        if f.get('Estado') != 'Activo':
            continue
        tipo = f.get('Tipo A,B o C', '')
        if tipo in ('A', 'B', 'C'):
            resultado[cod] = tipo
    return resultado

def get_pendientes_conteo_op():
    try:
        r = requests.get(
            f'{BACKEND_URL}/api/conteo-op/pendientes',
            headers={'X-Worker-Token': WORKER_TOKEN},
            params={'worker_id': WORKER_ID},
            timeout=20,
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        log(f'pendientes_conteo_op -> {e}', 'WARN')
    return []

def post_resultado_conteo_op(payload):
    try:
        r = requests.post(
            f'{BACKEND_URL}/api/conteo-op/resultado',
            headers={'X-Worker-Token': WORKER_TOKEN, 'Content-Type': 'application/json'},
            data=json.dumps(payload, default=str),
            timeout=60,
        )
        return r.status_code == 200
    except Exception as e:
        log(f'resultado_conteo_op -> {e}', 'ERROR')
    return False

def procesar_tarea_conteo_op(tarea, driver):
    """Descarga Excel de Contifico, obtiene productos del backend, inserta en BD.
    BP: 14 fijos | MP/Planta: 10 aleatorios semanales (seleccionados por el backend)."""
    ejec_id = tarea['id']
    bodega = tarea['bodega']
    fecha = tarea['fecha']
    log(f'>>> CONTEO OPERATIVO ejec_id={ejec_id} bodega={bodega} fecha={fecha}')

    cfg = BODEGAS.get(bodega)
    if not cfg:
        post_resultado_conteo_op({'id': ejec_id, 'estado': 'error', 'error_msg': f'bodega desconocida: {bodega}'})
        return True, False

    driver_ok = True
    try:
        # 1. Obtener productos seleccionados del backend
        log(f'  - Consultando productos seleccionados al backend...')
        try:
            r = requests.get(
                f'{BACKEND_URL}/api/conteo-op/productos-semana',
                params={'bodega': bodega, 'fecha': fecha},
                timeout=15
            )
            if r.status_code == 200:
                data_productos = r.json()
                codigos_seleccionados = [p['codigo'] for p in data_productos.get('productos', [])]
                log(f'    Backend devolvio {len(codigos_seleccionados)} productos')
            else:
                log(f'    Backend no tiene productos seleccionados (HTTP {r.status_code}), usando BD directa', 'WARN')
                codigos_seleccionados = []
        except Exception as e:
            log(f'    Error consultando backend: {e}, usando BD directa', 'WARN')
            codigos_seleccionados = []

        # Si el backend no contesta se leen los MISMOS productos de la semana
        # desde la BD. Antes, para MP y Planta, se sorteaban 10 al azar con
        # ORDER BY RANDOM(): ese dia la bodega habria contado unos productos
        # distintos a los de su semana y el cruce compararia cosas que no se
        # corresponden. Ahora se respeta la rotacion, y si no la hay se falla.
        if not codigos_seleccionados:
            marca = cfg.get('marca', bodega.upper())
            if bodega == 'bodega_principal':
                rows = db_query(
                    "SELECT codigo FROM goti.productos_por_marca "
                    "WHERE marca = %s AND tipo_conteo = 'fijo' AND activo = TRUE",
                    (marca,))
                codigos_seleccionados = [r['codigo'] for r in rows]
                log(f'    Sin backend: uso los {len(codigos_seleccionados)} fijos de la BD')
            else:
                rows = db_query(
                    "SELECT codigos FROM goti.rotacion_semanal_bodegas "
                    "WHERE bodega = %s AND %s BETWEEN semana_inicio AND semana_fin",
                    (bodega, fecha))
                codigos_seleccionados = list(rows[0]['codigos']) if rows and rows[0]['codigos'] else []
                log(f'    Sin backend: rotacion de la semana en BD -> '
                    f'{len(codigos_seleccionados)} productos')

        if not codigos_seleccionados:
            post_resultado_conteo_op({'id': ejec_id, 'estado': 'error', 'error_msg': 'No hay productos seleccionados'})
            return True, False

        # 2. Descargar Excel de Contifico (stock a la fecha del conteo)
        log(f'  - Descargando saldos de Contifico ({cfg["contifico"]}) al {fecha}...')
        driver.set_script_timeout(TIMEOUT_SCRIPT)
        archivo = descargar_saldos(driver, cfg['contifico'], fecha)
        log(f'    archivo: {os.path.basename(archivo)}')

        contifico = parsear_saldos(archivo, cfg['contifico'])
        log(f'    productos en Contifico: {len(contifico)}')

        if not contifico:
            post_resultado_conteo_op({'id': ejec_id, 'estado': 'error',
                                      'error_msg': 'No se encontraron productos en Contifico'})
            return True, False

        # 3. Cargar equivalencias para conversion de unidades
        equivalencias = cargar_equivalencias(bodega)

        # 4. Filtrar productos seleccionados — si no esta en Contifico, insertar con stock=0
        productos_final = []
        # Cargar datos de productos_por_marca para fallback de nombre/unidad
        marca = cfg.get('marca', bodega.upper())
        datos_marca = {}
        for r in db_query("SELECT codigo, nombre, unidad FROM goti.productos_por_marca WHERE marca = %s AND activo = TRUE", (marca,)):
            datos_marca[r['codigo']] = r

        sin_equivalencia = []
        for cod in codigos_seleccionados:
            if cod in contifico:
                c = contifico[cod]
                eq = equivalencias.get(cod)
                stock = c['stock']
                unidad = c['unidad']
                if eq:
                    factor = float(eq.get('factor', 1))
                    if factor and factor != 0:
                        # El stock viene en la unidad de Contifico y aqui se
                        # expresa en la unidad en que cuenta la persona, que es
                        # la direccion contraria a la del cruce: se DIVIDE.
                        # Multiplicando, CREMA DE LECHE quedaba en 74.480.000 Kg
                        # cuando Contifico tenia 74.480 g, o sea 74,48 Kg.
                        stock = stock / factor
                    unidad = eq.get('unidad_toma', unidad)
                else:
                    sin_equivalencia.append((cod, c['nombre']))
                productos_final.append({
                    'codigo': cod, 'nombre': c['nombre'],
                    'unidad': unidad, 'cantidad': stock, 'costo': c['costo']
                })
            else:
                # No esta en Contifico (stock 0) — insertar con cantidad=0
                pm = datos_marca.get(cod, {})
                eq = equivalencias.get(cod)
                unidad = eq.get('unidad_toma', pm.get('unidad', 'Unidad')) if eq else pm.get('unidad', 'Unidad')
                productos_final.append({
                    'codigo': cod, 'nombre': pm.get('nombre', cod),
                    'unidad': unidad, 'cantidad': 0, 'costo': 0
                })
                log(f'    {cod} no en Contifico -> insertado con stock=0')

        if sin_equivalencia:
            log(f'    SIN EQUIVALENCIA: {len(sin_equivalencia)} -> '
                + ', '.join(c for c, _ in sin_equivalencia[:10]), 'WARN')
            avisar_sin_equivalencia('Conteo operativo', bodega, sin_equivalencia)

        n_fijos = len(productos_final) if bodega == 'bodega_principal' else 0
        n_aleatorios = len(productos_final) if bodega != 'bodega_principal' else 0

        # 5. Insertar en inventario_ciego_conteos
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor(cursor_factory=RealDictCursor)
        insertados = 0
        for p in productos_final:
            cur.execute("""
                INSERT INTO goti.inventario_ciego_conteos
                (fecha, local, codigo, nombre, unidad, cantidad, costo_unitario)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (fecha, local, codigo) DO UPDATE SET
                    cantidad = EXCLUDED.cantidad, costo_unitario = EXCLUDED.costo_unitario
            """, (fecha, bodega, p['codigo'], p['nombre'], p['unidad'], p['cantidad'], p.get('costo', 0)))
            insertados += 1
        conn.commit()
        conn.close()

        log(f'  - CONTEO GENERADO: {n_fijos} fijos + {n_aleatorios} aleatorios = {insertados} total')

        post_resultado_conteo_op({
            'id': ejec_id, 'estado': 'completado',
            'total_productos': insertados, 'fijos': n_fijos,
            'aleatorios': n_aleatorios,
        })
        return True, True
    except Exception as e:
        tb = traceback.format_exc()
        log(f'ERROR conteo operativo {ejec_id}: {e}\n{tb}', 'ERROR')
        post_resultado_conteo_op({'id': ejec_id, 'estado': 'error', 'error_msg': str(e)[:500]})
        msg = str(e).lower()
        if any(k in msg for k in ('invalid session', 'disconnected', 'not connected',
                                   'no such window', 'connection refused', 'target closed',
                                   'chrome not reachable')):
            driver_ok = False
        return driver_ok, False


def driver_sano(driver):
    """Verifica que el driver siga respondiendo."""
    if driver is None:
        return False
    try:
        _ = driver.current_url
        return True
    except Exception:
        return False


# ============ INVENTARIO LOCALES (VENTAS) ============

def get_pendientes_inventario_locales():
    """Poll para tareas de inventario locales (actualizar cantidad / toma fisica)."""
    try:
        r = requests.get(
            f'{BACKEND_URL}/api/inventario-locales/pendientes',
            headers={'X-Worker-Token': WORKER_TOKEN},
            params={'worker_id': WORKER_ID},
            timeout=20,
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        log(f'pendientes_inventario_locales -> {e}', 'WARN')
    return []


def post_resultado_inventario_locales(payload):
    try:
        r = requests.post(
            f'{BACKEND_URL}/api/inventario-locales/resultado',
            headers={'X-Worker-Token': WORKER_TOKEN, 'Content-Type': 'application/json'},
            data=json.dumps(payload, default=str),
            timeout=60,
        )
        return r.status_code == 200
    except Exception as e:
        log(f'resultado_inventario_locales -> {e}', 'ERROR')
    return False


def obtener_productos_marca(marca):
    """Obtiene productos activos de la BD para una marca."""
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT codigo, nombre FROM goti.productos_por_marca
            WHERE marca = %s AND activo = TRUE
        """, (marca,))
        return {row[0]: row[1] for row in cur.fetchall()}
    except Exception as e:
        log(f'obtener_productos_marca({marca}) -> {e}', 'ERROR')
        return {}
    finally:
        if conn:
            conn.close()


def procesar_actualizar_cantidad(tarea, driver):
    """Descarga saldos de Contifico y UPDATE cantidad en BD sin tocar conteos."""
    ejec_id = tarea['id']
    bodega = tarea['bodega']
    fecha = tarea['fecha']
    fecha_iso = fecha if isinstance(fecha, str) else fecha.isoformat()
    fecha_dmY = datetime.strptime(fecha_iso, '%Y-%m-%d').strftime('%d/%m/%Y')

    log(f'>>> ACTUALIZAR CANTIDAD ejec_id={ejec_id} bodega={bodega} fecha={fecha_iso}')

    cfg = BODEGAS_LOCALES.get(bodega)
    if not cfg:
        post_resultado_inventario_locales({'id': ejec_id, 'estado': 'error', 'error_msg': f'bodega desconocida: {bodega}'})
        return True, False

    driver_ok = True
    conn = None
    try:
        # 1. Descargar Excel de Contifico
        log(f'  - Descargando saldos de {cfg["contifico"]} al {fecha_dmY}...')
        archivo = descargar_saldos(driver, cfg['contifico'], fecha_iso)
        log(f'  - Archivo descargado: {archivo}')

        # 2. Parsear Excel
        df = pd.read_excel(archivo, header=5)
        cols_map = {}
        for col in df.columns:
            cl = col.lower().strip()
            if ('digo' in cl or 'codigo' in cl) and 'cat' not in cl:
                cols_map['Codigo'] = col
            elif cl == 'stock':
                cols_map['Stock'] = col
            elif 'costo' in cl:
                cols_map['Costo'] = col

        # 3. Obtener productos de la marca
        productos = obtener_productos_marca(cfg['tipo'])
        if not productos:
            raise Exception(f'No hay productos activos para marca {cfg["tipo"]}')
        log(f'  - Productos a procesar: {len(productos)}')

        # 4. UPDATE cantidad y costo SIN TOCAR conteos
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        actualizados = 0
        for codigo, nombre in productos.items():
            fila = df[df[cols_map['Codigo']] == codigo]
            if len(fila) > 0:
                stock = float(fila[cols_map['Stock']].values[0])
                costo = float(fila[cols_map['Costo']].values[0]) if 'Costo' in cols_map else 0
            else:
                stock = 0
                costo = 0

            # UPSERT: insertar si no existe, actualizar solo cantidad/costo si existe
            cur.execute("""
                INSERT INTO goti.inventario_ciego_conteos
                    (local, fecha, codigo, nombre, cantidad, costo_unitario)
                VALUES (%s, %s::date, %s, %s, %s, %s)
                ON CONFLICT (local, fecha, codigo)
                DO UPDATE SET cantidad = EXCLUDED.cantidad, costo_unitario = EXCLUDED.costo_unitario
            """, (bodega, fecha_iso, codigo, nombre, stock, costo))
            actualizados += 1

        conn.commit()
        log(f'  - Registros actualizados: {actualizados}')

        # Limpiar archivo
        try:
            os.remove(archivo)
        except:
            pass

        post_resultado_inventario_locales({
            'id': ejec_id,
            'estado': 'completado',
            'total_productos': actualizados,
        })
        return True, True

    except Exception as e:
        tb = traceback.format_exc()
        log(f'ERROR actualizar_cantidad {ejec_id}: {e}\n{tb}', 'ERROR')
        post_resultado_inventario_locales({'id': ejec_id, 'estado': 'error', 'error_msg': str(e)[:500]})
        msg = str(e).lower()
        if any(k in msg for k in ('invalid session', 'disconnected', 'not connected',
                                   'no such window', 'connection refused', 'target closed')):
            driver_ok = False
        return driver_ok, False
    finally:
        if conn:
            conn.close()


# Nombre de bodega tal y como se guarda en goti.telegram_destinatarios. Las
# claves internas ('real_audiencia') no coinciden con las que se ven en el
# panel ('REAL'); sin traducir, la consulta no encuentra a nadie y el aviso se
# pierde en silencio.
BODEGA_TELEGRAM = {
    'real_audiencia': 'REAL',
    'floreana': 'FLOREANA',
    'portugal': 'PORTUGAL',
    'santo_cachon_real': 'SANTO CACHON REAL',
    'santo_cachon_portugal': 'SANTO CACHON PORTUGAL',
    'simon_bolon': 'SIMON BOLON',
    'bodega_principal': 'BODEGA PRINCIPAL',
    'materia_prima': 'BODEGA MATERIA PRIMA',
    'planta': 'PLANTA DE PRODUCCION',
    'bodega_pulmon': 'BODEGA PULMON',
}


def avisar_toma_fisica(bodega, fecha_dmY, num_doc, productos, url=None, error=None):
    """Avisa por Telegram del resultado de la toma fisica.

    Va con su propio try: que falle el aviso no puede tumbar una toma que ya
    quedo generada en Contifico. Perder un mensaje molesta; marcar como fallida
    una toma que si ajusto el inventario es peor.

    A quien le llega lo decide goti.telegram_destinatarios: se filtra por
    bodega y por operacion 'Toma Fisica', asi que cada local recibe la suya.
    """
    nombre = BODEGA_TELEGRAM.get(bodega, bodega.upper())
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from notificar_telegram import notificar_exito, notificar_error
        if error:
            notificar_error('Toma Fisica', nombre,
                            'Fecha: {}{}{} productos contados'.format(
                                fecha_dmY, '\n', productos),
                            error[:200])
            return
        partes = [
            'Fecha: {}'.format(fecha_dmY),
            '{} productos contados'.format(productos),
            'Estado: GENERADO (el inventario quedo ajustado)',
        ]
        if url:
            partes.append(url)
        notificar_exito('Toma Fisica', nombre, '\n'.join(partes), num_doc)
    except Exception as e:
        log('  aviso por Telegram fallo: {}'.format(str(e)[:120]), 'WARN')


def procesar_toma_fisica_local(tarea, driver):
    """Lee conteos de BD y registra la toma fisica en Contifico por CARGA MASIVA.

    Antes se llenaba el formulario producto por producto: 36 autocompletados
    encadenados que tardaban ~13 minutos, y bastaba que uno se descuadrara para
    que Contifico descartara el formulario al guardar. Peor: el codigo marcaba
    'completado' aunque no se hubiera creado nada, porque solo miraba que no
    saltara una excepcion. Asi se reportaron como subidas tomas del 17-ago que
    no existian.

    Ahora se sube el Excel por la pestana 'Carga Masiva' (input id_archivo), el
    camino que ofrece Contifico en su plantilla oficial
    (.../tomafisica/descargar_plantilla): dos columnas, 'Producto (Codigo)' y
    'Cantidad'. Tarda ~35s en vez de 13 min y Contifico valida el archivo.

    Y AHORA SE VERIFICA: no se reporta 'completado' si no aparece el numero TFI
    y el documento no queda en estado 'Generado', que es el paso que realmente
    ajusta el inventario.
    """
    ejec_id = tarea['id']
    bodega = tarea['bodega']
    fecha = tarea['fecha']
    fecha_iso = fecha if isinstance(fecha, str) else fecha.isoformat()
    fecha_dmY = datetime.strptime(fecha_iso, '%Y-%m-%d').strftime('%d/%m/%Y')

    log(f'>>> TOMA FISICA ejec_id={ejec_id} bodega={bodega} fecha={fecha_iso}')

    cfg = BODEGAS_LOCALES.get(bodega) or BODEGAS.get(bodega)
    if not cfg:
        post_resultado_inventario_locales({'id': ejec_id, 'estado': 'error',
                                           'error_msg': f'bodega desconocida: {bodega}'})
        return True, False

    driver_ok = True
    conn = None
    try:
        # 1. Conteos de la BD
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT codigo, COALESCE(cantidad_contada_2, cantidad_contada) AS cantidad
            FROM goti.inventario_ciego_conteos
            WHERE fecha = %s AND local = %s
              AND COALESCE(cantidad_contada_2, cantidad_contada) IS NOT NULL
              AND COALESCE(cantidad_contada_2, cantidad_contada) > 0
            ORDER BY nombre
        """, (fecha_iso, bodega))
        productos = [(r[0], float(r[1])) for r in cur.fetchall()]
        conn.close()
        conn = None
        if not productos:
            raise Exception(f'No hay conteos en BD para {bodega} fecha {fecha_iso}')
        log(f'  - Productos con conteo: {len(productos)}')

        # 2. Excel con el formato de la plantilla oficial
        import openpyxl
        os.makedirs(DOWNLOAD_DIR, exist_ok=True)
        ruta = os.path.join(DOWNLOAD_DIR, f'tf_{bodega}_{fecha_iso}.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Producto (Código)', 'Cantidad'])
        for cod, cant in productos:
            # El codigo va como TEXTO: la plantilla lo exige.
            ws.cell(row=ws.max_row + 1, column=1, value=str(cod)).number_format = '@'
            ws.cell(row=ws.max_row, column=2,
                    value=int(cant) if cant == int(cant) else cant)
        wb.save(ruta)
        log(f'  - Excel generado: {os.path.basename(ruta)}')

        # 3. Cabecera del formulario
        driver.get('https://1793168604001.contifico.com/sistema/inventario/tomafisica/registrar/')
        wait = WebDriverWait(driver, 60)
        wait.until(EC.presence_of_element_located((By.ID, 'id_fecha')))
        time.sleep(2)

        campo_fecha = driver.find_element(By.ID, 'id_fecha')
        driver.execute_script("arguments[0].value = '';", campo_fecha)
        campo_fecha.click()
        campo_fecha.clear()
        campo_fecha.send_keys(fecha_dmY)
        campo_fecha.send_keys(Keys.ESCAPE)
        time.sleep(1)

        campo_bodega = driver.find_element(
            By.CSS_SELECTOR, 'input.object-description[data_id="id_bodega"]')
        campo_bodega.click()
        campo_bodega.clear()
        campo_bodega.send_keys(cfg['contifico'])
        time.sleep(2)
        campo_bodega.send_keys(Keys.DOWN)
        campo_bodega.send_keys(Keys.ENTER)
        time.sleep(2)

        descripcion = f'AJUSTE DE INVENTARIO {cfg["contifico"]} - {fecha_dmY}'
        campo_desc = driver.find_element(By.ID, 'id_descripcion')
        driver.execute_script("arguments[0].focus();", campo_desc)
        campo_desc.clear()
        campo_desc.send_keys(descripcion)
        log(f'  - {fecha_dmY} | {cfg["contifico"]}')

        # 4. Carga masiva
        inp = driver.find_element(By.ID, 'id_archivo')
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        inp.send_keys(ruta)
        time.sleep(2)
        log('  - Archivo adjuntado, guardando...')
        driver.execute_script("registrarMovimiento();")
        time.sleep(9)

        # 5. VERIFICAR que se creo: sin numero TFI no hay documento
        cuerpo = driver.find_element(By.TAG_NAME, 'body').text
        m = re.search(r'TFI\s+\d+', cuerpo)
        if not m:
            resumen = ' | '.join(l.strip() for l in cuerpo.split('\n')[:12] if l.strip())
            raise Exception(f'Contifico no devolvio numero TFI. Pantalla: {resumen[:300]}')
        num_doc = m.group(0)
        url_doc = driver.current_url
        log(f'  - Registrado: {num_doc}  (Pendiente)')

        # 6. Generar: es el paso que ajusta el inventario
        driver.execute_script("mostrar_generar_movimientos();")
        time.sleep(3)
        try:
            WebDriverWait(driver, 15).until(
                EC.element_to_be_clickable((By.ID, 'btndlgContinuar'))).click()
        except Exception:
            driver.execute_script("generarMovimiento();")
        time.sleep(8)

        # 7. VERIFICAR el estado final releyendo el documento
        driver.get(url_doc)
        time.sleep(3)
        cuerpo = driver.find_element(By.TAG_NAME, 'body').text
        if 'Generado' not in cuerpo:
            raise Exception(f'{num_doc} se registro pero NO quedo Generado '
                            f'(no ajusta inventario). Revisar en Contifico.')
        log(f'  - {num_doc} GENERADO: el inventario quedo ajustado')

        avisar_toma_fisica(bodega, fecha_dmY, num_doc, len(productos), url_doc)

        post_resultado_inventario_locales({
            'id': ejec_id,
            'estado': 'completado',
            'total_productos': len(productos),
            'url_contifico': url_doc,
            'num_documento': num_doc,
        })
        return True, True

    except Exception as e:
        log(f'ERROR toma_fisica {ejec_id}: {e}', 'ERROR')
        log(traceback.format_exc()[:600], 'ERROR')
        post_resultado_inventario_locales({'id': ejec_id, 'estado': 'error',
                                           'error_msg': str(e)[:500]})
        msg = str(e).lower()
        if any(k in msg for k in ('invalid session', 'disconnected', 'not connected',
                                  'no such window', 'connection refused', 'target closed',
                                  'chrome not reachable')):
            driver_ok = False
        return driver_ok, False
    finally:
        if conn:
            conn.close()




def procesar_tarea_inventario_locales(tarea, driver):
    """Dispatcher para tareas de inventario locales."""
    accion = tarea.get('accion', '')
    if accion == 'actualizar_cantidad':
        return procesar_actualizar_cantidad(tarea, driver)
    elif accion == 'toma_fisica':
        return procesar_toma_fisica_local(tarea, driver)
    else:
        log(f'Accion desconocida: {accion}', 'ERROR')
        post_resultado_inventario_locales({'id': tarea['id'], 'estado': 'error', 'error_msg': f'accion desconocida: {accion}'})
        return True, False


# ============ PROCESAMIENTO PARALELO ============
def procesar_tarea_paralelo(tarea, tipo, worker_num):
    """Procesa una tarea en su propio thread con su propio Chrome."""
    driver = None
    try:
        log(f'Iniciando Chrome para {tipo}...', worker_id=worker_num)
        driver = make_chrome(parallel_id=worker_num)
        login_contifico(driver)

        if tipo == 'cruce':
            _, ok = procesar_tarea(tarea, driver)
        elif tipo == 'carga':
            _, ok = procesar_tarea_carga(tarea, driver)
        elif tipo == 'conteo':
            _, ok = procesar_tarea_conteo_op(tarea, driver)
        elif tipo == 'inventario':
            _, ok = procesar_tarea_inventario_locales(tarea, driver)
        else:
            ok = False

        return tarea.get('id'), ok
    except Exception as e:
        log(f'ERROR paralelo {tipo} tarea {tarea.get("id")}: {e}', 'ERROR', worker_id=worker_num)
        return tarea.get('id'), False
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass


def procesar_tareas_paralelo(tareas_por_tipo):
    """Procesa múltiples tareas en paralelo usando ThreadPoolExecutor."""
    todas = []
    for tipo, tareas in tareas_por_tipo.items():
        for t in tareas:
            todas.append((t, tipo))

    if not todas:
        return

    total = len(todas)
    log(f'Procesando {total} tarea(s) en paralelo (max {MAX_PARALLEL_WORKERS} simultáneas)')

    resultados = {'ok': 0, 'error': 0}
    with ThreadPoolExecutor(max_workers=MAX_PARALLEL_WORKERS) as executor:
        futures = {}
        for i, (tarea, tipo) in enumerate(todas):
            worker_num = i % MAX_PARALLEL_WORKERS + 1
            future = executor.submit(procesar_tarea_paralelo, tarea, tipo, worker_num)
            futures[future] = (tarea.get('id'), tipo)

        for future in as_completed(futures):
            ejec_id, tipo = futures[future]
            try:
                _, ok = future.result()
                if ok:
                    resultados['ok'] += 1
                    log(f'Tarea {tipo} #{ejec_id} completada OK')
                else:
                    resultados['error'] += 1
                    log(f'Tarea {tipo} #{ejec_id} completada con ERROR', 'WARN')
            except Exception as e:
                resultados['error'] += 1
                log(f'Tarea {tipo} #{ejec_id} EXCEPCION: {e}', 'ERROR')

    log(f'Paralelo terminado: {resultados["ok"]} OK, {resultados["error"]} errores')



# ============================================================
# TRASLADOS ENTRE BODEGAS (Selenium)
# ============================================================
# Los traslados son el unico movimiento que la API de Contifico NO puede crear:
# manda el mismo payload que un ingreso mas bodega_destino_id y responde
#     500 {"mensaje": "... 'NoneType' object has no attribute 'parametros'"}
# Pasa igual en v1 y en v2, con cualquier bodega, asi que no es el payload: es
# un fallo del lado de ellos. Por eso quedaban pendientes reintentandose para
# siempre sin llegar nunca a crearse.
#
# Se hace por el formulario web, que es el mismo camino que ya usan produccion
# y las tomas fisicas. Los selectores vienen de 'carga airtable bot/
# test_traslados.py', el bot que corria en la PC de Finanzas.
#
# NO se marca Hecho si el documento no aparece con su numero TRA: preferimos
# reintentar a dar por bueno algo que no existe.

AIRTABLE_TOKEN_GLOG = os.environ.get('AIRTABLE_TOKEN_GLOG', '')
AT_BASE_GLOG = 'appETTeYKD0DQpuN7'
AT_TABLA_TRASLADOS = 'tblpeKmVHSsMopxBQ'   # Egresos Emergentes Tiendas
AT_TABLA_GLOG_PRODUCTOS = 'tblOCyYpGJDFcGVvr'   # Matriz General de Productos
AT_TABLA_GLOG_CONTIFICO = 'tblxC58veM7i1UnYc'
# Campo de la Matriz General con el codigo real de Contifico. Es la fuente
# de verdad para cualquier movimiento: traslado, egreso o ingreso.
AT_CAMPO_CODIGO = 'Codigo Contifico'   # Matriz Contifico

CONTIFICO_MOVIMIENTOS_URL = ('https://1793168604001.contifico.com'
                             '/sistema/inventario/movimiento/registrar/')

MAPEO_BODEGAS_GLOG = {
    'recCypzc9E9uEhJYv': 'PLANTA DE PRODUCCION',
    'recGDd0jYLlVz9b6f': 'BODEGA SANTO CACHON PORTUGAL',
    'recKYprt4weEisem9': 'BODEGA SANTO CACHON REAL',
    'recM8vqHzgEMsff38': 'BODEGA SIMON BOLON',
    'reccM8WyxFZPhS7QL': 'BODEGA CHIOS REAL',
    'reco7xJnelmRE54f5': 'BODEGA CHIOS PORTUGAL',
    'recwIOf9ff2VU3IuS': 'BODEGA CHIOS FLOREANA',
    'recEgtaLkUBCT1fpj': 'BODEGA PRINCIPAL',
    'recQtytIc02x1pZWm': 'BODEGA MATERIA PRIMA',
    'recNUlLpZcSPD2TZt': 'BODEGA PULMON',
}
MAPEO_CENTROS_GLOG = {
    'recCypzc9E9uEhJYv': 'PLANTA DE PRODUCCION',
    'recGDd0jYLlVz9b6f': 'SANTO CACHON PORTUGAL',
    'recKYprt4weEisem9': 'SANTO CACHON REAL',
    'recM8vqHzgEMsff38': 'SIMON BOLON',
    'reccM8WyxFZPhS7QL': 'REAL',
    'reco7xJnelmRE54f5': 'PORTUGAL',
    'recwIOf9ff2VU3IuS': 'FLOREANA',
    'recEgtaLkUBCT1fpj': 'BODEGA PRINCIPAL',
    'recQtytIc02x1pZWm': 'BODEGA MATERIA PRIMA',
    'recNUlLpZcSPD2TZt': 'BODEGA PULMON',
}


def _at_glog(tabla, params=None):
    """Lee una tabla completa de la base GLOG, paginando.

    Sin paginar solo llegan los primeros 100 registros: con 'pendientes' eso
    daria una lista corta y aparentemente correcta.
    """
    filas, offset = [], None
    while True:
        p = dict(params or {})
        if offset:
            p['offset'] = offset
        r = requests.get(f'https://api.airtable.com/v0/{AT_BASE_GLOG}/{tabla}',
                         headers={'Authorization': f'Bearer {AIRTABLE_TOKEN_GLOG}'},
                         params=p, timeout=40)
        r.raise_for_status()
        d = r.json()
        filas.extend(d.get('records', []))
        offset = d.get('offset')
        if not offset:
            return filas


_catalogo_glog = None


def _norma_nombre(texto):
    """Normaliza para comparar: mayusculas y espacios de sobra fuera.

    NO se tocan los acentos ni se recorta nada mas. Los dos lados salen de la
    misma base de AirTable, asi que tienen que coincidir tal cual; cualquier
    "limpieza" adicional solo sirve para volver a juntar productos distintos.
    """
    return ' '.join((texto or '').upper().split())


def _sin_inventario_propio(codigo, contifico):
    """Devuelve el nombre si ese codigo NO tiene inventario propio; None si si.

    Un producto de venta formulado -una picana de menu, una hamburguesa, un
    combo de delivery- no se almacena: lo que se descuenta al venderlo son sus
    ingredientes. Trasladarlo mueve existencias que no existen, y la huella
    queda en Contifico como stock negativo enorme.

    Se mira el campo 'Inventariable' de la Matriz Contifico, que es el que se
    mantiene a mano. Bloquean 'No' y 'No aplica'; solo pasa el campo vacio, que
    si significa "sin clasificar": parar un traslado legitimo por falta de un
    dato seria peor que el problema que se quiere evitar.
    """
    cod = (codigo or '').strip().upper()
    if not cod:
        return None
    for ct in contifico:
        if (ct.get('Código') or '').strip().upper() != cod:
            continue
        marca = str(ct.get('Inventariable') or '').strip().upper()
        # 'No' y 'No aplica' bloquean los dos. Los 30 productos con 'No aplica'
        # son 'Proceso productivo' -PORCIONADO, LIMPIEZA, MEZCLADO-: no son
        # productos, son pasos de un proceso, y su stock negativo enorme es
        # justo la huella de haberlos movido.
        if marca.startswith('NO'):
            return ct.get('Nombre Producto') or cod
        return None
    return None      # no esta en la matriz: no se bloquea, no se sabe


def _codigo_contifico(producto_rec_id):
    """record_id de Matriz General -> codigo de Contifico. Solo coincidencia EXACTA.

    Antes se aceptaba coincidencia parcial y de ahi salio el desastre: pedir
    PAN DE PAPA devolvia el codigo de PAPA, y CHAMPINONES SALTEADOS devolvia el
    de SAL, porque SAL esta dentro de saLTEADOS. 69 de 208 traslados acabaron
    con el producto equivocado.

    Devuelve (codigo, nombre, motivo_del_fallo). Si no hay coincidencia exacta,
    o si hay mas de una, NO se elige nada: es preferible que el traslado falle y
    alguien lo mire, a que se cree un movimiento sobre otro producto.
    """
    global _catalogo_glog
    if _catalogo_glog is None:
        general = {r['id']: r['fields'] for r in _at_glog(AT_TABLA_GLOG_PRODUCTOS)}
        contifico = [r['fields'] for r in _at_glog(AT_TABLA_GLOG_CONTIFICO)]
        _catalogo_glog = (general, contifico)
        log(f'  catalogo GLOG: {len(general)} productos, {len(contifico)} en matriz Contifico')
    general, contifico = _catalogo_glog

    fila = general.get(producto_rec_id, {})
    nombre = (fila.get('Productos') or '')
    if not nombre:
        return None, '', 'el registro de AirTable no tiene producto'

    # Lo normal: el codigo viene escrito en la Matriz General y no hay nada que
    # deducir. Comparar nombres es lo que mando PAN DE PAPA como PAPA.
    codigo = (fila.get(AT_CAMPO_CODIGO) or '').strip()
    if codigo:
        ni = _sin_inventario_propio(codigo, contifico)
        if ni:
            return None, nombre, (f'"{ni}" ({codigo}) no es inventariable: es un '
                                  f'producto de venta formulado y no tiene '
                                  f'existencias propias que trasladar')
        return codigo, nombre, None

    # Respaldo para un producto recien creado al que aun no le llenaron el
    # campo. Coincidencia EXACTA: la parcial no vuelve.
    buscado = _norma_nombre(nombre)
    exactos = [ct for ct in contifico
               if _norma_nombre(ct.get('Nombre Producto')) == buscado]

    if not exactos:
        return None, nombre, (f'"{nombre}" no tiene codigo. Llenar el campo '
                              f'"{AT_CAMPO_CODIGO}" en la Matriz General de AirTable')
    if len(exactos) > 1:
        codigos = ', '.join(str(ct.get('Código') or '?') for ct in exactos[:5])
        return None, nombre, (f'"{nombre}" aparece {len(exactos)} veces en la Matriz '
                              f'Contifico ({codigos}): no se puede saber cual es')

    codigo = (exactos[0].get('Código') or '').strip()
    if not codigo:
        return None, nombre, f'"{nombre}" esta en la matriz pero sin codigo'
    ni = _sin_inventario_propio(codigo, contifico)
    if ni:
        return None, nombre, (f'"{ni}" ({codigo}) no es inventariable: es un '
                              f'producto de venta formulado y no tiene '
                              f'existencias propias que trasladar')
    return codigo, nombre, None


def traslados_pendientes():
    """Traslados de AirTable sin marcar como Hecho."""
    if not AIRTABLE_TOKEN_GLOG:
        log('AIRTABLE_TOKEN_GLOG no configurado: no se revisan traslados', 'WARN')
        return []
    try:
        regs = _at_glog(AT_TABLA_TRASLADOS)
    except Exception as e:
        log(f'No se pudo leer traslados de AirTable: {str(e)[:150]}', 'ERROR')
        return []
    pend = [r for r in regs if not r['fields'].get('Hecho', False)]
    if pend:
        log(f'{len(pend)} traslado(s) pendientes de {len(regs)}')
    return pend


def _marcar_traslado_hecho(record_id, num_doc):
    """Marca Hecho y guarda el numero. Si el update entero falla, reintenta con
    SOLO 'Hecho': perder el numero molesta, duplicar el traslado es grave."""
    url = (f'https://api.airtable.com/v0/{AT_BASE_GLOG}/{AT_TABLA_TRASLADOS}/{record_id}')
    cab = {'Authorization': f'Bearer {AIRTABLE_TOKEN_GLOG}',
           'Content-Type': 'application/json'}
    r = requests.patch(url, headers=cab,
                       json={'fields': {'Hecho': True, 'num_documento': num_doc}},
                       timeout=30)
    if r.status_code == 200:
        log(f'  [AT] Marcado: Hecho=True, num_documento={num_doc}')
        return True
    log(f'  [AT] update completo fallo ({r.status_code}): {r.text[:150]}', 'WARN')
    r2 = requests.patch(url, headers=cab, json={'fields': {'Hecho': True}}, timeout=30)
    if r2.status_code == 200:
        log(f'  [AT] Marcado solo Hecho (sin numero) para no duplicar', 'WARN')
        return True
    log(f'  [AT] NO se pudo marcar Hecho: {r2.text[:150]}', 'ERROR')
    return False


def _elegir_por_codigo(driver, elemento, codigo):
    """Escribe el codigo y escoge la sugerencia de ESE codigo, no la primera.

    El autocompletado normal acepta la primera opcion a ciegas. Con codigos eso
    no vale: si uno es el principio de otro -CONG001 y CONG0015- la primera que
    sale no tiene por que ser la que se pidio.

    Devuelve True si se pudo elegir la exacta. Si no, deja el trabajo hecho a
    medias a proposito y que lo resuelva la comprobacion posterior.
    """
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
    time.sleep(0.4)
    elemento.click()
    time.sleep(0.3)
    elemento.clear()
    elemento.send_keys(codigo)
    time.sleep(2.5)

    try:
        opciones = [li for li in driver.find_elements(
            By.CSS_SELECTOR, 'ul.ui-autocomplete li') if li.is_displayed()]
        patron = re.compile(r'(^|[^A-Za-z0-9])' + re.escape(codigo) + r'([^A-Za-z0-9]|$)',
                            re.IGNORECASE)
        for li in opciones:
            if patron.search(li.text or ''):
                li.click()
                time.sleep(0.8)
                _cerrar_sugerencias(driver, elemento)
                return True
        if opciones:
            log(f'  ninguna sugerencia coincide con el codigo {codigo}; '
                f'se ofrecian: ' + ' | '.join((li.text or '')[:40] for li in opciones[:4]),
                'WARN')
    except Exception as e:
        log(f'  no se pudo leer la lista de sugerencias: {str(e)[:80]}', 'WARN')

    # Camino de siempre, por si la lista no se pudo leer. Lo que valida de
    # verdad es la comprobacion de despues.
    elemento.send_keys(Keys.DOWN)
    time.sleep(0.5)
    elemento.send_keys(Keys.ENTER)
    time.sleep(0.8)
    _cerrar_sugerencias(driver, elemento)
    return False


def _autocompletar(driver, elemento, texto):
    """Escribe en un autocomplete de Contifico y escoge la primera sugerencia.

    Al final CIERRA la lista de sugerencias. Sin eso, el desplegable de la
    bodega se queda abierto flotando sobre el campo del producto y Selenium
    falla con 'element not interactable' al intentar escribir debajo. Por eso
    unos traslados entraban y otros no: dependia de si la lista se habia
    cerrado sola a tiempo.
    """
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", elemento)
    time.sleep(0.4)
    elemento.click()
    time.sleep(0.4)
    elemento.clear()
    elemento.send_keys(texto)
    time.sleep(2.5)
    elemento.send_keys(Keys.DOWN)
    time.sleep(0.5)
    elemento.send_keys(Keys.ENTER)
    time.sleep(0.8)
    _cerrar_sugerencias(driver, elemento)


def _cerrar_sugerencias(driver, elemento=None):
    """Cierra cualquier lista de autocompletado que haya quedado abierta."""
    if elemento is not None:
        try:
            elemento.send_keys(Keys.ESCAPE)
        except Exception:
            pass
    try:
        driver.execute_script(
            "document.querySelectorAll('ul.ui-autocomplete').forEach("
            "  function(u){ u.style.display='none'; });")
    except Exception:
        pass
    time.sleep(0.4)


def procesar_traslado(reg, driver):
    """Crea un traslado en Contifico por el formulario web.

    Devuelve (driver_sigue_vivo, salio_bien).
    """
    from selenium.webdriver.support.ui import Select

    rid = reg['id']
    f = reg['fields']
    origen_rec = (f.get('Tienda Origen') or [''])[0]
    destino_rec = (f.get('Tienda Destino') or [''])[0]
    prod_rec = (f.get('Productos') or [''])[0]
    cantidad = f.get('Cantidad') or 0

    bod_origen = MAPEO_BODEGAS_GLOG.get(origen_rec, '')
    bod_destino = MAPEO_BODEGAS_GLOG.get(destino_rec, '')
    centro = MAPEO_CENTROS_GLOG.get(origen_rec, bod_origen)

    fecha_raw = f.get('Fecha de Registro') or ''
    try:
        fecha = datetime.strptime(fecha_raw, '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        fecha = datetime.now().strftime('%d/%m/%Y')

    codigo, nombre_prod, motivo = _codigo_contifico(prod_rec)

    log(f'>>> TRASLADO {rid} | {codigo or "?"} x{cantidad} | '
        f'{bod_origen} -> {bod_destino} ({fecha})')

    # Datos malos: no se toca Contifico y el registro queda pendiente para que
    # alguien lo corrija en AirTable.
    if not bod_origen or not bod_destino:
        log(f'  [DATO] bodega no mapeada: origen={origen_rec!r} destino={destino_rec!r}', 'ERROR')
        return True, False
    if bod_origen == bod_destino:
        log('  [DATO] origen y destino son la misma bodega', 'ERROR')
        return True, False
    if not codigo:
        # Se deja el registro pendiente a proposito: sin codigo seguro, crear el
        # movimiento seria repetir el fallo que mando PAN DE PAPA como PAPA.
        log(f'  [DATO] no se puede resolver el producto: {motivo}', 'ERROR')
        return True, False
    if not cantidad:
        log('  [DATO] sin cantidad', 'ERROR')
        return True, False

    try:
        driver.get(CONTIFICO_MOVIMIENTOS_URL)
        wait = WebDriverWait(driver, 40)
        wait.until(EC.presence_of_element_located((By.NAME, 'tipo')))
        time.sleep(2)

        # 1. tipo TRASLADO
        sel = Select(driver.find_element(By.NAME, 'tipo'))
        opcion = next((o.text for o in sel.options if 'TRASLADO' in o.text.upper()), None)
        if not opcion:
            raise Exception('el formulario no ofrece el tipo TRASLADO')
        sel.select_by_visible_text(opcion)
        time.sleep(1.5)

        # 2. fecha
        campo = driver.find_element(By.NAME, 'fecha')
        campo.clear()
        campo.send_keys(fecha)
        campo.send_keys(Keys.ESCAPE)
        time.sleep(0.5)
        driver.find_element(By.TAG_NAME, 'body').click()
        time.sleep(0.5)

        # 3. bodegas
        _autocompletar(driver, wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "input[data_id='id_bodega_origen_id']"))), bod_origen)
        _autocompletar(driver, wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "input[data_id='id_bodega_destino_id']"))), bod_destino)

        # 4. descripcion
        desc = f'TRASLADO ENTRE BODEGAS {bod_origen}/{bod_destino}'
        cd = driver.find_element(By.NAME, 'descripcion')
        cd.clear()
        cd.send_keys(desc)

        # 5. producto y cantidad
        if not _fila_producto(driver, codigo, cantidad):
            raise Exception(f'no se pudo cargar el producto {codigo} en el detalle')

        # 6. guardar
        btn = wait.until(EC.element_to_be_clickable((By.ID, 'btn-guardar')))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(0.5)
        btn.click()
        time.sleep(6)

        # 7. VERIFICAR: sin numero TRA no hay documento, por muy verde que se
        #    vea la pantalla. Este es el control que faltaba en las tomas
        #    fisicas y por el que se reportaron como subidas sin existir.
        cuerpo = driver.find_element(By.TAG_NAME, 'body').text
        m = re.search(r'TRA\s+\d+', cuerpo)
        if not m:
            resumen = ' | '.join(l.strip() for l in cuerpo.split('\n')[:12] if l.strip())
            raise Exception(f'Contifico no devolvio numero TRA. Pantalla: {resumen[:300]}')
        num_doc = m.group(0)
        url_doc = driver.current_url
        estado = 'Generado' if 'Generado' in cuerpo else (
                 'Pendiente' if 'Pendiente' in cuerpo else '?')
        log(f'  [OK] {num_doc} creado (estado en pantalla: {estado})')

        _marcar_traslado_hecho(rid, num_doc)

        try:
            detalle = (f'📦 {nombre_prod or codigo} ({codigo}) x {cantidad}\n'
                       f'📅 {fecha}\n'
                       f'🔄 {bod_origen} ➜ {bod_destino}')
            notificar_traslado_ok(centro, detalle, num_doc)
        except Exception as e:
            log(f'  aviso por Telegram fallo: {str(e)[:100]}', 'WARN')
        return True, True

    except Exception as e:
        log(f'ERROR traslado {rid}: {str(e)[:250]}', 'ERROR')
        log(traceback.format_exc()[:500], 'ERROR')
        msg = str(e).lower()
        vivo = not any(k in msg for k in (
            'invalid session', 'disconnected', 'not connected', 'no such window',
            'connection refused', 'target closed', 'chrome not reachable'))
        return vivo, False


def _fila_producto(driver, codigo, cantidad):
    """Escribe codigo y cantidad en la fila de detalle del movimiento.

    La fila 1 ya viene creada por Contifico; sus campos son
    'id_detalle_1-producto_id' y 'id_detalle_1-cantidad'. Al elegir el producto
    la pagina agrega sola una fila 2 vacia, asi que NO hay que crear filas.

    La cantidad se pone por JS y no con send_keys: buscando "el primer input de
    cantidad visible" se acababa escribiendo en un campo que Selenium rechazaba
    con 'element not interactable' -pasa cuando la lista del autocompletado
    sigue abierta encima-. Yendo al id exacto de la fila y disparando los
    eventos a mano, Contifico recalcula igual y no depende de que el campo este
    despejado en pantalla.
    """
    # 1. Producto: el autocomplete de la fila de detalle, nunca los de cabecera.
    campos = [i for i in driver.find_elements(
        By.CSS_SELECTOR, "input.ui-autocomplete-input[data_id^='id_detalle_']")
        if i.get_attribute('data_id').endswith('-producto_id') and i.is_displayed()]
    if not campos:
        log('  no aparece el campo de producto en el detalle', 'ERROR')
        return False
    inp = campos[0]
    data_id = inp.get_attribute('data_id')           # id_detalle_1-producto_id
    fila = data_id.replace('id_detalle_', '').replace('-producto_id', '')

    # Por si el desplegable de la bodega destino sigue abierto sobre esta fila.
    _cerrar_sugerencias(driver)
    exacto = _elegir_por_codigo(driver, inp, codigo)
    valor = (inp.get_attribute('value') or '').strip()

    if exacto:
        # Se hizo clic en la sugerencia cuyo codigo coincide entero, asi que el
        # producto es el correcto. Solo falta confirmar que quedo SELECCIONADO y
        # no simplemente escrito, y eso lo dice el campo oculto con el id: el
        # texto visible muestra lo que Contifico quiera, normalmente solo el
        # nombre.
        oculto = ''
        for selector in (f"#id_detalle_{fila}-producto_id",
                         f"input[name='detalle_{fila}-producto_id']"):
            try:
                oculto = (driver.find_element(
                    By.CSS_SELECTOR, selector).get_attribute('value') or '').strip()
            except Exception:
                continue
            if oculto:
                break
        if not oculto:
            log(f'  el producto no quedo seleccionado: el campo oculto esta '
                f'vacio (texto {valor[:40]!r})', 'ERROR')
            return False
        log(f'  producto: {codigo} - {valor[:50]}')
    else:
        # No se pudo elegir por codigo y se cayo al camino de siempre, que acepta
        # la primera sugerencia. Ahi no se sabe cual quedo, asi que se exige ver
        # el codigo ENTERO en el texto: buscando "VER02" dentro de "VER020 PAPA"
        # se daria por bueno el producto equivocado.
        entero = re.compile(r'(^|[^A-Za-z0-9])' + re.escape(codigo) + r'([^A-Za-z0-9]|$)',
                            re.IGNORECASE)
        if not entero.search(valor):
            # 'no-results' significa que el buscador de Contifico no ofrece ese
            # codigo. Casi siempre es que el producto esta INACTIVO (estado I):
            # el formulario solo lista los activos. Lo arregla alguien en
            # Contifico o corrigiendo el producto en AirTable, no el worker.
            if 'no-results' in valor.lower():
                log(f'  [DATO] Contifico no encuentra el codigo {codigo}: '
                    f'lo normal es que este INACTIVO. Hay que activarlo o corregir '
                    f'el producto en AirTable.', 'ERROR')
            else:
                log(f'  el producto no quedo seleccionado (campo dice {valor[:60]!r})',
                    'ERROR')
            return False
        log(f'  producto: {valor[:60]}')

    # 2. Cantidad, por id exacto de la misma fila.
    cid = f'id_detalle_{fila}-cantidad'
    try:
        campo = driver.find_element(By.ID, cid)
    except Exception:
        log(f'  no existe el campo {cid}', 'ERROR')
        return False
    driver.execute_script(
        "arguments[0].value = arguments[1];"
        "arguments[0].dispatchEvent(new Event('input',  {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('keyup',  {bubbles:true}));"
        "arguments[0].dispatchEvent(new Event('blur',   {bubbles:true}));",
        campo, str(cantidad))
    time.sleep(1)
    puesto = (campo.get_attribute('value') or '').strip()
    if not puesto or float(puesto or 0) != float(cantidad):
        log(f'  la cantidad no se fijo: quedo {puesto!r} en vez de {cantidad}', 'ERROR')
        return False
    log(f'  cantidad: {puesto}')
    return True


def notificar_traslado_ok(centro, detalle, num_doc):
    """Avisa por Telegram reusando el modulo comun, el mismo que usan los
    demas bots, para que el mensaje salga con el formato de siempre."""
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from notificar_telegram import notificar_exito
    notificar_exito('Traslado', centro, detalle, num_doc)


# Clave arbitraria pero fija para el lock de Postgres. Tiene que ser la misma
# en todas las corridas: es lo que las identifica entre si.
LOCK_TRASLADOS = 918273645


def correr_traslados(driver_actual):
    """Atiende los traslados pendientes. Devuelve (driver, hechos, fallidos).

    Se le pasa el driver que ya tenga el worker para no abrir un segundo Chrome.

    Va con un lock de Postgres porque el cron dispara cada 3 minutos y cinco
    traslados tardan mas que eso: sin el lock, la corrida siguiente empezaria
    con los mismos registros -en AirTable no se marcan hasta despues de
    crearlos- y los duplicaria en Contifico. Es lo que paso con los 23 egresos
    del 15-ago. Si otra corrida lo tiene tomado, esta se salta los traslados y
    ya los cogera en 3 minutos.
    """
    con = lock = None
    try:
        con = psycopg2.connect(**DB_CONFIG)
        con.autocommit = True
        cur = con.cursor()
        cur.execute('SELECT pg_try_advisory_lock(%s)', (LOCK_TRASLADOS,))
        lock = bool(cur.fetchone()[0])
    except Exception as e:
        log(f'No se pudo tomar el lock de traslados: {str(e)[:120]}', 'WARN')
        if con:
            try: con.close()
            except Exception: pass
        return driver_actual, 0, 0

    if not lock:
        log('Otra corrida esta subiendo traslados; esta se los salta.')
        try: con.close()
        except Exception: pass
        return driver_actual, 0, 0

    try:
        return _correr_traslados(driver_actual)
    finally:
        try:
            cur.execute('SELECT pg_advisory_unlock(%s)', (LOCK_TRASLADOS,))
            con.close()
        except Exception:
            pass


def _correr_traslados(driver_actual):
    pend = traslados_pendientes()
    if not pend:
        return driver_actual, 0, 0

    driver = driver_actual
    hechos = fallidos = 0
    for reg in pend:
        if driver is None:
            log('Iniciando Chrome y login en Contifico (traslados)...')
            driver = make_chrome()
            login_contifico(driver)
        elif not driver_sano(driver):
            log('El driver murio, reiniciando...', 'WARN')
            try:
                driver.quit()
            except Exception:
                pass
            driver = make_chrome()
            login_contifico(driver)
        try:
            vivo, ok = procesar_traslado(reg, driver)
        except Exception as e:
            log(f'traslado {reg["id"]} reviento: {e}', 'ERROR')
            vivo, ok = False, False
        hechos += 1 if ok else 0
        fallidos += 0 if ok else 1
        if not vivo:
            try:
                driver.quit()
            except Exception:
                pass
            driver = None
    log(f'Traslados: {hechos} hechos, {fallidos} pendientes para la proxima')
    return driver, hechos, fallidos




# ============================================================
# TOMAS FISICAS PROGRAMADAS
# ============================================================
# En la PC de Finanzas las tomas fisicas se subian a hora fija: Simon Bolon a
# las 12:00 y los demas locales a las 16:00. Al migrar a Render eso se perdio y
# quedaron dependiendo de que alguien pulsara el boton del panel.
#
# Aqui se recupera: en cada pasada el worker mira la hora de Ecuador y, si toca,
# encola las tomas fisicas del dia. Encolar -y no subir directamente- deja el
# trabajo en la misma cola que usa el boton, con su mismo control de estados, y
# se ve igual en el panel.
#
# Se encola una sola vez por bodega y dia: antes de insertar se comprueba que no
# exista ya una tarea de toma fisica para esa bodega y esa fecha, sea del
# horario o pedida a mano. Sin eso, el cron dispara cada 3 minutos y crearia una
# tarea nueva en cada pasada durante toda la hora.

HORARIO_TOMAS = {
    12: ['simon_bolon'],
    16: ['real_audiencia', 'floreana', 'portugal',
         'santo_cachon_real', 'santo_cachon_portugal'],
}
TZ_ECUADOR = timezone(timedelta(hours=-5))

# Cuantos dias hacia atras se rescatan. Una semana cubre un puente o un fin de
# semana largo sin arrastrar historia vieja.
DIAS_ATRASO = int(os.environ.get('DIAS_ATRASO_TOMAS', '7'))
# Cuantas veces se reintenta un dia que falla antes de darlo por imposible.
MAX_INTENTOS_TOMA = int(os.environ.get('MAX_INTENTOS_TOMA', '3'))


def hora_ecuador():
    """Hora local de Ecuador. Render corre en UTC, asi que no vale datetime.now()
    a secas: daria las 21:00 cuando en el local son las 16:00."""
    return datetime.now(TZ_ECUADOR)


def programar_tomas_fisicas():
    """Encola las tomas fisicas que falten, de la mas vieja a la mas nueva.

    Tres cosas aprendidas sobre la marcha:

    1. La toma fisica que se sube hoy es la del DIA ANTERIOR. El local cuenta a
       lo largo del dia y al dia siguiente se carga a Contifico. La descarga de
       saldos si es del dia en curso, pero la carga no.

    2. NO encolar si no hay conteos. El 19-ago las cinco tareas de las 16:00
       fallaron con "No hay conteos en BD": el horario disparo puntual pero los
       locales aun no habian contado. Cinco errores falsos y cinco sustos por
       Telegram. Ahora, a partir de su hora, se revisa en cada pasada y se
       encola en cuanto aparecen los conteos, asi el que cuenta tarde tambien
       entra.

    3. RECUPERAR lo atrasado. Antes solo se miraba ayer: el dia que se perdia
       quedaba perdido para siempre. Paso con el 19-ago en cinco bodegas y hubo
       que subirlo a mano. Ahora se mira una semana hacia atras.
    """
    if os.environ.get('SIN_HORARIO_TOMAS', '0') == '1':
        return 0

    ahora = hora_ecuador()

    # Bodegas cuya hora ya paso hoy. A partir de ahi se sigue mirando hasta
    # que aparezcan los conteos o hasta que acabe el dia.
    pendientes = [b for hora, bods in HORARIO_TOMAS.items()
                  if ahora.hour >= hora for b in bods]
    if not pendientes:
        return 0

    # De ayer hacia atras. Una semana alcanza para cubrir un puente o un fin de
    # semana largo sin arrastrar meses de historia vieja.
    fechas = [(ahora.date() - timedelta(days=d)).isoformat()
              for d in range(DIAS_ATRASO, 0, -1)]

    creadas = 0
    con = None
    try:
        con = psycopg2.connect(**DB_CONFIG)
        cur = con.cursor()
        for bodega in pendientes:
            # Si la bodega ya tiene algo en cola o corriendo, no se le añade
            # nada mas: asi se ponen al dia en orden y de una en una.
            # Ojo con las zombis: hay tareas en 'en_proceso' desde julio,
            # de cuando la PC de Finanzas se apagaba a mitad de trabajo. Una
            # tarea que lleva horas asi no esta corriendo, esta muerta, y no
            # puede bloquear a su bodega para siempre.
            cur.execute("""
                SELECT COUNT(*) FROM goti.tareas_inventario_locales
                WHERE bodega = %s AND accion = 'toma_fisica'
                  AND (estado = 'pendiente'
                       OR (estado = 'en_proceso'
                           AND timestamp_inicio > NOW() - INTERVAL '2 hours'))
            """, (bodega,))
            if cur.fetchone()[0]:
                continue

            for fecha in fechas:
                cur.execute("""
                    SELECT estado, COALESCE(solicitado_por, '')
                    FROM goti.tareas_inventario_locales
                    WHERE bodega = %s AND fecha = %s AND accion = 'toma_fisica'
                    ORDER BY id
                """, (bodega, fecha))
                intentos = cur.fetchall()

                # Ya subida, o cancelada a proposito: no se toca. Cancelar
                # desde el panel tiene que significar "no lo hagas", si no el
                # worker la volveria a encolar a los tres minutos.
                if any(e in ('completado', 'cancelado') for e, _ in intentos):
                    continue

                # Un error se reintenta, pero no eternamente: si Contifico
                # rechaza ese dia una y otra vez, se deja constancia y se pasa
                # al siguiente en vez de atascar la cola.
                errores = sum(1 for e, _ in intentos if e == 'error')
                if errores >= MAX_INTENTOS_TOMA:
                    continue

                # Hay algo que subir?
                cur.execute("""
                    SELECT COUNT(*) FROM goti.inventario_ciego_conteos
                    WHERE fecha = %s AND local = %s
                      AND COALESCE(cantidad_contada_2, cantidad_contada) > 0
                """, (fecha, bodega))
                n = cur.fetchone()[0]
                if not n:
                    continue      # aun no han contado; se vuelve a mirar luego

                cur.execute("""
                    INSERT INTO goti.tareas_inventario_locales
                        (bodega, fecha, accion, solicitado_por)
                    VALUES (%s, %s, 'toma_fisica', 'horario')
                    RETURNING id
                """, (bodega, fecha))
                atraso = (ahora.date() - date.fromisoformat(fecha)).days
                log(f'  {bodega}: encolada toma fisica #{cur.fetchone()[0]} '
                    f'del {fecha} ({n} productos contados'
                    + (f', atrasada {atraso} dias' if atraso > 1 else '') + ')')
                creadas += 1
                break         # una por bodega y pasada: la mas vieja primero
        con.commit()
    except Exception as e:
        if con:
            con.rollback()
        log(f'No se pudieron programar las tomas fisicas: {str(e)[:150]}', 'ERROR')
        return 0
    finally:
        if con:
            try:
                con.close()
            except Exception:
                pass

    if creadas:
        log(f'Horario {ahora:%H:%M} Ecuador: {creadas} toma(s) fisica(s) encoladas')
    return creadas
BODEGAS_CONTEO_DIARIO = ('bodega_principal', 'materia_prima', 'planta')
HORA_CONTEO_DIARIO = int(os.environ.get('HORA_CONTEO_DIARIO', '7'))


def programar_conteo_operativo():
    """Crea la tarea de conteo diario de las bodegas operativas.

    Antes habia que pulsar un boton cada dia y, si nadie se acordaba, esa
    bodega se quedaba sin contar: Planta estuvo parada del 7 al 19 de agosto
    por eso.

    Se llama al endpoint del panel en vez de insertar a mano, porque es el que
    sabe elegir los productos de la semana -los 14 fijos de Bodega Principal o
    la rotacion de 10 de las otras- y el que evita repetir la seleccion.
    Domingo no: nadie cuenta.
    """
    if os.environ.get('SIN_CONTEO_DIARIO', '0') == '1':
        return 0
    ahora = hora_ecuador()
    if ahora.hour != HORA_CONTEO_DIARIO or ahora.weekday() == 6:
        return 0

    fecha = ahora.date().isoformat()
    creadas = 0
    for bodega in BODEGAS_CONTEO_DIARIO:
        try:
            r = requests.post(f'{BACKEND_URL}/api/inventario/generar-conteo-operativo',
                              json={'bodega': bodega, 'fecha': fecha}, timeout=40)
            if r.status_code == 409:
                log(f'  {bodega}: ya estaba generado para {fecha}')
                continue
            if r.status_code != 200:
                log(f'  {bodega}: el panel respondio {r.status_code}: {r.text[:120]}', 'WARN')
                continue
            d = r.json()
            log(f'  {bodega}: conteo #{d.get("id")} con {d.get("productos")} productos '
                f'(fijos={d.get("fijos")} aleatorios={d.get("aleatorios")})')
            creadas += 1
        except Exception as e:
            log(f'  {bodega}: no se pudo generar el conteo: {str(e)[:120]}', 'ERROR')
    if creadas:
        log(f'Conteo diario {ahora:%H:%M} Ecuador: {creadas} bodega(s) encoladas')
    return creadas


# ============ MAIN LOOP ============
def verificar_entorno():
    """Prueba de humo: Chrome + login + descarga del Excel de saldos + parseo.

    Hace falta porque con la cola vacia el navegador nunca se abre y la corrida
    sale verde sin haber probado nada. Ejercita todo el camino delicado
    -Chromium, login, reporte de Contifico, lectura del Excel- y NO escribe en
    la base ni toma tareas, asi que no compite con el worker de la PC.

    VERIFICAR_BODEGA y VERIFICAR_FECHA permiten elegir que descargar.
    """
    bodega = os.environ.get('VERIFICAR_BODEGA', 'bodega_principal')
    fecha = os.environ.get('VERIFICAR_FECHA') or datetime.now().strftime('%Y-%m-%d')
    cfg = BODEGAS.get(bodega) or BODEGAS_LOCALES.get(bodega)
    if not cfg:
        log(f'bodega desconocida: {bodega}', 'ERROR')
        return 1

    log('=' * 62)
    log('VERIFICACION DE ENTORNO (no escribe nada)')
    log(f'   CHROME_BIN   = {os.environ.get("CHROME_BIN")}')
    log(f'   CHROMEDRIVER = {os.environ.get("CHROMEDRIVER")}')
    log(f'   bodega={bodega} ({cfg["contifico"]})  fecha={fecha}')
    log('=' * 62)

    driver = None
    try:
        driver = make_chrome()
        log(f'[OK] Chromium arranco: {driver.capabilities.get("browserVersion")}')
        login_contifico(driver)
        log('[OK] Login en Contifico')

        # VERIFICAR_TOMA=1: ademas abre el formulario de toma fisica y comprueba
        # que sus campos sigan existiendo. NO guarda ni genera nada: es el unico
        # trozo del camino que no se puede probar de otra forma sin crear un
        # documento real de ajuste de inventario en Contifico.
        if os.environ.get('VERIFICAR_TOMA', '0') == '1':
            log('')
            log('--- formulario de TOMA FISICA (sin guardar) ---')
            driver.get('https://1793168604001.contifico.com/sistema/inventario/tomafisica/registrar/')
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, 'id_fecha')))
            time.sleep(2)
            campos = {
                'id_fecha': (By.ID, 'id_fecha'),
                'bodega (autocomplete)': (By.CSS_SELECTOR,
                    'input.object-description[data_id="id_bodega"]'),
                'btn-guardar': (By.ID, 'btn-guardar'),
            }
            for etq, (como, sel) in campos.items():
                driver.find_element(como, sel)
                log(f'[OK] campo presente: {etq}')
            # La fila de detalle se crea por JS; comprobar que la funcion existe.
            hay_js = driver.execute_script(
                "return typeof movimiento !== 'undefined' "
                "&& typeof movimiento.agregarDetalle === 'function';")
            log(f'[{"OK" if hay_js else "FALLO"}] movimiento.agregarDetalle() disponible')
            if not hay_js:
                raise RuntimeError('el formulario no expone movimiento.agregarDetalle()')
            log('[OK] formulario de toma fisica utilizable (no se guardo nada)')


        # exportarExcel() tarda y dispara una navegacion. El corte de 90s venia
        # del page_load_timeout, no del de script: hay que subir LOS DOS. Pasa
        # cuando el worker de la PC esta pidiendo el mismo reporte a la vez.
        archivo = descargar_saldos(driver, cfg['contifico'], fecha)
        log(f'[OK] Excel descargado: {os.path.basename(archivo)}')

        datos = parsear_saldos(archivo, cfg['contifico'])
        con_costo = sum(1 for v in datos.values() if float(v.get('costo') or 0) > 0)
        log(f'[OK] Excel parseado: {len(datos)} productos, {con_costo} con costo')
        if datos:
            cod, v = next(iter(datos.items()))
            log(f'     ejemplo: {cod} stock={v.get("stock")} costo={v.get("costo")}')
        if not datos:
            log('El Excel vino vacio', 'ERROR')
            return 1
        log('')
        log('ENTORNO CORRECTO')
        return 0
    except Exception as e:
        log(f'[FALLO] {type(e).__name__}: {str(e)[:300]}', 'ERROR')
        log(traceback.format_exc()[:800], 'ERROR')
        return 1
    finally:
        if driver is not None:
            try: driver.quit()
            except Exception: pass
            log('Navegador cerrado.')


def main():
    """Una sola pasada: atiende las tareas pendientes y termina.

    El worker original vive en la PC de Finanzas con un bucle infinito que
    consulta cada 15s. Como cron job de Render tiene que terminar, asi que se
    vacia la cola de una vez y se sale.

    Atiende las dos colas que dependian de esa PC:
      - conteo operativo   (/api/conteo-op/*)
      - inventario locales (/api/inventario-locales/*): actualizar cantidad y toma fisica

    Las tareas de CRUCE y de CARGA A CONTIFICO se dejan como estaban: las sigue
    tomando la PC si esta encendida. El backend reparte con FOR UPDATE SKIP
    LOCKED, asi que los dos pueden convivir sin pisarse ni duplicar trabajo.
    """
    inicio = datetime.now()
    log('=' * 62)
    log('WORKER OPERATIVO EN RENDER (una pasada)')
    log(f'backend={BACKEND_URL} | worker_id={WORKER_ID} | headless={USE_HEADLESS}')
    log('=' * 62)

    faltan = [n for n, v in (('DB_HOST', DB_CONFIG['host']),
                             ('DB_USER', DB_CONFIG['user']),
                             ('DB_PASSWORD', DB_CONFIG['password']),
                             ('CONTIFICO_WEB_USUARIO', CONTIFICO['usuario']),
                             ('CONTIFICO_WEB_PASSWORD', CONTIFICO['password'])) if not v]
    if faltan:
        log(f'Faltan variables de entorno: {", ".join(faltan)}', 'ERROR')
        return 1

    if os.environ.get('VERIFICAR', '0') == '1':
        return verificar_entorno()

    # Las CUATRO colas que atendia la PC de Finanzas. Todas las funciones
    # devuelven (driver_sigue_vivo, salio_bien), asi que se tratan igual.
    COLAS = (
        ('cruce',      get_pendientes,                     procesar_tarea),
        ('carga',      get_pendientes_carga,               procesar_tarea_carga),
        ('conteo',     get_pendientes_conteo_op,           procesar_tarea_conteo_op),
        ('inventario', get_pendientes_inventario_locales,  procesar_tarea_inventario_locales),
    )
    solo = os.environ.get('SOLO_COLA', '').strip().lower()
    activas = [c for c in COLAS if not solo or c[0] == solo]
    if solo and not activas:
        log(f'SOLO_COLA={solo!r} no coincide con ninguna cola', 'ERROR')
        return 1
    log(f'colas atendidas: {", ".join(c[0] for c in activas)}')

    try:
        programar_tomas_fisicas()
    except Exception as e:
        log(f'Fallo al programar tomas fisicas: {str(e)[:150]}', 'ERROR')

    try:
        programar_conteo_operativo()
    except Exception as e:
        log(f'Fallo al programar el conteo diario: {str(e)[:150]}', 'ERROR')

    driver = None
    hechas = fallidas = 0
    try:
        while True:
            trabajo = [(nombre, t, proc) for nombre, traer, proc in activas
                       for t in (traer() or [])]
            if not trabajo:
                break
            log(f'{len(trabajo)} tarea(s) tomadas: '
                + ', '.join(f'{n}#{t.get("id")}' for n, t, _ in trabajo))

            for nombre, tarea, procesar in trabajo:
                if driver is None:
                    log('Iniciando Chrome y login en Contifico...')
                    driver = make_chrome()
                    login_contifico(driver)
                elif not driver_sano(driver):
                    log('El driver murio, reiniciando...', 'WARN')
                    try: driver.quit()
                    except Exception: pass
                    driver = make_chrome()
                    login_contifico(driver)

                try:
                    driver_ok, ok = procesar(tarea, driver)
                except Exception as e:
                    log(f'{nombre}#{tarea.get("id")} reviento: {e}', 'ERROR')
                    log(traceback.format_exc()[:600], 'ERROR')
                    driver_ok, ok = False, False

                hechas += 1 if ok else 0
                fallidas += 0 if ok else 1
                if not driver_ok:
                    try: driver.quit()
                    except Exception: pass
                    driver = None
    except Exception as e:
        log(f'EXCEPCION en el bucle: {e}', 'ERROR')
        log(traceback.format_exc()[:800], 'ERROR')
        return 1
    finally:
        if driver is not None:
            try: driver.quit()
            except Exception: pass
            log('Navegador cerrado.')

    # Traslados de AirTable. Van al final y con su propio try para que un
    # fallo aqui no tire las cuatro colas del panel, que son lo urgente.
    # Se reusa el driver que quedo abierto en vez de arrancar otro Chrome.
    if os.environ.get('SALTAR_TRASLADOS', '0') != '1' and not solo:
        try:
            # El bloque de arriba ya cerro el navegador en su finally, asi que
            # se empieza con None: correr_traslados solo abre Chrome si de
            # verdad hay algo pendiente que subir.
            driver, t_ok, t_mal = correr_traslados(None)
            hechas += t_ok
            fallidas += t_mal
        except Exception as e:
            log(f'Traslados fallaron enteros: {str(e)[:200]}', 'ERROR')
            log(traceback.format_exc()[:500], 'ERROR')
        finally:
            if driver is not None:
                try: driver.quit()
                except Exception: pass
                driver = None

    fin = datetime.now()
    log('=' * 62)
    log(f'PROCESO COMPLETADO en {(fin - inicio).seconds}s')
    log(f'   tareas completadas: {hechas}')
    log(f'   tareas con error:   {fallidas}')
    log('=' * 62)

    # Una tarea que falla ya queda marcada como 'error' en el backend y no se
    # reintenta sola, asi que hacer fallar el cron por eso solo lo dejaria en
    # rojo cada corrida. Solo se falla si no se pudo ni arrancar.
    return 0


if __name__ == '__main__':
    sys.exit(main())
